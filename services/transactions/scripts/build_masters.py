"""
Reproducible builder for the two canonical HOTELVALORA institutional masters.

Run from repo root:
    python services/transactions/scripts/build_masters.py

What it produces:
    services/transactions/MASTER/HOTEL_TRANSACCIONES_MASTER.xlsx
    services/transactions/MASTER/HOTEL_PROYECTOS_MASTER.xlsx

Each workbook contains five sheets:
    1. DATA              — the canonical row table (transactions or projects)
    2. DICTIONARY        — column name + type + required + notes
    3. INGESTION_LOG     — one row per processed file (history of ingestion runs)
    4. SOURCES_REGISTRY  — labelled source kinds with reliability tiers
    5. README            — operator instructions

The schema lives here so it is regenerable. The .xlsx is the *frozen artefact*
that ships in git; this script regenerates it identically when run again. If a
schema column changes, update this script + re-run + commit both this script
and the regenerated workbook.

NORMALIZATION_VERSION bumps on every breaking schema change.
"""

from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

NORMALIZATION_VERSION = "v1.0"
BUILT_AT = datetime.now(timezone.utc).isoformat(timespec="seconds")

HERE = Path(__file__).resolve().parent
ROOT = HERE.parent  # services/transactions
MASTER_DIR = ROOT / "MASTER"


# ---------------------------------------------------------------------------
# Shared ingestion-meta block — IDENTICAL across both masters
# ---------------------------------------------------------------------------
INGESTION_META_COLUMNS: list[tuple[str, str, bool, str]] = [
    # (column_name, type, required, notes)
    ("canonical_id", "uuid", True, "Primary key in the master. Generated on first insertion. Stable across edits."),
    ("ingestion_id", "text", True, "Matches the ingestion log entry that landed this row."),
    ("source_file", "text", True, "Original filename verbatim — e.g. 'CoStar_2026Q1_Spain.xlsx'."),
    ("source_kind", "enum", True, "costar | brokerage | curated | news_extract | manual | press_release"),
    ("source_url", "url", False, "Public URL when the row was derived from an article or press release."),
    ("ingested_at", "timestamp", True, "ISO-8601 UTC. Auto-stamped by the ingestion run."),
    ("ingested_by", "email", True, "Operator email — audit trail for who triggered the import."),
    ("normalization_version", "text", True, f"Schema/rules version applied at ingestion time. Today: {NORMALIZATION_VERSION}."),
    ("dedup_key", "text", True, "sha256 of the canonicalised dedup fields (see transaction-schema.md / project-schema.md)."),
    ("review_required", "bool", True, "TRUE when validation flagged something. Pair with review_reason."),
    ("review_reason", "text", False, "Short label — e.g. 'missing_country', 'duplicate_candidate', 'ambiguous_price'."),
    ("ingestion_status", "enum", True, "ingested | under_review | superseded | rejected"),
    ("supersedes_id", "uuid", False, "Set when this row corrects/replaces a previously canonical row."),
    ("notes", "text", False, "Free-form operator notes."),
]


# ---------------------------------------------------------------------------
# Transactions-specific domain columns
# ---------------------------------------------------------------------------
TRANSACTION_COLUMNS: list[tuple[str, str, bool, str]] = [
    ("transaction_uid", "uuid", True, "Stable identifier of the deal (single source-of-truth for the transaction)."),
    ("category", "enum", True, "acquisition | sale | joint_venture | refinancing | distress"),
    ("asset_type", "enum", True, "single_property | portfolio | mixed_use"),
    ("asset_name", "text", True, "Property name — e.g. 'The Ritz-Carlton Madrid'."),
    ("portfolio_name", "text", False, "Set when asset_type='portfolio'."),
    ("portfolio_size_assets", "int", False, "Number of properties in the portfolio."),
    ("city", "text", True, "City of the asset (portfolio HQ if portfolio)."),
    ("country", "text", True, "ISO-3166-1 alpha-2."),
    ("market", "text", False, "Coarse market label — e.g. 'Costa del Sol', 'Baleares'."),
    ("submarket", "text", False, "Finer submarket — e.g. 'Marbella Golden Mile'."),
    ("address", "text", False, "Street address when known."),
    ("latitude", "numeric", False, "Decimal degrees, WGS84."),
    ("longitude", "numeric", False, "Decimal degrees, WGS84."),
    ("rooms", "int", False, "Total rooms — sum for portfolios."),
    ("hotel_segment", "enum", False, "luxury | upper_upscale | upscale | upper_midscale | midscale | economy | lifestyle | resort | boutique | mixed_use | serviced_apartments | unknown"),
    ("star_rating", "int", False, "1..5."),
    ("year_built", "int", False, "Original construction year."),
    ("year_renovated", "int", False, "Most recent full refurb."),
    ("gross_area_sqm", "numeric", False, "Built area in square metres."),
    ("price_eur", "numeric", True, "Total deal value in EUR — converted at announce-date rate when original was non-EUR."),
    ("price_currency_original", "text", False, "Source-reported currency (EUR/USD/GBP/CHF)."),
    ("price_per_key_eur", "numeric", False, "Derived when price_eur and rooms are both known."),
    ("cap_rate", "numeric", False, "Stabilised cap rate at closing (percentage points — 5.4 means 5.4%)."),
    ("gop_per_key_eur", "numeric", False, "Disclosed GOP per key — annual."),
    ("revpar_at_closing_eur", "numeric", False, "Trailing-12-month RevPAR at closing."),
    ("closed_at", "date", False, "Closing date when public."),
    ("announced_at", "date", False, "Announcement date — fallback when closed_at unknown."),
    ("buyer_name", "text", True, "Buyer entity name."),
    ("buyer_uid", "uuid", False, "Resolved canonical investor id — populated by the entity resolver."),
    ("buyer_country", "text", False, "HQ country (ISO-3166-1 alpha-2)."),
    ("buyer_kind", "enum", False, "pe | reit | sovereign | family_office | private_owner | bank | operator_owned | hospitality_fund | asset_manager | developer | unknown"),
    ("seller_name", "text", False, "Seller entity name. Optional in some markets."),
    ("seller_uid", "uuid", False, "Resolved canonical investor id."),
    ("seller_country", "text", False, "HQ country (ISO-3166-1 alpha-2)."),
    ("seller_kind", "enum", False, "Same vocabulary as buyer_kind."),
    ("broker", "text", False, "Lead broker / advisor."),
    ("operator_at_closing", "text", False, "Operator running the hotel at closing."),
    ("operator_post_closing", "text", False, "Operator post-closing if known."),
    ("brand_at_closing", "text", False, "Brand pre-closing."),
    ("brand_post_closing", "text", False, "Brand post-closing — set when reflagged."),
    ("financing_type", "enum", False, "equity | senior_loan | mezzanine | cmbs | mixed | unknown"),
    ("disclosed_terms", "enum", False, "full | partial | nondisclosed"),
    ("press_release_url", "url", False, "Buyer or seller official release."),
    ("news_url", "url", False, "Primary news article — preserve provenance."),
    ("linked_news_id", "uuid", False, "FK to public.market_news after the Intelligence Engine matches it."),
]


# ---------------------------------------------------------------------------
# Projects-specific domain columns
# ---------------------------------------------------------------------------
PROJECT_COLUMNS: list[tuple[str, str, bool, str]] = [
    ("project_uid", "uuid", True, "Stable identifier of the project."),
    ("category", "enum", True, "development | branded_residences | flex_living | pipeline_announcement | rebranding | operator_change"),
    ("asset_type", "enum", True, "new_build | conversion | refurb | extension | mixed_use"),
    ("project_name", "text", True, "Project name — e.g. 'The Madrid EDITION'."),
    ("city", "text", True, "City of the project."),
    ("country", "text", True, "ISO-3166-1 alpha-2."),
    ("market", "text", False, "Coarse market — e.g. 'Costa del Sol'."),
    ("submarket", "text", False, "Finer submarket label."),
    ("address", "text", False, "Street address when known."),
    ("latitude", "numeric", False, "Decimal degrees, WGS84."),
    ("longitude", "numeric", False, "Decimal degrees, WGS84."),
    ("rooms", "int", False, "Planned room count. 0 valid for non-room components."),
    ("gross_area_sqm", "numeric", False, "Built area in square metres."),
    ("capex_eur", "numeric", False, "Total CAPEX in EUR."),
    ("capex_per_key_eur", "numeric", False, "Derived when capex_eur and rooms both known."),
    ("estimated_opening", "date", False, "Operator-stated or analyst-derived opening date."),
    ("groundbreaking", "date", False, "Construction-start date."),
    ("announced_at", "date", True, "First public announcement date."),
    ("project_stage", "enum", True, "announced | permitting | under_construction | pre_opening | opened | cancelled | on_hold"),
    ("permitting_status", "text", False, "Notes on permits — license issued, planning approval, etc."),
    ("developer_name", "text", True, "Developer / promotor entity."),
    ("developer_uid", "uuid", False, "Resolved canonical investor id."),
    ("developer_country", "text", False, "ISO-3166-1 alpha-2."),
    ("developer_kind", "enum", False, "Same vocabulary as transactions.buyer_kind."),
    ("operator_name", "text", False, "Operator / brand manager."),
    ("operator_uid", "uuid", False, "Resolved canonical operator id."),
    ("operator_kind", "enum", False, "chain | independent | soft_brand | franchise | management_company | operator_owner | unknown"),
    ("brand", "text", False, "Brand name when known."),
    ("hotel_segment", "enum", False, "Same vocabulary as transactions.hotel_segment."),
    ("star_rating", "int", False, "1..5 — projected."),
    ("mixed_use_flag", "bool", False, "TRUE when project includes residential / retail / office components."),
    ("mixed_use_components", "text", False, "Comma-separated components — 'residential, retail, office'."),
    ("public_subsidy_flag", "bool", False, "TRUE when EU / state / municipal funds known to participate."),
    ("press_release_url", "url", False, "Developer / operator official release."),
    ("news_url", "url", False, "Primary news article URL."),
    ("linked_news_id", "uuid", False, "FK to public.market_news when matched."),
]


SOURCES_REGISTRY = [
    # (source_kind, label, reliability_tier, notes)
    ("costar", "CoStar Export", "A", "Authoritative — institutional, paid product. Always preferred when available."),
    ("brokerage", "Brokerage advisor list", "B", "CBRE, JLL, Cushman, Christie, Colliers — name in source_url meta."),
    ("curated", "Curated internal compilation", "A", "Hand-maintained HOTELVALORA spreadsheets. Treat as ground-truth."),
    ("press_release", "Official press release", "B", "Buyer/seller/operator-issued. Verify cross-reference where possible."),
    ("news_extract", "News extraction (Intelligence Engine)", "C", "Automatic extraction from market_news. Validate before merging into MASTER."),
    ("manual", "Manual operator entry", "C", "Operator-typed row. Audit trail in ingested_by."),
]


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill(start_color="FF003B2A", end_color="FF003B2A", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFD7F587", name="Calibri", size=11)
SUBHEADER_FILL = PatternFill(start_color="FFE8F4EC", end_color="FFE8F4EC", fill_type="solid")
SUBHEADER_FONT = Font(bold=True, color="FF003B2A", name="Calibri", size=10)
WRAP = Alignment(wrap_text=True, vertical="top")


def style_header(ws, header_row=1):
    for cell in ws[header_row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.freeze_panes = f"A{header_row + 1}"


def autosize(ws, max_width=44):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), max_width)


def build_data_sheet(ws, title: str, domain_columns, data_rows: list[dict] | None = None):
    ws.title = title
    columns = domain_columns + INGESTION_META_COLUMNS
    headers = [name for (name, _t, _r, _n) in columns]
    ws.append(headers)
    style_header(ws)
    if data_rows:
        for row in data_rows:
            ws.append([row.get(h) for h in headers])
    autosize(ws)


def build_dictionary_sheet(wb, title: str, domain_columns):
    ws = wb.create_sheet(title)
    ws.append(["section", "column", "type", "required", "notes"])
    style_header(ws)
    for (name, t, req, notes) in domain_columns:
        ws.append(["domain", name, t, "YES" if req else "no", notes])
    for (name, t, req, notes) in INGESTION_META_COLUMNS:
        ws.append(["ingestion_meta", name, t, "YES" if req else "no", notes])
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = WRAP
    autosize(ws, max_width=80)


def build_ingestion_log_sheet(wb):
    ws = wb.create_sheet("INGESTION_LOG")
    ws.append([
        "ingestion_id",
        "source_file",
        "source_kind",
        "started_at",
        "completed_at",
        "rows_seen",
        "rows_inserted",
        "rows_updated",
        "rows_skipped",
        "rows_flagged_review",
        "rows_failed",
        "operator_email",
        "normalization_version",
        "outcome",
        "notes",
    ])
    style_header(ws)
    autosize(ws)


def build_sources_registry_sheet(wb):
    ws = wb.create_sheet("SOURCES_REGISTRY")
    ws.append(["source_kind", "label", "reliability_tier", "notes"])
    style_header(ws)
    for row in SOURCES_REGISTRY:
        ws.append(row)
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = WRAP
    autosize(ws, max_width=80)


def build_readme_sheet(wb, dataset_label: str, domain_label: str, schema_doc: str):
    ws = wb.create_sheet("README")
    ws.append([f"HOTELVALORA · {dataset_label}"])
    ws.cell(row=1, column=1).font = Font(bold=True, size=16, color="FF003B2A")

    body = [
        "",
        f"Canonical institutional dataset for {domain_label}. Append-only — never overwrite an existing canonical row.",
        f"Built {BUILT_AT}. Normalization version: {NORMALIZATION_VERSION}.",
        "",
        "Files this workbook participates in:",
        "  • services/transactions/MASTER/HOTEL_TRANSACCIONES_MASTER.xlsx",
        "  • services/transactions/MASTER/HOTEL_PROYECTOS_MASTER.xlsx",
        "",
        "Operating contract:",
        "  1. The Data Ingestion Agent appends new rows. It NEVER edits in place.",
        "  2. Edits to an existing canonical row → insert a new row with",
        "     supersedes_id = <old canonical_id> and ingestion_status='superseded'.",
        "  3. Manual operator changes follow the same contract — track who, when, why",
        "     in ingested_by + notes.",
        "",
        "Sheets:",
        "  • DATA              — canonical row corpus (this sheet's named tab)",
        "  • DICTIONARY        — column schema + required flags + notes",
        "  • INGESTION_LOG     — one row per processed import file",
        "  • SOURCES_REGISTRY  — labels + reliability tiers per source_kind",
        "  • README            — this sheet",
        "",
        "Reference docs:",
        f"  • {schema_doc}",
        "  • docs/intelligence/master-dataset-architecture.md",
        "  • docs/intelligence/data-normalization-rules.md",
        "  • docs/intelligence/transaction-ingestion-workflow.md",
        "",
        "Regenerating this workbook:",
        "  python services/transactions/scripts/build_masters.py",
        "  (Idempotent. Run when the schema evolves; commit both the script and the regenerated .xlsx.)",
    ]
    for line in body:
        ws.append([line])
    ws.column_dimensions["A"].width = 110


def build_workbook(
    dataset_label: str,
    domain_columns,
    data_sheet_title: str,
    schema_doc: str,
    output_path: Path,
    data_rows: list[dict] | None = None,
):
    wb = Workbook()
    build_data_sheet(wb.active, data_sheet_title, domain_columns, data_rows)
    build_dictionary_sheet(wb, "DICTIONARY", domain_columns)
    build_ingestion_log_sheet(wb)
    build_sources_registry_sheet(wb)
    build_readme_sheet(wb, dataset_label, data_sheet_title.lower(), schema_doc)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    rows_written = len(data_rows) if data_rows else 0
    print(f"[build_masters] wrote {output_path.relative_to(ROOT.parent.parent)} · {rows_written} rows")


# ── Data extraction from snapshot.json ───────────────────────────────────
import json as _json
COSTAR_SNAPSHOT_PATH = ROOT.parent.parent / "services" / "costar" / "MASTER" / "snapshot.json"


def _load_snapshot() -> dict:
    if not COSTAR_SNAPSHOT_PATH.exists():
        return {}
    return _json.loads(COSTAR_SNAPSHOT_PATH.read_text(encoding="utf-8"))


def _meta_block(snapshot: dict, canonical_id: str, source_kind: str = "private") -> dict:
    return {
        "canonical_id": canonical_id,
        "ingestion_id": snapshot.get("ingestion_batch_id", ""),
        "source_file": "snapshot.json",
        "source_kind": source_kind,
        "source_url": "",
        "ingested_at": snapshot.get("generated_at", BUILT_AT),
        "ingested_by": "cli:build_masters",
        "normalization_version": (snapshot.get("batch") or {}).get("normalization_version", NORMALIZATION_VERSION),
        "dedup_key": "",
        "review_required": "FALSE",
        "review_reason": "",
        "ingestion_status": "ingested",
        "supersedes_id": "",
        "notes": "",
    }


def _tx_to_row(t: dict, hotels_by_id: dict, snapshot: dict) -> dict:
    """Map snapshot transactions[] row → HOTEL_TRANSACCIONES_MASTER row.
    Joins with hotels for geo/property context when hotel_id is linked."""
    h = hotels_by_id.get(t.get("hotel_id"), {}) if t.get("hotel_id") else {}
    return {
        "transaction_uid": t.get("transaction_id"),
        "category": t.get("category") or h.get("category"),
        "asset_type": "hotel",
        "asset_name": t.get("asset_name"),
        "portfolio_name": None,
        "portfolio_size_assets": t.get("portfolio_size_assets"),
        "city": t.get("city") or h.get("city") or t.get("location_label"),
        "country": t.get("country"),
        "market": t.get("market_name"),
        "submarket": t.get("submarket_name") or h.get("submarket_name"),
        "address": h.get("address_line"),
        "latitude": h.get("latitude"),
        "longitude": h.get("longitude"),
        "rooms": t.get("rooms_count") or h.get("rooms_count"),
        "hotel_segment": h.get("segment_type"),
        "star_rating": h.get("category"),
        "year_built": h.get("year_opened"),
        "year_renovated": h.get("year_last_renovated"),
        "gross_area_sqm": t.get("gross_area_sqm") or h.get("gross_building_sqm"),
        "price_eur": t.get("price_eur"),
        "price_currency_original": "EUR",
        "price_per_key_eur": t.get("price_per_key_eur"),
        "cap_rate": None,
        "gop_per_key_eur": None,
        "revpar_at_closing_eur": None,
        "closed_at": t.get("closed_at"),
        "announced_at": None,
        "buyer_name": t.get("buyer"),
        "buyer_uid": None,
        "buyer_country": None,
        "buyer_kind": None,
        "seller_name": t.get("seller"),
        "seller_uid": None,
        "seller_country": None,
        "seller_kind": None,
        "broker": t.get("broker"),
        "operator_at_closing": h.get("operator"),
        "operator_post_closing": None,
        "brand_at_closing": h.get("brand"),
        "brand_post_closing": None,
        "financing_type": None,
        "disclosed_terms": None,
        "press_release_url": None,
        "news_url": None,
        "linked_news_id": None,
        **_meta_block(snapshot, t.get("transaction_id", ""), source_kind=t.get("source", "private")),
        "notes": " · ".join(filter(None, [
            t.get("comments"),
            f"P.G={t.get('p_g_flag')}" if t.get('p_g_flag') else None,
            f"P.U={t.get('p_u_flag')}" if t.get('p_u_flag') else None,
            f"P.C={t.get('p_c_flag')}" if t.get('p_c_flag') else None,
            f"price/m²={t.get('price_per_sqm_eur')}" if t.get('price_per_sqm_eur') else None,
            "DUPLICATE_OF=" + str(t.get('duplicate_of')) if t.get('is_duplicate') and t.get('duplicate_of') else None,
        ])) or None,
    }


def _proj_to_row(p: dict, snapshot: dict) -> dict:
    """Map snapshot projects[] row → HOTEL_PROYECTOS_MASTER row."""
    return {
        "project_uid": p.get("project_id"),
        "category": "hotel_pipeline",
        "asset_type": "hotel",
        "project_name": p.get("project_name"),
        "city": p.get("city"),
        "country": p.get("country"),
        "market": p.get("market_name") or p.get("city"),
        "submarket": p.get("submarket_name"),
        "address": p.get("street"),
        "latitude": None,
        "longitude": None,
        "rooms": p.get("rooms_count"),
        "gross_area_sqm": None,
        "capex_eur": None,
        "capex_per_key_eur": None,
        "estimated_opening": p.get("opening_date"),
        "groundbreaking": None,
        "announced_at": None,
        "project_stage": p.get("phase"),
        "permitting_status": p.get("status"),
        "developer_name": p.get("office_company"),
        "developer_uid": None,
        "developer_country": p.get("office_country"),
        "developer_kind": p.get("office_role"),
        "operator_name": None,
        "operator_uid": None,
        "operator_kind": None,
        "brand": None,
        "hotel_segment": p.get("construction_type"),
        "star_rating": p.get("stars"),
        "mixed_use_flag": None,
        "mixed_use_components": None,
        "public_subsidy_flag": None,
        "press_release_url": None,
        "news_url": None,
        "linked_news_id": None,
        **_meta_block(snapshot, p.get("project_id", ""), source_kind="costar"),
        "notes": " · ".join(filter(None, [
            f"TBI={p.get('tbi')}" if p.get('tbi') else None,
            f"views={p.get('views')}" if p.get('views') else None,
            f"office_contact={p.get('office_contact_last_name')}" if p.get('office_contact_last_name') else None,
            f"office_postal={p.get('office_postal_code')}" if p.get('office_postal_code') else None,
        ])) or None,
    }


def main():
    snapshot = _load_snapshot()
    if snapshot:
        print(f"[build_masters] snapshot loaded · transactions={len(snapshot.get('transactions') or [])} · "
              f"projects={len(snapshot.get('projects') or [])}")
        hotels_by_id = {h["hotel_id"]: h for h in (snapshot.get("hotels") or [])}
        tx_rows = [_tx_to_row(t, hotels_by_id, snapshot) for t in (snapshot.get("transactions") or [])]
        proj_rows = [_proj_to_row(p, snapshot) for p in (snapshot.get("projects") or [])]
    else:
        print(f"[build_masters] no snapshot at {COSTAR_SNAPSHOT_PATH} · emitting headers-only templates")
        tx_rows = []
        proj_rows = []

    build_workbook(
        dataset_label="HOTEL_TRANSACCIONES_MASTER",
        domain_columns=TRANSACTION_COLUMNS,
        data_sheet_title="TRANSACTIONS",
        schema_doc="docs/intelligence/transaction-schema.md",
        output_path=MASTER_DIR / "HOTEL_TRANSACCIONES_MASTER.xlsx",
        data_rows=tx_rows,
    )
    build_workbook(
        dataset_label="HOTEL_PROYECTOS_MASTER",
        domain_columns=PROJECT_COLUMNS,
        data_sheet_title="PROJECTS",
        schema_doc="docs/intelligence/project-schema.md",
        output_path=MASTER_DIR / "HOTEL_PROYECTOS_MASTER.xlsx",
        data_rows=proj_rows,
    )


if __name__ == "__main__":
    main()
