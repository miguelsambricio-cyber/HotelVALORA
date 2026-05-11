"""Read + append the canonical MASTER workbooks safely.

Batch-in-memory pattern: open MASTER → accumulate changes in lists → save once
at the end. If the orchestrator crashes mid-run, the MASTER is unchanged on
disk (no half-written rows).
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook

# Column lists are imported from build_masters.py to keep a single source of truth.
# We mirror them here defensively so a future schema bump in build_masters is reflected.

from build_masters import (
    INGESTION_META_COLUMNS,
    TRANSACTION_COLUMNS,
    PROJECT_COLUMNS,
    NORMALIZATION_VERSION,
)


def _header_names(columns: list[tuple]) -> list[str]:
    return [c[0] for c in columns]


TRANSACTION_HEADERS = _header_names(TRANSACTION_COLUMNS) + _header_names(INGESTION_META_COLUMNS)
PROJECT_HEADERS = _header_names(PROJECT_COLUMNS) + _header_names(INGESTION_META_COLUMNS)
TRANSACTION_DOMAIN = _header_names(TRANSACTION_COLUMNS)
PROJECT_DOMAIN = _header_names(PROJECT_COLUMNS)
META_HEADERS = _header_names(INGESTION_META_COLUMNS)

INGESTION_LOG_HEADERS = [
    "ingestion_id", "source_file", "source_kind",
    "started_at", "completed_at",
    "rows_seen", "rows_inserted", "rows_updated", "rows_skipped",
    "rows_flagged_review", "rows_failed",
    "operator_email", "normalization_version", "outcome", "notes",
]


def open_master(path: Path) -> Workbook:
    if not path.exists():
        raise FileNotFoundError(
            f"MASTER not found: {path}. "
            "Regenerate via `python services/transactions/scripts/build_masters.py`."
        )
    return load_workbook(path)


def _data_sheet_name(target: str) -> str:
    return "TRANSACTIONS" if target == "transactions" else "PROJECTS"


def _domain_columns(target: str) -> list[str]:
    return TRANSACTION_DOMAIN if target == "transactions" else PROJECT_DOMAIN


def _headers(target: str) -> list[str]:
    return TRANSACTION_HEADERS if target == "transactions" else PROJECT_HEADERS


def load_existing_dedup_keys(wb: Workbook, target: str) -> dict[str, dict[str, Any]]:
    """Return {dedup_key: row_dict} for every existing canonical row (status=ingested).

    Empty corpus → empty dict. dedup_key is the master's institutional dedup field.
    """
    ws = wb[_data_sheet_name(target)]
    headers = [c.value for c in ws[1]]
    if not headers or headers[0] is None:
        return {}
    out: dict[str, dict[str, Any]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None or row[0] is None:
            continue
        rec = dict(zip(headers, row, strict=False))
        if rec.get("ingestion_status") != "ingested":
            continue
        dk = rec.get("dedup_key")
        if dk:
            out[dk] = rec
    return out


def append_master_rows(wb: Workbook, target: str, rows: list[dict[str, Any]]) -> int:
    """Append rows to the DATA sheet. Returns the count appended."""
    if not rows:
        return 0
    ws = wb[_data_sheet_name(target)]
    headers = [c.value for c in ws[1]]
    for row in rows:
        ws.append([row.get(h) for h in headers])
    return len(rows)


def mark_superseded(wb: Workbook, target: str, canonical_ids: list[str]) -> int:
    """Flip ingestion_status='superseded' on the listed canonical_ids.

    In-place update — the ONE non-append operation the master allows, recorded
    by the caller in the run's INGESTION_LOG.
    """
    if not canonical_ids:
        return 0
    ws = wb[_data_sheet_name(target)]
    headers = [c.value for c in ws[1]]
    cid_col = headers.index("canonical_id") + 1
    status_col = headers.index("ingestion_status") + 1
    targets = set(canonical_ids)
    updated = 0
    for row_idx in range(2, ws.max_row + 1):
        cid = ws.cell(row=row_idx, column=cid_col).value
        if cid in targets:
            ws.cell(row=row_idx, column=status_col).value = "superseded"
            updated += 1
    return updated


def append_ingestion_log(wb: Workbook, entry: dict[str, Any]) -> None:
    """One row per processed file."""
    ws = wb["INGESTION_LOG"]
    ws.append([entry.get(h) for h in INGESTION_LOG_HEADERS])


def save_master(wb: Workbook, path: Path) -> None:
    """Atomic-ish save: write to a sibling .tmp then rename."""
    tmp = path.with_suffix(path.suffix + ".tmp")
    wb.save(tmp)
    tmp.replace(path)  # rename is atomic on POSIX; best-effort on Windows


def get_headers_for(target: str) -> list[str]:
    return _headers(target)


def get_domain_columns(target: str) -> list[str]:
    return _domain_columns(target)


def get_normalization_version() -> str:
    return NORMALIZATION_VERSION
