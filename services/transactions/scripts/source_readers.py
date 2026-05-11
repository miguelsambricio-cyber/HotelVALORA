"""Lenient readers for XLSX + CSV operator drops.

Each reader emits a stream of `(row_index, raw_dict_with_canonical_keys)` tuples.
Canonical key folding lives in `normalization.fold_header`.
"""

from __future__ import annotations

import csv
from pathlib import Path
from typing import Iterator

from openpyxl import load_workbook

from normalization import fold_header


def _strip(v):
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        return v if v else None
    return v


def _fold_headers(target: str, raw_headers: list[str]) -> tuple[list[str | None], list[str]]:
    """Return (canonical_per_index, unknown_headers_list)."""
    canonical: list[str | None] = []
    unknown: list[str] = []
    for h in raw_headers:
        if h is None or str(h).strip() == "":
            canonical.append(None)
            continue
        c = fold_header(target, h)
        canonical.append(c)
        if c is None:
            unknown.append(str(h))
    return canonical, unknown


def read_xlsx(path: Path, target: str) -> tuple[Iterator[tuple[int, dict]], list[str]]:
    """Returns (row_iter, unknown_headers).

    Uses the FIRST sheet of the workbook by default. If a sheet named "DATA",
    "TRANSACTIONS" or "PROJECTS" exists, use that instead.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    preferred = next(
        (n for n in wb.sheetnames if n.upper() in ("DATA", "TRANSACTIONS", "PROJECTS")),
        wb.sheetnames[0],
    )
    ws = wb[preferred]
    rows = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration:
        return iter(()), []
    canonical, unknown = _fold_headers(target, list(header_row))

    def _iter():
        for idx, row in enumerate(rows, start=2):  # start=2 because header is row 1
            if row is None or all(v is None for v in row):
                continue
            rec: dict = {}
            for canon_key, value in zip(canonical, row, strict=False):
                if canon_key is None:
                    continue
                rec[canon_key] = _strip(value)
            if rec:
                yield idx, rec

    return _iter(), unknown


def read_csv(path: Path, target: str) -> tuple[Iterator[tuple[int, dict]], list[str]]:
    """Try utf-8-sig + utf-8 + latin-1 in order — operator files come from everywhere."""
    encodings = ("utf-8-sig", "utf-8", "latin-1")
    last_err = None
    for enc in encodings:
        try:
            with path.open("r", encoding=enc, newline="") as f:
                sniff = f.read(2048)
                f.seek(0)
                try:
                    dialect = csv.Sniffer().sniff(sniff, delimiters=",;\t")
                except csv.Error:
                    dialect = csv.excel
                reader = csv.reader(f, dialect)
                try:
                    header_row = next(reader)
                except StopIteration:
                    return iter(()), []
                canonical, unknown = _fold_headers(target, header_row)
                all_rows = list(reader)
            break
        except UnicodeDecodeError as e:
            last_err = e
            continue
    else:
        raise RuntimeError(f"Could not decode {path}: {last_err}")

    def _iter():
        for idx, row in enumerate(all_rows, start=2):
            if not row or all((v is None or str(v).strip() == "") for v in row):
                continue
            rec: dict = {}
            for canon_key, value in zip(canonical, row, strict=False):
                if canon_key is None:
                    continue
                rec[canon_key] = _strip(value)
            if rec:
                yield idx, rec

    return _iter(), unknown


def read_file(path: Path, target: str) -> tuple[Iterator[tuple[int, dict]], list[str]]:
    suffix = path.suffix.lower()
    if suffix in (".xlsx", ".xlsm"):
        return read_xlsx(path, target)
    if suffix in (".csv", ".tsv", ".txt"):
        return read_csv(path, target)
    raise ValueError(f"Unsupported file extension: {suffix}")
