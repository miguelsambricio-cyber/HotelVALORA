"""Reproducible builder for the two HOTELVALORA compset OPERATIONAL masters.

Run from repo root:
    python services/compset/scripts/build_masters.py

Produces:
    services/compset/MASTER/COMPSET_MASTER.xlsx              (subject + compset KPIs + indices)
    services/compset/MASTER/HOTEL_POSITIONING_MASTER.xlsx    (per-hotel underwriting positioning snapshots)

Why this workspace exists separately from services/costar/
==========================================================
CompSet analysis is OPERATIONAL, not warehouse:
  - hotel-specific (one row per target hotel × period)
  - on-demand (triggered by a deal, a valuation refresh, a quarterly review)
  - underwriting-critical (drives assumptions that flow into valuations + reports)
  - dynamic (the compset composition itself changes when CoStar reclassifies)

CoStar market warehouse (services/costar/) is administrative, macro, and
slow-changing. Mixing the two would force the same agent to span
fundamentally different operational rhythms. See
docs/architecture/market-vs-underwriting-separation.md for the rationale.
"""

from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

NORMALIZATION_VERSION = "v1.0"
BUILT_AT = datetime.now(timezone.utc).isoformat(timespec="seconds")

HERE = Path(__file__).resolve().parent
ROOT = HERE.parent  # services/compset
MASTER_DIR = ROOT / "MASTER"


# ---------------------------------------------------------------------------
# Shared ingestion-meta block — IDENTICAL to transactions/ and costar/
# ---------------------------------------------------------------------------
INGESTION_META_COLUMNS: list[tuple[str, str, bool, str]] = [
    ("canonical_id", "uuid", True, "Primary key in the master. Generated on first insertion. Stable across edits."),
    ("ingestion_id", "text", True, "Matches the ingestion log entry that landed this row."),
    ("source_file", "text", True, "Original filename verbatim — e.g. 'Ritz_Madrid_Compset_2026Q1.xlsx'."),
    ("source_kind", "enum", True, "costar | str | operator_brief | curated | manual"),
    ("source_url", "url", False, "Public URL when applicable."),
    ("ingested_at", "timestamp", True, "ISO-8601 UTC. Auto-stamped."),
    ("ingested_by", "email", True, "Operator email — audit trail."),
    ("normalization_version", "text", True, f"Schema/rules version applied. Today: {NORMALIZATION_VERSION}."),
    ("dedup_key", "text", True, "sha256 of canonicalised dedup fields (see compset-schema.md / hotel-positioning-schema.md)."),
    ("review_required", "bool", True, "TRUE when validation flagged something."),
    ("review_reason", "text", False, "Short label — 'compset_composition_change', 'fx_conversion_applied', 'restatement_candidate'."),
    ("ingestion_status", "enum", True, "ingested | under_review | superseded | rejected"),
    ("supersedes_id", "uuid", False, "Set when this row corrects/replaces a previously canonical row."),
    ("notes", "text", False, "Free-form operator notes."),
]


# ---------------------------------------------------------------------------
# COMPSET_MASTER — subject + compset KPIs + MPI/ARI/RGI time series
# (Ported from the v1.0 costar COMPSETS schema — same shape, different home.)
# ---------------------------------------------------------------------------
COMPSET_COLUMNS: list[tuple[str, str, bool, str]] = [
    ("compset_name", "text", True, "Operator-readable label — 'Madrid 5* Luxury Center'."),
    ("compset_uid", "uuid", False, "Resolved canonical compset id."),
    ("costar_compset_code", "text", False, "CoStar code for the comp set when issued."),
    ("target_hotel_name", "text", True, "The hotel being benchmarked."),
    ("target_hotel_uid", "uuid", False, "Resolved canonical hotel id — FK to public.valuations or future entity table."),
    ("compset_hotel_names", "text", False, "Comma-separated list of peers (display only — provenance)."),
    ("compset_size", "int", True, "Number of hotels in the compset. Range [3, 20]."),
    ("country", "text", True, "ISO-3166-1 alpha-2."),
    ("market_name", "text", False, "Parent market."),
    ("submarket_name", "text", False, "Parent submarket when applicable."),
    ("chain_scale", "enum", False, "Subject's chain scale at reporting time."),
    ("period_kind", "enum", True, "daily | weekly | monthly | quarterly | ytd | ltm | annual"),
    ("period_start", "date", True, "ISO-8601 start."),
    ("period_end", "date", True, "ISO-8601 end (inclusive)."),
    ("currency", "text", True, "ISO-4217."),
    # Subject hotel KPIs
    ("subject_occupancy_pct", "numeric", False, "Subject occupancy %."),
    ("subject_adr", "numeric", False, "Subject ADR in row's currency."),
    ("subject_revpar", "numeric", False, "Subject RevPAR in row's currency."),
    ("subject_rooms_available", "numeric", False, "Subject room-nights available."),
    ("subject_rooms_sold", "numeric", False, "Subject room-nights sold."),
    ("subject_revenue", "numeric", False, "Subject accommodation revenue."),
    # Compset (peer set) KPIs
    ("compset_occupancy_pct", "numeric", False, "Compset average occupancy %."),
    ("compset_adr", "numeric", False, "Compset average ADR."),
    ("compset_revpar", "numeric", False, "Compset average RevPAR."),
    ("compset_rooms_available", "numeric", False, "Compset total room-nights available."),
    ("compset_rooms_sold", "numeric", False, "Compset total room-nights sold."),
    # Performance indices — CoStar signature outputs
    ("mpi", "numeric", False, "Market Penetration Index = (subject_occ / compset_occ) × 100. 100 = on par."),
    ("ari", "numeric", False, "Average Rate Index = (subject_adr / compset_adr) × 100."),
    ("rgi", "numeric", False, "RevPAR Generation Index = (subject_revpar / compset_revpar) × 100."),
    ("mpi_yoy_pp", "numeric", False, "YoY change in MPI (pp of index — e.g. 102.4 → 105.1 = +2.7 pp)."),
    ("ari_yoy_pp", "numeric", False, "YoY change in ARI."),
    ("rgi_yoy_pp", "numeric", False, "YoY change in RGI."),
    ("fair_share_pct", "numeric", False, "Subject's fair share of compset demand by room count."),
    ("revpar_premium_eur", "numeric", False, "Subject RevPAR − compset RevPAR in row's currency."),
]


# ---------------------------------------------------------------------------
# HOTEL_POSITIONING_MASTER — per-hotel underwriting positioning snapshots
# (NEW in v1.0 — captures the CompSet Underwriting Agent's derived outputs.)
# ---------------------------------------------------------------------------
POSITIONING_COLUMNS: list[tuple[str, str, bool, str]] = [
    # Identification + trigger
    ("hotel_uid", "uuid", False, "Resolved canonical hotel id — FK to public.valuations or future entity table."),
    ("hotel_name", "text", True, "Target hotel display name."),
    ("snapshot_kind", "enum", True, "underwriting_baseline | quarterly_refresh | valuation_update | manual"),
    ("snapshot_at", "timestamp", True, "ISO-8601 UTC. When this positioning snapshot was generated."),
    ("trigger", "text", False, "What triggered the snapshot — 'operator_request' / 'scheduled_cron' / 'deal_pipeline_event' / etc."),

    # Geographic + segment context
    ("country", "text", True, "ISO-3166-1 alpha-2."),
    ("market_name", "text", False, "Parent market."),
    ("submarket_name", "text", False, "Parent submarket."),
    ("chain_scale", "enum", False, "Subject chain scale at snapshot time."),
    ("hotel_segment", "enum", False, "Operator-tagged segment (luxury / boutique / resort / ...)."),

    # Source data window — which CoStar period(s) the snapshot reads from
    ("as_of_period_start", "date", True, "Start of the source data window — typically the most recent closed quarter."),
    ("as_of_period_end", "date", True, "End of the source data window."),
    ("as_of_period_kind", "enum", True, "Period kind of the source data — monthly / quarterly / ytd / ltm / annual."),

    # Compset reference
    ("compset_uid", "uuid", False, "FK to COMPSET_MASTER.compset_uid when wired."),
    ("compset_name", "text", False, "Operator-readable compset label."),
    ("compset_size", "int", False, "Number of hotels in the compset used."),

    # Currency
    ("currency", "text", True, "ISO-4217 — all monetary values in row are denominated here."),

    # Subject vs compset benchmark
    ("subject_revpar", "numeric", False, "Subject RevPAR over as_of_period."),
    ("compset_revpar", "numeric", False, "Compset RevPAR over as_of_period."),
    ("revpar_premium_eur", "numeric", False, "Subject − compset RevPAR."),

    ("subject_adr", "numeric", False, "Subject ADR over as_of_period."),
    ("compset_adr", "numeric", False, "Compset ADR over as_of_period."),
    ("adr_premium_eur", "numeric", False, "Subject − compset ADR."),

    ("subject_occupancy_pct", "numeric", False, "Subject occupancy."),
    ("compset_occupancy_pct", "numeric", False, "Compset occupancy."),
    ("occupancy_premium_pp", "numeric", False, "Subject − compset occupancy (pp)."),

    # Performance indices
    ("mpi", "numeric", False, "Market Penetration Index over as_of_period."),
    ("ari", "numeric", False, "Average Rate Index over as_of_period."),
    ("rgi", "numeric", False, "RevPAR Generation Index over as_of_period."),

    # Underwriting forward assumptions (derived by the agent)
    ("adr_assumption_eur", "numeric", False, "Recommended underwriting ADR going forward."),
    ("occupancy_assumption_pct", "numeric", False, "Recommended underwriting occupancy going forward."),
    ("revpar_assumption_eur", "numeric", False, "Derived RevPAR assumption = adr × occupancy/100."),
    ("room_revenue_assumption_eur", "numeric", False, "Annualised room revenue assumption."),
    ("gop_per_key_assumption_eur", "numeric", False, "GOP-per-key assumption when derivable from operator + brand benchmarks."),

    # Valuation anchor
    ("valuation_anchor_eur_per_key", "numeric", False, "Suggested per-key valuation anchor."),
    ("cap_rate_assumption_pct", "numeric", False, "Cap rate used in valuation derivation."),
    ("multiple_of_revenue_assumption", "numeric", False, "Alternative valuation multiple."),

    # Confidence + reasoning
    ("confidence", "enum", False, "low | medium | high — agent's confidence in its assumptions."),
    ("assumptions_basis", "text", False, "Operator-readable narrative of what data backed the assumption."),
    ("risks", "text", False, "Operator-readable risk flags raised by the agent."),

    # Linkage
    ("valuation_outcome_uid", "uuid", False, "FK to public.valuations.id when the snapshot was consumed by a valuation."),
]


SOURCES_REGISTRY = [
    ("costar", "CoStar Hospitality Export", "A", "Authoritative — typical source for benchmarking inputs."),
    ("str", "STR (CoStar subsidiary)", "A", "STR-direct exports."),
    ("operator_brief", "Operator brief / management input", "B", "Brand + operator-supplied compset assertions."),
    ("curated", "Curated internal compilation", "A", "Hand-maintained HOTELVALORA spreadsheets."),
    ("manual", "Manual operator entry", "C", "Operator-typed row."),
]


# ---------------------------------------------------------------------------
# Sheet builders (identical pattern as transactions/ + costar/)
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill(start_color="FF003B2A", end_color="FF003B2A", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFD7F587", name="Calibri", size=11)
WRAP = Alignment(wrap_text=True, vertical="top")


def style_header(ws, header_row=1):
    for cell in ws[header_row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.freeze_panes = f"A{header_row + 1}"


def autosize(ws, max_width=44):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), max_width)


def build_data_sheet(ws, title: str, domain_columns, data_rows: list[dict] | None = None):
    ws.title = title
    columns = domain_columns + INGESTION_META_COLUMNS
    headers = [name for (name, _t, _r, _n) in columns]
    ws.append(headers)
    style_header(ws)
    if data_rows:
        for row in data_rows:
            ws.append([row.get(h) for h in headers])
    autosize(ws)


def build_dictionary_sheet(wb, title: str, domain_columns):
    ws = wb.create_sheet(title)
    ws.append(["section", "column", "type", "required", "notes"])
    style_header(ws)
    for (name, t, req, notes) in domain_columns:
        ws.append(["domain", name, t, "YES" if req else "no", notes])
    for (name, t, req, notes) in INGESTION_META_COLUMNS:
        ws.append(["ingestion_meta", name, t, "YES" if req else "no", notes])
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = WRAP
    autosize(ws, max_width=80)


def build_ingestion_log_sheet(wb):
    ws = wb.create_sheet("INGESTION_LOG")
    ws.append([
        "ingestion_id", "source_file", "source_kind",
        "started_at", "completed_at",
        "rows_seen", "rows_inserted", "rows_updated", "rows_skipped",
        "rows_flagged_review", "rows_failed",
        "operator_email", "normalization_version", "outcome", "notes",
    ])
    style_header(ws)
    autosize(ws)


def build_sources_registry_sheet(wb):
    ws = wb.create_sheet("SOURCES_REGISTRY")
    ws.append(["source_kind", "label", "reliability_tier", "notes"])
    style_header(ws)
    for row in SOURCES_REGISTRY:
        ws.append(row)
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = WRAP
    autosize(ws, max_width=80)


def build_readme_sheet(wb, dataset_label: str, domain_label: str, schema_doc: str):
    ws = wb.create_sheet("README")
    ws.append([f"HOTELVALORA · {dataset_label}"])
    ws.cell(row=1, column=1).font = Font(bold=True, size=16, color="FF003B2A")

    body = [
        "",
        f"Operational dataset for {domain_label}. Append-only — never overwrite an existing canonical row.",
        f"Built {BUILT_AT}. Normalization version: {NORMALIZATION_VERSION}.",
        "",
        "Sister workbook in services/compset/MASTER/:",
        "  • COMPSET_MASTER.xlsx              — subject + compset KPIs + MPI/ARI/RGI time series",
        "  • HOTEL_POSITIONING_MASTER.xlsx    — per-hotel underwriting positioning snapshots",
        "",
        "Owned by the CompSet Underwriting Agent. Hotel-specific, operational, on-demand —",
        "different rhythm from services/costar/ (which is the slower-changing market warehouse).",
        "",
        "Operating contract:",
        "  1. The CompSet Underwriting Agent appends new rows. It NEVER edits in place.",
        "  2. Corrections (e.g. CoStar restatements) → insert a new row with",
        "     supersedes_id = <old canonical_id> and ingestion_status='superseded'.",
        "  3. Manual edits follow the same contract — track who, when, why",
        "     in ingested_by + notes.",
        "",
        "Sheets:",
        "  • DATA              — canonical row corpus",
        "  • DICTIONARY        — column schema + required flags + notes",
        "  • INGESTION_LOG     — one row per processed import file",
        "  • SOURCES_REGISTRY  — labels + reliability tiers per source_kind",
        "  • README            — this sheet",
        "",
        "Reference docs:",
        f"  • {schema_doc}",
        "  • docs/agents/compset-underwriting-agent.md",
        "  • docs/architecture/market-vs-underwriting-separation.md",
        "",
        "Regenerating this workbook:",
        "  python services/compset/scripts/build_masters.py",
        "  (Idempotent. Run when the schema evolves; commit both the script and the regenerated .xlsx.)",
    ]
    for line in body:
        ws.append([line])
    ws.column_dimensions["A"].width = 110


def build_workbook(
    dataset_label: str,
    domain_columns,
    data_sheet_title: str,
    schema_doc: str,
    output_path: Path,
    data_rows: list[dict] | None = None,
):
    wb = Workbook()
    build_data_sheet(wb.active, data_sheet_title, domain_columns, data_rows)
    build_dictionary_sheet(wb, "DICTIONARY", domain_columns)
    build_ingestion_log_sheet(wb)
    build_sources_registry_sheet(wb)
    build_readme_sheet(wb, dataset_label, data_sheet_title.lower(), schema_doc)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    rows_written = len(data_rows) if data_rows else 0
    print(f"[build_masters] wrote {output_path.relative_to(ROOT.parent.parent)} · {rows_written} rows")


# ── Data extraction from snapshot.json ───────────────────────────────────
import json as _json
COSTAR_SNAPSHOT_PATH = ROOT.parent.parent / "services" / "costar" / "MASTER" / "snapshot.json"


def _load_snapshot() -> dict:
    if not COSTAR_SNAPSHOT_PATH.exists():
        return {}
    return _json.loads(COSTAR_SNAPSHOT_PATH.read_text(encoding="utf-8"))


def _meta_block(snapshot: dict, canonical_id: str, source_kind: str) -> dict:
    return {
        "canonical_id": canonical_id,
        "ingestion_id": snapshot.get("ingestion_batch_id", ""),
        "source_file": "snapshot.json",
        "source_kind": source_kind,
        "source_url": "",
        "ingested_at": snapshot.get("generated_at", BUILT_AT),
        "ingested_by": "cli:build_masters",
        "normalization_version": (snapshot.get("batch") or {}).get("normalization_version", NORMALIZATION_VERSION),
        "dedup_key": "",
        "review_required": "FALSE",
        "review_reason": "",
        "ingestion_status": "ingested",
        "supersedes_id": "",
        "notes": "",
    }


def _compset_to_row(c: dict, hotels_by_id: dict, snapshot: dict) -> dict:
    """Map snapshot synthetic_compsets[] row → COMPSET_MASTER row.
    Targets hotel inventory · uses member chain_scale + market for
    geographic context. Subject/compset performance KPIs (occupancy,
    ADR, RevPAR) come from operator-confirmed Smith Travel reports ·
    nullable until those reports are ingested."""
    target = hotels_by_id.get(c.get("target_hotel_id"), {})
    members = c.get("members") or []
    member_names = " · ".join(m.get("name", "") for m in members)
    # Pick chain_scale from the most-common member chain_scale
    scales = [m.get("chain_scale") for m in members if m.get("chain_scale")]
    chain_scale = max(set(scales), key=scales.count) if scales else target.get("chain_scale")
    return {
        "compset_name": f"{target.get('name') or c.get('target_name')} · compset",
        "compset_uid": c.get("compset_id"),
        "costar_compset_code": "",
        "target_hotel_name": target.get("name") or c.get("target_name"),
        "target_hotel_uid": c.get("target_hotel_id"),
        "compset_hotel_names": member_names,
        "compset_size": len(members),
        "country": target.get("country") or "ES",
        "market_name": c.get("market_name") or target.get("market_name"),
        "submarket_name": c.get("submarket_name") or target.get("submarket_name"),
        "chain_scale": chain_scale,
        "period_kind": "ltm",
        "period_start": "",
        "period_end": "",
        "currency": "EUR",
        "subject_occupancy_pct": None,
        "subject_adr": None,
        "subject_revpar": None,
        "subject_rooms_available": None,
        "subject_rooms_sold": None,
        "subject_revenue": None,
        "compset_occupancy_pct": None,
        "compset_adr": None,
        "compset_revpar": None,
        "compset_rooms_available": None,
        "compset_rooms_sold": None,
        "mpi": None,
        "ari": None,
        "rgi": None,
        "mpi_yoy_pp": None,
        "ari_yoy_pp": None,
        "rgi_yoy_pp": None,
        "fair_share_pct": None,
        "revpar_premium_eur": None,
        **_meta_block(snapshot, c.get("compset_id", ""), source_kind=c.get("provenance", "synthetic")),
        "review_required": "TRUE" if c.get("needs_operator_confirmation") else "FALSE",
        "review_reason": "synthetic_compset_pending_operator_confirmation" if c.get("needs_operator_confirmation") else "",
        "notes": f"algorithm={(c.get('algorithm') or {}).get('version', 'v1')} · members_with_similarity_scores",
    }


def _positioning_to_row(h: dict, compset_by_target: dict, market_snap: dict, snapshot: dict) -> dict:
    """Map snapshot hotels[] row → HOTEL_POSITIONING_MASTER row.
    One row per hotel · joins market KPIs (occupancy/ADR/RevPAR for the
    hotel's market) and the hotel's synthetic compset. Subject KPIs
    (the hotel's own occupancy/ADR/RevPAR) come from operator reports ·
    nullable until ingested · the structure of the master is there for
    institutional underwriting workflow."""
    cs = compset_by_target.get(h.get("hotel_id"))
    # Market lookup · CoStar uses "Madrid" on hotels but "Madrid ESP"
    # on KPI rows · try suffixed variant + suffix-strip
    hmkt = h.get("market_name") or ""
    market_kpis = (
        market_snap.get(hmkt)
        or market_snap.get(f"{hmkt} ESP")
        or market_snap.get(f"{hmkt} {(h.get('country') or 'ESP').upper()}")
        or {}
    )
    return {
        "hotel_uid": h.get("hotel_id"),
        "hotel_name": h.get("name"),
        "snapshot_kind": "ltm_market_average",
        "snapshot_at": snapshot.get("generated_at"),
        "trigger": "build_masters_run",
        "country": h.get("country"),
        "market_name": h.get("market_name"),
        "submarket_name": h.get("submarket_name"),
        "chain_scale": h.get("chain_scale"),
        "hotel_segment": h.get("segment_type"),
        "as_of_period_start": "",
        "as_of_period_end": "",
        "as_of_period_kind": "ltm",
        "compset_uid": cs.get("compset_id") if cs else None,
        "compset_name": f"{h.get('name')} · compset" if cs else None,
        "compset_size": len(cs.get("members") or []) if cs else None,
        "currency": "EUR",
        "subject_revpar": None,  # operator-reported
        "compset_revpar": (market_kpis.get("revpar_12m") or 0) if market_kpis.get("revpar_12m") else None,
        "revpar_premium_eur": None,
        "subject_adr": None,
        "compset_adr": (market_kpis.get("adr_12m") or 0) if market_kpis.get("adr_12m") else None,
        "adr_premium_eur": None,
        "subject_occupancy_pct": None,
        "compset_occupancy_pct": (market_kpis.get("occupancy_12m") or 0) * 100 if market_kpis.get("occupancy_12m") is not None else None,
        "occupancy_premium_pp": None,
        "mpi": None,
        "ari": None,
        "rgi": None,
        # Underwriting assumption fields · nullable until operator runs
        # the valuation pass
        "adr_assumption_eur": (market_kpis.get("adr_12m") or 0) if market_kpis.get("adr_12m") else None,
        "occupancy_assumption_pct": (market_kpis.get("occupancy_12m") or 0) * 100 if market_kpis.get("occupancy_12m") is not None else None,
        "revpar_assumption_eur": (market_kpis.get("revpar_12m") or 0) if market_kpis.get("revpar_12m") else None,
        "room_revenue_assumption_eur": None,
        "gop_per_key_assumption_eur": None,
        "valuation_anchor_eur_per_key": None,
        "cap_rate_assumption_pct": None,
        "multiple_of_revenue_assumption": None,
        "confidence": "low" if (h.get("_meta") or {}).get("confidence", 1) < 0.8 else "medium",
        "assumptions_basis": "market_average_via_costar_data_table",
        "risks": None,
        "valuation_outcome_uid": None,
        **_meta_block(snapshot, h.get("hotel_id", ""), source_kind="derived"),
        "notes": f"compset={'synthetic' if cs else 'none'}",
    }


def main():
    snapshot = _load_snapshot()
    if snapshot:
        print(f"[build_masters] snapshot loaded · hotels={len(snapshot.get('hotels') or [])} · "
              f"synthetic_compsets={len(snapshot.get('synthetic_compsets') or [])}")
        hotels = snapshot.get("hotels") or []
        hotels_by_id = {h["hotel_id"]: h for h in hotels}
        compsets_synth = snapshot.get("synthetic_compsets") or []
        compsets_real = snapshot.get("compset_membership") or snapshot.get("compsets") or []
        compset_by_target = {c.get("target_hotel_id"): c for c in compsets_synth}
        # Market snapshot index (LTM rolling per market). The source XLSX
        # carries multiple rows per market_name (overall + class-level
        # breakdowns). Prefer the row with populated occupancy_12m ·
        # that's the rolled-up market KPI row.
        market_snap = {}
        for r in (snapshot.get("market_snapshots") or []):
            if r.get("granularity") != "market" or not r.get("market_name"):
                continue
            key = r["market_name"]
            existing = market_snap.get(key)
            if existing is None or (r.get("occupancy_12m") is not None and existing.get("occupancy_12m") is None):
                market_snap[key] = r

        compset_rows = [_compset_to_row(c, hotels_by_id, snapshot) for c in compsets_synth]
        # Real compsets prepended (operator-confirmed wins over synthetic in display order)
        compset_rows = [_compset_to_row(c, hotels_by_id, snapshot) for c in compsets_real] + compset_rows
        positioning_rows = [_positioning_to_row(h, compset_by_target, market_snap, snapshot) for h in hotels]
    else:
        print(f"[build_masters] no snapshot at {COSTAR_SNAPSHOT_PATH} · emitting headers-only templates")
        compset_rows = []
        positioning_rows = []

    build_workbook(
        dataset_label="COMPSET_MASTER",
        domain_columns=COMPSET_COLUMNS,
        data_sheet_title="COMPSET",
        schema_doc="docs/intelligence/compset-schema.md",
        output_path=MASTER_DIR / "COMPSET_MASTER.xlsx",
        data_rows=compset_rows,
    )
    build_workbook(
        dataset_label="HOTEL_POSITIONING_MASTER",
        domain_columns=POSITIONING_COLUMNS,
        data_sheet_title="POSITIONING",
        schema_doc="docs/intelligence/hotel-positioning-schema.md",
        output_path=MASTER_DIR / "HOTEL_POSITIONING_MASTER.xlsx",
        data_rows=positioning_rows,
    )


if __name__ == "__main__":
    main()
