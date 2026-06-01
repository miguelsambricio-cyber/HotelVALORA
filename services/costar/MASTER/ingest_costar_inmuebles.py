#!/usr/bin/env python3
"""
ingest_costar_inmuebles.py — official, market-agnostic ingestor for a CoStar
"Inmuebles" property-list Excel into the per-hotel master (hotel_canonical).

Born from the Madrid M0/M1 work (2026-06-01). Reusable for any market's CoStar
export (Barcelona, NY, …): it does NOT hardcode column indices — it maps by
HEADER TEXT — and it never connects to the database. It produces:

  1. A match report (CSV): CLARO / DUDOSO / SIN GEMELO buckets for the existing
     canonical corpus vs the Excel, plus the Part-B promotion candidates.
  2. Staging SQL artifacts an operator applies via the Supabase MCP / psql:
       - <out>/backfill_claros.sql   · UPDATE costar_property_id + total_rooms
                                        on matched canonical rows (CLARO).
       - <out>/promote_partb.sql     · CREATE staging + the operator then runs
                                        the canonical INSERT (see README/M0).

Design rules baked in (see docs/changelog.md M0/M1 + memory project_master_hotel_plan):
  * Identity bridge = costar_property_id. Match by ADDRESS (street+number) + CP
    + submarket; NAME is a weak secondary signal (CoStar building-name ≠ Booking
    commercial name).
  * A building match REQUIRES a shared street-number UNLESS the name is near
    identical and CP+submarket agree (handles corner buildings / rebrands).
    This is what stops the "same street, different number" false positive.
  * Promotions are costar_only (no lat/lng) → excluded from the front by the
    loadSearchCorpus coords filter. Apartamentos are EXCLUDED by default.
  * Priority is manual/admin > CoStar > the rest; this script only PROPOSES —
    a human applies, and the merge step (M1) honours hotel_field_provenance
    override_by (sticky admin) and never clobbers coords/amenities/reviews.

USAGE
  python ingest_costar_inmuebles.py \
      --excel "1.1 CostarExport - INMUEBLES LISTA HOTELES MADRID.xlsx" \
      --corpus corpus_madrid.tsv \
      --market madrid \
      --out ./out

  --corpus is a TAB-separated export of the existing canonical rows, one per
  line:  canonical_name <TAB> postal_code <TAB> submarket_name <TAB> address
  (produce it from hotel_canonical; submarket joined by name).

  --submarket-map (optional) JSON {submarket_name: submarket_uuid} for the target
  market, used to fill submarket_id in the promotion staging. Without it, the
  staging leaves submarket_id NULL (operator maps later).
"""
from __future__ import annotations
import argparse, csv, json, re, sys, unicodedata
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl required: pip install openpyxl")

# --------------------------------------------------------------------------- #
# Normalisation helpers (identical semantics to the Madrid matcher v3)
# --------------------------------------------------------------------------- #
def strip_acc(s: str) -> str:
    s = unicodedata.normalize("NFD", s or "")
    return "".join(c for c in s if unicodedata.category(c) != "Mn")

STREET = {
    "calle", "c", "avenida", "avda", "av", "paseo", "po", "p", "plaza", "pza",
    "pl", "glorieta", "gta", "ronda", "rda", "camino", "cmno", "carretera",
    "ctra", "cuesta", "via", "travesia", "del", "de", "la", "los", "las", "el",
    "y", "s", "n", "sn", "bajo", "km",
}

def addr_parts(a: str):
    """Return (street-token set, number set). Numbers are the strong signal."""
    a = strip_acc((a or "").lower()).replace("s/n", " ")
    nums = set(re.findall(r"\d+", a))
    toks = {t for t in re.split(r"[^a-z0-9]+", a)
            if t and not t.isdigit() and t not in STREET and len(t) > 2}
    return toks, nums

# brand/marketing suffixes stripped before name comparison
NSUF = [
    " - the leading hotels of the world", ", autograph collection",
    ", a luxury collection hotel, madrid", ", a luxury collection hotel",
    ", small luxury hotels", ", a small luxury hotel of the world",
    ", a member of design hotels", ", a member of preferred hotels & resorts",
    ", tapestry collection by hilton", ", curio collection by hilton",
    ", affiliated by melia", " by ihg", " by marriott", " by hilton",
    " - evok collection", " - adults only", " - new opening",
    " member of melia collection", " a gran melia hotel", " by leonardo hotels",
    " by hyatt", "(adults only)",
]
# generic words (incl. street-type words) that must NOT inflate name similarity
NSTOP = {
    "hotel", "hoteles", "hospedaje", "madrid", "de", "del", "la", "el", "los",
    "las", "by", "the", "a", "of", "world", "collection", "spa", "suites",
    "apartments", "apartamentos", "boutique", "edificio", "and", "y", "con",
    "servicios", "aparthotel", "hostal", "hostel", "grand", "gran", "via",
    "plaza", "paseo", "calle", "avenida", "avda", "ronda", "cuesta",
}

def nname(s: str):
    s = (s or "").lower()
    for x in NSUF:
        s = s.replace(x, "")
    s = strip_acc(s)
    s = re.sub(r"[^a-z0-9 ]", " ", s)
    return {t for t in s.split() if t not in NSTOP and len(t) > 2}

def jac(a: set, b: set) -> float:
    return len(a & b) / len(a | b) if (a and b) else 0.0

# CoStar "Clase" → hotel_segment enum
SCALE = {
    "luxury": "luxury", "upper upscale": "upper_upscale", "upscale": "upscale",
    "upper midscale": "upper_midscale", "midscale": "midscale", "economy": "economy",
}
def scale_of(clase: str) -> str:
    return SCALE.get(strip_acc((clase or "").lower()).strip(), "unknown")

# CoStar "Estado de construcción" → hotel_lifecycle_enum
def status_of(estado: str) -> str:
    e = strip_acc((estado or "").lower()).strip()
    if e in ("construido", "en reformas"):
        return "active"
    if e == "en construccion":
        return "under_construction"
    if e in ("suelo urbano", "planificacion final", "proyecto"):
        return "planned"
    if e in ("demolido", "proyecto abandonado"):
        return "closed"
    return "unverified"

def slugify(s: str) -> str:
    s = strip_acc((s or "").lower())
    s = re.sub(r"[^a-z0-9]+", "-", s).strip("-")
    return s or "hotel"

def sql_str(s) -> str:
    return "NULL" if s in (None, "") else "'" + str(s).replace("'", "''") + "'"

# --------------------------------------------------------------------------- #
# CoStar column resolution by header text (robust across market exports)
# --------------------------------------------------------------------------- #
# Maps a logical field → list of accepted header substrings (accent-insensitive,
# lowercased). First matching column wins.
HEADER_ALIASES = {
    "name":        ["nombre del edificio", "building name", "nombre"],
    "rooms":       ["habitaciones", "rooms", "number of rooms"],
    "submarket":   ["submercado", "submarket"],
    "clase":       ["clase", "class"],
    "escala":      ["escala", "scale"],
    "estado":      ["estado de construccion", "construction status"],
    "address":     ["direccion", "address"],
    "cp":          ["codigo postal", "postal code", "zip"],
    "tsec":        ["tipo secundario", "secondary type"],
    "costar_id":   ["id del inmueble", "property id", "costar id"],
    "stars":       ["clasificacion por estrellas", "star rating", "estrellas"],
    "year_built":  ["ano de construccion", "year built", "ano construccion"],
    "year_reno":   ["ano de reform", "year renovated", "ano reforma"],
    "sqm":         ["superficie alquilable", "sba", "rentable area"],
    "floors":      ["plantas", "floors", "stories"],
    "owner_real":  ["propietario real", "true owner"],
    "owner_reg":   ["propietario registrado", "registered owner"],
    "marca":       ["marca", "brand"],
    "matriz":      ["empresa matriz", "parent company"],
    "operador":    ["operador del hotel", "operator"],
}

def resolve_columns(header: list[str]) -> dict[str, int]:
    norm = [strip_acc(str(h or "").lower()).strip() for h in header]
    out: dict[str, int] = {}
    for field, aliases in HEADER_ALIASES.items():
        for i, h in enumerate(norm):
            if any(a in h for a in aliases):
                out[field] = i
                break
    missing = [f for f in ("name", "costar_id", "address", "cp", "tsec") if f not in out]
    if missing:
        sys.exit(f"CoStar Excel missing required columns: {missing}")
    return out

# --------------------------------------------------------------------------- #
# Matcher v3 (address-first, number-mandatory unless strong name + area)
# --------------------------------------------------------------------------- #
def feats(c: dict, e: dict):
    aj = jac(c["at"], e["at"])
    num = bool(c["an"] & e["an"])
    cp = bool(c["cp"] and e["cp"] and c["cp"] == e["cp"])
    sm = bool(c["sm"] and e["sm"] and c["sm"] == e["sm"])
    nj = jac(c["nt"], e["nt"])
    return aj, num, cp, sm, nj

def rankf(aj, num, cp, sm, nj):
    strong = aj >= 0.5 and num
    return (aj if num else aj * 0.25) + (0.5 if strong else 0) + (0.5 if cp else 0) \
        + (0.3 if sm else 0) + nj * 0.4

def is_claro(aj, num, cp, sm, nj):
    strong = aj >= 0.5 and num
    area = cp and sm
    # number still mandatory UNLESS name near-identical + same CP&submarket
    return (strong and (cp or sm or nj >= 0.3)) or (nj >= 0.85 and area) \
        or (nj >= 0.7 and area and aj >= 0.34)

def is_dud(aj, num, cp, sm, nj):
    area = cp and sm
    nconf = num is False  # placeholder; computed by caller with real conflict
    return (aj >= 0.5 and area and nj >= 0.3 and not num) \
        or (nj >= 0.55 and area) or (nj >= 0.6 and area) \
        or (aj >= 0.34 and num and cp)

def load_corpus(path: Path) -> list[dict]:
    rows = []
    for line in path.read_text(encoding="utf-8").splitlines():
        p = line.split("\t")
        if len(p) < 4 or not p[0].strip():
            continue
        at, an = addr_parts(p[3])
        rows.append({"name": p[0].strip(), "cp": p[1].strip(), "sm": p[2].strip(),
                     "addr": p[3].strip(), "at": at, "an": an, "nt": nname(p[0])})
    return rows

def load_excel(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    col = resolve_columns(list(next(it)))
    g = lambda r, f: ("" if col.get(f) is None or col[f] >= len(r) or r[col[f]] is None
                      else str(r[col[f]]).strip())
    out = []
    for r in it:
        if not g(r, "name"):
            continue
        at, an = addr_parts(g(r, "address"))
        rec = {k: g(r, k) for k in HEADER_ALIASES}
        # uniform matcher keys (corpus uses sm/cp/addr/at/an/nt)
        rec.update({"sm": rec.get("submarket", ""), "addr": rec.get("address", ""),
                    "at": at, "an": an, "nt": nname(rec.get("name", ""))})
        out.append(rec)
    wb.close()
    return out

def match(corpus: list[dict], excel: list[dict], reject: set[str]):
    """Two-pass: CLARO claims first, DUDOSO recall over unclaimed rows."""
    claro, dud, none, claimed, claro_for = [], [], [], set(), set()
    for ci, c in enumerate(corpus):
        bi, bf = -1, None
        br = -1.0
        for i, e in enumerate(excel):
            f = feats(c, e)
            r = rankf(*f)
            if r > br:
                br, bi, bf = r, i, f
        if c["name"] in reject or bi < 0:
            none.append((c, None, *(bf or (0, 0, 0, 0, 0))))
            continue
        if is_claro(*bf):
            claro.append((c, excel[bi], *bf)); claimed.add(bi); claro_for.add(ci)
    for ci, c in enumerate(corpus):
        if ci in claro_for or c["name"] in reject:
            if ci not in claro_for and not any(c is x[0] for x in none):
                none.append((c, None, 0, 0, 0, 0, 0))
            continue
        bi, bf, br = -1, None, -1.0
        for i, e in enumerate(excel):
            if i in claimed:
                continue
            f = feats(c, e); r = rankf(*f)
            if r > br:
                br, bi, bf = r, i, f
        if bi >= 0 and is_dud(*bf):
            dud.append((c, excel[bi], *bf))
        else:
            none.append((c, excel[bi] if bi >= 0 else None, *(bf or (0, 0, 0, 0, 0))))
    return claro, dud, none, claimed

# --------------------------------------------------------------------------- #
# Emitters
# --------------------------------------------------------------------------- #
def write_reports(out: Path, claro, dud, none, partB):
    out.mkdir(parents=True, exist_ok=True)
    with open(out / "match_claro.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f); w.writerow(["corpus_name", "excel_name", "costar_id", "rooms", "addrJ", "num", "cp", "sm", "nameJ"])
        for c, e, aj, num, cp, sm, nj in claro:
            w.writerow([c["name"], e["name"], e["costar_id"], e["rooms"], f"{aj:.2f}", num, cp, sm, f"{nj:.2f}"])
    with open(out / "match_dudoso.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f); w.writerow(["corpus_name", "corpus_addr", "excel_name", "excel_addr", "addrJ", "num", "cp", "sm", "nameJ"])
        for c, e, aj, num, cp, sm, nj in dud:
            w.writerow([c["name"], c["addr"], e["name"], e["address"], f"{aj:.2f}", num, cp, sm, f"{nj:.2f}"])
    with open(out / "match_sin_gemelo.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f); w.writerow(["corpus_name", "corpus_addr", "corpus_cp", "submarket"])
        for c, *_ in none:
            w.writerow([c["name"], c["addr"], c["cp"], c["sm"]])
    with open(out / "partB_promociones.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f); w.writerow(["tipo", "excel_name", "costar_id", "rooms", "stars", "clase", "submarket", "cp", "addr"])
        for e in partB:
            w.writerow([e["tsec"], e["name"], e["costar_id"], e["rooms"], e["stars"], e["clase"], e["submarket"], e["cp"], e["address"]])

def write_backfill_sql(out: Path, claro):
    vals = []
    for c, e, *_ in claro:
        rooms = e["rooms"] if str(e["rooms"]).isdigit() else "NULL"
        vals.append(f"  ({sql_str(c['name'])}, {sql_str(e['costar_id'])}, {rooms})")
    sql = ("-- back-fill costar_property_id + total_rooms on CLARO matches.\n"
           "-- total_rooms uses COALESCE so it never nulls an existing value;\n"
           "-- review manual/admin-curated rooms BEFORE applying (manual > CoStar).\n"
           "UPDATE public.hotel_canonical AS h\n"
           "SET costar_property_id = v.cid,\n"
           "    total_rooms = COALESCE(v.rooms, h.total_rooms)\n"
           "FROM (VALUES\n" + ",\n".join(vals) +
           "\n) AS v(name, cid, rooms)\nWHERE h.canonical_name = v.name;\n")
    (out / "backfill_claros.sql").write_text(sql, encoding="utf-8")

def write_promote_sql(out: Path, partB, submarket_map: dict[str, str]):
    seen, vals = set(), []
    for e in partB:
        base = slugify(e["name"]); slug, k = base, 2
        while slug in seen:
            slug = f"{base}-{k}"; k += 1
        seen.add(slug)
        pk = "hotel" if e["tsec"].lower() == "hotel" else ("hostel" if "hostel" in e["tsec"].lower() else "apartment")
        rooms = e["rooms"] if str(e["rooms"]).isdigit() else "NULL"
        stars = e["stars"] if str(e["stars"]).isdigit() else "NULL"
        yb = e["year_built"] if str(e["year_built"]).isdigit() else "NULL"
        yr = e["year_reno"] if str(e["year_reno"]).isdigit() else "NULL"
        try:
            sqm = str(round(float(e["sqm"]), 2)) if e["sqm"] else "NULL"
        except ValueError:
            sqm = "NULL"
        smid = submarket_map.get(e["submarket"])
        vals.append("  (" + ", ".join([
            sql_str(e["costar_id"]), sql_str(e["name"]), sql_str(slug), str(rooms), str(stars),
            sql_str(scale_of(e["clase"])), sql_str(pk), sql_str(e["address"]), sql_str(e["cp"]),
            (sql_str(smid) if smid else "NULL"), str(yb), str(yr), str(sqm),
            sql_str(status_of(e["estado"]))]) + ")")
    sql = ("-- Part-B promotion STAGING. Load this, then run the canonical INSERT\n"
           "-- (see docs/changelog.md M0) which sets data_quality_tier='costar_only',\n"
           "-- no lat/lng, and the anti-duplicate guards (costar_id / slug / address).\n"
           "DROP TABLE IF EXISTS public._costar_partb_stage;\n"
           "CREATE TABLE public._costar_partb_stage (costar_id text, name text, base_slug text, "
           "rooms int, stars int, scale text, prop_kind text, addr text, cp text, submarket_id uuid, "
           "year_opened int, year_renovated int, sqm numeric, status text);\n"
           "INSERT INTO public._costar_partb_stage VALUES\n" + ",\n".join(vals) + ";\n")
    (out / "promote_partb.sql").write_text(sql, encoding="utf-8")

# --------------------------------------------------------------------------- #
def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--excel", required=True, type=Path)
    ap.add_argument("--corpus", required=True, type=Path, help="TSV: name<TAB>cp<TAB>submarket<TAB>address")
    ap.add_argument("--market", default="market")
    ap.add_argument("--out", default=Path("./out"), type=Path)
    ap.add_argument("--submarket-map", type=Path, help="JSON {submarket_name: uuid}")
    ap.add_argument("--reject", type=Path, help="optional file, one corpus name per line, to force SIN GEMELO")
    ap.add_argument("--include-apartments", action="store_true",
                    help="include 'Apartamento con servicios' in Part B (default: excluded)")
    args = ap.parse_args()

    reject = set()
    if args.reject and args.reject.exists():
        reject = {l.strip() for l in args.reject.read_text(encoding="utf-8").splitlines() if l.strip()}
    submarket_map = {}
    if args.submarket_map and args.submarket_map.exists():
        submarket_map = json.loads(args.submarket_map.read_text(encoding="utf-8"))

    corpus = load_corpus(args.corpus)
    excel = load_excel(args.excel)
    claro, dud, none, claimed = match(corpus, excel, reject)

    def is_apt(t): return "apartamento" in strip_acc((t or "").lower())
    partB = [e for i, e in enumerate(excel) if i not in claimed
             and not is_apt(e["tsec"]) and (args.include_apartments or e["tsec"].lower() in ("hotel", "hostel"))]
    if args.include_apartments:
        partB = [e for i, e in enumerate(excel) if i not in claimed]

    write_reports(args.out, claro, dud, none, partB)
    write_backfill_sql(args.out, claro)
    write_promote_sql(args.out, partB, submarket_map)

    H = sum(1 for e in partB if e["tsec"].lower() == "hotel")
    HS = sum(1 for e in partB if "hostel" in e["tsec"].lower())
    print(f"[{args.market}] corpus={len(corpus)} excel={len(excel)}")
    print(f"  CLARO={len(claro)}  DUDOSO={len(dud)}  SIN GEMELO={len(none)}")
    print(f"  Part B (promociones)={len(partB)}  (Hotel {H} / Hostel {HS})")
    print(f"  → {args.out}/  (match_*.csv, backfill_claros.sql, promote_partb.sql)")
    print("  NOTE: read-only proposal. Apply via Supabase MCP after human review;")
    print("        respect manual/admin provenance (override_by) — manual > CoStar.")

if __name__ == "__main__":
    main()
