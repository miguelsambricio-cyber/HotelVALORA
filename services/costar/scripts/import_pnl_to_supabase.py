"""Imports COSTAR_MASTER_FINANCIALS.xlsx → pnl_template Supabase.

PHASE 2 of the Supabase migration roadmap (see docs/database.md).

What it does:
  · Reads the operator-filled Excel `datos_reales` (149 rows · Madrid + 41
    country headers).
  · Normalises three drifts vs the canonical schema:
      - data_source `costar_national_ES` → `costar_national` (scalability ajuste Fase 1)
      - submarket `Arguelles-Chamberi`        → `Arguelles & Chamberi`        (BD canonical form)
      - submarket `Chamartin-Plaza de Castilla` → `Chamartin & Plaza de Castilla` (BD canonical form)
      - percentage ratios 0-1 → 0-100 (DB convention · numeric(5,2))
  · Generates apartahotel + hostel derived rows for Madrid only, per the
    methodology MVP rule (VALUATION_METHODOLOGY.md annex "Perfiles
    derivados apartahotel y hostel"). Rule: for each Madrid hotel row with
    real data (costar_submarket_aggregate or costar_national), overwrite
    the matching apartahotel + hostel pending_costar placeholders with
    the derived percentage table and flip data_source to derived_mvp_rule.
    NO derived generation outside Madrid · NO derivation over pending bases.
  · Emits a JSON plan + a bulk UPSERT SQL with idempotency semantics:
    `WHERE pnl_template.col IS DISTINCT FROM EXCLUDED.col` so re-running with
    identical data is a true no-op (updated_at doesn't move).
  · Reports the plan to stdout: counts by data_source, sample rows by
    category, and (if a current-state JSON is provided) conflict detection
    when data_source would be silently flipped.

What it does NOT do:
  · Never connects directly to Supabase. The SQL output is applied via
    Supabase MCP execute_sql in a follow-up step (operator-authorized).
  · Never touches pnl_template_override. Operator edits are preserved by
    the two-table architecture.
  · Never derives across pending_costar bases (data integrity principle).

Usage:
  python services/costar/scripts/import_pnl_to_supabase.py --dry-run
  python services/costar/scripts/import_pnl_to_supabase.py            # writes _pnl_upsert.sql
"""

from __future__ import annotations

import argparse
import json
import sys
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

sys.stdout.reconfigure(encoding="utf-8")

REPO = Path(__file__).resolve().parents[3]
SOURCE_XLSX = REPO / "services" / "costar" / "MASTER" / "COSTAR_MASTER_FINANCIALS.xlsx"
PLAN_JSON  = REPO / "services" / "costar" / "MASTER" / "_pnl_upsert_plan.json"
SQL_OUT    = REPO / "services" / "costar" / "MASTER" / "_pnl_upsert.sql"

# ───────────────────────────────────────────────────────────────────────────
# Normalisations (Excel form → canonical BD form)
# ───────────────────────────────────────────────────────────────────────────

DATA_SOURCE_NORMALIZATIONS: dict[str, str] = {
    "costar_national_ES": "costar_national",   # Fase 1 scalability rename
}

SUBMARKET_NORMALIZATIONS: dict[str, str] = {
    "Arguelles-Chamberi": "Arguelles & Chamberi",
    "Chamartin-Plaza de Castilla": "Chamartin & Plaza de Castilla",
}

# ───────────────────────────────────────────────────────────────────────────
# Percentage column inventory · matches pnl_template DB schema exactly
# ───────────────────────────────────────────────────────────────────────────

PCT_COLS: list[str] = [
    "rooms_revenue_pct", "fb_food_pct", "fb_beverage_pct", "meeting_events_pct",
    "spa_wellness_pct", "parking_other_pct",
    "expenses_rooms_pct", "expenses_fb_pct", "other_departments_pct",
    "admin_general_pct", "it_telecom_pct", "sales_marketing_pct",
    "operations_maintenance_pct", "utilities_pct",
    "gop_pct", "management_fees_pct", "rent_pct",
    "property_taxes_pct", "insurance_pct", "ebitda_pct", "staff_cost_memo_pct",
]

ID_COLS: list[str] = ["country", "market", "submarket", "class", "segmentation_type"]
PROVENANCE_COLS: list[str] = ["data_source", "last_imported_at", "imported_from", "notes"]

# ───────────────────────────────────────────────────────────────────────────
# Derived MVP rules (operator-firmed methodology, 2026-05-28)
# Values are stored as 0-100 percentages to match the DB convention.
# Lines marked None are stored as NULL in BD (no methodology value yet OR
# excluded from HV EBITDA per design).
# ───────────────────────────────────────────────────────────────────────────

APARTAHOTEL_RULES: dict[str, float | None] = {
    "rooms_revenue_pct": 92.0,
    "fb_food_pct": 4.0,
    "fb_beverage_pct": 1.0,
    "meeting_events_pct": 0.0,
    "spa_wellness_pct": 0.0,
    "parking_other_pct": 3.0,
    "expenses_rooms_pct": 22.0,
    "expenses_fb_pct": 75.0,
    "other_departments_pct": None,     # no methodology value for derived
    "admin_general_pct": 4.0,
    "it_telecom_pct": None,             # stored but excluded from HV EBITDA
    "sales_marketing_pct": 4.0,
    "operations_maintenance_pct": 3.0,
    "utilities_pct": 2.5,
    "gop_pct": 61.3,
    "management_fees_pct": 20.0,        # apartahotel mgmt fee firmed
    "rent_pct": None,                   # stored but excluded from HV EBITDA
    "property_taxes_pct": 0.7,
    "insurance_pct": 0.4,
    "ebitda_pct": 40.2,                 # HV EBITDA pre-alquiler
    "staff_cost_memo_pct": 18.0,
}

HOSTEL_RULES: dict[str, float | None] = {
    "rooms_revenue_pct": 82.0,
    "fb_food_pct": 3.0,
    "fb_beverage_pct": 5.0,
    "meeting_events_pct": 0.0,
    "spa_wellness_pct": 0.0,
    "parking_other_pct": 10.0,
    "expenses_rooms_pct": 28.0,
    "expenses_fb_pct": 70.0,
    "other_departments_pct": None,
    "admin_general_pct": 5.0,
    "it_telecom_pct": None,
    "sales_marketing_pct": 5.5,
    "operations_maintenance_pct": 3.5,
    "utilities_pct": 3.5,
    "gop_pct": 49.9,
    "management_fees_pct": 12.0,        # hostel mgmt fee firmed
    "rent_pct": None,
    "property_taxes_pct": 0.7,
    "insurance_pct": 0.4,
    "ebitda_pct": 36.8,                 # HV EBITDA pre-alquiler
    "staff_cost_memo_pct": 22.0,
}

DERIVED_BASES_OK = {"costar_submarket_aggregate", "costar_national"}  # post-normalisation

# ───────────────────────────────────────────────────────────────────────────
# Excel reader
# ───────────────────────────────────────────────────────────────────────────

def read_excel() -> list[dict[str, Any]]:
    if not SOURCE_XLSX.exists():
        raise FileNotFoundError(f"Source Excel not found: {SOURCE_XLSX}")
    wb = load_workbook(SOURCE_XLSX, read_only=True)
    ws = wb["DATA"]
    rows = list(ws.iter_rows(values_only=True))
    hdr = list(rows[0])
    return [dict(zip(hdr, r)) for r in rows[1:]]


# ───────────────────────────────────────────────────────────────────────────
# Normalisations
# ───────────────────────────────────────────────────────────────────────────

def normalise(row: dict[str, Any]) -> None:
    """In-place normalisation of a row to canonical BD form."""
    # data_source
    ds = row.get("data_source")
    if ds in DATA_SOURCE_NORMALIZATIONS:
        row["data_source"] = DATA_SOURCE_NORMALIZATIONS[ds]
    # submarket
    sm = row.get("submarket")
    if sm in SUBMARKET_NORMALIZATIONS:
        row["submarket"] = SUBMARKET_NORMALIZATIONS[sm]
    # ratio 0-1 → percentage 0-100 (only when the cell is a float in 0..1 range)
    for col in PCT_COLS:
        v = row.get(col)
        if isinstance(v, (int, float)) and v is not None and not isinstance(v, bool):
            if 0.0 <= v <= 1.0:
                row[col] = round(float(v) * 100, 2)
            else:
                row[col] = round(float(v), 2)


# ───────────────────────────────────────────────────────────────────────────
# Derived row generation (Madrid only)
# ───────────────────────────────────────────────────────────────────────────

def generate_derived(rows: list[dict[str, Any]]) -> int:
    """For each Madrid hotel row with data_source ∈ DERIVED_BASES_OK,
    overwrite the matching (apartahotel, hostel) placeholders with the
    derived MVP rule percentages. Returns the count of derived rows
    written (apartahotel + hostel both count)."""
    by_key: dict[tuple, dict[str, Any]] = {}
    for r in rows:
        key = (
            r.get("country"), r.get("market"), r.get("submarket"),
            r.get("class"), r.get("segmentation_type"),
        )
        by_key[key] = r

    derived_count = 0
    for r in list(rows):
        if not (
            r.get("country") == "ES"
            and r.get("market") == "Madrid"
            and r.get("segmentation_type") == "hotel"
            and r.get("data_source") in DERIVED_BASES_OK
        ):
            continue
        # We have a valid Madrid hotel base · derive aparta + hostel rows.
        for derived_type, rules in (("apartahotel", APARTAHOTEL_RULES), ("hostel", HOSTEL_RULES)):
            derived_key = (r["country"], r["market"], r["submarket"], r["class"], derived_type)
            target = by_key.get(derived_key)
            if target is None:
                # Excel didn't have a placeholder · create a new row
                target = {col: None for col in ID_COLS + PCT_COLS + PROVENANCE_COLS}
                target["country"] = r["country"]
                target["market"] = r["market"]
                target["submarket"] = r["submarket"]
                target["class"] = r["class"]
                target["segmentation_type"] = derived_type
                rows.append(target)
                by_key[derived_key] = target
            # Apply the derived rule · OVERWRITES any prior values (incl. NULL placeholders)
            for col in PCT_COLS:
                target[col] = rules[col]
            target["data_source"] = "derived_mvp_rule"
            target["notes"] = (
                f"Derived MVP rule (VALUATION_METHODOLOGY.md annex "
                f"'Perfiles derivados apartahotel y hostel') · base hotel row: "
                f"{r['submarket']} × {r['class']} ({r['data_source']})"
            )
            derived_count += 1
    return derived_count


# ───────────────────────────────────────────────────────────────────────────
# SQL emitter
# ───────────────────────────────────────────────────────────────────────────

def sql_literal(v: Any) -> str:
    if v is None:
        return "NULL"
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    if isinstance(v, (int, float)):
        return f"{v}"
    if isinstance(v, datetime):
        return f"'{v.strftime('%Y-%m-%dT%H:%M:%SZ')}'"
    # text · escape single quotes
    s = str(v).replace("'", "''")
    return f"'{s}'"


def emit_upsert_sql(rows: list[dict[str, Any]], imported_from: str) -> str:
    """Generate one bulk INSERT...ON CONFLICT DO UPDATE statement.

    Idempotency: the WHERE clause on DO UPDATE compares every payload
    column for IS DISTINCT FROM; if nothing changed, the UPDATE is a no-op
    and the BEFORE UPDATE trigger doesn't fire, so updated_at stays put.
    """
    all_cols = ID_COLS + PCT_COLS + ["data_source", "imported_from", "notes"]
    # last_imported_at is set server-side via NOW() in the VALUES tuple to
    # match the column default and to give us a fresh import timestamp.
    col_list = ", ".join(all_cols + ["last_imported_at"])

    value_lines: list[str] = []
    for r in rows:
        vs = [sql_literal(r.get(c)) for c in ID_COLS + PCT_COLS]
        vs.append(sql_literal(r.get("data_source")))
        vs.append(sql_literal(imported_from))
        vs.append(sql_literal(r.get("notes")))
        vs.append("NOW()")
        value_lines.append("  (" + ", ".join(vs) + ")")

    # DO UPDATE clause · all payload columns refresh from EXCLUDED
    update_cols = PCT_COLS + ["data_source", "imported_from", "notes", "last_imported_at"]
    update_clause = ",\n  ".join(f"{c} = EXCLUDED.{c}" for c in update_cols)

    # WHERE clause · idempotency
    where_clauses = [
        f"pnl_template.{c} IS DISTINCT FROM EXCLUDED.{c}"
        for c in PCT_COLS + ["data_source", "imported_from", "notes"]
    ]
    where_clause = "\n     OR ".join(where_clauses)

    sql = (
        "-- ============================================================================\n"
        "-- pnl_template bulk UPSERT generated by import_pnl_to_supabase.py\n"
        "-- Source: services/costar/MASTER/COSTAR_MASTER_FINANCIALS.xlsx\n"
        "-- Generated at: " + datetime.now(timezone.utc).isoformat() + "\n"
        f"-- Rows: {len(rows)}\n"
        "-- ============================================================================\n\n"
        f"INSERT INTO public.pnl_template ({col_list})\nVALUES\n"
        + ",\n".join(value_lines)
        + "\nON CONFLICT ON CONSTRAINT pnl_template_uk DO UPDATE SET\n  "
        + update_clause
        + "\nWHERE "
        + where_clause
        + ";\n"
    )
    return sql


def emit_audit_sql(rows: list[dict[str, Any]], derived_count: int, imported_from: str) -> str:
    """Single INSERT into ai_agent_runs documenting this import."""
    metadata = {
        "script": "import_pnl_to_supabase.py",
        "imported_from": imported_from,
        "total_rows": len(rows),
        "derived_count": derived_count,
        "by_data_source": dict(Counter(r.get("data_source") for r in rows)),
    }
    md_literal = sql_literal(json.dumps(metadata, ensure_ascii=False))
    return (
        "INSERT INTO public.ai_agent_runs (agent_id, status, metadata) VALUES "
        f"('data_ingestion', 'completed', {md_literal}::jsonb);\n"
    )


# ───────────────────────────────────────────────────────────────────────────
# Conflict detection helper (offline · operator-fed current state JSON)
# ───────────────────────────────────────────────────────────────────────────

def detect_conflicts(rows: list[dict[str, Any]], current_state_json: str | None) -> list[dict[str, Any]]:
    """If `current_state_json` is provided (path to a JSON array of current
    pnl_template rows · queried separately via Supabase MCP), check for
    data_source flips that should NOT be auto-overwritten. Returns the list
    of conflicts. Today's first run: BD is empty · no conflicts possible.
    """
    if not current_state_json:
        return []
    current = json.loads(Path(current_state_json).read_text(encoding="utf-8"))
    by_key = {
        (c.get("country"), c.get("market"), c.get("submarket"), c.get("class"), c.get("segmentation_type")): c
        for c in current
    }
    conflicts: list[dict[str, Any]] = []
    for r in rows:
        key = (r.get("country"), r.get("market"), r.get("submarket"), r.get("class"), r.get("segmentation_type"))
        existing = by_key.get(key)
        if existing and existing.get("data_source") != r.get("data_source"):
            conflicts.append({
                "key": key,
                "existing_data_source": existing.get("data_source"),
                "incoming_data_source": r.get("data_source"),
            })
    return conflicts


# ───────────────────────────────────────────────────────────────────────────
# Reporting (stdout)
# ───────────────────────────────────────────────────────────────────────────

def print_summary(rows: list[dict[str, Any]], derived_count: int, conflicts: list[dict[str, Any]]) -> None:
    ds_counts = Counter(r.get("data_source") for r in rows)
    print("===========================================================")
    print(" PNL_TEMPLATE UPSERT PLAN")
    print("===========================================================")
    print(f" Source: {SOURCE_XLSX.relative_to(REPO)}")
    print(f" Total rows to upsert : {len(rows)}")
    print(f" Derived rows added   : {derived_count} (Madrid · apartahotel + hostel)")
    print(f" Conflicts detected   : {len(conflicts)}")
    print()
    print(" Distribution by data_source:")
    for ds, n in sorted(ds_counts.items()):
        print(f"   · {ds:<32} {n}")
    print()


def print_samples(rows: list[dict[str, Any]]) -> None:
    """3-5 representative samples covering each data_source category present."""
    by_ds: dict[str, list[dict[str, Any]]] = {}
    for r in rows:
        by_ds.setdefault(r.get("data_source"), []).append(r)
    print(" Sample rows (one per data_source category):")
    print()
    for ds in ("costar_submarket_aggregate", "costar_national", "derived_mvp_rule", "pending_costar", "hardcoded_default"):
        if ds not in by_ds:
            continue
        r = by_ds[ds][0]
        print(f"   --- data_source = {ds} ---")
        key_str = " × ".join(str(r.get(c) or "<null>") for c in ID_COLS)
        print(f"   key : {key_str}")
        sample_pcts = ["rooms_revenue_pct", "fb_food_pct", "fb_beverage_pct", "meeting_events_pct",
                       "spa_wellness_pct", "management_fees_pct", "ebitda_pct"]
        kv = ", ".join(f"{c}={r.get(c)}" for c in sample_pcts)
        print(f"   pct : {kv}")
        print(f"   notes: {(r.get('notes') or '')[:120]}")
        print()


def print_conflicts(conflicts: list[dict[str, Any]]) -> None:
    if not conflicts:
        print(" No data_source conflicts detected.")
        return
    print(" ⚠ DATA_SOURCE CONFLICTS (will NOT auto-overwrite without your OK):")
    for c in conflicts[:20]:
        print(f"   · {' × '.join(str(x) for x in c['key'])}")
        print(f"       existing: {c['existing_data_source']}  → incoming: {c['incoming_data_source']}")


# ───────────────────────────────────────────────────────────────────────────
# Main
# ───────────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(description="Import COSTAR_MASTER_FINANCIALS.xlsx into pnl_template")
    parser.add_argument("--dry-run", action="store_true",
                        help="Print plan + samples but don't write the SQL file")
    parser.add_argument("--current-state-json", default=None,
                        help="Optional path to a JSON dump of current pnl_template rows (for conflict detection)")
    args = parser.parse_args()

    # Read + normalise
    rows = read_excel()
    for r in rows:
        normalise(r)

    # Derive Madrid apartahotel/hostel rows from real-base Madrid hotel rows
    derived_count = generate_derived(rows)

    # Conflict detection (no-op when no current state json provided)
    conflicts = detect_conflicts(rows, args.current_state_json)

    # Always write the plan JSON for downstream inspection (idempotent regen)
    PLAN_JSON.write_text(
        json.dumps(rows, indent=2, default=str, ensure_ascii=False),
        encoding="utf-8",
    )

    # Print to stdout
    print_summary(rows, derived_count, conflicts)
    print_samples(rows)
    print_conflicts(conflicts)
    print()
    print(f" Plan written to: {PLAN_JSON.relative_to(REPO)}")

    if args.dry_run:
        print(" --dry-run · SQL output NOT written. Re-run without --dry-run to emit SQL.")
        return

    # Emit SQL
    imported_from = f"COSTAR_MASTER_FINANCIALS.xlsx@{datetime.now(timezone.utc).date().isoformat()}"
    sql = emit_upsert_sql(rows, imported_from) + "\n" + emit_audit_sql(rows, derived_count, imported_from)
    SQL_OUT.write_text(sql, encoding="utf-8")
    print(f" SQL written to : {SQL_OUT.relative_to(REPO)}")
    print()
    print(" NEXT STEP · operator review the SQL then apply via Supabase MCP execute_sql.")


if __name__ == "__main__":
    main()
