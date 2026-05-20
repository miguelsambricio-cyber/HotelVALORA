"""Reproducible builder for the canonical HOTELVALORA CoStar masters.

Run from repo root:
    python services/costar/scripts/build_masters.py

Produces (v1.2 · 2026-05-14):
    services/costar/MASTER/COSTAR_MASTER_PAIS.xlsx
    services/costar/MASTER/COSTAR_MASTER_MERCADOS.xlsx
    services/costar/MASTER/COSTAR_MASTER_SUBMERCADOS.xlsx
    services/costar/MASTER/COSTAR_MASTER_HOTELES_POR_MERCADO.xlsx

Retired in v1.2:
    services/costar/MASTER/COSTAR_MASTER_CLASS.xlsx — chain-scale aggregates
    are no longer a separate granularity. chain_scale becomes an attribute
    on each hotel record in HOTELESperMARKET. The legacy file stays in
    MASTER/ for archival but is no longer regenerated.

CompSet datasets live in services/compset/ — operationally distinct
(hotel-specific underwriting workflows, not warehouse ingestion). See
docs/architecture/market-vs-underwriting-separation.md.

For per-file ingestion (sweep INPUT/, normalise, dedup, emit snapshot.json),
use `ingest.py` not this script. `build_masters.py` regenerates the empty
template workbooks with the canonical schema in DICTIONARY sheets.

Each workbook contains five sheets:
    1. DATA              — the canonical row table (one per granularity)
    2. DICTIONARY        — column name + type + required + notes
    3. INGESTION_LOG     — one row per processed file (history of ingestion runs)
    4. SOURCES_REGISTRY  — labelled source kinds with reliability tiers
    5. README            — operator instructions

The schema lives here so it is regenerable. The .xlsx is the frozen artefact
in git; this script regenerates it identically when run again. NORMALIZATION_VERSION
bumps on every breaking schema change.

Why four masters, not one
=========================
country / market / submarket / class have different schemas, different
granularity, different KPIs, different aggregation logic, and different
underwriting relevance. Mixing them once would force endless filters on every
analyst query. They share infrastructure (workspace + ingestion-meta block +
SOURCES_REGISTRY) but never share a DATA sheet. CompSet datasets live in
services/compset/ (operational workspace, not warehouse) since they are
hotel-specific underwriting outputs rather than market aggregates.
"""

from __future__ import annotations

import json
import os
import sys
import urllib.error
import urllib.request
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

NORMALIZATION_VERSION = "v1.2"
BUILT_AT = datetime.now(timezone.utc).isoformat(timespec="seconds")

HERE = Path(__file__).resolve().parent
ROOT = HERE.parent  # services/costar
MASTER_DIR = ROOT / "MASTER"


# ---------------------------------------------------------------------------
# Shared ingestion-meta block — IDENTICAL across all 4 masters and across
# the entire services/* family of ingestion workspaces.
# ---------------------------------------------------------------------------
INGESTION_META_COLUMNS: list[tuple[str, str, bool, str]] = [
    ("canonical_id", "uuid", True, "Primary key in the master. Generated on first insertion. Stable across edits."),
    ("ingestion_id", "text", True, "Matches the ingestion log entry that landed this row."),
    ("source_file", "text", True, "Original filename verbatim — e.g. 'CoStar_ES_Country_2026Q1.xlsx'."),
    ("source_kind", "enum", True, "costar | str | kalibri | curated | manual"),
    ("source_url", "url", False, "Public URL when applicable."),
    ("ingested_at", "timestamp", True, "ISO-8601 UTC. Auto-stamped by the ingestion run."),
    ("ingested_by", "email", True, "Operator email — audit trail."),
    ("normalization_version", "text", True, f"Schema/rules version applied at ingestion time. Today: {NORMALIZATION_VERSION}."),
    ("dedup_key", "text", True, "sha256 of the canonicalised dedup fields (see costar-normalization-rules.md §4)."),
    ("review_required", "bool", True, "TRUE when validation flagged something."),
    ("review_reason", "text", False, "Short label — e.g. 'missing_period', 'duplicate_candidate', 'fx_conversion_applied'."),
    ("ingestion_status", "enum", True, "ingested | under_review | superseded | rejected"),
    ("supersedes_id", "uuid", False, "Set when this row corrects/replaces a previously canonical row."),
    ("notes", "text", False, "Free-form operator notes."),
]


# ---------------------------------------------------------------------------
# Country-level schema (COSTAR_MASTER_PAIS)
# ---------------------------------------------------------------------------
COUNTRY_COLUMNS: list[tuple[str, str, bool, str]] = [
    ("country", "text", True, "ISO-3166-1 alpha-2 uppercase. Anchors the row to a nation."),
    ("country_name_display", "text", False, "Display form — 'España', 'Spain', 'Portugal'."),
    ("period_kind", "enum", True, "daily | weekly | monthly | quarterly | ytd | ltm | annual"),
    ("period_start", "date", True, "ISO-8601 YYYY-MM-DD. Start of the reporting window."),
    ("period_end", "date", True, "ISO-8601 YYYY-MM-DD. End of the reporting window (inclusive)."),
    ("currency", "text", True, "ISO-4217 — all monetary KPIs in this row are denominated here. EUR canonical; others trigger fx review."),
    ("occupancy_pct", "numeric", False, "Country-level occupancy %. Range [0, 100]."),
    ("adr", "numeric", False, "Average Daily Rate in the row's currency."),
    ("revpar", "numeric", False, "Revenue Per Available Room in the row's currency."),
    ("supply_rooms", "numeric", False, "Total rooms available in the period (room-nights)."),
    ("demand_rooms", "numeric", False, "Total rooms sold in the period (room-nights)."),
    ("revenue", "numeric", False, "Total accommodation revenue in the row's currency."),
    ("supply_yoy_pct", "numeric", False, "YoY % change in supply."),
    ("demand_yoy_pct", "numeric", False, "YoY % change in demand."),
    ("occupancy_yoy_pp", "numeric", False, "YoY change in occupancy in percentage points (+/-)."),
    ("adr_yoy_pct", "numeric", False, "YoY % change in ADR."),
    ("revpar_yoy_pct", "numeric", False, "YoY % change in RevPAR."),
    ("hotel_count", "int", False, "Number of hotels in scope for the period."),
    ("room_count_total", "int", False, "Snapshot room count at the start of the period."),
    ("pipeline_rooms", "int", False, "Rooms in pipeline (planned/under construction/pre-opening) — when reported."),
    ("pipeline_hotels", "int", False, "Hotels in pipeline."),
    ("tourism_arrivals", "numeric", False, "Inbound international tourist arrivals — when reported by the source."),
    ("tourism_arrivals_yoy_pct", "numeric", False, "YoY % change in international tourist arrivals."),
    ("gdp_growth_pct", "numeric", False, "GDP growth — macro context column, optional."),
    ("inflation_rate_pct", "numeric", False, "CPI inflation rate — macro context, optional."),
    # v1.3 passthrough · CoStar PAIS source columns
    ("rooms_under_construction", "int", False, "Rooms under construction — 'Habitaciones construcción'."),
    ("rooms_delivered_12m", "int", False, "Rooms delivered last 12 months — 'Habitaciones entregadas 12 m.'."),
]


# ---------------------------------------------------------------------------
# Market-level schema (COSTAR_MASTER_MERCADOS)
# ---------------------------------------------------------------------------
MARKET_COLUMNS: list[tuple[str, str, bool, str]] = [
    ("country", "text", True, "ISO-3166-1 alpha-2 uppercase."),
    ("market_name", "text", True, "Canonical market name — 'Madrid', 'Costa del Sol', 'Baleares'."),
    ("costar_market_code", "text", False, "CoStar / STR market code when available — preserves cross-product joinability."),
    ("market_uid", "uuid", False, "Resolved canonical market id — populated by future entity resolver."),
    ("region", "text", False, "Region or autonomous community — 'Madrid', 'Andalucía'."),
    ("period_kind", "enum", True, "daily | weekly | monthly | quarterly | ytd | ltm | annual"),
    ("period_start", "date", True, "ISO-8601 start."),
    ("period_end", "date", True, "ISO-8601 end (inclusive)."),
    ("currency", "text", True, "ISO-4217 — row's monetary denomination."),
    ("occupancy_pct", "numeric", False, "Market occupancy %. Range [0, 100]."),
    ("adr", "numeric", False, "Average Daily Rate in the row's currency."),
    ("revpar", "numeric", False, "RevPAR in the row's currency."),
    ("supply_rooms", "numeric", False, "Total rooms available (room-nights)."),
    ("demand_rooms", "numeric", False, "Total rooms sold (room-nights)."),
    ("revenue", "numeric", False, "Total accommodation revenue in the row's currency."),
    ("supply_yoy_pct", "numeric", False, "YoY % change in supply."),
    ("demand_yoy_pct", "numeric", False, "YoY % change in demand."),
    ("occupancy_yoy_pp", "numeric", False, "YoY change in occupancy (pp)."),
    ("adr_yoy_pct", "numeric", False, "YoY % change in ADR."),
    ("revpar_yoy_pct", "numeric", False, "YoY % change in RevPAR."),
    ("revpar_index_vs_country", "numeric", False, "Market RevPAR as % of national RevPAR — positioning indicator."),
    ("hotel_count", "int", False, "Hotels in market in scope for the period."),
    ("room_count_total", "int", False, "Room snapshot at period start."),
    ("pipeline_rooms", "int", False, "Pipeline rooms in this market."),
    ("pipeline_hotels", "int", False, "Pipeline hotels in this market."),
    ("seasonality_index", "numeric", False, "Period RevPAR / annual average RevPAR — context for sub-annual rows."),
    # v1.3 passthrough · CoStar MERCADO source columns
    ("rooms_under_construction", "int", False, "Rooms under construction — 'Habitaciones construcción'."),
    ("rooms_delivered_12m", "int", False, "Rooms delivered last 12m — 'Habitaciones entregadas 12 m.'."),
    ("buildings_built", "int", False, "Total buildings built — 'Edificios construidos'."),
    ("buildings_under_construction", "int", False, "Buildings under construction — 'Edificios construcción'."),
    ("avg_rooms_per_building", "numeric", False, "Avg rooms per building — 'Media habitaciones por edificio'."),
    ("inventory_growth_12m", "numeric", False, "Inventory growth last 12m — 'Crecimiento inventario 12 m.'."),
    ("rooms_opened_12m", "int", False, "Rooms opened last 12m — 'Habitaciones abiertas 12 m.'."),
    ("buildings_opened_12m", "int", False, "Buildings opened last 12m — 'Edificios abiertos 12 m.'."),
    # Spot (current-month) KPIs · separate from 12m-rolling
    ("occupancy_spot_pct", "numeric", False, "Current-month occupancy % — 'Ocupación' (spot)."),
    ("occupancy_yoy_spot_pp", "numeric", False, "YoY change in spot occupancy (pp) — 'Camb. ocupación (interanual)'."),
    ("adr_spot", "numeric", False, "Current-month ADR — 'ADR' (spot)."),
    ("adr_yoy_spot_pct", "numeric", False, "YoY % change in spot ADR — 'Camb. ADR (interanual)'."),
    ("revpar_spot", "numeric", False, "Current-month RevPAR — 'RevPAR' (spot)."),
    ("revpar_yoy_spot_pct", "numeric", False, "YoY % change in spot RevPAR — 'Camb. RevPAR (interanual)'."),
    # MERCADO sale + yield (CoStar institutional)
    ("market_sale_price_per_room", "numeric", False, "Market sale price per room — 'Precio venta de mercado/habitación'."),
    ("sales_volume_12m", "numeric", False, "Sales volume last 12m — 'Volumen ventas 12 m.'."),
    ("market_yield", "numeric", False, "Market yield/return — 'Rentabilidad mercado'."),
    # DataTable-only
    ("period_iso", "text", False, "ISO period for time-series rows · YYYY-MM derived from CoStar 'Periodo'."),
]


# ---------------------------------------------------------------------------
# Submarket-level schema (COSTAR_MASTER_SUBMERCADOS)
# ---------------------------------------------------------------------------
SUBMARKET_COLUMNS: list[tuple[str, str, bool, str]] = [
    ("country", "text", True, "ISO-3166-1 alpha-2."),
    ("market_name", "text", True, "Parent market — 'Madrid'."),
    ("submarket_name", "text", True, "Submarket — 'Salamanca', 'Madrid Centro', 'Aeropuerto'."),
    ("costar_submarket_code", "text", False, "CoStar / STR submarket code when available."),
    ("submarket_uid", "uuid", False, "Resolved canonical submarket id."),
    ("chain_scale", "enum", False, "luxury | upper_upscale | upscale | upper_midscale | midscale | economy | independent | all (when row aggregates the submarket overall)"),
    ("segment_type", "enum", False, "transient | group | contract | combined — KPI breakdown by guest segment."),
    ("period_kind", "enum", True, "daily | weekly | monthly | quarterly | ytd | ltm | annual"),
    ("period_start", "date", True, "ISO-8601 start."),
    ("period_end", "date", True, "ISO-8601 end (inclusive)."),
    ("currency", "text", True, "ISO-4217."),
    ("occupancy_pct", "numeric", False, "Submarket occupancy %."),
    ("adr", "numeric", False, "ADR in the row's currency."),
    ("revpar", "numeric", False, "RevPAR in the row's currency."),
    ("supply_rooms", "numeric", False, "Room-nights available."),
    ("demand_rooms", "numeric", False, "Room-nights sold."),
    ("revenue", "numeric", False, "Period accommodation revenue."),
    ("supply_yoy_pct", "numeric", False, "YoY % change in supply."),
    ("demand_yoy_pct", "numeric", False, "YoY % change in demand."),
    ("occupancy_yoy_pp", "numeric", False, "YoY change in occupancy (pp)."),
    ("adr_yoy_pct", "numeric", False, "YoY % change in ADR."),
    ("revpar_yoy_pct", "numeric", False, "YoY % change in RevPAR."),
    ("revpar_index_vs_market", "numeric", False, "Submarket RevPAR as % of parent market RevPAR."),
    ("hotel_count", "int", False, "Hotels in submarket in scope."),
    ("room_count_total", "int", False, "Room snapshot."),
    ("pipeline_rooms", "int", False, "Submarket pipeline rooms."),
    ("pipeline_hotels", "int", False, "Submarket pipeline hotels."),
    # v1.3 passthrough · CoStar SUBMERCADO source columns
    ("rooms_under_construction", "int", False, "Rooms under construction — 'Habitaciones construcción'."),
    ("rooms_delivered_12m", "int", False, "Rooms delivered last 12m — 'Habitaciones entregadas 12 m.'."),
]


# ---------------------------------------------------------------------------
# Class-level schema (COSTAR_MASTER_CLASS) — chain-scale time series
# ---------------------------------------------------------------------------
# Class (chain_scale) is its own granularity because:
#  - operators report KPIs by class within country AND within market
#  - aggregate by-class series are critical for sub-segment positioning
#  - operationally simpler than carrying chain_scale on every submarket row
# A class row either anchors at country level (market_name=null) OR at market
# level (market_name='Madrid'). Both legitimate; dedup_key distinguishes.
CLASS_COLUMNS: list[tuple[str, str, bool, str]] = [
    ("country", "text", True, "ISO-3166-1 alpha-2 uppercase."),
    ("market_name", "text", False, "Parent market — null = country-level class aggregate."),
    ("submarket_name", "text", False, "Parent submarket — typically null at this granularity."),
    ("chain_scale", "enum", True, "luxury | upper_upscale | upscale | upper_midscale | midscale | economy | independent | all_classes (aggregate)"),
    ("class_label_display", "text", False, "Operator-facing label — 'Luxury', 'Upper Upscale', 'Independent'."),
    ("segment_type", "enum", False, "transient | group | contract | combined — KPI breakdown by guest segment when reported."),
    ("period_kind", "enum", True, "daily | weekly | monthly | quarterly | ytd | ltm | annual"),
    ("period_start", "date", True, "ISO-8601 start."),
    ("period_end", "date", True, "ISO-8601 end (inclusive)."),
    ("currency", "text", True, "ISO-4217 — row's monetary denomination."),
    ("occupancy_pct", "numeric", False, "Class occupancy %. Range [0, 100]."),
    ("adr", "numeric", False, "ADR in row's currency. Range [10, 5000]."),
    ("revpar", "numeric", False, "RevPAR in row's currency. Range [5, 5000]."),
    ("supply_rooms", "numeric", False, "Room-nights available for this class within scope."),
    ("demand_rooms", "numeric", False, "Room-nights sold."),
    ("revenue", "numeric", False, "Period accommodation revenue."),
    ("supply_yoy_pct", "numeric", False, "YoY % change in supply."),
    ("demand_yoy_pct", "numeric", False, "YoY % change in demand."),
    ("occupancy_yoy_pp", "numeric", False, "YoY change in occupancy in percentage points."),
    ("adr_yoy_pct", "numeric", False, "YoY % change in ADR."),
    ("revpar_yoy_pct", "numeric", False, "YoY % change in RevPAR."),
    ("revpar_index_vs_country", "numeric", False, "When market_name=null: ignored. When market_name set: this class's RevPAR vs national class RevPAR."),
    ("revpar_index_vs_market", "numeric", False, "Class RevPAR as % of parent market overall RevPAR (`market_name` row in COSTAR_MASTER_MERCADOS)."),
    ("hotel_count", "int", False, "Hotels in this class within scope."),
    ("room_count_total", "int", False, "Snapshot room count at period start."),
    ("pipeline_rooms", "int", False, "Pipeline rooms for this class."),
    ("pipeline_hotels", "int", False, "Pipeline hotels for this class."),
]


# ---------------------------------------------------------------------------
# Hotel-by-market inventory schema (COSTAR_MASTER_HOTELES_POR_MERCADO · v1.2)
# ---------------------------------------------------------------------------
# Mirrors the schema doc at docs/intelligence/costar-hotels-by-market-schema.md.
# This is Dataset B (slowly-changing hotel inventory) — NOT a time series.
# Primary key: (country, market_name, hotel_id).
HOTELS_BY_MARKET_COLUMNS: list[tuple[str, str, bool, str]] = [
    # Identification
    ("country", "text", True, "ISO-3166-1 alpha-2 uppercase."),
    ("market_name", "text", True, "Canonical market — matches COSTAR_MASTER_MERCADOS.market_name."),
    ("submarket_name", "text", False, "Optional submarket — canonical neighborhood where the hotel sits."),
    ("hotel_id", "text", True, "Canonical hotel identifier. costar_<PROPERTY_ID> when present; otherwise h_<sha256[:16]>(country|market|name)."),
    ("hotel_id_synthetic", "boolean", True, "True when hotel_id was computed because the export lacked a CoStar PROPERTY ID."),
    ("name", "text", True, "Hotel display name."),
    ("brand", "text", False, "Brand affiliation. null for independent."),
    ("operator", "text", False, "Operating company. May differ from brand (white-label management)."),
    ("owner", "text", False, "Ownership entity when disclosed in the CoStar export."),
    # Property characteristics
    ("chain_scale", "enum", False, "luxury | upper_upscale | upscale | upper_midscale | midscale | economy | independent"),
    ("category", "text", False, "Star rating or local equivalent (e.g. '5-star', '4-star superior')."),
    ("segment_type", "enum", False, "business | leisure | extended_stay | resort | convention"),
    ("rooms_count", "int", False, "Total guest rooms."),
    ("year_opened", "int", False, "Original opening year."),
    ("year_last_renovated", "int", False, "Most recent renovation year."),
    ("total_floors", "int", False, "Building floors."),
    # Location
    ("address_line", "text", False, "Street + number."),
    ("postal_code", "text", False, "Local postal code."),
    ("latitude", "numeric", False, "Decimal degrees."),
    ("longitude", "numeric", False, "Decimal degrees."),
    ("neighborhood", "text", False, "Free-text neighborhood when finer than submarket."),
    # Facilities · amenities · scoring
    ("facilities", "text[]", False, "Normalised facility codes — meeting_space · pool · spa · fitness · restaurant · bar · parking · pet_friendly · business_center · accessibility · kids_club."),
    ("amenities", "text[]", False, "Free-text amenities not in the canonical facility enum."),
    ("meeting_space_sqm", "numeric", False, "Total meeting/event space when reported."),
    ("parking_spaces", "int", False, "Total parking spaces."),
    ("score_costar", "numeric", False, "CoStar property score when present (0–5 scale)."),
    ("score_external", "jsonb", False, "External-platform scores (Booking · Tripadvisor) keyed by source."),
    # ── v1.3 · CoStar institutional passthrough (2026-05-14) ──
    # Every CoStar source column has a slot in the master so administrator
    # review hits column-parity with the source export.
    ("costar_property_id", "text", False, "CoStar property id — 'ID del inmueble'."),
    ("city", "text", False, "City name — 'Ciudad' (typically Madrid in our drop)."),
    ("state_province", "text", False, "State or province — 'Estado o provincia'."),
    ("county", "text", False, "County — 'Condado'."),
    ("continent", "text", False, "Continent — 'Continente'."),
    ("subcontinent", "text", False, "Subcontinent — 'Subcontinente'."),
    ("submarket_cluster", "text", False, "Submarket cluster — 'Cluster de submercado'."),
    ("location_type", "text", False, "Hotel-location type — 'Tipo de ubicación del hotel' (airport · resort · urban · highway · suburban · etc)."),
    ("operation_type", "text", False, "Operation type — 'Tipo de operación'."),
    ("construction_state", "text", False, "Construction state — 'Estado de construcción'."),
    ("operating_state", "text", False, "Operating state — 'Estado de funcionamiento'."),
    ("operating_status", "text", False, "Status — 'Estado'."),
    ("expansion_rooms", "int", False, "Expansion rooms in pipeline — 'Habitaciones de expansión'."),
    ("expansion_status", "text", False, "Expansion status — 'Estado de expansión'."),
    ("expansion_status_date", "date", False, "Expansion status date — 'Fecha de estado de expansión'."),
    ("data_participation", "text", False, "Data participation flag — 'Participación en los datos'."),
    ("meeting_contig_sqm", "numeric", False, "Contiguous meeting space max — 'Espacio de reunión contig. máx.'."),
    ("all_inclusive_price", "numeric", False, "All-inclusive price — 'Precio todo incluido'."),
    ("modules", "text", False, "All modules — 'Todos los módulos'."),
    ("main_axes", "text", False, "Main axes — 'Ejes principales'."),
    ("parking_per_room", "numeric", False, "Parking spaces per room — 'Plazas de parking/habitación'."),
    ("parking_ratio", "numeric", False, "Parking ratio — 'Ratio de parking'."),
    ("fund", "text", False, "Fund — 'Fondo'."),
    ("property_type", "text", False, "Property type — 'Tipo'."),
    ("eco_rating", "text", False, "Eco/sustainability rating — 'Clasificación ecológica'."),
    ("rentable_sqm_available", "numeric", False, "Available rentable sqm — 'm² disp.'."),
    ("owner_rep", "text", False, "Owner representative — 'Repres. del propietario'."),
    ("owner_rep_contact", "text", False, "Owner-rep contact — 'Contacto repres. del propietario'."),
    ("sales_company", "text", False, "Sales company — 'Empresa de ventas'."),
    ("sales_contact", "text", False, "Sales contact — 'Contacto de ventas'."),
    ("asking_sale_price", "numeric", False, "Asking sale price — 'Precio de venta'."),
    ("price_per_sqm_min", "numeric", False, "Min price per sqm — 'Precio mín./m²'."),
    ("price_per_sqm_max", "numeric", False, "Max price per sqm — 'Precio máx./m²'."),
    ("price_per_room", "numeric", False, "Price per room — 'Precio/habitación'."),
    ("sale_status", "text", False, "Sale status — 'Estado de venta'."),
    ("leased_pct", "numeric", False, "Leased percentage — '% alquilado'."),
    ("building_park", "text", False, "Building park — 'Parque de edificios'."),
    ("construction_month", "text", False, "Construction month — 'Mes de construcción'."),
    ("renovation_month", "text", False, "Renovation month — 'Mes de reform.'."),
    ("zoning", "text", False, "Zoning — 'Zonificación'."),
    ("registered_owner", "text", False, "Registered owner — 'Propietario registrado'."),
    ("construction_start", "date", False, "Construction start — 'Inicio de construcción'."),
    ("loan_origination_amount", "numeric", False, "Loan origination amount — 'Importe de originación'."),
    ("loan_origination_date", "date", False, "Loan origination date — 'Fecha de originación'."),
    ("loan_maturity_date", "date", False, "Loan maturity date — 'Fecha de vencimiento'."),
    ("lender", "text", False, "Lender — 'Prestamista'."),
    ("interest_rate_type", "text", False, "Interest rate type — 'Tipo de interés'."),
    ("collateral_type", "text", False, "Collateral type — 'Tipo de garantía'."),
    ("loan_type", "text", False, "Loan type — 'Tipo de préstamo'."),
    ("co_owners", "text", False, "Co-owners — 'Multipropietarios'."),
    ("rent_estimated", "numeric", False, "Estimated rent — 'Alquiler estimado'."),
    ("rent_listing", "numeric", False, "Listing rent — 'Alquiler'."),
    ("rent_min_per_sqm_month", "numeric", False, "Min rent per sqm/month — 'Precio de alquiler mín./m²/mes'."),
    ("rent_max_per_sqm_month", "numeric", False, "Max rent per sqm/month — 'Precio de alquiler máx./m²/mes'."),
    ("rent_undisclosed", "text", False, "Rent undisclosed flag — 'Precio de alquiler no divulgado'."),
    ("sale_price_undisclosed", "text", False, "Sale-price undisclosed flag — 'Precio de venta no divulgado'."),
    ("multi_owner_sale_only", "text", False, "Multi-owner sale only flag — 'Venta solo de multipropietarios'."),
    ("price_per_sqm_undisclosed", "text", False, "Price/sqm undisclosed flag — 'Precio/superficie no divulgado(a)'."),
    ("map_page", "text", False, "Map page — 'Página de mapa'."),
    ("rent_listing_min_month", "numeric", False, "Listing rent min/month — 'Precio de salida de alquiler mín./mes'."),
    ("rent_listing_max_month", "numeric", False, "Listing rent max/month — 'Precio de salida de alquiler máx./mes'."),
    ("rent_listing_undisclosed", "text", False, "Listing rent undisclosed flag — 'Precio de salida de alquiler no divulgado'."),
    # FEMA flood-plain (US-only · always null in Spain · captured for institutional parity)
    ("flood_risk", "text", False, "FEMA flood risk — 'Riesgo de inundación' (US-only)."),
    ("flood_zone", "text", False, "FEMA flood zone — 'Zona de inundación' (US-only)."),
    ("fema_map_date", "date", False, "FEMA map date — 'Fecha del mapa de la FEMA' (US-only)."),
    ("fema_map_id", "text", False, "FEMA map id — 'Identificador de mapa de la FEMA' (US-only)."),
    ("firm_map_id", "text", False, "FIRM map id — 'ID del mapa FIRM' (US-only)."),
    ("firm_panel_number", "text", False, "FIRM panel number — 'Número de panel de FIRM' (US-only)."),
    ("flood_special_risk_zone", "text", False, "Special flood risk zone — 'Zona especial de riesgo de inundación' (US-only)."),
    ("flood_plain_zone", "text", False, "Floodplain zone — 'Zona de llanura inundable' (US-only)."),
    # Property additions (v1.3 · 2026-05-14)
    ("floors_above_ground", "int", False, "Floors above grade — CoStar 'Plantas sobre rasante'."),
    ("floors_below_ground", "int", False, "Floors below grade (basements) — CoStar 'Plantas bajo rasante'."),
    ("gross_building_sqm", "numeric", False, "Gross built area (m²) — CoStar 'Superficie alquilable del inmueble/SBA'."),
    ("lot_size_sqm", "numeric", False, "Plot/lot area (m²) — CoStar 'Terreno (m²)'."),
    ("typical_floor_sqm", "numeric", False, "Typical-floor area (m²) — CoStar 'Planta tipo (m²)'."),
    ("meeting_rooms_count", "int", False, "Count of distinct meeting rooms — CoStar 'Salas de reuniones'."),
    ("last_sale_date", "date", False, "Date of last asset sale — CoStar 'Fecha de la última venta'."),
    ("last_sale_price_eur", "numeric", False, "Last sale price in EUR — CoStar 'Último precio de venta'."),
    ("catastro_id", "text", False, "Spanish cadastre identifier (manual entry today; Catastro API later)."),
    # Booking enrichment merge (v1.3 · sourced from RapidAPI booking-com15)
    ("booking_hotel_id", "int", False, "Booking.com property id from manual_enrichment merge."),
    ("booking_url", "text", False, "Booking.com canonical URL when matched."),
    ("review_score", "numeric", False, "Booking overall review score (0–10)."),
    ("review_count", "int", False, "Booking total review count."),
    ("location_score", "numeric", False, "Booking sub-score · hotel_location."),
    ("comfort_score", "numeric", False, "Booking sub-score · hotel_comfort."),
    ("cleanliness_score", "numeric", False, "Booking sub-score · hotel_clean."),
    ("staff_score", "numeric", False, "Booking sub-score · hotel_staff."),
    ("value_score", "numeric", False, "Booking sub-score · hotel_value."),
    ("facilities_score", "numeric", False, "Booking sub-score · hotel_facilities."),
    ("has_bar", "bool", False, "Booking · bar present."),
    ("has_restaurant", "bool", False, "Booking · restaurant present."),
    ("has_rooftop", "bool", False, "Booking · rooftop present."),
    ("has_gym", "bool", False, "Booking · gym present."),
    ("has_spa", "bool", False, "Booking · spa present."),
    ("has_pool", "bool", False, "Booking · pool present."),
    ("has_parking", "bool", False, "Booking · parking present."),
    ("has_meeting", "bool", False, "Booking · meeting rooms present."),
    ("booking_facilities_count", "int", False, "Count of raw Booking facility strings captured."),
    ("booking_room_types_count", "int", False, "Count of distinct Booking room types."),
    ("check_in_time", "text", False, "Booking policy · check-in time."),
    ("check_out_time", "text", False, "Booking policy · check-out time."),
    ("pet_policy", "text", False, "Booking policy · pets."),
    ("cancellation_policy", "text", False, "Booking policy · cancellation/prepayment."),
    ("smoking_policy", "text", False, "Booking policy · smoking."),
    ("coords_source", "enum", False, "Provenance of latitude/longitude · CoStar | Booking | null."),
    ("enrichment_sources", "text", False, "Comma-separated provenance · rapidapi_booking · manual_operator · google_places."),
    ("enrichment_confidence", "numeric", False, "0–1 match confidence at enrichment time."),
    ("profile_completeness_score", "numeric", False, "0–100 % of priority enrichment fields populated."),
    ("last_scraped_at", "timestamp", False, "ISO timestamp of last enrichment scrape."),
    # ── HotelVALORA enrichment passthrough (v1.4 · 2026-05-20) ──
    # Supabase hotel_canonical bridge + Google Places / Wikidata enrichment fields.
    ("phone", "text", False, "Hotel contact phone (E.164 when normalized) · sourced from Google Places fallback."),
    ("website_url", "text", False, "Hotel website URL · Google Places fallback or chain registry."),
    ("google_place_id", "text", False, "Google Places id (Places API New v1) · self-authoritative."),
    ("wikidata_qid", "text", False, "Wikidata Q-id when matched via SPARQL · cross-reference utility."),
    ("canonical_id_supabase", "text", False, "UUID v4 bridging to Supabase public.hotel_canonical(id) — institutional canonical layer."),
    ("data_quality_tier", "text", False, "Data quality tier · gold / silver / bronze / quarantined · per Supabase hotel_canonical."),
    # Commercial context
    ("competitive_set_ids", "text[]", False, "Sibling hotel_ids in the property's competitive set."),
    ("transactions_history_ref", "text", False, "Foreign key into HOTEL_TRANSACCIONES_MASTER when the hotel has known transaction history."),
    ("notes", "text", False, "Operator-curated free-text notes."),
]


SOURCES_REGISTRY = [
    ("costar", "CoStar Hospitality Export", "A", "Authoritative — institutional, paid product. The canonical source."),
    ("str", "STR (CoStar subsidiary)", "A", "STR-direct exports — same provenance as CoStar; preserve attribution."),
    ("kalibri", "Kalibri Labs export", "B", "Operator analytics — useful for cross-validation."),
    ("curated", "Curated internal compilation", "A", "Hand-maintained HOTELVALORA spreadsheets — treat as ground-truth."),
    ("manual", "Manual operator entry", "C", "Operator-typed row. Audit trail in ingested_by."),
]


# ---------------------------------------------------------------------------
# Sheet builders (identical pattern across the four masters)
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill(start_color="FF003B2A", end_color="FF003B2A", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFD7F587", name="Calibri", size=11)
WRAP = Alignment(wrap_text=True, vertical="top")


def style_header(ws, header_row=1):
    for cell in ws[header_row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.freeze_panes = f"A{header_row + 1}"


def autosize(ws, max_width=44):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), max_width)


def build_data_sheet(ws, title: str, domain_columns, data_rows: list[dict] | None = None):
    ws.title = title
    columns = domain_columns + INGESTION_META_COLUMNS
    column_names = [name for (name, _t, _r, _n) in columns]
    ws.append(column_names)
    style_header(ws)
    if data_rows:
        for row in data_rows:
            ws.append([row.get(col) for col in column_names])
    autosize(ws)


def build_dictionary_sheet(wb, title: str, domain_columns):
    ws = wb.create_sheet(title)
    ws.append(["section", "column", "type", "required", "notes"])
    style_header(ws)
    for (name, t, req, notes) in domain_columns:
        ws.append(["domain", name, t, "YES" if req else "no", notes])
    for (name, t, req, notes) in INGESTION_META_COLUMNS:
        ws.append(["ingestion_meta", name, t, "YES" if req else "no", notes])
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = WRAP
    autosize(ws, max_width=80)


def build_ingestion_log_sheet(wb):
    ws = wb.create_sheet("INGESTION_LOG")
    ws.append([
        "ingestion_id", "source_file", "source_kind",
        "started_at", "completed_at",
        "rows_seen", "rows_inserted", "rows_updated", "rows_skipped",
        "rows_flagged_review", "rows_failed",
        "operator_email", "normalization_version", "outcome", "notes",
    ])
    style_header(ws)
    autosize(ws)


def build_sources_registry_sheet(wb):
    ws = wb.create_sheet("SOURCES_REGISTRY")
    ws.append(["source_kind", "label", "reliability_tier", "notes"])
    style_header(ws)
    for row in SOURCES_REGISTRY:
        ws.append(row)
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = WRAP
    autosize(ws, max_width=80)


def build_readme_sheet(wb, dataset_label: str, domain_label: str, schema_doc: str):
    ws = wb.create_sheet("README")
    ws.append([f"HOTELVALORA · {dataset_label}"])
    ws.cell(row=1, column=1).font = Font(bold=True, size=16, color="FF003B2A")

    body = [
        "",
        f"Canonical institutional dataset for {domain_label}. Append-only — never overwrite an existing canonical row.",
        f"Built {BUILT_AT}. Normalization version: {NORMALIZATION_VERSION}.",
        "",
        "Sister workbooks (services/costar/MASTER/):",
        "  • COSTAR_MASTER_PAIS.xlsx                  — country aggregates (Dataset A)",
        "  • COSTAR_MASTER_MERCADOS.xlsx              — market aggregates (Dataset A)",
        "  • COSTAR_MASTER_SUBMERCADOS.xlsx           — submarket aggregates (Dataset A)",
        "  • COSTAR_MASTER_HOTELES_POR_MERCADO.xlsx   — hotel inventory (Dataset B · v1.2)",
        "  • COSTAR_MASTER_CLASS.xlsx                 — LEGACY (chain-scale aggregates · retired v1.2)",
        "",
        "  CompSet workbooks now live in the OPERATIONAL workspace at services/compset/",
        "  (different purpose — hotel-specific underwriting outputs, not warehouse ingestion).",
        "",
        "Operating contract:",
        "  1. The Data Ingestion Agent appends new rows. It NEVER edits in place.",
        "  2. Corrections → insert a new row with",
        "     supersedes_id = <old canonical_id> and ingestion_status='superseded'.",
        "  3. Manual edits follow the same contract — track who, when, why",
        "     in ingested_by + notes.",
        "",
        "Sheets:",
        "  • DATA              — canonical row corpus",
        "  • DICTIONARY        — column schema + required flags + notes",
        "  • INGESTION_LOG     — one row per processed import file",
        "  • SOURCES_REGISTRY  — labels + reliability tiers per source_kind",
        "  • README            — this sheet",
        "",
        "Reference docs:",
        f"  • {schema_doc}",
        "  • docs/intelligence/costar-master-dataset-architecture.md",
        "  • docs/intelligence/costar-normalization-rules.md",
        "  • docs/intelligence/costar-ingestion-workflow.md",
        "",
        "Regenerating this workbook:",
        "  python services/costar/scripts/build_masters.py",
        "  (Idempotent. Run when the schema evolves; commit both the script and the regenerated .xlsx.)",
    ]
    for line in body:
        ws.append([line])
    ws.column_dimensions["A"].width = 110


def build_workbook(
    dataset_label: str,
    domain_columns,
    data_sheet_title: str,
    schema_doc: str,
    output_path: Path,
    data_rows: list[dict] | None = None,
):
    wb = Workbook()
    build_data_sheet(wb.active, data_sheet_title, domain_columns, data_rows)
    build_dictionary_sheet(wb, "DICTIONARY", domain_columns)
    build_ingestion_log_sheet(wb)
    build_sources_registry_sheet(wb)
    build_readme_sheet(wb, dataset_label, data_sheet_title.lower(), schema_doc)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    rows_written = len(data_rows) if data_rows else 0
    print(f"[build_masters] wrote {output_path.relative_to(ROOT.parent.parent)} · {rows_written} rows")


# ---------------------------------------------------------------------------
# Data extraction from snapshot.json + Supabase Storage enrichment
# ---------------------------------------------------------------------------
SNAPSHOT_PATH = ROOT / "MASTER" / "snapshot.json"
SUPABASE_BUCKET = "costar-master"
ENRICHMENT_PREFIX = "manual_enrichment"


def _read_env_local() -> dict[str, str]:
    """Lightweight parser for apps/web/.env.local · KEY=VALUE per line.

    Strips surrounding double-quotes from values (Vercel CLI dumps quote
    secrets) and treats literal empty strings ('""', "''") as absent so
    downstream code doesn't try to fetch from a quoted-empty URL.
    """
    env_path = ROOT.parent.parent / "apps" / "web" / ".env.local"
    if not env_path.exists():
        return {}
    out: dict[str, str] = {}
    for line in env_path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        v = v.strip()
        if (len(v) >= 2 and v[0] == v[-1] and v[0] in ('"', "'")):
            v = v[1:-1]
        if v == "":
            continue
        out[k.strip()] = v
    return out


def _supabase_env() -> tuple[str | None, str | None]:
    url = os.environ.get("NEXT_PUBLIC_SUPABASE_URL")
    key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
    if url and key:
        return url, key
    env = _read_env_local()
    return env.get("NEXT_PUBLIC_SUPABASE_URL"), env.get("SUPABASE_SERVICE_ROLE_KEY")


def _load_enrichment_from_storage() -> dict[str, dict]:
    """Returns { hotel_id: payload } · empty dict if Supabase not reachable."""
    url, key = _supabase_env()
    if not url or not key:
        print("[build_masters] Supabase env not set · skipping Booking enrichment merge")
        return {}
    try:
        req = urllib.request.Request(
            f"{url}/storage/v1/object/list/{SUPABASE_BUCKET}",
            data=json.dumps({"prefix": ENRICHMENT_PREFIX, "limit": 1000, "offset": 0}).encode("utf-8"),
            headers={"Authorization": f"Bearer {key}", "apikey": key, "Content-Type": "application/json"},
            method="POST",
        )
        with urllib.request.urlopen(req, timeout=30) as r:
            items = json.loads(r.read())
    except urllib.error.URLError as e:
        print(f"[build_masters] Supabase list failed: {e} · skipping enrichment merge")
        return {}
    out: dict[str, dict] = {}
    for item in items:
        name = item.get("name", "")
        if not name.endswith(".json"):
            continue
        hotel_id = name[: -len(".json")]
        try:
            dreq = urllib.request.Request(
                f"{url}/storage/v1/object/{SUPABASE_BUCKET}/{ENRICHMENT_PREFIX}/{name}",
                headers={"Authorization": f"Bearer {key}", "apikey": key},
            )
            with urllib.request.urlopen(dreq, timeout=30) as r:
                out[hotel_id] = json.loads(r.read())
        except (urllib.error.URLError, json.JSONDecodeError):
            continue
    print(f"[build_masters] loaded {len(out)} Booking enrichment records from Supabase Storage")
    return out


def _round_int(v: Any) -> int | None:
    if v is None:
        return None
    try:
        return round(float(v))
    except (TypeError, ValueError):
        return None


def _yn(v: Any) -> str:
    return "Y" if v else "N"


def _meta_block(snapshot: dict, hotel_id_or_index: Any, source_kind: str = "costar") -> dict:
    """Common ingestion-meta block · same shape across all masters."""
    return {
        "canonical_id": str(hotel_id_or_index),
        "ingestion_id": snapshot.get("ingestion_batch_id", ""),
        "source_file": "snapshot.json",
        "source_kind": source_kind,
        "source_url": "",
        "ingested_at": snapshot.get("generated_at", BUILT_AT),
        "ingested_by": "cli:build_masters",
        "normalization_version": (snapshot.get("batch") or {}).get("normalization_version", NORMALIZATION_VERSION),
        "dedup_key": "",
        "review_required": "TRUE" if hotel_id_or_index else "FALSE",
        "review_reason": "",
        "ingestion_status": "ingested",
        "supersedes_id": "",
        "notes": "",
    }


def _market_row_to_pais(r: dict, snapshot: dict) -> dict:
    """Map snapshot market_snapshots row (granularity=country_listing) →
    COSTAR_MASTER_PAIS canonical row."""
    return {
        "country": r.get("country"),
        "country_name_display": r.get("country") or r.get("market_name"),
        "period_kind": "ltm",
        "period_start": "",
        "period_end": "",
        "currency": "EUR",
        "occupancy_pct": (r.get("occupancy_12m") or 0) * 100 if r.get("occupancy_12m") is not None else None,
        "adr": r.get("adr_12m"),
        "revpar": r.get("revpar_12m"),
        "supply_rooms": _round_int(r.get("supply_12m")),
        "demand_rooms": _round_int(r.get("demand_12m")),
        "revenue": _round_int(r.get("revenue_12m")),
        "supply_yoy_pct": (r.get("supply_yoy_12m") or 0) * 100 if r.get("supply_yoy_12m") is not None else None,
        "demand_yoy_pct": (r.get("demand_yoy_12m") or 0) * 100 if r.get("demand_yoy_12m") is not None else None,
        "occupancy_yoy_pp": (r.get("occupancy_yoy_12m") or 0) * 100 if r.get("occupancy_yoy_12m") is not None else None,
        "adr_yoy_pct": (r.get("adr_yoy_12m") or 0) * 100 if r.get("adr_yoy_12m") is not None else None,
        "revpar_yoy_pct": (r.get("revpar_yoy_12m") or 0) * 100 if r.get("revpar_yoy_12m") is not None else None,
        "hotel_count": None,
        "room_count_total": _round_int(r.get("rooms_inventory")),
        "pipeline_rooms": _round_int(r.get("rooms_under_construction")),
        "pipeline_hotels": None,
        # v1.3 passthrough
        "rooms_under_construction": _round_int(r.get("rooms_under_construction")),
        "rooms_delivered_12m": _round_int(r.get("rooms_delivered_12m")),
        **_meta_block(snapshot, r.get("country"), "costar"),
    }


def _market_row_to_mercado(r: dict, snapshot: dict) -> dict:
    return {
        "country": r.get("country"),
        "market_name": r.get("market_name"),
        "costar_market_code": "",
        "market_uid": "",
        "region": "",
        "period_kind": "ltm" if not r.get("period") else "monthly",
        "period_start": r.get("period") or "",
        "period_end": r.get("period") or "",
        "currency": "EUR",
        "occupancy_pct": (r.get("occupancy_12m") or 0) * 100 if r.get("occupancy_12m") is not None else None,
        "adr": r.get("adr_12m"),
        "revpar": r.get("revpar_12m"),
        "supply_rooms": _round_int(r.get("supply_12m")),
        "demand_rooms": _round_int(r.get("demand_12m")),
        "revenue": _round_int(r.get("revenue_12m")),
        "supply_yoy_pct": (r.get("supply_yoy_12m") or 0) * 100 if r.get("supply_yoy_12m") is not None else None,
        "demand_yoy_pct": (r.get("demand_yoy_12m") or 0) * 100 if r.get("demand_yoy_12m") is not None else None,
        "occupancy_yoy_pp": (r.get("occupancy_yoy_12m") or 0) * 100 if r.get("occupancy_yoy_12m") is not None else None,
        "adr_yoy_pct": (r.get("adr_yoy_12m") or 0) * 100 if r.get("adr_yoy_12m") is not None else None,
        "revpar_yoy_pct": (r.get("revpar_yoy_12m") or 0) * 100 if r.get("revpar_yoy_12m") is not None else None,
        "revpar_index_vs_country": None,
        "hotel_count": None,
        "room_count_total": _round_int(r.get("rooms_inventory")),
        "pipeline_rooms": _round_int(r.get("rooms_under_construction")),
        "pipeline_hotels": None,
        "seasonality_index": None,
        # v1.3 passthrough
        "rooms_under_construction": _round_int(r.get("rooms_under_construction")),
        "rooms_delivered_12m": _round_int(r.get("rooms_delivered_12m")),
        "buildings_built": _round_int(r.get("buildings_built")),
        "buildings_under_construction": _round_int(r.get("buildings_under_construction")),
        "avg_rooms_per_building": r.get("avg_rooms_per_building"),
        "inventory_growth_12m": (r.get("inventory_growth_12m") or 0) * 100 if r.get("inventory_growth_12m") is not None else None,
        "rooms_opened_12m": _round_int(r.get("rooms_opened_12m")),
        "buildings_opened_12m": _round_int(r.get("buildings_opened_12m")),
        "occupancy_spot_pct": (r.get("occupancy_spot") or 0) * 100 if r.get("occupancy_spot") is not None else None,
        "occupancy_yoy_spot_pp": (r.get("occupancy_yoy_spot") or 0) * 100 if r.get("occupancy_yoy_spot") is not None else None,
        "adr_spot": r.get("adr_spot"),
        "adr_yoy_spot_pct": (r.get("adr_yoy_spot") or 0) * 100 if r.get("adr_yoy_spot") is not None else None,
        "revpar_spot": r.get("revpar_spot"),
        "revpar_yoy_spot_pct": (r.get("revpar_yoy_spot") or 0) * 100 if r.get("revpar_yoy_spot") is not None else None,
        "market_sale_price_per_room": r.get("market_sale_price_per_room"),
        "sales_volume_12m": _round_int(r.get("sales_volume_12m")),
        "market_yield": (r.get("market_yield") or 0) * 100 if r.get("market_yield") is not None else None,
        "period_iso": r.get("period"),
        **_meta_block(snapshot, f"{r.get('country')}|{r.get('market_name')}|{r.get('period') or ''}", "costar"),
    }


def _market_row_to_submercado(r: dict, snapshot: dict) -> dict:
    return {
        "country": r.get("country"),
        "market_name": r.get("market_name") or "Madrid",  # Submarket parent
        "submarket_name": r.get("submarket_name"),
        "costar_submarket_code": "",
        "submarket_uid": "",
        "chain_scale": "",
        "segment_type": "",
        "period_kind": "ltm",
        "period_start": "",
        "period_end": "",
        "currency": "EUR",
        "occupancy_pct": (r.get("occupancy_12m") or 0) * 100 if r.get("occupancy_12m") is not None else None,
        "adr": r.get("adr_12m"),
        "revpar": r.get("revpar_12m"),
        "supply_rooms": _round_int(r.get("supply_12m")),
        "demand_rooms": _round_int(r.get("demand_12m")),
        "revenue": _round_int(r.get("revenue_12m")),
        "supply_yoy_pct": (r.get("supply_yoy_12m") or 0) * 100 if r.get("supply_yoy_12m") is not None else None,
        "demand_yoy_pct": (r.get("demand_yoy_12m") or 0) * 100 if r.get("demand_yoy_12m") is not None else None,
        "occupancy_yoy_pp": (r.get("occupancy_yoy_12m") or 0) * 100 if r.get("occupancy_yoy_12m") is not None else None,
        "adr_yoy_pct": (r.get("adr_yoy_12m") or 0) * 100 if r.get("adr_yoy_12m") is not None else None,
        "revpar_yoy_pct": (r.get("revpar_yoy_12m") or 0) * 100 if r.get("revpar_yoy_12m") is not None else None,
        "revpar_index_vs_market": None,
        "hotel_count": None,
        "room_count_total": _round_int(r.get("rooms_inventory")),
        "pipeline_rooms": _round_int(r.get("rooms_under_construction")),
        "pipeline_hotels": None,
        # v1.3 passthrough
        "rooms_under_construction": _round_int(r.get("rooms_under_construction")),
        "rooms_delivered_12m": _round_int(r.get("rooms_delivered_12m")),
        **_meta_block(snapshot, f"{r.get('submarket_name')}", "costar"),
    }


def _hotel_to_row(h: dict, enrichment: dict, snapshot: dict) -> dict:
    """Map snapshot hotel + Booking enrichment → COSTAR_MASTER_HOTELESperMARKET row."""
    e = enrichment.get(h.get("hotel_id"), {}) or {}
    p = e.get("profile") or {}
    meta = e.get("_enrichment_meta") or {}
    fnb = p.get("fnb") or {}
    gym = p.get("gym") or {}
    spa = p.get("spa") or {}
    pool = p.get("pool") or {}
    parking = p.get("parking") or {}
    meeting = p.get("meeting_rooms") or {}
    rooftop = p.get("rooftop") or {}
    hmeta = h.get("_meta") or {}
    needs_review = hmeta.get("needs_review") or []

    lat = h.get("latitude") if h.get("latitude") is not None else p.get("latitude")
    lng = h.get("longitude") if h.get("longitude") is not None else p.get("longitude")
    coords_source = (
        "CoStar" if h.get("latitude") is not None
        else "Booking" if p.get("latitude") is not None
        else ""
    )

    return {
        # Identification
        "country": h.get("country"),
        "market_name": h.get("market_name"),
        "submarket_name": h.get("submarket_name"),
        "hotel_id": h.get("hotel_id"),
        "hotel_id_synthetic": "TRUE" if h.get("hotel_id_synthetic") else "FALSE",
        "name": h.get("name"),
        "brand": h.get("brand"),
        "operator": h.get("operator"),
        "owner": h.get("owner"),
        # Property characteristics
        "chain_scale": h.get("chain_scale"),
        "category": h.get("category"),
        "segment_type": h.get("segment_type"),
        "rooms_count": h.get("rooms_count"),
        "year_opened": h.get("year_opened"),
        "year_last_renovated": h.get("year_last_renovated"),
        "total_floors": h.get("total_floors"),
        # Location
        "address_line": h.get("address_line"),
        "postal_code": h.get("postal_code"),
        "latitude": lat,
        "longitude": lng,
        "neighborhood": h.get("neighborhood"),
        # Facilities · amenities · scoring (CoStar)
        "facilities": ", ".join(h.get("facilities") or []),
        "amenities": ", ".join(h.get("amenities") or []),
        "meeting_space_sqm": _round_int(h.get("meeting_space_sqm")),
        "parking_spaces": h.get("parking_spaces"),
        "score_costar": h.get("score_costar"),
        "score_external": json.dumps(h.get("score_external") or {}) if h.get("score_external") else "",
        # v1.3 · CoStar institutional passthrough · direct from row
        "costar_property_id": h.get("costar_property_id"),
        "city": h.get("city"),
        "state_province": h.get("state_province"),
        "county": h.get("county"),
        "continent": h.get("continent"),
        "subcontinent": h.get("subcontinent"),
        "submarket_cluster": h.get("submarket_cluster"),
        "location_type": h.get("location_type"),
        "operation_type": h.get("operation_type"),
        "construction_state": h.get("construction_state"),
        "operating_state": h.get("operating_state"),
        "operating_status": h.get("operating_status"),
        "expansion_rooms": h.get("expansion_rooms"),
        "expansion_status": h.get("expansion_status"),
        "expansion_status_date": h.get("expansion_status_date"),
        "data_participation": h.get("data_participation"),
        "meeting_contig_sqm": _round_int(h.get("meeting_contig_sqm")),
        "all_inclusive_price": h.get("all_inclusive_price"),
        "modules": h.get("modules"),
        "main_axes": h.get("main_axes"),
        "parking_per_room": h.get("parking_per_room"),
        "parking_ratio": h.get("parking_ratio"),
        "fund": h.get("fund"),
        "property_type": h.get("property_type"),
        "eco_rating": h.get("eco_rating"),
        "rentable_sqm_available": _round_int(h.get("rentable_sqm_available")),
        "owner_rep": h.get("owner_rep"),
        "owner_rep_contact": h.get("owner_rep_contact"),
        "sales_company": h.get("sales_company"),
        "sales_contact": h.get("sales_contact"),
        "asking_sale_price": h.get("asking_sale_price"),
        "price_per_sqm_min": h.get("price_per_sqm_min"),
        "price_per_sqm_max": h.get("price_per_sqm_max"),
        "price_per_room": h.get("price_per_room"),
        "sale_status": h.get("sale_status"),
        "leased_pct": h.get("leased_pct"),
        "building_park": h.get("building_park"),
        "construction_month": h.get("construction_month"),
        "renovation_month": h.get("renovation_month"),
        "zoning": h.get("zoning"),
        "registered_owner": h.get("registered_owner"),
        "construction_start": h.get("construction_start"),
        "loan_origination_amount": h.get("loan_origination_amount"),
        "loan_origination_date": h.get("loan_origination_date"),
        "loan_maturity_date": h.get("loan_maturity_date"),
        "lender": h.get("lender"),
        "interest_rate_type": h.get("interest_rate_type"),
        "collateral_type": h.get("collateral_type"),
        "loan_type": h.get("loan_type"),
        "co_owners": h.get("co_owners"),
        "rent_estimated": h.get("rent_estimated"),
        "rent_listing": h.get("rent_listing"),
        "rent_min_per_sqm_month": h.get("rent_min_per_sqm_month"),
        "rent_max_per_sqm_month": h.get("rent_max_per_sqm_month"),
        "rent_undisclosed": h.get("rent_undisclosed"),
        "sale_price_undisclosed": h.get("sale_price_undisclosed"),
        "multi_owner_sale_only": h.get("multi_owner_sale_only"),
        "price_per_sqm_undisclosed": h.get("price_per_sqm_undisclosed"),
        "map_page": h.get("map_page"),
        "rent_listing_min_month": h.get("rent_listing_min_month"),
        "rent_listing_max_month": h.get("rent_listing_max_month"),
        "rent_listing_undisclosed": h.get("rent_listing_undisclosed"),
        "flood_risk": h.get("flood_risk"),
        "flood_zone": h.get("flood_zone"),
        "fema_map_date": h.get("fema_map_date"),
        "fema_map_id": h.get("fema_map_id"),
        "firm_map_id": h.get("firm_map_id"),
        "firm_panel_number": h.get("firm_panel_number"),
        "flood_special_risk_zone": h.get("flood_special_risk_zone"),
        "flood_plain_zone": h.get("flood_plain_zone"),
        # v1.3 additions
        "floors_above_ground": h.get("floors_above_ground"),
        "floors_below_ground": h.get("floors_below_ground"),
        "gross_building_sqm": _round_int(h.get("gross_building_sqm")),
        "lot_size_sqm": _round_int(h.get("lot_size_sqm")),
        "typical_floor_sqm": _round_int(h.get("typical_floor_sqm")),
        "meeting_rooms_count": h.get("meeting_rooms_count"),
        "last_sale_date": h.get("last_sale_date"),
        "last_sale_price_eur": h.get("last_sale_price_eur"),
        "catastro_id": h.get("catastro_id"),
        # Booking enrichment
        "booking_hotel_id": meta.get("booking_hotel_id"),
        "booking_url": p.get("booking_url"),
        "review_score": p.get("review_score"),
        "review_count": p.get("review_count"),
        "location_score": p.get("location_score"),
        "comfort_score": p.get("comfort_score"),
        "cleanliness_score": p.get("cleanliness_score"),
        "staff_score": p.get("staff_score"),
        "value_score": p.get("value_score"),
        "facilities_score": p.get("facilities_score"),
        "has_bar": _yn((fnb.get("bars_count") or 0) > 0),
        "has_restaurant": _yn((fnb.get("restaurants_count") or 0) > 0),
        "has_rooftop": _yn(rooftop.get("has_rooftop")),
        "has_gym": _yn(gym.get("has_gym")),
        "has_spa": _yn(spa.get("has_spa")),
        "has_pool": _yn(pool.get("has_pool")),
        "has_parking": _yn(parking.get("has_parking")),
        "has_meeting": _yn((meeting.get("count") or 0) > 0),
        "booking_facilities_count": len(p.get("facilities_detailed") or []),
        "booking_room_types_count": len(p.get("room_types") or []),
        "check_in_time": p.get("check_in_time"),
        "check_out_time": p.get("check_out_time"),
        "pet_policy": p.get("pet_policy"),
        "cancellation_policy": p.get("cancellation_policy"),
        "smoking_policy": p.get("smoking_policy"),
        "coords_source": coords_source,
        "enrichment_sources": h.get("enrichment_sources") or ", ".join(meta.get("enrichment_sources") or []),
        "enrichment_confidence": meta.get("enrichment_confidence"),
        "profile_completeness_score": meta.get("profile_completeness_score"),
        "last_scraped_at": h.get("last_scraped_at") or meta.get("last_scraped_at"),
        # ── HotelVALORA enrichment passthrough (v1.4 · 2026-05-20) ──
        "phone": h.get("phone"),
        "website_url": h.get("website_url"),
        "google_place_id": h.get("google_place_id"),
        "wikidata_qid": h.get("wikidata_qid"),
        "canonical_id_supabase": h.get("canonical_id_supabase"),
        "data_quality_tier": h.get("data_quality_tier"),
        # Commercial context
        "competitive_set_ids": ", ".join(h.get("competitive_set_ids") or []),
        "transactions_history_ref": h.get("transactions_history_ref"),
        "notes": h.get("notes"),
        # Audit meta
        **_meta_block(snapshot, h.get("hotel_id"), "costar+rapidapi_booking" if e else "costar"),
        "review_required": "TRUE" if needs_review else "FALSE",
        "review_reason": ", ".join(needs_review),
    }


def main():
    """Build all institutional masters · populated with snapshot data
    when snapshot.json exists. Behaviour:

      - snapshot present + Supabase reachable: full populate including
        Booking enrichment merge on HOTELESperMARKET
      - snapshot present + Supabase unreachable: CoStar fields only ·
        Booking enrichment columns empty
      - no snapshot: emit headers-only templates (legacy v1.2 behaviour)
    """
    snapshot: dict = {}
    if SNAPSHOT_PATH.exists():
        snapshot = json.loads(SNAPSHOT_PATH.read_text(encoding="utf-8"))
        print(f"[build_masters] snapshot loaded · hotels={len(snapshot.get('hotels') or [])} · "
              f"market_snapshots={len(snapshot.get('market_snapshots') or [])} · "
              f"market_timeseries={len(snapshot.get('market_timeseries') or [])}")
    else:
        print(f"[build_masters] no snapshot at {SNAPSHOT_PATH} · emitting headers-only templates")

    market_snapshots = snapshot.get("market_snapshots") or []
    market_timeseries = snapshot.get("market_timeseries") or []
    hotels = snapshot.get("hotels") or []

    # ── PAIS · country_listing rows from snapshot ──
    pais_rows = [
        _market_row_to_pais(r, snapshot)
        for r in market_snapshots if r.get("granularity") == "country_listing"
    ]
    build_workbook(
        dataset_label="COSTAR_MASTER_PAIS",
        domain_columns=COUNTRY_COLUMNS,
        data_sheet_title="COUNTRY",
        schema_doc="docs/intelligence/costar-country-schema.md",
        output_path=MASTER_DIR / "COSTAR_MASTER_PAIS.xlsx",
        data_rows=pais_rows,
    )

    # ── MERCADOS · market snapshots + time-series ──
    mercados_rows = [
        _market_row_to_mercado(r, snapshot)
        for r in market_snapshots if r.get("granularity") == "market"
    ] + [
        _market_row_to_mercado(r, snapshot) for r in market_timeseries
    ]
    build_workbook(
        dataset_label="COSTAR_MASTER_MERCADOS",
        domain_columns=MARKET_COLUMNS,
        data_sheet_title="MARKET",
        schema_doc="docs/intelligence/costar-market-schema.md",
        output_path=MASTER_DIR / "COSTAR_MASTER_MERCADOS.xlsx",
        data_rows=mercados_rows,
    )

    # ── SUBMERCADOS · submarket rows ──
    submercados_rows = [
        _market_row_to_submercado(r, snapshot)
        for r in market_snapshots if r.get("granularity") == "submarket"
    ]
    build_workbook(
        dataset_label="COSTAR_MASTER_SUBMERCADOS",
        domain_columns=SUBMARKET_COLUMNS,
        data_sheet_title="SUBMARKET",
        schema_doc="docs/intelligence/costar-submarket-schema.md",
        output_path=MASTER_DIR / "COSTAR_MASTER_SUBMERCADOS.xlsx",
        data_rows=submercados_rows,
    )

    # ── CLASS · chain-scale aggregates derived from hotel inventory ──
    class_rows = []
    by_class: dict[tuple, dict] = {}
    for h in hotels:
        scale = h.get("chain_scale")
        if not scale:
            continue
        key = (h.get("country") or "", h.get("market_name") or "", scale)
        d = by_class.setdefault(key, {"hotel_count": 0, "rooms_total": 0})
        d["hotel_count"] += 1
        d["rooms_total"] += int(h.get("rooms_count") or 0)
    for (country, market, scale), agg in sorted(by_class.items()):
        class_rows.append({
            "country": country,
            "market_name": market,
            "submarket_name": "",
            "chain_scale": scale,
            "class_label_display": scale.replace("_", " ").title(),
            "segment_type": "",
            "period_kind": "snapshot",
            "period_start": "",
            "period_end": "",
            "currency": "EUR",
            "occupancy_pct": None, "adr": None, "revpar": None,
            "supply_rooms": None, "demand_rooms": None, "revenue": None,
            "supply_yoy_pct": None, "demand_yoy_pct": None, "occupancy_yoy_pp": None,
            "adr_yoy_pct": None, "revpar_yoy_pct": None,
            "revpar_index_vs_country": None, "revpar_index_vs_market": None,
            "hotel_count": agg["hotel_count"],
            "room_count_total": agg["rooms_total"],
            "pipeline_rooms": None, "pipeline_hotels": None,
            **_meta_block(snapshot, f"{country}|{market}|{scale}", "derived"),
            "notes": "Derived from hotel inventory · CoStar doesn't ship class-level KPIs in this drop",
        })
    build_workbook(
        dataset_label="COSTAR_MASTER_CLASS",
        domain_columns=CLASS_COLUMNS,
        data_sheet_title="CLASS",
        schema_doc="docs/intelligence/costar-class-schema.md",
        output_path=MASTER_DIR / "COSTAR_MASTER_CLASS.xlsx",
        data_rows=class_rows,
    )

    # ── HOTELESperMARKET · 364 hotels + Booking enrichment merged ──
    enrichment = _load_enrichment_from_storage()
    hotel_rows = [_hotel_to_row(h, enrichment, snapshot) for h in hotels]
    # Rename to match workspace directory naming · the legacy
    # COSTAR_MASTER_HOTELES_POR_MERCADO.xlsx stays in MASTER/ for
    # backward compatibility but the canonical filename going forward
    # mirrors the INPUT directory (HOTELESperMARKET).
    build_workbook(
        dataset_label="COSTAR_MASTER_HOTELESperMARKET",
        domain_columns=HOTELS_BY_MARKET_COLUMNS,
        data_sheet_title="HOTELS",
        schema_doc="docs/intelligence/costar-hotels-by-market-schema.md",
        output_path=MASTER_DIR / "COSTAR_MASTER_HOTELESperMARKET.xlsx",
        data_rows=hotel_rows,
    )


if __name__ == "__main__":
    main()
