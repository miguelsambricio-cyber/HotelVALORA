"""XLSX/CSV readers + header-alias folding for the COSTAR v1.2 pipeline.

The operator drops files into the various INPUT/ folders. This module:
  1. Locates every readable file under those folders (xlsx + csv).
  2. Reads each file's rows.
  3. Folds source-column headers to the canonical schema via alias maps.

Each `read_*` function returns `list[dict[str, Any]]` — one dict per
operator row, keyed by canonical column names. Unrecognised source columns
are dropped (operator can extend the alias maps in `normalization.py`).
"""

from __future__ import annotations

import csv
import re
import sys
from pathlib import Path
from typing import Any, Iterable

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None  # type: ignore[assignment]

from dedup import strip_diacritics


def _hkey(s: str) -> str:
    s = strip_diacritics(str(s)).lower().strip()
    return re.sub(r"[^a-z0-9]+", "_", s).strip("_")


def fold_headers(headers: list[str], alias_map: dict[str, str]) -> dict[int, str]:
    """Map header column index → canonical column name (or skip)."""
    result: dict[int, str] = {}
    for idx, h in enumerate(headers):
        if h is None:
            continue
        canonical = alias_map.get(_hkey(h))
        if canonical:
            result[idx] = canonical
    return result


def iter_input_files(root: Path) -> Iterable[Path]:
    """All .xlsx and .csv files under `root` excluding archives + dotfiles.

    Excluded path segments:
      - `.*`           dotfiles + dot-prefixed dirs
      - `old.*`        legacy per-workspace archives (old.class, old.mercado, …)
      - `OLD` / `old`  canonical post-2026-05-14 governance archive

    Without the OLD exclusion the recursive scan would re-process files
    that earlier ingest runs already moved INPUT → OLD, producing
    duplicate transactions and ghost archive_failed events.
    """
    if not root.exists():
        return []
    out: list[Path] = []
    excluded_dir_names = {"OLD", "old"}
    for p in root.rglob("*"):
        if not p.is_file():
            continue
        if p.name.startswith("."):
            continue
        if any(seg.startswith("old.") for seg in p.parts):
            continue
        if any(seg in excluded_dir_names for seg in p.parts):
            continue
        if p.suffix.lower() in (".xlsx", ".csv"):
            out.append(p)
    return out


def read_xlsx_rows(path: Path, sheet: str | None = None) -> list[list[Any]]:
    if load_workbook is None:
        print("✗ openpyxl not installed — `pip install -r services/costar/scripts/requirements.txt`", file=sys.stderr)
        return []
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
    if ws is None:
        return []
    return [list(row) for row in ws.iter_rows(values_only=True)]


def read_csv_rows(path: Path) -> list[list[Any]]:
    rows: list[list[Any]] = []
    with path.open(encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        for r in reader:
            rows.append(list(r))
    return rows


def read_table(path: Path) -> list[list[Any]]:
    if path.suffix.lower() == ".xlsx":
        return read_xlsx_rows(path)
    if path.suffix.lower() == ".csv":
        return read_csv_rows(path)
    return []


def read_rows_with_aliases(
    path: Path,
    alias_map: dict[str, str],
) -> list[dict[str, Any]]:
    """Read a file and fold its first row as headers into the canonical schema."""
    table = read_table(path)
    if len(table) < 2:
        return []
    headers = [str(h) if h is not None else "" for h in table[0]]
    col_map = fold_headers(headers, alias_map)
    if not col_map:
        return []
    out: list[dict[str, Any]] = []
    for raw in table[1:]:
        if all(cell in (None, "") for cell in raw):
            continue
        row: dict[str, Any] = {}
        for idx, canonical in col_map.items():
            if idx < len(raw):
                row[canonical] = raw[idx]
        if row:
            out.append(row)
    return out
