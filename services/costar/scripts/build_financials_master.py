"""Reproducible builder for an EMPTY scaffold template + JSON snapshot.

LEGACY · superseded by import_pnl_to_supabase.py for the production data
path (FASE 2 of Supabase migration · 2026-05-28). This script remains
useful for BOOTSTRAPPING new countries: it generates a methodology-firmed
empty template the operator can copy values into. Its output now writes
to COSTAR_MASTER_FINANCIALS.scaffold.xlsx (NOT .xlsx) so it can never
overwrite the operator-curated canonical file at COSTAR_MASTER_FINANCIALS.xlsx.

Run from repo root:
    python services/costar/scripts/build_financials_master.py

Produces two artefacts:

  1. services/costar/MASTER/COSTAR_MASTER_FINANCIALS.xlsx
     · operator-readable institutional record · 5 sheets (DATA · DICTIONARY ·
     INGESTION_LOG · SOURCES_REGISTRY · README) following the existing master
     conventions.

  2. apps/web/src/lib/report/financials/costar-financials-master.generated.json
     · runtime lookup consumed by `isProvisionalTemplate` in coverage.ts ·
     small footprint (~220 rows × 6 fields) · Next.js imports it natively.

Why two artefacts: Excel is the human-editable source of truth for the
institutional record (carries the full 21-column USALI template per row,
plus provenance + notes). The generated JSON is the minimal runtime
projection (5 identification columns + data_source) — fast to import, no
xlsx parser needed at cold-start.

Both regenerate identically on rerun. Edits to the Excel survive only when
manually reflected back into this script (today the script is the source).
A future "sync-from-edited-Excel" companion lives in backlog.

Scope today (2026-05-28):

  · Spain · 10 submarkets × 6 classes × 3 segmentation_types = 180 rows.
    All percentages populated from the methodology-firmed Madrid Centre
    Upper-Upscale plantilla (VALUATION_METHODOLOGY.md §3.1) ·
    data_source = "hardcoded_default" everywhere · EXCEPT the unique
    combination (Spain, Madrid, Madrid Centre, Upper Upscale, hotel) which
    carries data_source = "costar_real" — the one real CoStar template
    operationally loaded today.

  · 41 institutional hotel markets pending CoStar contract · one header
    row per country with empty USALI cells and data_source = "pending_costar".
    The list intentionally covers the world's institutional hospitality
    investment markets and is not exhaustive.
"""

from __future__ import annotations

import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

REPO = Path(__file__).resolve().parents[3]
MASTER_DIR = REPO / "services" / "costar" / "MASTER"
OUTPUT_XLSX = MASTER_DIR / "COSTAR_MASTER_FINANCIALS.scaffold.xlsx"
OUTPUT_JSON = (
    REPO / "apps" / "web" / "src" / "lib" / "report" / "financials"
    / "costar-financials-master.generated.json"
)

GENERATOR_VERSION = "1.0.0"

# ───────────────────────────────────────────────────────────────────────────
# USALI plantilla · methodology-firmed values from VALUATION_METHODOLOGY.md §3.1
# (Madrid Centro Upper-Upscale baseline). Values are percentages 0-100.
# ───────────────────────────────────────────────────────────────────────────

DEFAULT_USALI: dict[str, float] = {
    "rooms_revenue_pct": 67.5,
    "fb_food_pct": 17.7,
    "fb_beverage_pct": 6.8,
    "meeting_events_pct": 3.8,
    "spa_wellness_pct": 2.2,
    "parking_other_pct": 1.9,
    "expenses_rooms_pct": 25.7,
    "expenses_fb_pct": 79.2,
    "other_departments_pct": 85.8,
    "admin_general_pct": 7.2,
    "it_telecom_pct": 1.3,
    "sales_marketing_pct": 6.5,
    "operations_maintenance_pct": 3.8,
    "utilities_pct": 2.8,
    "gop_pct": 36.7,
    "management_fees_pct": 4.6,
    "rent_pct": 7.8,
    "property_taxes_pct": 0.7,
    "insurance_pct": 0.4,
    "ebitda_pct": 23.3,
    "staff_cost_memo_pct": 31.7,
}

ID_COLS = ["country", "market", "submarket", "class", "segmentation_type"]
PCT_COLS = list(DEFAULT_USALI.keys())
PROVENANCE_COLS = ["data_source", "last_updated", "notes"]
ALL_COLS = ID_COLS + PCT_COLS + PROVENANCE_COLS

SEGMENTATION_TYPES = ["hotel", "apartahotel", "hostel"]

# CoStar market names carry the country suffix (e.g. "Madrid ESP") so the
# raw data are unambiguous worldwide. The DB normalises to bare names
# ("Madrid"). For the JSON snapshot we emit the normalised form so the
# runtime matcher (coverage.ts) can compare against `hotel.market_name`
# directly. The Excel keeps the raw CoStar form for institutional record.
def normalise_market_name(raw: str | None) -> str | None:
    if raw is None:
        return None
    s = raw.strip()
    if s.endswith(" ESP"):
        return s[:-4].strip()
    return s


# The 41 countries flagged "pending_costar". One header row per country ·
# all USALI cells empty. ISO 3166-1 alpha-2.
PENDING_COUNTRIES: list[tuple[str, str]] = [
    ("US", "United States"),
    ("GB", "United Kingdom"),
    ("PT", "Portugal"),
    ("FR", "France"),
    ("DE", "Germany"),
    ("BE", "Belgium"),
    ("NL", "Netherlands"),
    ("IT", "Italy"),
    ("CH", "Switzerland"),
    ("AT", "Austria"),
    ("PL", "Poland"),
    ("HU", "Hungary"),
    ("CZ", "Czech Republic"),
    ("GR", "Greece"),
    ("NO", "Norway"),
    ("SE", "Sweden"),
    ("DK", "Denmark"),
    ("BG", "Bulgaria"),
    ("CA", "Canada"),
    ("MX", "Mexico"),
    ("BR", "Brazil"),
    ("AR", "Argentina"),
    ("CO", "Colombia"),
    ("PE", "Peru"),
    ("AE", "United Arab Emirates"),
    ("SA", "Saudi Arabia"),
    ("TR", "Turkey"),
    ("IL", "Israel"),
    ("CN", "China"),
    ("JP", "Japan"),
    ("KR", "South Korea"),
    ("IN", "India"),
    ("AU", "Australia"),
    ("ZA", "South Africa"),
    ("IE", "Ireland"),
    ("FI", "Finland"),
    ("RO", "Romania"),
    ("HR", "Croatia"),
    ("EG", "Egypt"),
    ("SG", "Singapore"),
    ("NZ", "New Zealand"),
]
assert len(PENDING_COUNTRIES) == 41

# The single combination valued today with a fully-loaded CoStar template.
LOADED_COSTAR_MATCH = {
    "country": "ES",
    "market_raw": "Madrid ESP",        # raw form (matches CoStar master)
    "submarket": "Madrid Centre",
    "class": "Upper Upscale",
    "segmentation_type": "hotel",
}

# ───────────────────────────────────────────────────────────────────────────
# Source readers
# ───────────────────────────────────────────────────────────────────────────

def read_spain_submarkets() -> list[tuple[str, str]]:
    """Return (raw_market_name, submarket_name) tuples for country=ES."""
    wb = load_workbook(MASTER_DIR / "COSTAR_MASTER_SUBMERCADOS.xlsx", read_only=True)
    ws = wb["SUBMARKET"]
    rows = list(ws.iter_rows(values_only=True))
    hdr = rows[0]
    i_country = hdr.index("country")
    i_market = hdr.index("market_name")
    i_submarket = hdr.index("submarket_name")
    out: list[tuple[str, str]] = []
    for r in rows[1:]:
        if r[i_country] == "ES" and r[i_market] and r[i_submarket]:
            out.append((str(r[i_market]), str(r[i_submarket])))
    return out


def read_chain_scales() -> list[tuple[str, str]]:
    """Return (chain_scale_enum, class_label_display) tuples from CLASS master."""
    wb = load_workbook(MASTER_DIR / "COSTAR_MASTER_CLASS.xlsx", read_only=True)
    ws = wb["CLASS"]
    rows = list(ws.iter_rows(values_only=True))
    hdr = rows[0]
    i_scale = hdr.index("chain_scale")
    i_label = hdr.index("class_label_display")
    out: list[tuple[str, str]] = []
    seen: set[str] = set()
    for r in rows[1:]:
        scale = r[i_scale]
        label = r[i_label]
        if scale and label and scale not in seen:
            out.append((str(scale), str(label)))
            seen.add(str(scale))
    return out


# ───────────────────────────────────────────────────────────────────────────
# Row generation
# ───────────────────────────────────────────────────────────────────────────

def build_rows(now_iso: str) -> list[dict[str, Any]]:
    """Build the full list of rows for both the Excel and the JSON snapshot."""
    rows: list[dict[str, Any]] = []

    # ── Spain · detailed coverage ──
    submarkets = read_spain_submarkets()
    classes = read_chain_scales()
    for market_raw, submarket in submarkets:
        for _scale_enum, class_label in classes:
            for seg in SEGMENTATION_TYPES:
                is_loaded_match = (
                    market_raw == LOADED_COSTAR_MATCH["market_raw"]
                    and submarket == LOADED_COSTAR_MATCH["submarket"]
                    and class_label == LOADED_COSTAR_MATCH["class"]
                    and seg == LOADED_COSTAR_MATCH["segmentation_type"]
                )
                data_source = "costar_real" if is_loaded_match else "hardcoded_default"
                row = {
                    "country": "ES",
                    "market": market_raw,
                    "submarket": submarket,
                    "class": class_label,
                    "segmentation_type": seg,
                    **DEFAULT_USALI,
                    "data_source": data_source,
                    "last_updated": now_iso,
                    "notes": (
                        "Live CoStar template (single combination loaded operationally)"
                        if is_loaded_match
                        else "Methodology-firmed defaults applied until CoStar segmented data lands"
                    ),
                }
                rows.append(row)

    # ── 41 countries · pending ──
    for iso, name in PENDING_COUNTRIES:
        row = {
            "country": iso,
            "market": None,
            "submarket": None,
            "class": None,
            "segmentation_type": None,
            **{k: None for k in PCT_COLS},
            "data_source": "pending_costar",
            "last_updated": now_iso,
            "notes": f"{name} · awaiting CoStar contract + ingestion",
        }
        rows.append(row)

    return rows


# ───────────────────────────────────────────────────────────────────────────
# Excel writer
# ───────────────────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill(start_color="FF0F2A1F", end_color="FF0F2A1F", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFFFF")
CELL_FONT = Font(name="Calibri", size=10)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=False)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")


def write_excel(rows: list[dict[str, Any]], now_iso: str) -> None:
    wb = Workbook()
    # Remove default sheet · we create our own in canonical order.
    default = wb.active
    wb.remove(default)

    # ── DATA sheet ──
    ws = wb.create_sheet("DATA")
    for col_idx, col in enumerate(ALL_COLS, 1):
        c = ws.cell(row=1, column=col_idx, value=col)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = ALIGN_LEFT
    for row_idx, r in enumerate(rows, 2):
        for col_idx, col in enumerate(ALL_COLS, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=r.get(col))
            cell.font = CELL_FONT
            cell.alignment = ALIGN_RIGHT if col in PCT_COLS else ALIGN_LEFT
    # Column widths · institutional readability
    width_for = {**{c: 14 for c in PCT_COLS}, "country": 9, "market": 22, "submarket": 30, "class": 16, "segmentation_type": 16, "data_source": 18, "last_updated": 22, "notes": 60}
    for col_idx, col in enumerate(ALL_COLS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width_for.get(col, 12)
    ws.freeze_panes = "A2"

    # ── DICTIONARY sheet ──
    ws = wb.create_sheet("DICTIONARY")
    ws.append(["column", "type", "required", "notes"])
    dict_rows = [
        ("country", "text", "yes", "ISO 3166-1 alpha-2."),
        ("market", "text", "no", "CoStar raw market name (carries the country suffix, e.g. 'Madrid ESP'). Null for country-only pending rows."),
        ("submarket", "text", "no", "CoStar submarket. Null for country-only pending rows."),
        ("class", "text", "no", "CoStar class label · Title-Case ('Upper Upscale'). Null for country-only pending rows."),
        ("segmentation_type", "text", "no", "hotel · apartahotel · hostel. Null for country-only pending rows."),
        *[(c, "number", "no", "USALI percentage 0-100 · null when no data") for c in PCT_COLS],
        ("data_source", "text", "yes", "costar_real | hardcoded_default | pending_costar."),
        ("last_updated", "text", "yes", "ISO-8601 UTC timestamp of last write."),
        ("notes", "text", "no", "Free-text provenance note."),
    ]
    for r in dict_rows:
        ws.append(r)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 80

    # ── INGESTION_LOG sheet ──
    ws = wb.create_sheet("INGESTION_LOG")
    ws.append(["timestamp", "operator", "source", "rows_written", "generator_version", "notes"])
    ws.append([now_iso, "build_financials_master.py", "scripted", len(rows), GENERATOR_VERSION, "Initial seed · methodology-firmed defaults + 1 costar_real + 41 pending_costar countries."])

    # ── SOURCES_REGISTRY sheet ──
    ws = wb.create_sheet("SOURCES_REGISTRY")
    ws.append(["data_source", "reliability_tier", "description"])
    ws.append(["costar_real", "S", "Live CoStar template ingested for this exact (submarket × class × segmentation_type)."])
    ws.append(["hardcoded_default", "C", "Methodology-firmed defaults (Madrid Centre Upper-Upscale) applied because no segmented CoStar data exists yet for this combination."])
    ws.append(["pending_costar", "Z", "No data loaded. Country-level placeholder pending CoStar contract."])

    # ── README sheet ──
    ws = wb.create_sheet("README")
    readme_lines = [
        "COSTAR_MASTER_FINANCIALS · USALI percentage records per (country × market × submarket × class × segmentation_type)",
        "",
        "Generated by services/costar/scripts/build_financials_master.py · regenerable.",
        "",
        "Runtime consumption:",
        "  apps/web/src/lib/report/financials/costar-financials-master.generated.json carries a minimal projection",
        "  (5 identification columns + data_source). The Next.js runtime imports it directly · isProvisionalTemplate",
        "  in coverage.ts reads from it to decide when to show the 'plantilla provisional' banner.",
        "",
        "Source-of-truth model:",
        "  The script is the source of truth today. Editing this xlsx directly is allowed but the next script run",
        "  will overwrite. A future 'sync-from-edited-xlsx' companion will preserve edits.",
        "",
        "Adding a CoStar-real combination:",
        "  Update LOADED_COSTAR_MATCH (or expand to a list) in the script and rerun. The corresponding row will",
        "  flip from hardcoded_default to costar_real. Future: CoStar real percentages overwrite the defaults",
        "  per-row when a segmented row lands.",
    ]
    for line in readme_lines:
        ws.append([line])
    ws.column_dimensions["A"].width = 110

    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_XLSX)


# ───────────────────────────────────────────────────────────────────────────
# JSON snapshot writer (runtime consumption)
# ───────────────────────────────────────────────────────────────────────────

def write_json_snapshot(rows: list[dict[str, Any]], now_iso: str) -> None:
    """Emit the minimal lookup table the runtime needs.

    Only id-columns + data_source are projected · percentages are not needed
    at runtime today (the engine uses getDefaultAssumptions). The market
    name is normalised (no ESP suffix) so it matches `hotel.market_name`
    from the BD join directly.
    """
    projected = []
    for r in rows:
        projected.append({
            "country": r["country"],
            "market": normalise_market_name(r["market"]),
            "submarket": r["submarket"],
            "class": r["class"],
            "segmentation_type": r["segmentation_type"],
            "data_source": r["data_source"],
        })
    payload = {
        "generated_at": now_iso,
        "generator_version": GENERATOR_VERSION,
        "row_count": len(projected),
        "rows": projected,
    }
    OUTPUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_JSON.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


# ───────────────────────────────────────────────────────────────────────────
# Main
# ───────────────────────────────────────────────────────────────────────────

def main() -> None:
    now_iso = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    rows = build_rows(now_iso)
    write_excel(rows, now_iso)
    write_json_snapshot(rows, now_iso)
    # Summary
    by_source: dict[str, int] = {}
    for r in rows:
        by_source[r["data_source"]] = by_source.get(r["data_source"], 0) + 1
    print(f"wrote {OUTPUT_XLSX.relative_to(REPO)} · {len(rows)} rows")
    print(f"wrote {OUTPUT_JSON.relative_to(REPO)} · {len(rows)} rows")
    print(f"by data_source: {by_source}")


if __name__ == "__main__":
    main()
