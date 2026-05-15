#!/usr/bin/env python3
"""
audit_master_alignment.py — read-only diagnostic.

Confirms whether the Master xlsx headers and data rows are aligned
column-for-column, or whether the data is offset by N positions
relative to the headers.

Method:
  1. Load Master headers + all data rows (openpyxl, read_only).
  2. Load Supabase relationship_contacts (master_id, full_name, email,
     investor_type) — these values are known-canonical (promoted before
     the corruption).
  3. For each Master row, locate its master_id under three hypotheses:
       H0 · no shift                       (data[col] = header[col])
       H1 · data shifted left by 1         (data[col] = header[col+1])
       H2 · data shifted left by 2         (data[col] = header[col+2])
     Score each hypothesis by counting how many Master rows match
     Supabase on full_name + email + investor_type when read under
     that hypothesis.
  4. Print a winning hypothesis ONLY if dominant by ≥95% match rate.
     Otherwise refuse to declare and dump per-hypothesis stats.
  5. For the two Phase 2.B.3 replaced rows (master_id f193186dd9eb0c22
     · 596a76514db8d527) print the full per-column dump under both raw
     and winning-hypothesis interpretations.
  6. Write JSON report to:
       CONTACTOS DATASITE/reports/master-alignment-audit_<TS>.jsonl

NO MUTATIONS. Safe to run repeatedly.
"""
from __future__ import annotations

import json
import os
import sys
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

sys.stdout.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parent.parent.parent
DEFAULT_MASTER = ROOT / "CONTACTOS DATASITE" / "master" / "metcub-contacts-master.xlsx"
# Allow CLI override for validating repaired or alternate copies:
#   python audit_master_alignment.py <path-to-xlsx>
MASTER_PATH = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_MASTER
REPORTS_DIR = ROOT / "CONTACTOS DATASITE" / "reports"
REPORTS_DIR.mkdir(parents=True, exist_ok=True)

# Two Phase 2.B.3 replacement rows we want a full dump of
PHASE_2B3_MASTER_IDS = {
    "f193186dd9eb0c22": "crocher@bancsabadell.com → prietose@bancsabadell.com",
    "596a76514db8d527": "p.j.rodera@reyalurbis.com → gestiondeactivos2@reyalurbis.com",
}


def normalize(value: object) -> str | None:
    """Strip + lowercase + treat blank as None for comparison."""
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    return s.lower()


def read_master() -> tuple[list[str | None], list[list[object]]]:
    print(f"→ Reading Master from {MASTER_PATH}")
    wb = load_workbook(MASTER_PATH, read_only=True, data_only=True)
    ws = wb["Master"]
    header = [c.value for c in ws[1]]
    rows: list[list[object]] = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        rows.append(list(r))
    wb.close()
    print(f"  · header cols: {len(header)} · data rows: {len(rows)}")
    return header, rows


def read_supabase() -> dict[str, dict[str, str | None]]:
    """Pull canonical reference from Supabase via psycopg2 if available,
    else from a previously-cached JSONL snapshot. Falls back gracefully.
    """
    cache = REPORTS_DIR / "supabase-canonical-snapshot.jsonl"
    if cache.exists():
        print(f"→ Reading Supabase reference from cache: {cache.name}")
        out: dict[str, dict[str, str | None]] = {}
        with cache.open("r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                obj = json.loads(line)
                mid = obj.get("master_id")
                if mid:
                    out[mid] = {
                        "full_name": normalize(obj.get("full_name")),
                        "email": normalize(obj.get("email")),
                        "investor_type": normalize(obj.get("investor_type")),
                    }
        print(f"  · loaded {len(out)} Supabase reference rows")
        return out

    print("  · no Supabase cache present — audit will run in self-only mode")
    print("    (build cache first by exporting `master_id, full_name, email, investor_type`")
    print("    from relationship_contacts to reports/supabase-canonical-snapshot.jsonl)")
    return {}


def get_at_shift(row: list[object], header: list[str | None], col_name: str, shift: int) -> object | None:
    """Resolve `col_name` under shift hypothesis (data[col] = header[col+shift])."""
    try:
        header_idx = header.index(col_name)
    except ValueError:
        return None
    data_idx = header_idx - shift
    if 0 <= data_idx < len(row):
        return row[data_idx]
    return None


def score_hypothesis(
    rows: list[list[object]],
    header: list[str | None],
    sb_ref: dict[str, dict[str, str | None]],
    shift: int,
) -> dict[str, int]:
    """Score how many rows match Supabase under given shift."""
    counts = {
        "rows_with_master_id": 0,
        "rows_in_sb": 0,
        "match_full_name": 0,
        "match_email": 0,
        "match_investor_type": 0,
        "match_all_three": 0,
    }
    for row in rows:
        mid_val = get_at_shift(row, header, "master_id", shift)
        mid = normalize(mid_val)
        if not mid:
            continue
        counts["rows_with_master_id"] += 1
        ref = sb_ref.get(mid)
        if not ref:
            continue
        counts["rows_in_sb"] += 1
        fn = normalize(get_at_shift(row, header, "full_name", shift))
        em = normalize(get_at_shift(row, header, "email", shift))
        it = normalize(get_at_shift(row, header, "investor_type", shift))
        ok_fn = ref["full_name"] is not None and fn == ref["full_name"]
        ok_em = ref["email"] is not None and em == ref["email"]
        ok_it = ref["investor_type"] is not None and it == ref["investor_type"]
        counts["match_full_name"] += int(ok_fn)
        counts["match_email"] += int(ok_em)
        counts["match_investor_type"] += int(ok_it)
        counts["match_all_three"] += int(ok_fn and ok_em and ok_it)
    return counts


def dump_row_under_hypothesis(
    row: list[object], header: list[str | None], shift: int, label: str
) -> dict[str, object]:
    """Return a {semantic_name: value} dict under given shift, omitting empty cells."""
    out: dict[str, object] = {"_hypothesis": label, "_shift": shift}
    for header_idx, name in enumerate(header):
        if not name:
            continue
        data_idx = header_idx - shift
        if 0 <= data_idx < len(row):
            v = row[data_idx]
            if v not in (None, "", " "):
                out[str(name)] = v
    return out


def find_phase_2b3_rows(
    rows: list[list[object]], header: list[str | None]
) -> dict[str, tuple[int, list[object]]]:
    """Locate the 2 Phase 2.B.3 replacement rows by scanning ALL columns
    for known master_id strings (since we don't yet know which col is
    actually master_id)."""
    out: dict[str, tuple[int, list[object]]] = {}
    targets = set(PHASE_2B3_MASTER_IDS.keys())
    for row_idx, row in enumerate(rows, start=2):
        for v in row:
            if v in targets:
                out[str(v)] = (row_idx, row)
                break
    return out


def main() -> int:
    if not MASTER_PATH.exists():
        print(f"ERROR: Master not found at {MASTER_PATH}")
        return 1

    header, rows = read_master()
    sb_ref = read_supabase()
    if not sb_ref:
        print("\n!! Supabase cache missing · run again after caching reference data")
        return 2

    print(f"\n→ Scoring shift hypotheses ({len(rows)} rows × 3 hypotheses)")
    scores: dict[int, dict[str, int]] = {}
    for shift in (0, 1, 2, -1):
        counts = score_hypothesis(rows, header, sb_ref, shift)
        scores[shift] = counts
        rows_in_sb = counts["rows_in_sb"]
        match_pct = (counts["match_all_three"] / rows_in_sb * 100) if rows_in_sb else 0.0
        label = (
            "no shift"
            if shift == 0
            else f"data shifted LEFT by {shift}"
            if shift > 0
            else f"data shifted RIGHT by {-shift}"
        )
        print(f"\n  H[shift={shift:+d}] · {label}")
        print(f"    rows with master_id resolved: {counts['rows_with_master_id']}")
        print(f"    rows present in Supabase    : {rows_in_sb}")
        print(f"    match · full_name           : {counts['match_full_name']}")
        print(f"    match · email               : {counts['match_email']}")
        print(f"    match · investor_type       : {counts['match_investor_type']}")
        print(f"    match · ALL THREE           : {counts['match_all_three']} ({match_pct:.1f}% of rows-in-sb)")

    winner = max(scores.items(), key=lambda kv: kv[1]["match_all_three"])
    win_shift, win_counts = winner
    win_pct = (
        win_counts["match_all_three"] / win_counts["rows_in_sb"] * 100
        if win_counts["rows_in_sb"]
        else 0.0
    )
    runner_up = sorted(scores.items(), key=lambda kv: kv[1]["match_all_three"], reverse=True)[1]
    ru_pct = (
        runner_up[1]["match_all_three"] / runner_up[1]["rows_in_sb"] * 100
        if runner_up[1]["rows_in_sb"]
        else 0.0
    )
    margin = win_pct - ru_pct

    print(f"\n→ Winning hypothesis: shift={win_shift:+d} ({win_pct:.1f}% match)")
    print(f"  margin over runner-up (shift={runner_up[0]:+d}, {ru_pct:.1f}%): {margin:.1f}pp")
    confident = win_pct >= 95.0 and margin >= 50.0
    print(f"  confidence: {'HIGH · safe to fix' if confident else 'LOW · do not auto-fix'}")

    # Phase 2.B.3 row dumps
    print("\n→ Phase 2.B.3 replacement-row dumps (raw + winning-hypothesis)")
    p2b3 = find_phase_2b3_rows(rows, header)
    p2b3_dumps: dict[str, dict[str, object]] = {}
    for mid, intent in PHASE_2B3_MASTER_IDS.items():
        if mid not in p2b3:
            print(f"\n  !! master_id {mid} ({intent}) NOT FOUND in any column of any row")
            continue
        sheet_row, row = p2b3[mid]
        print(f"\n  master_id {mid} · sheet_row {sheet_row}")
        print(f"  intent: {intent}")
        raw_dump = dump_row_under_hypothesis(row, header, 0, "no shift (raw)")
        win_dump = dump_row_under_hypothesis(row, header, win_shift, f"winning shift={win_shift:+d}")
        print("\n  RAW (header[N] = data[N]):")
        for k, v in raw_dump.items():
            if k.startswith("_"):
                continue
            print(f"    {k:32s} = {v!r}")
        print(f"\n  UNDER WINNING HYPOTHESIS (data[col-{win_shift}] = header[col]):")
        for k, v in win_dump.items():
            if k.startswith("_"):
                continue
            print(f"    {k:32s} = {v!r}")
        ref = sb_ref.get(mid)
        if ref:
            print("\n  SUPABASE reference for this master_id:")
            for k, v in ref.items():
                print(f"    {k:32s} = {v!r}")
        p2b3_dumps[mid] = {
            "intent": intent,
            "sheet_row": sheet_row,
            "raw": raw_dump,
            "winning": win_dump,
            "supabase_ref": ref,
        }

    # Sample 5 random rows for human eyeballing
    print("\n→ Sample 5 random rows under winning hypothesis (non-Phase-2.B.3)")
    import random

    random.seed(42)
    sample_rows = random.sample(range(len(rows)), min(5, len(rows)))
    samples: list[dict[str, object]] = []
    for idx in sample_rows:
        row = rows[idx]
        sheet_row = idx + 2
        win_dump = dump_row_under_hypothesis(row, header, win_shift, f"winning shift={win_shift:+d}")
        mid = win_dump.get("master_id")
        ref = sb_ref.get(normalize(mid)) if mid else None
        print(f"\n  sheet_row {sheet_row} · master_id={mid!r}")
        for k in ("master_id", "full_name", "email", "investor_type", "company"):
            v = win_dump.get(k, "<missing>")
            print(f"    {k:16s} = {v!r}")
        if ref:
            print(f"    SUPABASE.full_name      = {ref['full_name']!r}")
            print(f"    SUPABASE.email          = {ref['email']!r}")
            print(f"    SUPABASE.investor_type  = {ref['investor_type']!r}")
        samples.append({"sheet_row": sheet_row, "winning": win_dump, "supabase_ref": ref})

    # Write report
    ts = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    report_path = REPORTS_DIR / f"master-alignment-audit_{ts}.json"
    with report_path.open("w", encoding="utf-8") as f:
        json.dump(
            {
                "generated_at": datetime.now(timezone.utc).isoformat(),
                "master_path": str(MASTER_PATH),
                "header_count": len(header),
                "row_count": len(rows),
                "supabase_ref_count": len(sb_ref),
                "scores_by_shift": {str(k): v for k, v in scores.items()},
                "winning_shift": win_shift,
                "winning_match_pct": win_pct,
                "runner_up_shift": runner_up[0],
                "runner_up_match_pct": ru_pct,
                "margin_pp": margin,
                "confidence": "HIGH" if confident else "LOW",
                "phase_2b3_rows": p2b3_dumps,
                "samples": samples,
            },
            f,
            ensure_ascii=False,
            indent=2,
            default=str,
        )
    print(f"\n→ Report written: {report_path}")
    return 0 if confident else 3


if __name__ == "__main__":
    sys.exit(main())
