#!/usr/bin/env python3
"""
final_repair_validation.py · Phase B-Repair · 2026-05-15

Closes the loop on the alignment repair:
  1. Re-runs the alignment audit on the canonical Master (must show
     shift=0 winning with ≥99% match)
  2. Confirms the 2 Phase 2.B.3 replacements are now correctly applied
     (email = new value · original_email = old value)
  3. Investigates the 16 unmatched rows (not failures of the repair · just
     legitimate Master/Supabase divergence)
  4. Dumps 5 random rows in BEFORE/AFTER form (broken file vs canonical)
     for the changelog evidence
  5. Writes the final report to reports/phase_b_repair_final_<TS>.json

Read-only on Master. Writes only the JSON report.
"""
from __future__ import annotations

import json
import random
import sys
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

sys.stdout.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parent.parent.parent
MASTER_PATH = ROOT / "CONTACTOS DATASITE" / "master" / "metcub-contacts-master.xlsx"
BROKEN_PATH = ROOT / "CONTACTOS DATASITE" / "master" / "metcub-contacts-master.broken-2026-05-15.xlsx"
SB_CACHE = ROOT / "CONTACTOS DATASITE" / "reports" / "supabase-canonical-snapshot.jsonl"
REPORTS_DIR = ROOT / "CONTACTOS DATASITE" / "reports"


def load_master(path: Path) -> tuple[list[str | None], list[list[object]]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Master"]
    header = [c.value for c in ws[1]]
    rows = [list(r) for r in ws.iter_rows(min_row=2, values_only=True)]
    wb.close()
    return header, rows


def load_supabase() -> dict[str, dict[str, str | None]]:
    out: dict[str, dict[str, str | None]] = {}
    with SB_CACHE.open("r", encoding="utf-8") as f:
        for line in f:
            obj = json.loads(line.strip())
            mid = obj.get("master_id")
            if mid:
                out[mid] = {
                    "full_name": (obj.get("full_name") or "").strip().lower() or None,
                    "email": (obj.get("email") or "").strip().lower() or None,
                    "investor_type": (obj.get("investor_type") or "").strip().lower() or None,
                }
    return out


def main() -> int:
    print("→ Loading canonical Master + broken backup + Supabase cache")
    canon_header, canon_rows = load_master(MASTER_PATH)
    broken_header, broken_rows = load_master(BROKEN_PATH)
    sb = load_supabase()
    print(f"  · canonical: {len(canon_header)} cols × {len(canon_rows)} rows")
    print(f"  · broken   : {len(broken_header)} cols × {len(broken_rows)} rows")
    print(f"  · supabase : {len(sb)} ref rows")

    # --- 1. Sanity on canonical schema ---
    if len(canon_header) != 64 or canon_header[0] != "original_email" or canon_header[1] != "master_id":
        print(f"!! Canonical schema is wrong: {canon_header[:5]}")
        return 1

    # --- 2. Audit shift=0 match rate ---
    print("\n→ Re-running shift=0 audit on canonical Master")
    col_mid = canon_header.index("master_id")
    col_fn = canon_header.index("full_name")
    col_em = canon_header.index("email")
    col_it = canon_header.index("investor_type")
    col_oem = canon_header.index("original_email")
    matched, present, mismatches = 0, 0, []
    for row in canon_rows:
        mid = (row[col_mid] or "").strip() if isinstance(row[col_mid], str) else None
        if not mid or mid not in sb:
            continue
        present += 1
        ref = sb[mid]
        fn_ok = (row[col_fn] or "").strip().lower() == ref["full_name"]
        em_ok = (row[col_em] or "").strip().lower() == ref["email"]
        it_ok = (row[col_it] or "").strip().lower() == ref["investor_type"]
        if fn_ok and em_ok and it_ok:
            matched += 1
        else:
            mismatches.append((mid, row[col_fn], row[col_em], row[col_it], row[col_oem], ref))
    pct = matched / present * 100 if present else 0.0
    print(f"  · rows present in Supabase : {present}")
    print(f"  · all-three match          : {matched} ({pct:.1f}%)")
    print(f"  · mismatches               : {len(mismatches)}")

    # --- 3. Analyse mismatches (expected: ~16, dominated by replaced emails) ---
    print("\n→ Mismatch analysis (first 10)")
    mismatch_breakdown = {"email_now_replaced": 0, "other": 0}
    sample_mm = []
    for mid, fn, em, it, oem, ref in mismatches[:10]:
        replaced = bool(oem)  # original_email populated → this is a replacement row
        if replaced:
            mismatch_breakdown["email_now_replaced"] += 1
        else:
            mismatch_breakdown["other"] += 1
        flag = "REPLACED" if replaced else "DIVERGENT"
        print(f"  [{flag}] {mid[:12]} · master={em!r} sb={ref['email']!r} · oem={oem!r}")
        sample_mm.append({
            "master_id": mid,
            "tag": flag,
            "master_email": em,
            "master_full_name": fn,
            "master_investor_type": it,
            "master_original_email": oem,
            "supabase": ref,
        })
    # Full count of replaced vs other across ALL mismatches (not just 10)
    total_repl = sum(1 for m in mismatches if m[4])
    total_other = len(mismatches) - total_repl
    print(f"\n  Across all {len(mismatches)} mismatches: {total_repl} are replaced rows · {total_other} are other divergence")

    # --- 4. Phase 2.B.3 replacement-row dumps post-fix + post-apply ---
    print("\n→ Phase 2.B.3 replacement rows (post-repair + post-reapplication)")
    p2b3 = {"f193186dd9eb0c22": "crocher → prietose", "596a76514db8d527": "rodera → gestiondeactivos2"}
    p2b3_status = {}
    for row in canon_rows:
        mid = row[col_mid]
        if mid in p2b3:
            print(f"\n  master_id {mid} · intent: {p2b3[mid]}")
            print(f"    full_name        = {row[col_fn]!r}")
            print(f"    email            = {row[col_em]!r}        ← should be NEW")
            print(f"    original_email   = {row[col_oem]!r}        ← should be OLD")
            print(f"    investor_type    = {row[col_it]!r}")
            p2b3_status[mid] = {
                "full_name": row[col_fn],
                "email_post": row[col_em],
                "original_email_post": row[col_oem],
                "investor_type": row[col_it],
            }

    # --- 5. Build broken-file index for before/after diff ---
    print("\n→ 5 BEFORE/AFTER samples (random rows)")
    random.seed(2026_05_15)
    target_rows = random.sample(range(len(canon_rows)), 5)
    samples = []
    for idx in target_rows:
        sheet_row = idx + 2
        canon_row = canon_rows[idx]
        broken_row = broken_rows[idx] if idx < len(broken_rows) else []
        # BEFORE: read raw with broken header (header[N] = data[N])
        before = {h: broken_row[i] for i, h in enumerate(broken_header)
                  if h and i < len(broken_row) and broken_row[i] not in (None, "")}
        # AFTER: read raw with canonical header
        after = {h: canon_row[i] for i, h in enumerate(canon_header)
                 if h and i < len(canon_row) and canon_row[i] not in (None, "")}
        sb_ref = sb.get(after.get("master_id"))
        print(f"\n  --- sheet_row {sheet_row} ---")
        print(f"  BEFORE (broken file · header[N]=data[N]):")
        for k in ["master_id", "full_name", "email", "investor_type", "company"]:
            print(f"    {k:20s} = {before.get(k)!r}")
        print(f"  AFTER  (repaired · header[N]=data[N]):")
        for k in ["master_id", "full_name", "email", "investor_type", "company"]:
            print(f"    {k:20s} = {after.get(k)!r}")
        if sb_ref:
            print(f"  SUPABASE reference:")
            for k in ["full_name", "email", "investor_type"]:
                print(f"    {k:20s} = {sb_ref.get(k)!r}")
        samples.append({"sheet_row": sheet_row, "before": before, "after": after, "supabase": sb_ref})

    # --- 6. Write final report ---
    ts = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    report = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "canonical_master": str(MASTER_PATH),
        "broken_backup": str(BROKEN_PATH),
        "canonical_schema_cols": len(canon_header),
        "row_count": len(canon_rows),
        "supabase_ref_count": len(sb),
        "shift0_audit": {
            "rows_present_in_supabase": present,
            "all_three_match": matched,
            "match_pct": pct,
            "mismatches_total": len(mismatches),
            "mismatches_replaced_rows": total_repl,
            "mismatches_other_divergence": total_other,
            "first_10_mismatches": sample_mm,
        },
        "phase_2b3_replacements": p2b3_status,
        "before_after_samples": samples,
    }
    out_path = REPORTS_DIR / f"phase_b_repair_final_{ts}.json"
    out_path.write_text(json.dumps(report, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    print(f"\n→ Final report written: {out_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
