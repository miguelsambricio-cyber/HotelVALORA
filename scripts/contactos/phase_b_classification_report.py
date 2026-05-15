#!/usr/bin/env python3
"""
phase_b_classification_report.py · Phase B · 2026-05-15

Read-only report comparing v1 (`contact_category`) vs v2
(`contact_category_v2`) classification on the canonical Master xlsx.

Outputs:
  - JSON report at reports/phase_b_classification_report_<TS>.json
  - Stdout summary: distribution side-by-side · migration matrix ·
    operator-split breakdown · IA Supply discoveries · top ambiguous
    mappings · 20 random rows for human review.

NO Master mutations.
"""
from __future__ import annotations

import json
import random
import sys
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

sys.stdout.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parent.parent.parent
MASTER_PATH = ROOT / "CONTACTOS DATASITE" / "master" / "metcub-contacts-master.xlsx"
REPORTS_DIR = ROOT / "CONTACTOS DATASITE" / "reports"
REPORTS_DIR.mkdir(parents=True, exist_ok=True)


def main() -> int:
    if not MASTER_PATH.exists():
        print(f"ERROR: Master not found at {MASTER_PATH}")
        return 1

    print(f"→ Reading Master from {MASTER_PATH}")
    wb = load_workbook(MASTER_PATH, read_only=True, data_only=True)
    ws = wb["Master"]
    header = [c.value for c in ws[1]]
    needed = ["master_id", "full_name", "email", "company", "investor_type",
              "investor_subtype", "gmail_labels", "title", "role",
              "contact_category", "contact_category_v2",
              "original_category_raw", "original_category_source"]
    missing = [c for c in needed if c not in header]
    if missing:
        print(f"!! Missing required columns: {missing}")
        return 2
    idx = {c: header.index(c) for c in needed}

    rows: list[dict] = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        d = {c: raw[idx[c]] if idx[c] < len(raw) else None for c in needed}
        rows.append(d)
    wb.close()
    print(f"  · loaded {len(rows)} rows")

    # ── Distribution side-by-side ───────────────────────────────────
    v1_counts: defaultdict[str, int] = defaultdict(int)
    v2_counts: defaultdict[str, int] = defaultdict(int)
    for r in rows:
        v1_counts[(r.get("contact_category") or "<empty>")] += 1
        v2_counts[(r.get("contact_category_v2") or "<empty>")] += 1

    v1_buckets = ["Principal", "Broker", "Lender", "Developer",
                  "Proveedor", "IA aplicaciones", "Uncategorized", "<empty>"]
    v2_buckets = ["Principal", "Broker", "Lender", "Operator",
                  "Developer", "Hotel Supply", "IA Supply", "Uncategorized", "<empty>"]

    print("\n" + "=" * 80)
    print("DISTRIBUTION · v1 vs v2 side-by-side")
    print("=" * 80)
    print(f"{'bucket':25s} {'v1':>10s} {'v2':>10s}    delta")
    print("-" * 80)
    all_buckets = sorted(set(v1_buckets) | set(v2_buckets))
    for b in all_buckets:
        v1 = v1_counts.get(b, 0)
        v2 = v2_counts.get(b, 0)
        delta = v2 - v1
        marker = "  ←" if abs(delta) > 50 else ""
        print(f"  {b:23s} {v1:>10d} {v2:>10d}   {delta:+5d}{marker}")
    total = len(rows)
    v1_uncat_pct = v1_counts.get("Uncategorized", 0) / total * 100
    v2_uncat_pct = v2_counts.get("Uncategorized", 0) / total * 100
    print(f"\n  % uncategorized · v1: {v1_uncat_pct:.1f}% · v2: {v2_uncat_pct:.1f}%")

    # ── Migration matrix (cross-tab) ──────────────────────────────
    matrix: defaultdict[tuple[str, str], int] = defaultdict(int)
    for r in rows:
        v1c = r.get("contact_category") or "<empty>"
        v2c = r.get("contact_category_v2") or "<empty>"
        matrix[(v1c, v2c)] += 1

    print("\n" + "=" * 80)
    print("MIGRATION MATRIX · v1 → v2 (rows = v1 · cols = v2)")
    print("=" * 80)
    v2_present = sorted({k[1] for k in matrix.keys()})
    print(f"{'v1 \\ v2':25s} " + " ".join(f"{c[:10]:>10s}" for c in v2_present))
    print("-" * (25 + 11 * len(v2_present)))
    for v1b in sorted({k[0] for k in matrix.keys()}):
        row_label = f"  {v1b:23s} "
        cells = []
        for v2b in v2_present:
            n = matrix.get((v1b, v2b), 0)
            cells.append(f"{n:>10d}" if n else f"{'.':>10s}")
        print(row_label + " ".join(cells))

    # ── Operator-split breakdown (v1 Principal → v2 buckets) ─────
    op_split: defaultdict[str, int] = defaultdict(int)
    for r in rows:
        if (r.get("contact_category") or "") == "Principal":
            op_split[r.get("contact_category_v2") or "<empty>"] += 1
    print("\n" + "=" * 80)
    print(f"OPERATOR SPLIT · {sum(op_split.values())} rows previously v1=Principal · v2 buckets:")
    print("=" * 80)
    for k, v in sorted(op_split.items(), key=lambda kv: -kv[1]):
        print(f"  v2={k:25s} : {v:5d}")

    # ── IA Supply newly classified ────────────────────────────────
    ia_new: list[dict] = []
    for r in rows:
        if (r.get("contact_category_v2") or "") == "IA Supply":
            ia_new.append({
                "master_id": r["master_id"],
                "full_name": r["full_name"],
                "email": r["email"],
                "company": r["company"],
                "investor_type": r["investor_type"],
                "v1_category": r["contact_category"],
            })
    print("\n" + "=" * 80)
    print(f"IA SUPPLY · {len(ia_new)} rows newly classified")
    print("=" * 80)
    for r in ia_new[:25]:
        print(f"  {r['email'] or '<no-email>':40s} · {(r['company'] or '<no-co>'):35s} · v1={r['v1_category']!r}")
    if len(ia_new) > 25:
        print(f"  ... and {len(ia_new) - 25} more")

    # ── Hotel Supply newly classified ─────────────────────────────
    hs_new: list[dict] = []
    for r in rows:
        if (r.get("contact_category_v2") or "") == "Hotel Supply":
            hs_new.append({
                "master_id": r["master_id"],
                "full_name": r["full_name"],
                "email": r["email"],
                "company": r["company"],
                "investor_type": r["investor_type"],
                "v1_category": r["contact_category"],
            })
    print("\n" + "=" * 80)
    print(f"HOTEL SUPPLY · {len(hs_new)} rows classified")
    print("=" * 80)
    for r in hs_new[:25]:
        print(f"  {r['email'] or '<no-email>':40s} · {(r['company'] or '<no-co>'):35s} · v1={r['v1_category']!r}")

    # ── Operator (v2) breakdown ──────────────────────────────────
    op_v2: list[dict] = []
    for r in rows:
        if (r.get("contact_category_v2") or "") == "Operator":
            op_v2.append({
                "company": r["company"],
                "investor_type": r["investor_type"],
                "v1_category": r["contact_category"],
            })
    print("\n" + "=" * 80)
    print(f"OPERATOR (v2) · {len(op_v2)} total · investor_type distribution")
    print("=" * 80)
    op_inv: defaultdict[str, int] = defaultdict(int)
    for r in op_v2:
        op_inv[r["investor_type"] or "<empty>"] += 1
    for k, v in sorted(op_inv.items(), key=lambda kv: -kv[1]):
        print(f"  investor_type={k:30s} : {v:5d}")

    # ── Top ambiguous (v2 Uncategorized) ──────────────────────────
    uncat_v2: list[dict] = []
    for r in rows:
        if (r.get("contact_category_v2") or "") == "Uncategorized":
            uncat_v2.append({
                "master_id": r["master_id"],
                "full_name": r["full_name"],
                "email": r["email"],
                "company": r["company"],
                "investor_type": r["investor_type"],
                "v1_category": r["contact_category"],
            })
    print("\n" + "=" * 80)
    print(f"TOP AMBIGUOUS · {len(uncat_v2)} rows v2=Uncategorized · first 20:")
    print("=" * 80)
    for r in uncat_v2[:20]:
        print(f"  {(r['email'] or '<no-email>'):40s} · {(r['company'] or '<no-co>'):28s} "
              f"· investor_type={r['investor_type']!r:20s} · v1={r['v1_category']!r}")

    # ── 20 random rows · human-eyeball sample ────────────────────
    random.seed(20260515)
    sample_idx = random.sample(range(len(rows)), 20)
    samples: list[dict] = []
    print("\n" + "=" * 80)
    print("20 RANDOM ROWS · human-eyeball sample")
    print("=" * 80)
    print(f"{'#':>2}  {'investor_type':18s} {'v1':12s} {'v2':14s} {'company':30s} {'email'}")
    print("-" * 120)
    for n, i in enumerate(sample_idx, start=1):
        r = rows[i]
        sheet_row = i + 2
        v1 = r.get("contact_category") or "<empty>"
        v2 = r.get("contact_category_v2") or "<empty>"
        flag = " ⚠" if v1 != v2 and v1 != "<empty>" else ""
        print(f"  {n:>2}  {(r['investor_type'] or '<empty>'):18s} {v1:12s} {v2:14s} "
              f"{(r['company'] or '<no-co>')[:30]:30s} {r['email'] or '<no-email>'}{flag}")
        samples.append({
            "row_n": n,
            "sheet_row": sheet_row,
            "master_id": r["master_id"],
            "investor_type": r["investor_type"],
            "contact_category_v1": r["contact_category"],
            "contact_category_v2": r["contact_category_v2"],
            "original_category_raw": r["original_category_raw"],
            "company": r["company"],
            "email": r["email"],
            "full_name": r["full_name"],
            "v1_v2_changed": v1 != v2 and v1 != "<empty>",
        })

    # ── Write JSON ───────────────────────────────────────────────
    ts = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    report = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "master_path": str(MASTER_PATH),
        "row_count": len(rows),
        "v1_distribution": dict(v1_counts),
        "v2_distribution": dict(v2_counts),
        "v1_pct_uncategorized": v1_uncat_pct,
        "v2_pct_uncategorized": v2_uncat_pct,
        "migration_matrix": {f"{k[0]}__TO__{k[1]}": v for k, v in matrix.items()},
        "operator_split_from_v1_principal": dict(op_split),
        "ia_supply_classifications": ia_new,
        "hotel_supply_classifications": hs_new,
        "operator_investor_type_breakdown": dict(op_inv),
        "uncategorized_v2_first_20": uncat_v2[:20],
        "random_sample_20": samples,
    }
    out = REPORTS_DIR / f"phase_b_classification_report_{ts}.json"
    out.write_text(json.dumps(report, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    print(f"\n→ Report written: {out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
