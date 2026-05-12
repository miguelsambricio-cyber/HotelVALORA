#!/usr/bin/env python3
"""
CONTACTOS DATASITE · Google Contacts enrichment pipeline.

Purpose
-------
Cross-reference Google Contacts (personal/professional address book) against
the canonical Datasite Master to:

  - identify which of your personal contacts already exist in the
    institutional deal-flow graph
  - surface NEW institutional contacts that aren't yet in the Master
  - classify the operator's relationship universe (investor · lender ·
    broker · operator · brand · consultant · advisor · personal · unknown)
  - generate suggested joins WITHOUT auto-mutating the Master

By design, this script does NOT modify the canonical Master. Every output
lands under `CONTACTOS DATASITE/google-contacts/` so the operator can
review the analysis before deciding what to merge.

Operator workflow
-----------------
1. In Google Contacts (contacts.google.com) → select all → Export → choose
   "Google CSV". Save the .csv to:
       CONTACTOS DATASITE/incoming/google-contacts/

2. Run:
       python scripts/contactos/ingest_google.py

3. Pipeline behaviour:
     - parse the Google CSV (Google's column schema is multi-value:
       Email 1 - Value · Email 2 - Value · Phone 1 - Value · etc.)
     - build GoogleContacts_Raw (verbatim copy, encoding-normalised)
     - build GoogleContacts_Normalized (canonical-shape rows · single
       primary email/phone + secondaries · LinkedIn extracted from
       websites · all metadata preserved)
     - classify each contact (high-level taxonomy · 9 buckets)
     - load existing Datasite Master
     - run identity resolution against the Master with the same
       priority used by the Master ingester
     - emit Relationship_Enriched workbook + 4 reports + a markdown
       analysis at google-contacts/relationship-enrichment-report.md
     - move the source CSV → old/google-contacts/

Re-running with an empty `incoming/google-contacts/` does nothing
(unlike the Master script which regenerates the Summary).

Privacy
-------
The whole `CONTACTOS DATASITE/` tree is .gitignore'd. Nothing under
google-contacts/ is ever committed to git.

Dependencies: openpyxl. The rest is stdlib.
"""

from __future__ import annotations

import csv
import difflib
import json
import re
import shutil
import sys
import unicodedata
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

# Windows console UTF-8 (matches ingest.py)
try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except (AttributeError, ValueError):
    pass

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl", file=sys.stderr)
    sys.exit(2)


# ── paths ───────────────────────────────────────────────────────────────────

REPO = Path(__file__).resolve().parent.parent.parent
ROOT = REPO / "CONTACTOS DATASITE"
INCOMING = ROOT / "incoming" / "google-contacts"
OLD = ROOT / "old" / "google-contacts"
GOOGLE_DIR = ROOT / "google-contacts"
RAW_DIR = GOOGLE_DIR / "raw"
NORM_DIR = GOOGLE_DIR / "normalized"
ENRICHED_DIR = GOOGLE_DIR / "enriched"
REPORTS = ROOT / "reports"
MASTER_FILE = ROOT / "master" / "metcub-contacts-master.xlsx"
ENRICHMENT_REPORT = GOOGLE_DIR / "relationship-enrichment-report.md"

for d in (INCOMING, OLD, GOOGLE_DIR, RAW_DIR, NORM_DIR, ENRICHED_DIR, REPORTS):
    d.mkdir(parents=True, exist_ok=True)


# ── helpers shared with ingest.py ───────────────────────────────────────────


def norm_text(value: Any) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    if not s:
        return ""
    s = unicodedata.normalize("NFC", s)
    s = re.sub(r"\s+", " ", s)
    return s


def norm_email(value: Any) -> str:
    s = norm_text(value).lower()
    s = re.sub(r"^<+|>+$", "", s)
    s = re.sub(r"^mailto:", "", s, flags=re.IGNORECASE)
    if "@" not in s or "." not in s.split("@")[-1]:
        return ""
    return s


def norm_phone(value: Any) -> str:
    s = norm_text(value)
    if not s:
        return ""
    plus = "+" if s.startswith("+") else ""
    digits = re.sub(r"\D", "", s)
    if not digits:
        return ""
    # Strip leading 00 (common ES export prefix) when no '+' was present
    if not plus and digits.startswith("00"):
        digits = digits[2:]
        plus = "+"
    return plus + digits


def norm_linkedin(value: Any) -> str:
    s = norm_text(value).lower()
    if not s:
        return ""
    if "linkedin" not in s:
        return ""
    s = re.sub(r"^https?://(www\.)?", "", s)
    s = re.sub(r"/$", "", s)
    return s


def company_key(value: Any) -> str:
    s = norm_text(value).lower()
    if not s:
        return ""
    s = re.sub(r"[,\.]", "", s)
    s = re.sub(r"\b(llc|inc|ltd|sl|sa|sarl|gmbh|kg|ag|nv|bv|plc|limited|llp)\b\s*$", "", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


# ── Google CSV parsing ──────────────────────────────────────────────────────
# Google Contacts exports use a multi-value column pattern:
#   "E-mail 1 - Label", "E-mail 1 - Value", "E-mail 2 - Value", ...
#   "Phone 1 - Label", "Phone 1 - Value", "Phone 2 - Value", ...
#   "Website 1 - Label", "Website 1 - Value", ...
#   "Organization 1 - Name", "Organization 1 - Title", ...
# Number of repetitions is unbounded · we detect them dynamically.
#
# Newer exports use "E-mail" / "Phone" without the space, or "Emails" /
# "Phones" combined into a single semicolon-separated column. We tolerate
# both shapes.


PERSONAL_EMAIL_DOMAINS = {
    "gmail.com", "googlemail.com", "hotmail.com", "outlook.com", "outlook.es",
    "live.com", "yahoo.com", "yahoo.es", "yahoo.co.uk", "icloud.com",
    "me.com", "mac.com", "protonmail.com", "proton.me", "yandex.com",
    "yandex.ru", "mail.ru", "aol.com", "msn.com", "ymail.com",
    "hotmail.es", "hotmail.fr", "hotmail.co.uk", "telefonica.net",
    "terra.es", "ono.com", "wanadoo.es", "movistar.es",
}


# Classification rules · ordered most-specific first.
# Matches against (company + title + notes + email_domain + labels)
CLASSIFICATION_RULES: list[tuple[re.Pattern[str], str]] = [
    # Investor-side: funds, family offices, sovereign, REIT
    (re.compile(r"\b(REIT|SOCIMI|inmobiliaria cotizada|listed property)\b", re.IGNORECASE), "investor"),
    (re.compile(r"\b(family\s*office|FO|patrimonio familiar)\b", re.IGNORECASE), "investor"),
    (re.compile(r"\b(private\s*equity|PE\s*fund|venture\s*capital|VC|fondo de inversi[óo]n|investment fund|asset management|capital partners|equity partners|sovereign wealth|pension fund)\b", re.IGNORECASE), "investor"),
    (re.compile(r"\b(institutional investor|inversor institucional|investor relations)\b", re.IGNORECASE), "investor"),
    # Lenders
    (re.compile(r"\b(banco|bank|financiador|prestamista|debt fund|debt|mortgage|lender|crédito)\b", re.IGNORECASE), "lender"),
    # Brokers / RE advisory firms
    (re.compile(r"\b(colliers|JLL|cushman|wakefield|CBRE|savills|knight frank|christie|HVS hodges)\b", re.IGNORECASE), "broker"),
    (re.compile(r"\b(broker|intermediario|asesor inmobiliario|real estate advisor|RE advisor|hotel broker)\b", re.IGNORECASE), "broker"),
    # Operators / hotel chains (operator first → falls to brand only if explicit franchise)
    (re.compile(r"\b(operator|operador|gestor hotelero|management company|hotel management|cadena hotelera|cadenas hotel)\b", re.IGNORECASE), "operator"),
    # Brand / franchise
    (re.compile(r"\b(brand|marca|franchisor|franchise)\b", re.IGNORECASE), "brand"),
    # Advisors / consultants
    (re.compile(r"\b(advisor|advisory|asesor[íi]a)\b", re.IGNORECASE), "advisor"),
    (re.compile(r"\b(consultor|consultora|consultancy|consulting)\b", re.IGNORECASE), "consultant"),
]


HOSPITALITY_HINT: re.Pattern[str] = re.compile(
    r"\b(hotel|hospitality|resort|hostel|aparthotel|short[\s-]?stay|extended[\s-]?stay|"
    r"hotelero|hotelera|hoteles|RevPAR|ADR|GOPPAR|key[s]?|room[s]?|llaves|habitaciones|"
    r"hospitality real estate|hotel real estate)\b",
    re.IGNORECASE,
)


def classify_high_level(company: str, title: str, notes: str, email: str, labels: str) -> str:
    """Return one of: investor · lender · broker · operator · brand · consultant · advisor · personal · unknown."""
    corpus = " ".join(s for s in (company, title, notes, labels) if s)
    for pattern, label in CLASSIFICATION_RULES:
        if pattern.search(corpus):
            return label
    # No institutional signal · decide personal vs unknown by email domain
    if email:
        dom = email.split("@", 1)[-1].lower()
        if dom in PERSONAL_EMAIL_DOMAINS and not company.strip():
            return "personal"
        if dom in PERSONAL_EMAIL_DOMAINS:
            return "personal"  # gmail+org could be either · default personal
    if not company.strip():
        return "personal" if email else "unknown"
    return "unknown"


def email_domain_kind(email: str) -> str:
    if not email or "@" not in email:
        return "unknown"
    dom = email.split("@", 1)[-1].lower()
    return "personal" if dom in PERSONAL_EMAIL_DOMAINS else "institutional"


def hotel_focus(corpus_parts: list[str]) -> str:
    corpus = " ".join(s for s in corpus_parts if s)
    if not corpus.strip():
        return "Unknown"
    hits = HOSPITALITY_HINT.findall(corpus)
    if len(hits) >= 2:
        return "Yes"
    if len(hits) == 1:
        return "Likely"
    return "No"


# ── Google CSV reader ───────────────────────────────────────────────────────


@dataclass
class GoogleRow:
    raw: dict[str, str]                       # verbatim CSV row · field → string
    full_name: str = ""
    emails: list[str] = field(default_factory=list)        # canonical (lowercased)
    email_labels: list[str] = field(default_factory=list)  # parallel labels
    phones: list[str] = field(default_factory=list)
    phone_labels: list[str] = field(default_factory=list)
    websites: list[str] = field(default_factory=list)
    linkedin: str = ""
    company: str = ""
    title: str = ""
    department: str = ""
    notes: str = ""
    labels: str = ""                          # Google "Labels" column (categories)
    birthday: str = ""
    address: str = ""
    nickname: str = ""


def _split_multi(raw: dict[str, str], pattern: re.Pattern[str]) -> list[tuple[str, str]]:
    """Find columns like 'E-mail 1 - Value' / 'E-mail 2 - Value' and return [(label, value)]."""
    out: list[tuple[str, str]] = []
    indices: set[int] = set()
    for col in raw:
        m = pattern.match(col)
        if m:
            indices.add(int(m.group(1)))
    for i in sorted(indices):
        label = ""
        value = ""
        # Try a few label/value column-name variations
        for label_col in (f"E-mail {i} - Label", f"Phone {i} - Label",
                          f"Email {i} - Label", f"Website {i} - Label",
                          f"Organization {i} - Type"):
            if label_col in raw and raw[label_col]:
                label = raw[label_col]
                break
        for value_col in (f"E-mail {i} - Value", f"Phone {i} - Value",
                          f"Email {i} - Value", f"Website {i} - Value"):
            if value_col in raw and raw[value_col]:
                value = raw[value_col]
                break
        if value:
            out.append((label, value))
    return out


_EMAIL_COL_RE = re.compile(r"^E-?mail\s+(\d+)\s*-\s*Value$", re.IGNORECASE)
_PHONE_COL_RE = re.compile(r"^Phone\s+(\d+)\s*-\s*Value$", re.IGNORECASE)
_WEBSITE_COL_RE = re.compile(r"^Website\s+(\d+)\s*-\s*Value$", re.IGNORECASE)
_ORG_NAME_RE = re.compile(r"^Organization\s+(\d+)\s*-\s*Name$", re.IGNORECASE)


def parse_google_csv(path: Path) -> list[GoogleRow]:
    """Read a Google Contacts CSV · tolerant of column-name variations."""
    # Google CSV is UTF-8 BOM. csv.DictReader handles BOM with utf-8-sig.
    rows: list[GoogleRow] = []
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for raw in reader:
            r = {k: norm_text(v) for k, v in raw.items() if k}
            row = GoogleRow(raw=r)

            # Name · prefer composed Full Name, else assemble
            full = r.get("Name") or r.get("Full Name") or ""
            if not full:
                first = r.get("Given Name") or r.get("First Name") or ""
                middle = r.get("Additional Name") or r.get("Middle Name") or ""
                last = r.get("Family Name") or r.get("Last Name") or ""
                full = " ".join(p for p in (first, middle, last) if p)
            row.full_name = norm_text(full)
            row.nickname = norm_text(r.get("Nickname"))

            # Emails (multi-value)
            for label, value in _split_multi(r, _EMAIL_COL_RE):
                ne = norm_email(value)
                if ne:
                    row.emails.append(ne)
                    row.email_labels.append(label or "")
            # Newer combined "E-mail" single column
            if not row.emails:
                single = r.get("E-mail") or r.get("Email") or r.get("Emails")
                if single:
                    for part in re.split(r"[;,\s]+", single):
                        ne = norm_email(part)
                        if ne:
                            row.emails.append(ne)
                            row.email_labels.append("")

            # Phones (multi-value)
            for label, value in _split_multi(r, _PHONE_COL_RE):
                np = norm_phone(value)
                if np:
                    row.phones.append(np)
                    row.phone_labels.append(label or "")
            if not row.phones:
                single = r.get("Phone") or r.get("Phones")
                if single:
                    for part in re.split(r"[;,]+", single):
                        np = norm_phone(part)
                        if np:
                            row.phones.append(np)
                            row.phone_labels.append("")

            # Websites · LinkedIn extracted as priority
            for _label, value in _split_multi(r, _WEBSITE_COL_RE):
                w = norm_text(value)
                if not w:
                    continue
                row.websites.append(w)
                if "linkedin" in w.lower() and not row.linkedin:
                    row.linkedin = norm_linkedin(w)

            # Organization · use the FIRST org (modern Google CSV uses one
            # Organization slot; older multi-org pattern still works).
            for col_name, dst in (("Organization 1 - Name", "company"),
                                  ("Organization Name", "company"),
                                  ("Organization 1 - Title", "title"),
                                  ("Organization Title", "title"),
                                  ("Organization 1 - Department", "department"),
                                  ("Organization Department", "department")):
                if col_name in r and r[col_name]:
                    setattr(row, dst, r[col_name])

            row.notes = norm_text(r.get("Notes"))
            row.labels = norm_text(r.get("Labels") or r.get("Group Membership") or "")
            row.birthday = norm_text(r.get("Birthday"))

            # Address · combine the first address block (Type/Street/City/Region/Postal/Country)
            addr_parts = []
            for k in ("Address 1 - Street", "Address 1 - Extended Address",
                     "Address 1 - City", "Address 1 - Region",
                     "Address 1 - Postal Code", "Address 1 - Country"):
                if r.get(k):
                    addr_parts.append(r[k])
            row.address = ", ".join(addr_parts)

            # Reject pure-empty rows
            if not (row.full_name or row.emails or row.phones or row.company):
                continue
            rows.append(row)
    return rows


# ── Normalize Google rows to canonical-shape dicts ──────────────────────────


def synth_google_id(row: GoogleRow) -> str:
    import hashlib
    primary_email = row.emails[0] if row.emails else ""
    primary_phone = row.phones[0] if row.phones else ""
    if primary_email:
        key = f"email:{primary_email}"
    elif row.linkedin:
        key = f"li:{row.linkedin}"
    elif primary_phone:
        key = f"phone:{primary_phone}"
    else:
        key = f"name:{row.full_name.lower()}|company:{company_key(row.company)}"
    return hashlib.md5(key.encode("utf-8")).hexdigest()[:16]


def normalize_google_row(row: GoogleRow, source_file: str, batch_id: str) -> dict[str, Any]:
    primary_email = row.emails[0] if row.emails else ""
    secondary_emails = "; ".join(row.emails[1:])
    primary_phone = row.phones[0] if row.phones else ""
    secondary_phones = "; ".join(row.phones[1:])
    classification = classify_high_level(
        row.company, row.title, row.notes, primary_email, row.labels,
    )
    hf = hotel_focus([row.company, row.title, row.notes, row.labels])
    return {
        "google_id": synth_google_id(row),
        "full_name": row.full_name,
        "primary_email": primary_email,
        "secondary_emails": secondary_emails,
        "all_emails": "; ".join(row.emails),
        "primary_phone": primary_phone,
        "secondary_phones": secondary_phones,
        "all_phones": "; ".join(row.phones),
        "linkedin": row.linkedin,
        "websites": "; ".join(row.websites),
        "company": row.company,
        "title": row.title,
        "department": row.department,
        "labels": row.labels,
        "notes": row.notes,
        "birthday": row.birthday,
        "address": row.address,
        "nickname": row.nickname,
        "classification": classification,
        "email_domain_kind": email_domain_kind(primary_email),
        "hotel_focus": hf,
        "has_company": "yes" if row.company.strip() else "no",
        "email_count": len(row.emails),
        "phone_count": len(row.phones),
        "source_file": source_file,
        "batch_id": batch_id,
        "ingested_at": now_iso(),
    }


# ── Master loader (read-only, for identity resolution) ──────────────────────


def load_master() -> list[dict[str, Any]]:
    if not MASTER_FILE.exists():
        print(f"  ⚠  Master not found at {MASTER_FILE} · enrichment will only flag unique-in-Google rows")
        return []
    wb = load_workbook(MASTER_FILE, data_only=True, read_only=True)
    if "Master" not in wb.sheetnames:
        return []
    ws = wb["Master"]
    if ws.max_row < 2:
        return []
    header = [norm_text(c.value) for c in ws[1]]
    out = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        if all(cell in (None, "") for cell in raw):
            continue
        rec: dict[str, Any] = {}
        for i, h in enumerate(header):
            if i >= len(raw):
                break
            val = raw[i]
            rec[h] = "" if val is None else val
        out.append(rec)
    return out


# ── Identity resolution ─────────────────────────────────────────────────────


def build_master_indices(master: list[dict[str, Any]]) -> dict[str, dict[Any, int]]:
    """Pre-compute lookup indices on the Master for fast matching."""
    by_email: dict[str, int] = {}
    by_phone: dict[str, int] = {}
    by_linkedin: dict[str, int] = {}
    by_name_company: dict[tuple[str, str], int] = {}
    for i, row in enumerate(master):
        e = norm_email(row.get("email"))
        if e:
            by_email[e] = i
        p = norm_phone(row.get("phone"))
        if p:
            by_phone[p] = i
        li = norm_linkedin(row.get("linkedin"))
        if li:
            by_linkedin[li] = i
        fn = norm_text(row.get("full_name")).lower()
        ck = company_key(row.get("company"))
        if fn and ck:
            by_name_company[(fn, ck)] = i
    return {
        "email": by_email, "phone": by_phone, "linkedin": by_linkedin,
        "name_company": by_name_company,
    }


def resolve_identity(g: dict[str, Any], indices: dict[str, dict[Any, int]],
                     master: list[dict[str, Any]]) -> tuple[int | None, str, float]:
    """Returns (master_idx, strategy, score). None when no match."""
    # 1. Email exact (across primary + secondaries)
    candidate_emails = [g["primary_email"]] + ([s.strip() for s in (g["secondary_emails"] or "").split(";") if s.strip()])
    for e in candidate_emails:
        if not e:
            continue
        idx = indices["email"].get(e.lower())
        if idx is not None:
            return idx, "email_exact", 1.0
    # 2. Phone exact
    candidate_phones = [g["primary_phone"]] + ([s.strip() for s in (g["secondary_phones"] or "").split(";") if s.strip()])
    for p in candidate_phones:
        if not p:
            continue
        idx = indices["phone"].get(p)
        if idx is not None:
            return idx, "phone_exact", 0.98
    # 3. LinkedIn exact
    if g["linkedin"]:
        idx = indices["linkedin"].get(g["linkedin"])
        if idx is not None:
            return idx, "linkedin_exact", 0.95
    # 4. Name + company exact
    name_key = (g["full_name"].lower(), company_key(g["company"]))
    if name_key[0] and name_key[1]:
        idx = indices["name_company"].get(name_key)
        if idx is not None:
            return idx, "name_company_exact", 0.92
    # 5. Fuzzy fallback · same company_key + name similarity ≥ 0.88
    target_company = name_key[1]
    if target_company and name_key[0]:
        best_idx, best_score = None, 0.0
        for i, row in enumerate(master):
            if company_key(row.get("company")) != target_company:
                continue
            score = difflib.SequenceMatcher(
                None, name_key[0], norm_text(row.get("full_name")).lower()
            ).ratio()
            if score > best_score:
                best_score = score
                best_idx = i
        if best_idx is not None and best_score >= 0.88:
            return best_idx, "fuzzy", best_score
    return None, "none", 0.0


def recommend_action(strategy: str, score: float, g: dict[str, Any],
                     master_row: dict[str, Any] | None) -> tuple[str, str]:
    """(action, reason). action ∈ {MERGE, INSERT, REVIEW, NO_OP}."""
    if master_row is None:
        # No match · candidate for INSERT depending on quality
        if g["classification"] == "personal" and not g["has_company"] == "yes":
            return "NO_OP", "personal contact · no company · not institutional signal"
        if not g["primary_email"] and not g["linkedin"] and not (g["full_name"] and g["company"]):
            return "REVIEW", "missing identifier · cannot dedup against future ingests"
        if g["classification"] in ("investor", "lender", "broker", "operator", "brand", "advisor", "consultant"):
            return "INSERT", f"institutional ({g['classification']}) not in Master"
        return "REVIEW", f"unclassified · {g['classification']} · consider adding to Master if hospitality-relevant"
    # Match found
    if strategy == "email_exact":
        return "MERGE", "exact email match · safe to merge LinkedIn/phone/notes"
    if strategy == "phone_exact":
        return "MERGE", "exact phone match · email differs · merge contact info"
    if strategy == "linkedin_exact":
        return "MERGE", "exact LinkedIn match · merge email/phone"
    if strategy == "name_company_exact":
        return "MERGE", "name + company exact match · merge metadata"
    if strategy == "fuzzy":
        return "REVIEW", f"fuzzy name match ({score:.2f}) within same company · manual review"
    return "NO_OP", "no match"


# ── output writers ──────────────────────────────────────────────────────────


def write_raw_csv(rows: list[GoogleRow], batch_id: str, source_file: str) -> Path:
    out = RAW_DIR / f"google_raw_{batch_id}.csv"
    # Union of all observed columns across rows
    cols = list({k for r in rows for k in r.raw.keys()})
    cols.sort()
    with out.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=["__source_file__", "__batch_id__"] + cols)
        w.writeheader()
        for r in rows:
            rec = {k: v for k, v in r.raw.items()}
            rec["__source_file__"] = source_file
            rec["__batch_id__"] = batch_id
            w.writerow(rec)
    return out


NORM_FIELDS = [
    "google_id", "full_name", "primary_email", "secondary_emails", "all_emails",
    "primary_phone", "secondary_phones", "all_phones", "linkedin", "websites",
    "company", "title", "department", "labels", "notes", "birthday", "address",
    "nickname", "classification", "email_domain_kind", "hotel_focus",
    "has_company", "email_count", "phone_count",
    "source_file", "batch_id", "ingested_at",
]


def write_normalized_csv(normalized: list[dict[str, Any]], batch_id: str) -> Path:
    out = NORM_DIR / f"google_normalized_{batch_id}.csv"
    with out.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=NORM_FIELDS)
        w.writeheader()
        for r in normalized:
            w.writerow({k: r.get(k, "") for k in NORM_FIELDS})
    return out


def write_enriched_workbook(normalized: list[dict[str, Any]],
                            enriched: list[dict[str, Any]],
                            within_google_dups: list[dict[str, Any]],
                            suggested_joins: list[dict[str, Any]],
                            new_unique: list[dict[str, Any]],
                            batch_id: str) -> Path:
    out = ENRICHED_DIR / f"google_enriched_{batch_id}.xlsx"
    wb = Workbook()
    wb.remove(wb.active)

    HEADER_FILL = PatternFill("solid", fgColor="0F172A")
    HEADER_FONT = Font(bold=True, color="A3E635", size=10)

    def write_sheet(name: str, cols: list[str], rows: list[dict[str, Any]]) -> None:
        ws = wb.create_sheet(name)
        ws.append(cols)
        for c in range(1, len(cols) + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 22
        ws.freeze_panes = "A2"
        for r in rows:
            ws.append([r.get(c, "") for c in cols])
        for c, col in enumerate(cols, 1):
            ws.column_dimensions[get_column_letter(c)].width = min(40, max(12, len(col) + 4))

    # GoogleContacts_Normalized · the canonical-shape view of Google input
    write_sheet("GoogleContacts_Normalized", NORM_FIELDS, normalized)

    # Relationship_Enriched · per-Google-row resolution outcome
    enriched_cols = [
        "google_id", "google_full_name", "google_primary_email", "google_company",
        "google_classification", "google_hotel_focus",
        "match_status", "match_strategy", "match_score",
        "master_id", "master_full_name", "master_email", "master_company",
        "master_investor_type", "master_pipeline_state",
        "recommended_action", "recommendation_reason",
        "batch_id",
    ]
    write_sheet("Relationship_Enriched", enriched_cols, enriched)

    # Suggested-Joins · only rows recommended for MERGE or INSERT
    write_sheet("Suggested-Joins", enriched_cols, suggested_joins)

    # New-Unique-Contacts · in Google but not in Master · candidate INSERTs
    write_sheet("New-Unique-Contacts", NORM_FIELDS, new_unique)

    # Within-Google-Duplicates
    write_sheet("Within-Google-Duplicates", [
        "duplicate_strategy", "google_id_a", "google_id_b",
        "full_name_a", "full_name_b", "email_a", "email_b",
        "company_a", "company_b", "similarity",
    ], within_google_dups)

    wb.save(out)
    return out


def detect_within_google_duplicates(normalized: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Find duplicates inside the Google CSV itself (before joining Master)."""
    dups: list[dict[str, Any]] = []

    # 1. Same email
    by_email: dict[str, list[int]] = defaultdict(list)
    for i, r in enumerate(normalized):
        if r["primary_email"]:
            by_email[r["primary_email"]].append(i)
    for email, idxs in by_email.items():
        if len(idxs) < 2:
            continue
        a = normalized[idxs[0]]
        for b_idx in idxs[1:]:
            b = normalized[b_idx]
            dups.append({
                "duplicate_strategy": "same_email",
                "google_id_a": a["google_id"], "google_id_b": b["google_id"],
                "full_name_a": a["full_name"], "full_name_b": b["full_name"],
                "email_a": a["primary_email"], "email_b": b["primary_email"],
                "company_a": a["company"], "company_b": b["company"],
                "similarity": "1.0",
            })

    # 2. Same LinkedIn
    by_li: dict[str, list[int]] = defaultdict(list)
    for i, r in enumerate(normalized):
        if r["linkedin"]:
            by_li[r["linkedin"]].append(i)
    for li, idxs in by_li.items():
        if len(idxs) < 2:
            continue
        a = normalized[idxs[0]]
        for b_idx in idxs[1:]:
            b = normalized[b_idx]
            dups.append({
                "duplicate_strategy": "same_linkedin",
                "google_id_a": a["google_id"], "google_id_b": b["google_id"],
                "full_name_a": a["full_name"], "full_name_b": b["full_name"],
                "email_a": a["primary_email"], "email_b": b["primary_email"],
                "company_a": a["company"], "company_b": b["company"],
                "similarity": "1.0",
            })

    # 3. Same name+company
    by_nc: dict[tuple[str, str], list[int]] = defaultdict(list)
    for i, r in enumerate(normalized):
        key = (r["full_name"].lower(), company_key(r["company"]))
        if key[0] and key[1]:
            by_nc[key].append(i)
    for _key, idxs in by_nc.items():
        if len(idxs) < 2:
            continue
        a = normalized[idxs[0]]
        for b_idx in idxs[1:]:
            b = normalized[b_idx]
            dups.append({
                "duplicate_strategy": "name_company_exact",
                "google_id_a": a["google_id"], "google_id_b": b["google_id"],
                "full_name_a": a["full_name"], "full_name_b": b["full_name"],
                "email_a": a["primary_email"], "email_b": b["primary_email"],
                "company_a": a["company"], "company_b": b["company"],
                "similarity": "1.0",
            })
    return dups


# ── Markdown report ─────────────────────────────────────────────────────────


def write_enrichment_report(normalized: list[dict[str, Any]],
                            enriched: list[dict[str, Any]],
                            within_dups: list[dict[str, Any]],
                            master_count: int,
                            batch_id: str,
                            source_files: list[str]) -> Path:
    n_total = len(normalized)
    by_class = Counter(r["classification"] for r in normalized)
    by_status = Counter(r["match_status"] for r in enriched)
    by_strategy = Counter(r["match_strategy"] for r in enriched)
    by_action = Counter(r["recommended_action"] for r in enriched)
    by_hotel = Counter(r["hotel_focus"] for r in normalized)
    by_domain = Counter(r["email_domain_kind"] for r in normalized)

    companies = Counter()
    contacts_per_company: dict[str, int] = defaultdict(int)
    for r in normalized:
        c = r["company"].strip()
        if c:
            companies[c] += 1
            contacts_per_company[c] += 1

    no_company = sum(1 for r in normalized if r["has_company"] == "no")
    no_email = sum(1 for r in normalized if not r["primary_email"])
    no_phone = sum(1 for r in normalized if not r["primary_phone"])
    no_linkedin = sum(1 for r in normalized if not r["linkedin"])

    # Network clusters = companies with multiple Google contacts AND existing in Master
    enriched_by_google_id = {r["google_id"]: r for r in enriched}
    clusters: dict[str, dict[str, Any]] = {}
    for r in normalized:
        c = r["company"].strip()
        if not c or contacts_per_company[c] < 2:
            continue
        slot = clusters.setdefault(c, {"google_contacts": 0, "matched_to_master": 0})
        slot["google_contacts"] += 1
        e = enriched_by_google_id.get(r["google_id"])
        if e and e["match_status"] == "matched":
            slot["matched_to_master"] += 1
    cluster_rows = sorted(clusters.items(), key=lambda kv: (-kv[1]["google_contacts"], kv[0].lower()))[:20]

    # Overlap with Datasite Master · top companies that appear in both
    google_companies_set = {c.lower() for c in companies if c}
    master_companies_in_enriched: Counter[str] = Counter()
    for r in enriched:
        if r["match_status"] == "matched" and r["master_company"]:
            master_companies_in_enriched[r["master_company"]] += 1
    overlap_top = master_companies_in_enriched.most_common(20)

    # New-unique top companies (only companies whose Google contacts ALL had no Master match)
    company_match_count: dict[str, int] = defaultdict(int)
    company_total_count: dict[str, int] = defaultdict(int)
    for r in normalized:
        c = r["company"].strip()
        if not c:
            continue
        company_total_count[c] += 1
        e = enriched_by_google_id.get(r["google_id"])
        if e and e["match_status"] == "matched":
            company_match_count[c] += 1
    new_unique_companies = sorted(
        ((c, company_total_count[c]) for c in company_total_count if company_match_count[c] == 0),
        key=lambda kv: (-kv[1], kv[0].lower()),
    )[:20]

    md = []
    md.append(f"# Relationship Enrichment Report · Google × Datasite Master\n")
    md.append(f"**Batch:** `{batch_id}`")
    md.append(f"**Generated:** {now_iso()}")
    md.append(f"**Sources:** {', '.join(source_files) if source_files else '(no incoming files)'}")
    md.append(f"**Datasite Master loaded:** {master_count} canonical contacts\n")

    md.append("## 1 · Totals\n")
    md.append(f"- **Google contacts processed:** {n_total}")
    md.append(f"- **Within-Google duplicates detected:** {len(within_dups)}")
    md.append(f"- **Matched to Datasite Master:** {by_status.get('matched', 0)}")
    md.append(f"- **Unmatched (new to Master):** {by_status.get('unmatched', 0)}")
    md.append(f"- **Multi-match (review):** {by_status.get('multi-match', 0)}\n")

    md.append("## 2 · Recommended Actions\n")
    md.append("| Action | Count | Meaning |")
    md.append("|---|---|---|")
    md.append(f"| MERGE | {by_action.get('MERGE', 0)} | safe merge into existing Master row · enrich email/phone/LinkedIn/notes |")
    md.append(f"| INSERT | {by_action.get('INSERT', 0)} | institutional contact NOT in Master · candidate for new row |")
    md.append(f"| REVIEW | {by_action.get('REVIEW', 0)} | fuzzy match OR unclassified institutional · manual decision |")
    md.append(f"| NO_OP | {by_action.get('NO_OP', 0)} | personal contact without company · skip |\n")

    md.append("## 3 · Match strategies used\n")
    for s, n in sorted(by_strategy.items(), key=lambda kv: -kv[1]):
        md.append(f"- `{s}` · {n}")
    md.append("")

    md.append("## 4 · High-level classification (Google taxonomy)\n")
    md.append("| Bucket | Count |")
    md.append("|---|---|")
    for k in ("investor", "lender", "broker", "operator", "brand", "consultant", "advisor", "personal", "unknown"):
        md.append(f"| {k} | {by_class.get(k, 0)} |")
    md.append("")

    md.append("## 5 · Hotel focus density\n")
    for k, n in sorted(by_hotel.items(), key=lambda kv: -kv[1]):
        md.append(f"- `{k}` · {n}")
    md.append("")

    md.append("## 6 · Email domain kind\n")
    md.append(f"- institutional · {by_domain.get('institutional', 0)}")
    md.append(f"- personal · {by_domain.get('personal', 0)}")
    md.append(f"- unknown / missing · {by_domain.get('unknown', 0)}\n")

    md.append("## 7 · Overlap with Datasite Master\n")
    md.append(f"Companies where at least one Google contact matched a Master row · top 20 by matched-contact count:\n")
    if overlap_top:
        md.append("| Master Company | Matched Google contacts |")
        md.append("|---|---|")
        for c, n in overlap_top:
            md.append(f"| {c} | {n} |")
    else:
        md.append("_(none)_")
    md.append("")

    md.append("## 8 · New unique institutional contacts in Google\n")
    md.append(f"Companies where Google contacts did NOT match any Master row · top 20 by Google contact count (candidates for INSERT):\n")
    if new_unique_companies:
        md.append("| Company | Google contacts |")
        md.append("|---|---|")
        for c, n in new_unique_companies:
            md.append(f"| {c} | {n} |")
    else:
        md.append("_(none — every Google contact mapped to a known Master company)_")
    md.append("")

    md.append("## 9 · Relationship density\n")
    if companies:
        top_companies = companies.most_common(25)
        md.append("Companies with the most Google contacts (relationship coverage):\n")
        md.append("| Company | Google contacts |")
        md.append("|---|---|")
        for c, n in top_companies:
            md.append(f"| {c} | {n} |")
        md.append("")

    md.append("## 10 · Inferred network clusters\n")
    md.append("Companies with multiple Google contacts where at least one matches the Datasite Master · these are the live institutional relationships you have personal coverage on:\n")
    if cluster_rows:
        md.append("| Company | Google contacts | Matched to Master |")
        md.append("|---|---|---|")
        for c, info in cluster_rows:
            md.append(f"| {c} | {info['google_contacts']} | {info['matched_to_master']} |")
    else:
        md.append("_(no multi-contact clusters)_")
    md.append("")

    md.append("## 11 · Missing metadata\n")
    md.append(f"- Contacts without company affiliation · **{no_company}** ({100*no_company/n_total:.1f}%)" if n_total else "")
    md.append(f"- Contacts without primary email · **{no_email}** ({100*no_email/n_total:.1f}%)" if n_total else "")
    md.append(f"- Contacts without phone · **{no_phone}** ({100*no_phone/n_total:.1f}%)" if n_total else "")
    md.append(f"- Contacts without LinkedIn · **{no_linkedin}** ({100*no_linkedin/n_total:.1f}%)" if n_total else "")
    md.append("")

    md.append("---\n")
    md.append("## Operator next steps\n")
    md.append("- **MERGE rows** (action=MERGE): review the `Suggested-Joins` sheet in `google_enriched_<batch_id>.xlsx` → cherry-pick the rows you trust → manually update Master (or wait for Phase 2.A · automated merge tool)")
    md.append("- **INSERT rows** (action=INSERT): institutional contacts NOT yet in Master · likely candidates to add to a future Datasite outreach push")
    md.append("- **REVIEW rows** (action=REVIEW): fuzzy matches OR unclassified contacts with company info · 5-min triage each")
    md.append("- **NO_OP rows**: personal contacts without institutional relevance · safely ignored\n")

    md.append("This report does NOT modify the Master · all changes require an explicit operator step.\n")

    out = ENRICHMENT_REPORT
    out.write_text("\n".join(md), encoding="utf-8")
    return out


# ── main ─────────────────────────────────────────────────────────────────────


def main() -> int:
    incoming_files = sorted(
        [p for p in INCOMING.iterdir() if p.is_file() and p.suffix.lower() == ".csv"],
        key=lambda p: p.stat().st_mtime,
    )
    print(f"→ Google Contacts enrichment · {len(incoming_files)} CSV(s) in incoming/google-contacts/")

    if not incoming_files:
        print("  nothing to do · drop a CSV from contacts.google.com (Export → Google CSV) into")
        print(f"  {INCOMING}")
        return 0

    master = load_master()
    print(f"  Master loaded: {len(master)} contacts")
    indices = build_master_indices(master)

    all_normalized: list[dict[str, Any]] = []
    all_enriched: list[dict[str, Any]] = []
    all_raw_rows: list[GoogleRow] = []
    processed_files: list[str] = []
    last_batch_id = ""

    for src in incoming_files:
        batch_id = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
        last_batch_id = batch_id
        print(f"\n→ processing {src.name} · batch_id={batch_id}")

        try:
            raw_rows = parse_google_csv(src)
        except Exception as err:
            print(f"  ✗ parse failed: {err}")
            continue
        print(f"  parsed {len(raw_rows)} rows from Google CSV")
        all_raw_rows.extend(raw_rows)

        # Normalize
        normalized = [normalize_google_row(r, src.name, batch_id) for r in raw_rows]
        all_normalized.extend(normalized)

        # Identity resolution against Master
        for g in normalized:
            idx, strategy, score = resolve_identity(g, indices, master)
            master_row = master[idx] if idx is not None else None
            action, reason = recommend_action(strategy, score, g, master_row)
            all_enriched.append({
                "google_id": g["google_id"],
                "google_full_name": g["full_name"],
                "google_primary_email": g["primary_email"],
                "google_company": g["company"],
                "google_classification": g["classification"],
                "google_hotel_focus": g["hotel_focus"],
                "match_status": "matched" if idx is not None else "unmatched",
                "match_strategy": strategy,
                "match_score": f"{score:.3f}" if idx is not None else "",
                "master_id": (master_row or {}).get("master_id", ""),
                "master_full_name": (master_row or {}).get("full_name", ""),
                "master_email": (master_row or {}).get("email", ""),
                "master_company": (master_row or {}).get("company", ""),
                "master_investor_type": (master_row or {}).get("investor_type", ""),
                "master_pipeline_state": (master_row or {}).get("pipeline_state", ""),
                "recommended_action": action,
                "recommendation_reason": reason,
                "batch_id": batch_id,
            })

        # Write raw + normalized for this batch
        write_raw_csv(raw_rows, batch_id, src.name)
        write_normalized_csv(normalized, batch_id)

        # Move source → old/google-contacts/
        dest = OLD / src.name
        if dest.exists():
            stem = dest.stem
            dest = OLD / f"{stem}.{batch_id}{src.suffix}"
        shutil.move(str(src), str(dest))
        processed_files.append(src.name)
        print(f"  moved {src.name} → old/google-contacts/{dest.name}")

    # Within-Google duplicates across the full normalized set
    within_dups = detect_within_google_duplicates(all_normalized)
    print(f"\n  within-google duplicates: {len(within_dups)}")

    # Slice enriched for "Suggested-Joins" and "New-Unique-Contacts" sheets
    suggested = [r for r in all_enriched if r["recommended_action"] in ("MERGE", "INSERT", "REVIEW")]
    new_unique_ids = {r["google_id"] for r in all_enriched if r["match_status"] == "unmatched"}
    new_unique_rows = [r for r in all_normalized if r["google_id"] in new_unique_ids]

    # Workbook
    enriched_path = write_enriched_workbook(
        all_normalized, all_enriched, within_dups, suggested, new_unique_rows, last_batch_id,
    )
    print(f"  enriched workbook → {enriched_path}")

    # Per-batch reports
    write_jsonl_line(REPORTS / "google-ingestion-log.jsonl", {
        "batch_id": last_batch_id,
        "source_files": processed_files,
        "rows_total": len(all_normalized),
        "matched": sum(1 for r in all_enriched if r["match_status"] == "matched"),
        "unmatched": sum(1 for r in all_enriched if r["match_status"] == "unmatched"),
        "within_google_duplicates": len(within_dups),
        "suggested_merge": sum(1 for r in all_enriched if r["recommended_action"] == "MERGE"),
        "suggested_insert": sum(1 for r in all_enriched if r["recommended_action"] == "INSERT"),
        "completed_at": now_iso(),
    })

    # Per-batch CSV reports (operator-readable)
    write_csv_report(REPORTS / f"google-identity-resolution_{last_batch_id}.csv", all_enriched, [
        "google_id", "google_full_name", "google_primary_email", "google_company",
        "google_classification", "match_status", "match_strategy", "match_score",
        "master_id", "master_full_name", "master_company", "master_investor_type",
        "recommended_action", "recommendation_reason",
    ])
    write_csv_report(REPORTS / f"google-overlap-analysis_{last_batch_id}.csv", [
        r for r in all_enriched if r["match_status"] == "matched"
    ], [
        "google_full_name", "google_primary_email", "google_company",
        "master_full_name", "master_email", "master_company",
        "master_investor_type", "master_pipeline_state",
        "match_strategy", "match_score", "recommended_action",
    ])
    write_csv_report(REPORTS / f"google-within-duplicates_{last_batch_id}.csv", within_dups, [
        "duplicate_strategy", "google_id_a", "google_id_b",
        "full_name_a", "full_name_b", "email_a", "email_b",
        "company_a", "company_b", "similarity",
    ])
    write_csv_report(REPORTS / f"google-suggested-joins_{last_batch_id}.csv", suggested, [
        "google_id", "google_full_name", "google_primary_email", "google_company",
        "google_classification", "match_status", "match_strategy", "match_score",
        "master_id", "master_full_name", "master_company", "master_investor_type",
        "recommended_action", "recommendation_reason",
    ])

    # Markdown report
    report_path = write_enrichment_report(
        all_normalized, all_enriched, within_dups, len(master), last_batch_id, processed_files,
    )
    print(f"  enrichment report → {report_path}")

    # Summary console output
    print()
    print(f"✓ done · {len(all_normalized)} Google contacts processed across {len(processed_files)} file(s)")
    matched = sum(1 for r in all_enriched if r["match_status"] == "matched")
    print(f"  · {matched} matched to Master ({100*matched/max(1,len(all_normalized)):.1f}%)")
    print(f"  · {len(all_normalized)-matched} unmatched")
    by_action = Counter(r["recommended_action"] for r in all_enriched)
    print(f"  · actions: MERGE={by_action['MERGE']} · INSERT={by_action['INSERT']} · REVIEW={by_action['REVIEW']} · NO_OP={by_action['NO_OP']}")
    print()
    print(f"  Open: {ENRICHMENT_REPORT}")
    return 0


def write_jsonl_line(path: Path, record: dict[str, Any]) -> None:
    with path.open("a", encoding="utf-8") as f:
        f.write(json.dumps(record, ensure_ascii=False) + "\n")


def write_csv_report(path: Path, rows: list[dict[str, Any]], fieldnames: list[str]) -> None:
    with path.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        for r in rows:
            w.writerow({k: ("" if r.get(k) is None else r.get(k)) for k in fieldnames})


if __name__ == "__main__":
    sys.exit(main())
