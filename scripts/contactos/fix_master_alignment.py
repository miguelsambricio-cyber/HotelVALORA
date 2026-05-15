#!/usr/bin/env python3
"""
fix_master_alignment.py · Phase B-Repair · 2026-05-15

Re-aligns the Master xlsx after the Phase 2.B.3 --apply offset corruption
(headers shifted RIGHT by 1 vs data; audit confirmed shift=-1 with 99.6%
match against Supabase).

Strategy (per operator decision · option A · target schema 64 cols):
  - Drop the spurious None at position 1 of each data row
  - Preserve position 0 (original_email · audit · empty for all rows
    today since Phase 2.B.3 --apply never wrote to it)
  - Shift remaining cells LEFT by 1 so each value lands under its
    correct header label
  - Truncate to 64 cells per row (= original_email + 63 canonical)

Header target (64 cols):
  col  1: original_email
  col  2: master_id
  cols 3..64: 62 remaining canonical fields ending in contact_category

Atomicity:
  - Reads from CURRENT Master (the broken one)
  - Writes to a NEW file: metcub-contacts-master.repaired-2026-05-15.xlsx
  - DOES NOT touch the canonical filename (operator does atomic swap
    after running audit_master_alignment.py against the repaired file)
  - Pre-cleanup backup + broken backup remain untouched

Diff report:
  - reports/master-alignment-fix_<TS>.json
  - 5 sample rows × pre+post side by side
  - per-column counts of before-vs-after non-empty cells

Pre-conditions:
  - Phase B-Repair sentinel exists (this script does NOT abort on it ·
    it's the script that DOES the repair)
"""
from __future__ import annotations

import json
import sys
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook, load_workbook

sys.stdout.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parent.parent.parent
MASTER_PATH = ROOT / "CONTACTOS DATASITE" / "master" / "metcub-contacts-master.xlsx"
REPAIRED_PATH = ROOT / "CONTACTOS DATASITE" / "master" / "metcub-contacts-master.repaired-2026-05-15.xlsx"
REPORTS_DIR = ROOT / "CONTACTOS DATASITE" / "reports"
REPORTS_DIR.mkdir(parents=True, exist_ok=True)

# Canonical 63-col header (from pre-cleanup backup · 2026-05-15 01:21)
CANONICAL_HEADER = [
    "master_id", "full_name", "email", "phone", "linkedin", "title", "role",
    "is_primary_contact", "company", "investor_type", "investor_subtype", "tier",
    "industry", "fund_size", "investment_preference", "investment_min",
    "investment_max", "association", "hotel_focus", "geography", "city", "state",
    "country", "continent", "latest_deal_stage", "last_activity_type",
    "last_activity_date", "buyer_added_date", "pipeline_state", "ioi_bid_low",
    "ioi_bid_high", "loi_bid_low", "loi_bid_high", "revised_bid_low",
    "revised_bid_high", "relationship_status", "relationship_manager",
    "coverage_officer", "calling_lead", "notes_consolidated",
    "datasite_contact_number", "datasite_company_number", "client_contact_id",
    "client_company_id", "source_file", "first_seen_batch_id",
    "last_seen_batch_id", "last_updated_at", "relationship_strength",
    "last_email_date", "active_threads", "gmail_labels",
    "inferred_relationship_stage", "email_directionality", "gmail_signal_source",
    "relationship_band", "collaboration_potential_score", "email_validity",
    "bounce_count", "last_bounce_date", "flagged_for_correction", "bucket",
    "contact_category",
]
TARGET_HEADER = ["original_email"] + CANONICAL_HEADER  # 64 cols

# Sample rows for diff report (chosen for diversity · canonical Phase 2.B.3 + 5 random)
SAMPLE_ROW_INDICES = [504, 3443, 914, 206, 2255, 2008, 1830]


def fix_row(row: list[object]) -> list[object]:
    """Apply the shift-left-by-1 fix to a single data row.

    Input  layout: [original_email, spurious_None, master_id, full_name, ..., contact_category, trailing_None_cells]
    Output layout: [original_email, master_id, full_name, ..., contact_category]   # 64 cells
    """
    if len(row) < 65:
        # Pad short rows so slicing is safe
        row = list(row) + [None] * (65 - len(row))
    new_row: list[object] = [row[0]]            # preserve original_email
    new_row.extend(row[2:65])                   # master_id..contact_category (63 cells)
    return new_row


def cell_is_empty(v: object) -> bool:
    return v is None or (isinstance(v, str) and not v.strip())


def main() -> int:
    if not MASTER_PATH.exists():
        print(f"ERROR: Master not found at {MASTER_PATH}")
        return 1

    print(f"→ Reading current (broken) Master from {MASTER_PATH}")
    wb_in = load_workbook(MASTER_PATH, read_only=True, data_only=True)
    ws_in = wb_in["Master"]
    src_header = [c.value for c in ws_in[1]]
    src_rows: list[list[object]] = []
    for r in ws_in.iter_rows(min_row=2, values_only=True):
        src_rows.append(list(r))
    src_sheets = wb_in.sheetnames
    wb_in.close()
    print(f"  · source header cols: {len(src_header)}")
    print(f"  · source data rows  : {len(src_rows)}")
    print(f"  · source sheets     : {src_sheets}")

    if src_header[0] != "original_email":
        print(f"!! UNEXPECTED · header[0] = {src_header[0]!r} (expected 'original_email')")
        print("   Aborting · the broken-file shape doesn't match what Phase 2.B.3 produced.")
        return 2

    # Sanity: confirm canonical 63 starts at header[1] (master_id)
    if src_header[1] != "master_id":
        print(f"!! UNEXPECTED · header[1] = {src_header[1]!r} (expected 'master_id')")
        return 2

    print(f"\n→ Applying shift-left-by-1 fix on {len(src_rows)} rows")
    fixed_rows = [fix_row(r) for r in src_rows]

    # Diff stats
    pre_filled = sum(1 for r in src_rows for v in r if not cell_is_empty(v))
    post_filled = sum(1 for r in fixed_rows for v in r if not cell_is_empty(v))
    print(f"  · pre-fix  populated cells: {pre_filled:,}")
    print(f"  · post-fix populated cells: {post_filled:,}")
    print(f"  · cells dropped           : {pre_filled - post_filled:,}")

    # Sanity check: all dropped cells should be empty (we drop position 1 + cols 65-68 which were None)
    dropped_cells_per_row = []
    for src, fixed in zip(src_rows, fixed_rows):
        # Positions in src that we DROPPED:
        #   - src[1] (the spurious None)
        #   - src[65..end] (trailing cells)
        dropped = [src[1]] + (src[65:] if len(src) > 65 else [])
        dropped_cells_per_row.append(dropped)

    nonempty_dropped = [
        (idx + 2, dropped)
        for idx, dropped in enumerate(dropped_cells_per_row)
        if any(not cell_is_empty(v) for v in dropped)
    ]
    if nonempty_dropped:
        print(f"\n!! WARNING · {len(nonempty_dropped)} rows have non-empty cells in dropped positions")
        for sheet_row, dropped in nonempty_dropped[:10]:
            print(f"   row {sheet_row}: dropped values = {[v for v in dropped if not cell_is_empty(v)]}")
        if len(nonempty_dropped) > 10:
            print(f"   ... and {len(nonempty_dropped) - 10} more")
        print("   ABORTING · refuse to drop populated data without operator review.")
        return 3
    else:
        print(f"  · ZERO non-empty cells dropped (all dropped positions were empty for all 4398 rows)")

    print(f"\n→ Writing repaired file to {REPAIRED_PATH}")
    wb_out = Workbook()
    # Replace default sheet name
    ws_out = wb_out.active
    ws_out.title = "Master"
    ws_out.append(TARGET_HEADER)
    for fixed in fixed_rows:
        ws_out.append(fixed)
    # Preserve any side-sheets from the source (read-only mode loses them ·
    # we re-load in normal mode just to copy them across)
    print(f"  · checking for side-sheets to preserve …")
    wb_aux = load_workbook(MASTER_PATH, read_only=False, data_only=True)
    for sheet_name in wb_aux.sheetnames:
        if sheet_name == "Master":
            continue
        print(f"  · preserving sheet: {sheet_name!r}")
        src_ws = wb_aux[sheet_name]
        new_ws = wb_out.create_sheet(title=sheet_name)
        for row in src_ws.iter_rows(values_only=True):
            new_ws.append(list(row))
    wb_aux.close()

    wb_out.save(REPAIRED_PATH)
    wb_out.close()
    print(f"  · saved · file size: {REPAIRED_PATH.stat().st_size:,} bytes")

    # Build diff report with sample rows
    print(f"\n→ Building diff report with sample rows")
    samples = []
    for sheet_row in SAMPLE_ROW_INDICES:
        if sheet_row - 2 < 0 or sheet_row - 2 >= len(src_rows):
            continue
        src = src_rows[sheet_row - 2]
        fixed = fixed_rows[sheet_row - 2]
        # Build pre dict (raw — under broken header) and post dict (under fixed header)
        pre = {h: src[i] for i, h in enumerate(src_header) if h and i < len(src) and not cell_is_empty(src[i])}
        post = {h: fixed[i] for i, h in enumerate(TARGET_HEADER) if i < len(fixed) and not cell_is_empty(fixed[i])}
        samples.append({
            "sheet_row": sheet_row,
            "pre_fix_under_broken_header": pre,
            "post_fix_under_canonical_header": post,
        })

    ts = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    report = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "source_master": str(MASTER_PATH),
        "repaired_master": str(REPAIRED_PATH),
        "fix_strategy": "shift LEFT by 1 · drop position 1 spurious None · drop trailing cells beyond col 64",
        "source_header_cols": len(src_header),
        "target_header_cols": len(TARGET_HEADER),
        "row_count": len(src_rows),
        "pre_fix_populated_cells": pre_filled,
        "post_fix_populated_cells": post_filled,
        "non_empty_dropped_rows": len(nonempty_dropped),
        "samples": samples,
    }
    report_path = REPORTS_DIR / f"master-alignment-fix_{ts}.json"
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    print(f"  · report written: {report_path}")

    print("\n✓ Repair complete · run audit_master_alignment.py against the repaired file")
    print(f"  before swapping it to the canonical filename.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
