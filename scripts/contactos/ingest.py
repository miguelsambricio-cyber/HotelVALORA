#!/usr/bin/env python3
"""
CONTACTOS DATASITE · institutional relationship intelligence ingester.

Operator workflow
-----------------
1. Drop a Datasite Outreach Full Report (`.xlsm`) into:
       CONTACTOS DATASITE/incoming/

2. Run:
       python scripts/contactos/ingest.py

3. Pipeline behaviour (per file, oldest-first):
     - parse the Datasite export (Contacts · Companies · Activities sheets)
     - clean + normalise (encoding · whitespace · phone · email)
     - map source columns to the canonical institutional schema
     - LEFT JOIN Contacts → Companies → Activities on Company name to
       enrich every person row with deal stage + investor metadata
     - deduplicate against the existing Master:
         1. exact email
         2. LinkedIn URL
         3. full name + company
         4. fuzzy fallback (Levenshtein on full_name within same company)
     - merge in-place when a match is found (notes concatenated · latest
       state wins · provenance + first_seen_batch_id preserved)
     - update the four canonical sheets (Master · Contacts · Companies ·
       Activities) and rebuild the Summary dashboard
     - write four reports under reports/
     - move the source file → CONTACTOS DATASITE/old/

Outputs (deterministic, idempotent)
-----------------------------------
  CONTACTOS DATASITE/master/metcub-contacts-master.xlsx
  CONTACTOS DATASITE/reports/ingestion-log.jsonl     (append-only)
  CONTACTOS DATASITE/reports/duplicate-resolution_<batch_id>.csv
  CONTACTOS DATASITE/reports/schema-mapping_<batch_id>.csv
  CONTACTOS DATASITE/reports/invalid-missing_<batch_id>.csv

Re-runs are idempotent: an empty incoming/ directory just regenerates the
Summary sheet against the existing Master and exits.

Dependencies: openpyxl (only). All other tools are Python stdlib.
"""

from __future__ import annotations

import csv
import difflib
import json
import re
import shutil
import sys
import unicodedata

# Windows console defaults to cp1252 · this script logs unicode (→ · accents)
# so force UTF-8 stdout to avoid UnicodeEncodeError on print().
try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except (AttributeError, ValueError):
    pass
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl", file=sys.stderr)
    sys.exit(2)

# ── paths ───────────────────────────────────────────────────────────────────

REPO = Path(__file__).resolve().parent.parent.parent
ROOT = REPO / "CONTACTOS DATASITE"
INCOMING = ROOT / "incoming"
OLD = ROOT / "old"
MASTER_DIR = ROOT / "master"
REPORTS = ROOT / "reports"
MASTER_FILE = MASTER_DIR / "metcub-contacts-master.xlsx"

for d in (INCOMING, OLD, MASTER_DIR, REPORTS):
    d.mkdir(parents=True, exist_ok=True)


# ── canonical schema ────────────────────────────────────────────────────────
# The Master sheet's column order is the institutional contract — every
# downstream consumer (UI, deal-flow joins, future graph) references it.
# Adding a column means appending to MASTER_SCHEMA; never re-ordering.

MASTER_SCHEMA: list[str] = [
    # identity
    "master_id",                   # synthetic stable id (md5 of lowest-priority match key)
    "full_name",
    "email",
    "phone",
    "linkedin",
    "title",
    "role",
    "is_primary_contact",          # "Yes" / "No"
    # company + investor frame
    "company",
    "investor_type",               # canonical (Fund / Family Office / Operator / Brand / Lender / Broker / REIT-SOCIMI / Investor / Developer / Other)
    "investor_subtype",            # raw Datasite Subtype value
    "tier",                        # Datasite-assigned tier
    "industry",
    "fund_size",
    "investment_preference",
    "investment_min",
    "investment_max",
    "association",
    "hotel_focus",                 # derived flag (Yes/No/Likely) from keywords in description+notes
    # geography
    "geography",                   # canonical short string (City, Country)
    "city",
    "state",
    "country",
    "continent",
    # deal-state (latest activity for the contact's company)
    "latest_deal_stage",
    "last_activity_type",
    "last_activity_date",
    "buyer_added_date",
    "pipeline_state",              # Active | On Hold | Removed | Declined | (empty)
    "ioi_bid_low",
    "ioi_bid_high",
    "loi_bid_low",
    "loi_bid_high",
    "revised_bid_low",
    "revised_bid_high",
    # relationship + notes
    "relationship_status",         # Active | Inactive (Datasite "Active" column)
    "relationship_manager",
    "coverage_officer",
    "calling_lead",
    "notes_consolidated",
    # provenance + lineage
    "datasite_contact_number",
    "datasite_company_number",
    "client_contact_id",
    "client_company_id",
    "source_file",                 # last file that touched this row
    "first_seen_batch_id",
    "last_seen_batch_id",
    "last_updated_at",
    # Phase 2.B · Gmail-derived relationship intelligence (institutional graph)
    "relationship_strength",       # 0–100 numeric · derived score (see compute_relationship_strength)
    "last_email_date",             # ISO date · most recent inbound/outbound thread
    "active_threads",              # rolling 12-month Gmail thread count
    "gmail_labels",                # semicolon-joined list of institutional Gmail labels touching this email
    "inferred_relationship_stage", # canonical stage derived from Gmail labels + Datasite pipeline_state
    "email_directionality",        # inbound | outbound | bidirectional | none
    "gmail_signal_source",         # filename of the Gmail signal snapshot that populated these fields
]


# ── source column mapping ───────────────────────────────────────────────────
# Datasite Outreach export schema (verified 2026-05-12). When future
# exports add or rename columns, extend these maps and rerun ingest.

CONTACT_COL_MAP: dict[str, str] = {
    "Number": "datasite_contact_number",
    "Company": "company",
    "Title": "title",
    "First Name": "first_name",
    "Last Name": "last_name",
    "Email": "email",
    "Primary": "is_primary_contact",
    "Role": "role",
    "Industry": "industry",
    "Office Phone": "office_phone",
    "Relationship Manager": "relationship_manager",
    "Notes": "contact_notes",
    "Address": "address",
    "City": "city",
    "State": "state",
    "Zip Code": "zip_code",
    "Country": "country",
    "Linked In": "linkedin",
    "Active": "relationship_status",
    "Client Contact ID": "client_contact_id",
    "Cell Phone": "cell_phone",
}

COMPANY_COL_MAP: dict[str, str] = {
    "Number": "datasite_company_number",
    "Company": "company",
    "Type": "company_type",
    "Location": "company_location",
    "Industry": "company_industry",
    "Coverage Officer": "coverage_officer",
    "Notes": "company_notes",
    "Tier": "tier",
    "Calling Lead": "calling_lead",
    "Subtype": "investor_subtype",
    "Type of Investor": "investor_type_raw",
    "Continente": "continent",
    "Palabra Clave": "keyword",
    "Description": "description",
    "Fund Size": "fund_size",
    "Investment Preference": "investment_preference",
    "Internal Notes": "internal_notes",
    "External Notes": "external_notes",
    "Client Company ID": "client_company_id",
    "Association": "association",
    "Investment Minimum": "investment_min",
    "Investment Maximum": "investment_max",
}

ACTIVITY_COL_MAP: dict[str, str] = {
    "Number": "datasite_activity_number",
    "Company": "company",
    "Buyer Added": "buyer_added_date",
    "Initial Contact": "initial_contact_date",
    "Teaser Sent": "teaser_sent_date",
    "NDA Initial Sent": "nda_initial_sent_date",
    "NDA Signed": "nda_signed_date",
    "NDA Executed": "nda_executed_date",
    "CIM Sent": "cim_sent_date",
    "IOI Process Letter": "ioi_process_letter_date",
    "IOI Bid Received": "ioi_bid_received_date",
    "LOI Process Letter": "loi_process_letter_date",
    "LOI Bid Received": "loi_bid_received_date",
    "Management Presentation": "management_presentation_date",
    "Declined": "declined_date",
    "Declined Reason": "declined_reason",
    "Declined Comments": "declined_comments",
    "Last Activity": "last_activity_date",
    "Last Activity Comments": "last_activity_comments",
    "Last Activity Name": "last_activity_type",
    "Latest Stage Name": "latest_deal_stage",
    "IOI Bid Low": "ioi_bid_low",
    "IOI Bid High": "ioi_bid_high",
    "LOI Bid Low": "loi_bid_low",
    "LOI Bid High": "loi_bid_high",
    "Revised Bid Received": "revised_bid_received_date",
    "Revised Bid Low": "revised_bid_low",
    "Revised Bid High": "revised_bid_high",
    "Removed": "removed_flag",
    "On Hold": "on_hold_flag",
    "Client Buyer Id": "client_buyer_id",
}


# ── investor type canonicalisation ──────────────────────────────────────────
# Datasite's "Type of Investor" field is free-text Spanish + English mix.
# We collapse to a stable institutional taxonomy for the Master + Summary.

# Ordered most-specific first. Patterns are applied IN ORDER against the
# concatenated corpus (investor_type_raw + subtype + description + notes) so
# strong tokens (REIT/SOCIMI, Family Office) capture before broader ones
# (Investor). Add new canonical buckets here when Datasite invents new
# raw labels · NEVER collapse two distinct institutional categories.
INVESTOR_TYPE_CANON: list[tuple[re.Pattern[str], str]] = [
    (re.compile(r"\b(REIT|SOCIMI|inmobiliaria cotizada)\b", re.IGNORECASE), "REIT/SOCIMI"),
    (re.compile(r"\b(family\s*office|FO|patrimonio familiar)\b", re.IGNORECASE), "Family Office"),
    (re.compile(r"\b(sovereign|wealth fund|fondo soberano)\b", re.IGNORECASE), "Sovereign Wealth"),
    (re.compile(r"\b(pension|pensiones)\b", re.IGNORECASE), "Pension Fund"),
    (re.compile(r"\b(insurance|asegurador|insurer|aseguradora)\b", re.IGNORECASE), "Insurance"),
    (re.compile(r"\b(private\s*equity|PE\s*fund|venture\s*capital|VC|fondo de inversi[óo]n|investment fund)\b", re.IGNORECASE), "Fund"),
    # Lender / Financiador (banks, debt providers, mortgage)
    (re.compile(r"\b(financiador|financiadora|banco|prestamista|debt fund|debt|mortgage|lender)\b", re.IGNORECASE), "Lender"),
    # Hotel Chain (Cadena Hotelera) — distinct from Operator
    # Datasite Spanish label "Cadenas Hotel" is the canonical signal.
    (re.compile(r"\b(cadena hotelera|cadenas hotel|hotel chain|chain)\b", re.IGNORECASE), "Hotel Chain"),
    (re.compile(r"\b(operator|operador|gestor hotelero|management company)\b", re.IGNORECASE), "Operator"),
    (re.compile(r"\b(franchisor|brand|marca|franchise)\b", re.IGNORECASE), "Brand"),
    # Owner (Propietario) — single-asset / portfolio owners, distinct from Investor
    (re.compile(r"\b(propietario|owner|propietaria)\b", re.IGNORECASE), "Owner"),
    (re.compile(r"\b(broker|intermediario|asesor inmobiliario|real estate advisor|RE broker)\b", re.IGNORECASE), "Broker"),
    # Advisor / Consultancy
    (re.compile(r"\b(consultor|consultora|consultancy|advisory|asesor[íi]a)\b", re.IGNORECASE), "Advisor"),
    # Developer / Promotor
    (re.compile(r"\b(developer|promotor|desarrollador|desarrolladora|promotora)\b", re.IGNORECASE), "Developer"),
    # Architect / Arquitecto
    (re.compile(r"\b(architect|arquitecto|arquitectura|arquitecta|estudio de arquitectura)\b", re.IGNORECASE), "Architect"),
    # Service Provider (Proveedor de Servicios)
    (re.compile(r"\b(proveedor|provider|service provider|servicios)\b", re.IGNORECASE), "Service Provider"),
    # F&B
    (re.compile(r"\bF&B|food\s*&\s*beverage\b", re.IGNORECASE), "F&B Operator"),
    # Media / Press
    (re.compile(r"\b(medios|media|press|prensa|periodista|publicaci[óo]n)\b", re.IGNORECASE), "Media"),
    # Corporate / Industrial (non-RE corporates)
    (re.compile(r"\b(corporate|industrial)\b", re.IGNORECASE), "Corporate"),
    (re.compile(r"\b(institutional|institucional)\b", re.IGNORECASE), "Institutional Investor"),
    (re.compile(r"\b(inversor|investor)\b", re.IGNORECASE), "Investor"),
]

HOTEL_FOCUS_HINT: re.Pattern[str] = re.compile(
    r"\b(hotel|hospitality|resort|hostel|aparthotel|short[\s-]?stay|extended[\s-]?stay|"
    r"hotelero|hotelera|hoteles|RevPAR|ADR|GOPPAR|key[s]?|room[s]?|llaves|habitaciones)\b",
    re.IGNORECASE,
)


# ── helpers ─────────────────────────────────────────────────────────────────


def canonical_investor_type(raw: Any, subtype: Any, description: Any, notes: Any) -> str:
    """Map raw Datasite values to canonical institutional taxonomy."""
    corpus = " ".join(str(x or "") for x in (raw, subtype, description, notes))
    for pattern, canon in INVESTOR_TYPE_CANON:
        if pattern.search(corpus):
            return canon
    if raw:
        return str(raw).strip()
    return "Unknown"


def detect_hotel_focus(industry: Any, description: Any, notes: Any, association: Any) -> str:
    """Yes / Likely / No based on hospitality keyword density."""
    corpus = " ".join(str(x or "") for x in (industry, description, notes, association))
    if not corpus.strip():
        return "Unknown"
    matches = HOTEL_FOCUS_HINT.findall(corpus)
    if len(matches) >= 2:
        return "Yes"
    if len(matches) == 1:
        return "Likely"
    return "No"


def norm_text(value: Any) -> str:
    """NFC normalise · strip · collapse internal whitespace · empty→''."""
    if value is None:
        return ""
    s = str(value).strip()
    if not s:
        return ""
    s = unicodedata.normalize("NFC", s)
    s = re.sub(r"\s+", " ", s)
    return s


def norm_email(value: Any) -> str:
    s = norm_text(value).lower()
    # Strip surrounding < > and mailto: prefix that some exports add.
    s = re.sub(r"^<+|>+$", "", s)
    s = re.sub(r"^mailto:", "", s, flags=re.IGNORECASE)
    if "@" not in s or "." not in s.split("@")[-1]:
        return ""
    return s


def norm_phone(value: Any) -> str:
    s = norm_text(value)
    if not s:
        return ""
    # Keep leading +, drop other non-digits, collapse internal duplicates
    plus = "+" if s.startswith("+") else ""
    digits = re.sub(r"\D", "", s)
    if not digits:
        return ""
    return plus + digits


def norm_linkedin(value: Any) -> str:
    s = norm_text(value).lower()
    if not s:
        return ""
    s = re.sub(r"^https?://(www\.)?", "", s)
    s = re.sub(r"/$", "", s)
    return s


def norm_company(value: Any) -> str:
    """Normalise company name for join keys · case-insensitive · punctuation light."""
    s = norm_text(value)
    if not s:
        return ""
    # Strip ", LLC" / ", Inc" / " S.L." style suffixes for matching (preserve original elsewhere)
    return s


def company_key(value: Any) -> str:
    """Aggressive normalisation for joining: lowercase + strip legal suffixes + alphanumerics only."""
    s = norm_company(value).lower()
    s = re.sub(r"[,\.]", "", s)
    s = re.sub(r"\b(llc|inc|ltd|sl|sa|sarl|gmbh|kg|ag|nv|bv|plc|limited|llp)\b\s*$", "", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def to_iso_date(value: Any) -> str:
    """ISO date string · datetime → YYYY-MM-DD · else best-effort."""
    if value is None or value == "":
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    s = norm_text(value)
    if not s:
        return ""
    # Try a few common formats
    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s.split("T")[0].split(" ")[0], fmt).date().isoformat()
        except ValueError:
            continue
    return s  # leave as-is if unrecognised


def synthesize_master_id(email: str, linkedin: str, full_name: str, company: str) -> str:
    """Stable synthetic id · md5 of the lowest-priority discriminator that's present."""
    import hashlib
    if email:
        key = f"email:{email}"
    elif linkedin:
        key = f"li:{linkedin}"
    else:
        key = f"name:{full_name.lower()}|company:{company_key(company)}"
    return hashlib.md5(key.encode("utf-8")).hexdigest()[:16]


def now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def consolidate_notes(*parts: Any) -> str:
    """Join non-empty unique note strings with ' | ' separator."""
    seen = set()
    pieces = []
    for p in parts:
        s = norm_text(p)
        if not s or s.lower() in seen:
            continue
        seen.add(s.lower())
        pieces.append(s)
    return " | ".join(pieces)


# ── sheet reading ───────────────────────────────────────────────────────────


def read_sheet(ws, col_map: dict[str, str]) -> list[dict[str, Any]]:
    """Read a sheet into list of dicts using col_map · ignores unknown columns."""
    if ws.max_row < 2:
        return []
    header = [norm_text(c.value) for c in ws[1]]
    rows: list[dict[str, Any]] = []
    for raw in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if all(v in (None, "") for v in raw):
            continue
        record: dict[str, Any] = {}
        for i, value in enumerate(raw):
            if i >= len(header):
                continue
            source_col = header[i]
            canon = col_map.get(source_col)
            if canon is None:
                continue
            record[canon] = value
        rows.append(record)
    return rows


# ── join + canonicalise ────────────────────────────────────────────────────


@dataclass
class ParseResult:
    contacts: list[dict[str, Any]]
    companies: list[dict[str, Any]]
    activities: list[dict[str, Any]]
    schema_map_rows: list[dict[str, str]]
    invalid_rows: list[dict[str, str]] = field(default_factory=list)


def parse_workbook(path: Path, batch_id: str) -> ParseResult:
    wb = load_workbook(path, data_only=True, read_only=True, keep_vba=False)
    schema_rows: list[dict[str, str]] = []

    def collect_schema(sheet_name: str, col_map: dict[str, str]) -> None:
        if sheet_name not in wb.sheetnames:
            return
        ws = wb[sheet_name]
        if ws.max_row < 1:
            return
        for cell in ws[1]:
            v = norm_text(cell.value)
            if not v:
                continue
            schema_rows.append({
                "source_file": path.name,
                "sheet": sheet_name,
                "source_column": v,
                "canonical_field": col_map.get(v, ""),
                "mapped": "yes" if v in col_map else "no",
            })

    collect_schema("Contacts", CONTACT_COL_MAP)
    collect_schema("Companies", COMPANY_COL_MAP)
    collect_schema("Activities", ACTIVITY_COL_MAP)

    contacts = read_sheet(wb["Contacts"], CONTACT_COL_MAP) if "Contacts" in wb.sheetnames else []
    companies = read_sheet(wb["Companies"], COMPANY_COL_MAP) if "Companies" in wb.sheetnames else []
    activities = read_sheet(wb["Activities"], ACTIVITY_COL_MAP) if "Activities" in wb.sheetnames else []
    wb.close()
    return ParseResult(contacts=contacts, companies=companies, activities=activities, schema_map_rows=schema_rows)


def companies_index(companies: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    """Index by company_key (joining key)."""
    idx: dict[str, dict[str, Any]] = {}
    for c in companies:
        key = company_key(c.get("company"))
        if not key:
            continue
        # If duplicate, last write wins (Datasite shouldn't dup but defensive).
        idx[key] = c
    return idx


def activities_index(activities: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    idx: dict[str, dict[str, Any]] = {}
    for a in activities:
        key = company_key(a.get("company"))
        if not key:
            continue
        idx[key] = a
    return idx


def pipeline_state(activity: dict[str, Any] | None) -> str:
    """Derive the active state of the deal from flags + declined."""
    if not activity:
        return ""
    if activity.get("removed_flag"):
        return "Removed"
    if activity.get("on_hold_flag"):
        return "On Hold"
    if activity.get("declined_date"):
        return "Declined"
    return "Active"


def build_master_row(
    contact: dict[str, Any],
    company: dict[str, Any] | None,
    activity: dict[str, Any] | None,
    source_file: str,
    batch_id: str,
) -> dict[str, Any] | None:
    """Compose a single Master row · returns None when the row is unidentifiable."""
    first = norm_text(contact.get("first_name"))
    last = norm_text(contact.get("last_name"))
    full_name = " ".join(p for p in (first, last) if p)
    email = norm_email(contact.get("email"))
    cell = norm_phone(contact.get("cell_phone"))
    office = norm_phone(contact.get("office_phone"))
    phone = cell or office
    linkedin = norm_linkedin(contact.get("linkedin"))
    company_name = norm_company(contact.get("company") or (company or {}).get("company"))

    # Must have at least: email OR linkedin OR (full_name AND company)
    if not (email or linkedin or (full_name and company_name)):
        return None

    raw_investor = (company or {}).get("investor_type_raw")
    subtype = (company or {}).get("investor_subtype")
    description = (company or {}).get("description")
    company_notes = (company or {}).get("company_notes")
    internal = (company or {}).get("internal_notes")
    external = (company or {}).get("external_notes")
    keyword = (company or {}).get("keyword")
    association = (company or {}).get("association")
    industry = norm_text(contact.get("industry") or (company or {}).get("company_industry"))

    notes = consolidate_notes(
        contact.get("contact_notes"),
        company_notes,
        internal,
        external,
        keyword,
    )

    location = norm_text((company or {}).get("company_location"))
    city = norm_text(contact.get("city"))
    state = norm_text(contact.get("state"))
    country = norm_text(contact.get("country"))
    geography = ", ".join(p for p in (city, country) if p) or location

    master_id = synthesize_master_id(email, linkedin, full_name, company_name)

    return {
        "master_id": master_id,
        "full_name": full_name,
        "email": email,
        "phone": phone,
        "linkedin": linkedin,
        "title": norm_text(contact.get("title")),
        "role": norm_text(contact.get("role")),
        "is_primary_contact": norm_text(contact.get("is_primary_contact")),
        "company": company_name,
        "investor_type": canonical_investor_type(raw_investor, subtype, description, notes),
        "investor_subtype": norm_text(subtype),
        "tier": norm_text((company or {}).get("tier")),
        "industry": industry,
        "fund_size": norm_text((company or {}).get("fund_size")),
        "investment_preference": norm_text((company or {}).get("investment_preference")),
        "investment_min": norm_text((company or {}).get("investment_min")),
        "investment_max": norm_text((company or {}).get("investment_max")),
        "association": norm_text(association),
        "hotel_focus": detect_hotel_focus(industry, description, notes, association),
        "geography": geography,
        "city": city,
        "state": state,
        "country": country,
        "continent": norm_text((company or {}).get("continent")),
        "latest_deal_stage": norm_text((activity or {}).get("latest_deal_stage")),
        "last_activity_type": norm_text((activity or {}).get("last_activity_type")),
        "last_activity_date": to_iso_date((activity or {}).get("last_activity_date")),
        "buyer_added_date": to_iso_date((activity or {}).get("buyer_added_date")),
        "pipeline_state": pipeline_state(activity),
        "ioi_bid_low": norm_text((activity or {}).get("ioi_bid_low")),
        "ioi_bid_high": norm_text((activity or {}).get("ioi_bid_high")),
        "loi_bid_low": norm_text((activity or {}).get("loi_bid_low")),
        "loi_bid_high": norm_text((activity or {}).get("loi_bid_high")),
        "revised_bid_low": norm_text((activity or {}).get("revised_bid_low")),
        "revised_bid_high": norm_text((activity or {}).get("revised_bid_high")),
        "relationship_status": norm_text(contact.get("relationship_status")),
        "relationship_manager": norm_text(contact.get("relationship_manager")),
        "coverage_officer": norm_text((company or {}).get("coverage_officer")),
        "calling_lead": norm_text((company or {}).get("calling_lead")),
        "notes_consolidated": notes,
        "datasite_contact_number": norm_text(contact.get("datasite_contact_number")),
        "datasite_company_number": norm_text((company or {}).get("datasite_company_number")),
        "client_contact_id": norm_text(contact.get("client_contact_id")),
        "client_company_id": norm_text((company or {}).get("client_company_id")),
        "source_file": source_file,
        "first_seen_batch_id": batch_id,
        "last_seen_batch_id": batch_id,
        "last_updated_at": now_iso(),
        # Phase 2.B Gmail-derived fields · default empty until ingest_gmail.py populates them
        "relationship_strength": 0,
        "last_email_date": "",
        "active_threads": 0,
        "gmail_labels": "",
        "inferred_relationship_stage": "",
        "email_directionality": "",
        "gmail_signal_source": "",
    }


# ── deduplication ───────────────────────────────────────────────────────────


def find_match(new_row: dict[str, Any], by_email: dict[str, int], by_linkedin: dict[str, int],
               by_name_company: dict[tuple[str, str], int], existing: list[dict[str, Any]]
               ) -> tuple[int | None, str, float]:
    """Return (index_in_existing, strategy, score) or (None, "", 0.0)."""
    email = new_row["email"]
    if email and email in by_email:
        return by_email[email], "email_exact", 1.0
    linkedin = new_row["linkedin"]
    if linkedin and linkedin in by_linkedin:
        return by_linkedin[linkedin], "linkedin_exact", 1.0
    name_key = (new_row["full_name"].lower(), company_key(new_row["company"]))
    if name_key[0] and name_key[1] and name_key in by_name_company:
        return by_name_company[name_key], "name_company_exact", 0.95
    # Fuzzy: same company_key, name within 0.88 similarity
    target_company = name_key[1]
    if not target_company or not name_key[0]:
        return None, "", 0.0
    best_idx, best_score = None, 0.0
    for i, row in enumerate(existing):
        if company_key(row["company"]) != target_company:
            continue
        score = difflib.SequenceMatcher(None, name_key[0], row["full_name"].lower()).ratio()
        if score > best_score:
            best_score = score
            best_idx = i
    if best_idx is not None and best_score >= 0.88:
        return best_idx, "fuzzy", best_score
    return None, "", 0.0


def merge_existing(existing: dict[str, Any], incoming: dict[str, Any]) -> tuple[dict[str, Any], list[str]]:
    """Latest non-empty value wins · provenance fields preserved · notes concat."""
    changed: list[str] = []
    out = dict(existing)
    for field_name in MASTER_SCHEMA:
        if field_name == "first_seen_batch_id":
            continue  # preserve original
        if field_name == "master_id":
            continue  # never change the synthetic id
        new_val = incoming.get(field_name, "")
        old_val = existing.get(field_name, "")
        if field_name == "notes_consolidated":
            # Always merge note text · dedup substrings
            merged = consolidate_notes(old_val, new_val)
            if merged != old_val:
                out[field_name] = merged
                changed.append(field_name)
            continue
        if new_val and new_val != old_val:
            out[field_name] = new_val
            changed.append(field_name)
    out["last_seen_batch_id"] = incoming["last_seen_batch_id"]
    out["last_updated_at"] = incoming["last_updated_at"]
    out["source_file"] = incoming["source_file"]
    return out, changed


# ── master xlsx I/O ─────────────────────────────────────────────────────────


def load_existing_master() -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    """Returns (master_rows, contacts_rows, companies_rows, activities_rows)."""
    if not MASTER_FILE.exists():
        return [], [], [], []
    wb = load_workbook(MASTER_FILE, data_only=True, read_only=True)

    def read(sheet_name: str) -> list[dict[str, Any]]:
        if sheet_name not in wb.sheetnames:
            return []
        ws = wb[sheet_name]
        if ws.max_row < 2:
            return []
        header = [norm_text(c.value) for c in ws[1]]
        out = []
        for raw in ws.iter_rows(min_row=2, values_only=True):
            if all(cell in (None, "") for cell in raw):
                continue
            record = {}
            for i, h in enumerate(header):
                if i >= len(raw):
                    break
                val = raw[i]
                record[h] = "" if val is None else val
            out.append(record)
        return out

    return read("Master"), read("Contacts"), read("Companies"), read("Activities")


def write_master_workbook(
    master_rows: list[dict[str, Any]],
    contacts_rows: list[dict[str, Any]],
    companies_rows: list[dict[str, Any]],
    activities_rows: list[dict[str, Any]],
    summary: dict[str, Any],
) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    HEADER_FILL = PatternFill("solid", fgColor="0F172A")
    HEADER_FONT = Font(bold=True, color="A3E635", size=10)
    LABEL_FONT = Font(bold=True, color="0F172A")

    def write_sheet(name: str, columns: list[str], rows: list[dict[str, Any]]) -> None:
        ws = wb.create_sheet(name)
        ws.append(columns)
        for c, _ in enumerate(columns, 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 22
        ws.freeze_panes = "A2"
        for row in rows:
            ws.append([row.get(col, "") for col in columns])
        # Column widths · light heuristic
        for c, col in enumerate(columns, 1):
            ws.column_dimensions[get_column_letter(c)].width = min(40, max(12, len(col) + 4))

    # Master
    write_sheet("Master", MASTER_SCHEMA, master_rows)

    # Contacts (raw normalised — every Datasite contact row)
    contact_cols = ["datasite_contact_number", "company", "first_name", "last_name", "email",
                    "office_phone", "cell_phone", "linkedin", "title", "role",
                    "is_primary_contact", "industry", "relationship_status",
                    "relationship_manager", "address", "city", "state", "zip_code", "country",
                    "contact_notes", "client_contact_id"]
    write_sheet("Contacts", contact_cols, contacts_rows)

    # Companies (canonical company table)
    company_cols = ["datasite_company_number", "company", "investor_type_canonical",
                    "investor_type_raw", "investor_subtype", "tier", "company_type",
                    "company_industry", "continent", "company_location", "fund_size",
                    "investment_preference", "investment_min", "investment_max",
                    "association", "description", "keyword",
                    "coverage_officer", "calling_lead",
                    "internal_notes", "external_notes", "company_notes", "client_company_id"]
    write_sheet("Companies", company_cols, companies_rows)

    # Activities (timeline)
    activity_cols = ["datasite_activity_number", "company", "buyer_added_date",
                     "initial_contact_date", "teaser_sent_date", "nda_initial_sent_date",
                     "nda_signed_date", "nda_executed_date", "cim_sent_date",
                     "ioi_process_letter_date", "ioi_bid_received_date",
                     "loi_process_letter_date", "loi_bid_received_date",
                     "management_presentation_date", "revised_bid_received_date",
                     "declined_date", "declined_reason", "declined_comments",
                     "last_activity_type", "last_activity_date", "last_activity_comments",
                     "latest_deal_stage", "pipeline_state",
                     "ioi_bid_low", "ioi_bid_high", "loi_bid_low", "loi_bid_high",
                     "revised_bid_low", "revised_bid_high",
                     "removed_flag", "on_hold_flag", "client_buyer_id"]
    write_sheet("Activities", activity_cols, activities_rows)

    # Summary
    ws = wb.create_sheet("Summary")
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 24

    def section(title: str) -> None:
        ws.append([title, ""])
        cell = ws.cell(row=ws.max_row, column=1)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        ws.cell(row=ws.max_row, column=2).fill = HEADER_FILL

    def line(label: str, value: Any) -> None:
        ws.append([label, value])
        ws.cell(row=ws.max_row, column=1).font = LABEL_FONT

    section("Totals")
    line("Total contacts (Master)", summary["totals"]["contacts"])
    line("Unique companies (Master)", summary["totals"]["unique_companies"])
    line("Companies in Companies sheet", summary["totals"]["company_records"])
    line("Activities records", summary["totals"]["activity_records"])
    line("Last regenerated", now_iso())
    line("Last batch id", summary["totals"]["last_batch_id"])
    ws.append([])

    section("Pipeline · contacts by latest deal stage")
    for stage, n in summary["by_stage"]:
        line(stage or "(none)", n)
    ws.append([])

    section("Pipeline state distribution")
    for state, n in summary["by_pipeline_state"]:
        line(state or "(none)", n)
    ws.append([])

    section("Contacts by canonical investor type")
    for t, n in summary["by_investor_type"]:
        line(t or "(unknown)", n)
    ws.append([])

    section("Contacts by continent")
    for cont, n in summary["by_continent"]:
        line(cont or "(unknown)", n)
    ws.append([])

    section("Contacts by hotel focus")
    for hf, n in summary["by_hotel_focus"]:
        line(hf or "(unknown)", n)
    ws.append([])

    section("Coverage")
    line("Companies WITH at least 1 contact", summary["coverage"]["companies_with_contacts"])
    line("Companies WITHOUT contacts", summary["coverage"]["companies_without_contacts"])
    line("Contacts with email", summary["coverage"]["contacts_with_email"])
    line("Contacts with LinkedIn", summary["coverage"]["contacts_with_linkedin"])
    line("Contacts with phone", summary["coverage"]["contacts_with_phone"])
    line("Primary contacts (Yes)", summary["coverage"]["primary_yes"])
    ws.append([])

    section("Top 15 active investors (by Master contacts count)")
    for company, n in summary["top_active_investors"]:
        line(company, n)
    ws.append([])

    section("Seniority (heuristic, from title/role)")
    for level, n in summary["by_seniority"]:
        line(level, n)

    ws.freeze_panes = "A2"

    MASTER_FILE.parent.mkdir(parents=True, exist_ok=True)
    wb.save(MASTER_FILE)


# ── summary builder ─────────────────────────────────────────────────────────


SENIORITY_PATTERNS: list[tuple[re.Pattern[str], str]] = [
    (re.compile(r"\b(CEO|chief executive|managing partner|managing director|chairman|founder|owner|president)\b", re.IGNORECASE), "C-Suite / Owner"),
    (re.compile(r"\b(CFO|CIO|COO|CTO|CRO|chief)\b", re.IGNORECASE), "C-Suite / Owner"),
    (re.compile(r"\b(managing principal|partner|principal)\b", re.IGNORECASE), "Partner / Principal"),
    (re.compile(r"\b(head of|director|VP|vice president|executive director)\b", re.IGNORECASE), "Director / VP"),
    (re.compile(r"\b(senior associate|senior manager|senior analyst)\b", re.IGNORECASE), "Senior"),
    (re.compile(r"\b(associate|analyst|manager)\b", re.IGNORECASE), "Associate / Analyst"),
]


def detect_seniority(title: str, role: str) -> str:
    corpus = f"{title or ''} {role or ''}"
    for pat, level in SENIORITY_PATTERNS:
        if pat.search(corpus):
            return level
    return "Other / Unknown"


def build_summary(master: list[dict[str, Any]], companies: list[dict[str, Any]], activities: list[dict[str, Any]], last_batch_id: str) -> dict[str, Any]:
    def top_counts(it: Iterable[str], limit: int | None = None) -> list[tuple[str, int]]:
        from collections import Counter
        c = Counter(it)
        items = sorted(c.items(), key=lambda kv: (-kv[1], kv[0].lower() if kv[0] else ""))
        return items[:limit] if limit else items

    contact_companies = {company_key(r.get("company")) for r in master if r.get("company")}
    company_keys_set = {company_key(c.get("company")) for c in companies if c.get("company")}
    companies_with = company_keys_set & contact_companies
    companies_without = company_keys_set - contact_companies

    by_stage = top_counts(r.get("latest_deal_stage", "") or "" for r in master)
    by_pipeline = top_counts(r.get("pipeline_state", "") or "" for r in master)
    by_investor = top_counts(r.get("investor_type", "") or "Unknown" for r in master)
    by_continent = top_counts(r.get("continent", "") or "" for r in master)
    by_hotel = top_counts(r.get("hotel_focus", "") or "" for r in master)
    by_seniority = top_counts(detect_seniority(r.get("title", ""), r.get("role", "")) for r in master)

    contacts_per_company: dict[str, int] = {}
    for r in master:
        c = r.get("company") or ""
        if c:
            contacts_per_company[c] = contacts_per_company.get(c, 0) + 1
    top_active = sorted(contacts_per_company.items(), key=lambda kv: -kv[1])[:15]

    coverage = {
        "companies_with_contacts": len(companies_with),
        "companies_without_contacts": len(companies_without),
        "contacts_with_email": sum(1 for r in master if r.get("email")),
        "contacts_with_linkedin": sum(1 for r in master if r.get("linkedin")),
        "contacts_with_phone": sum(1 for r in master if r.get("phone")),
        "primary_yes": sum(1 for r in master if str(r.get("is_primary_contact", "")).lower() in ("yes", "true", "1")),
    }

    return {
        "totals": {
            "contacts": len(master),
            "unique_companies": len(contact_companies),
            "company_records": len(companies),
            "activity_records": len(activities),
            "last_batch_id": last_batch_id,
        },
        "by_stage": by_stage[:30],
        "by_pipeline_state": by_pipeline,
        "by_investor_type": by_investor[:30],
        "by_continent": by_continent[:30],
        "by_hotel_focus": by_hotel,
        "by_seniority": by_seniority,
        "coverage": coverage,
        "top_active_investors": top_active,
    }


# ── report writers ──────────────────────────────────────────────────────────


def write_jsonl(path: Path, record: dict[str, Any]) -> None:
    with path.open("a", encoding="utf-8") as f:
        f.write(json.dumps(record, ensure_ascii=False) + "\n")


def write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str]) -> None:
    with path.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        for r in rows:
            w.writerow({k: ("" if r.get(k) is None else r.get(k)) for k in fieldnames})


# ── main pipeline ───────────────────────────────────────────────────────────


def process_one_file(path: Path, batch_id: str, master_state: dict[str, list[dict[str, Any]]]) -> dict[str, Any]:
    """Mutates master_state in place. Returns per-file ingestion summary."""
    parsed = parse_workbook(path, batch_id)
    company_idx = companies_index(parsed.companies)
    activity_idx = activities_index(parsed.activities)

    # Build per-person master rows from this file
    new_rows: list[dict[str, Any]] = []
    invalid_rows: list[dict[str, str]] = []
    for contact in parsed.contacts:
        ckey = company_key(contact.get("company"))
        company = company_idx.get(ckey)
        activity = activity_idx.get(ckey)
        master_row = build_master_row(contact, company, activity, path.name, batch_id)
        if master_row is None:
            invalid_rows.append({
                "source_file": path.name,
                "datasite_contact_number": norm_text(contact.get("datasite_contact_number")),
                "reason": "missing_identifier (no email, no LinkedIn, no name+company)",
                "company": norm_text(contact.get("company")),
                "first_name": norm_text(contact.get("first_name")),
                "last_name": norm_text(contact.get("last_name")),
            })
            continue
        new_rows.append(master_row)

    # Dedupe against existing master
    existing = master_state["master"]
    by_email = {r["email"]: i for i, r in enumerate(existing) if r.get("email")}
    by_linkedin = {r["linkedin"]: i for i, r in enumerate(existing) if r.get("linkedin")}
    by_name_company = {(r["full_name"].lower(), company_key(r["company"])): i
                        for i, r in enumerate(existing) if r.get("full_name") and r.get("company")}

    inserted = 0
    updated = 0
    unchanged = 0
    dup_decisions: list[dict[str, Any]] = []

    for new_row in new_rows:
        idx, strategy, score = find_match(new_row, by_email, by_linkedin, by_name_company, existing)
        if idx is None:
            existing.append(new_row)
            # Keep indices fresh
            i = len(existing) - 1
            if new_row["email"]:
                by_email[new_row["email"]] = i
            if new_row["linkedin"]:
                by_linkedin[new_row["linkedin"]] = i
            if new_row["full_name"] and new_row["company"]:
                by_name_company[(new_row["full_name"].lower(), company_key(new_row["company"]))] = i
            inserted += 1
            dup_decisions.append({
                "decision": "INSERT",
                "strategy": "new",
                "score": "",
                "master_id": new_row["master_id"],
                "email": new_row["email"],
                "full_name": new_row["full_name"],
                "company": new_row["company"],
                "fields_changed": "",
                "source_file": path.name,
            })
        else:
            merged, changed = merge_existing(existing[idx], new_row)
            existing[idx] = merged
            if changed:
                updated += 1
            else:
                unchanged += 1
            dup_decisions.append({
                "decision": "UPDATE" if changed else "UNCHANGED",
                "strategy": strategy,
                "score": f"{score:.3f}",
                "master_id": merged["master_id"],
                "email": merged["email"],
                "full_name": merged["full_name"],
                "company": merged["company"],
                "fields_changed": "|".join(changed),
                "source_file": path.name,
            })

    # Update canonical Contacts/Companies/Activities sheets (also incremental)
    # Contacts: keep ALL rows · dedupe by (company, first_name, last_name, email)
    contacts_existing = master_state["contacts"]
    contacts_seen = {(norm_text(r.get("company")).lower(),
                      norm_text(r.get("first_name")).lower(),
                      norm_text(r.get("last_name")).lower(),
                      norm_email(r.get("email"))) for r in contacts_existing}
    for c in parsed.contacts:
        norm_c = {
            "datasite_contact_number": norm_text(c.get("datasite_contact_number")),
            "company": norm_text(c.get("company")),
            "first_name": norm_text(c.get("first_name")),
            "last_name": norm_text(c.get("last_name")),
            "email": norm_email(c.get("email")),
            "office_phone": norm_phone(c.get("office_phone")),
            "cell_phone": norm_phone(c.get("cell_phone")),
            "linkedin": norm_linkedin(c.get("linkedin")),
            "title": norm_text(c.get("title")),
            "role": norm_text(c.get("role")),
            "is_primary_contact": norm_text(c.get("is_primary_contact")),
            "industry": norm_text(c.get("industry")),
            "relationship_status": norm_text(c.get("relationship_status")),
            "relationship_manager": norm_text(c.get("relationship_manager")),
            "address": norm_text(c.get("address")),
            "city": norm_text(c.get("city")),
            "state": norm_text(c.get("state")),
            "zip_code": norm_text(c.get("zip_code")),
            "country": norm_text(c.get("country")),
            "contact_notes": norm_text(c.get("contact_notes")),
            "client_contact_id": norm_text(c.get("client_contact_id")),
        }
        key = (norm_c["company"].lower(), norm_c["first_name"].lower(), norm_c["last_name"].lower(), norm_c["email"])
        if key in contacts_seen:
            continue
        contacts_seen.add(key)
        contacts_existing.append(norm_c)

    # Companies · dedupe by company_key
    companies_existing = master_state["companies"]
    companies_seen = {company_key(r.get("company")) for r in companies_existing}
    for c in parsed.companies:
        ckey = company_key(c.get("company"))
        if not ckey or ckey in companies_seen:
            continue
        companies_seen.add(ckey)
        notes_field = consolidate_notes(c.get("company_notes"), c.get("internal_notes"),
                                        c.get("external_notes"), c.get("keyword"))
        canon = canonical_investor_type(c.get("investor_type_raw"),
                                        c.get("investor_subtype"),
                                        c.get("description"),
                                        notes_field)
        companies_existing.append({
            "datasite_company_number": norm_text(c.get("datasite_company_number")),
            "company": norm_text(c.get("company")),
            "investor_type_canonical": canon,
            "investor_type_raw": norm_text(c.get("investor_type_raw")),
            "investor_subtype": norm_text(c.get("investor_subtype")),
            "tier": norm_text(c.get("tier")),
            "company_type": norm_text(c.get("company_type")),
            "company_industry": norm_text(c.get("company_industry")),
            "continent": norm_text(c.get("continent")),
            "company_location": norm_text(c.get("company_location")),
            "fund_size": norm_text(c.get("fund_size")),
            "investment_preference": norm_text(c.get("investment_preference")),
            "investment_min": norm_text(c.get("investment_min")),
            "investment_max": norm_text(c.get("investment_max")),
            "association": norm_text(c.get("association")),
            "description": norm_text(c.get("description")),
            "keyword": norm_text(c.get("keyword")),
            "coverage_officer": norm_text(c.get("coverage_officer")),
            "calling_lead": norm_text(c.get("calling_lead")),
            "internal_notes": norm_text(c.get("internal_notes")),
            "external_notes": norm_text(c.get("external_notes")),
            "company_notes": norm_text(c.get("company_notes")),
            "client_company_id": norm_text(c.get("client_company_id")),
        })

    # Activities · dedupe by company_key (one timeline per company; latest export wins)
    activities_existing = master_state["activities"]
    by_company_act = {company_key(r.get("company")): i for i, r in enumerate(activities_existing)}
    for a in parsed.activities:
        ckey = company_key(a.get("company"))
        if not ckey:
            continue
        row = {
            "datasite_activity_number": norm_text(a.get("datasite_activity_number")),
            "company": norm_text(a.get("company")),
            "buyer_added_date": to_iso_date(a.get("buyer_added_date")),
            "initial_contact_date": to_iso_date(a.get("initial_contact_date")),
            "teaser_sent_date": to_iso_date(a.get("teaser_sent_date")),
            "nda_initial_sent_date": to_iso_date(a.get("nda_initial_sent_date")),
            "nda_signed_date": to_iso_date(a.get("nda_signed_date")),
            "nda_executed_date": to_iso_date(a.get("nda_executed_date")),
            "cim_sent_date": to_iso_date(a.get("cim_sent_date")),
            "ioi_process_letter_date": to_iso_date(a.get("ioi_process_letter_date")),
            "ioi_bid_received_date": to_iso_date(a.get("ioi_bid_received_date")),
            "loi_process_letter_date": to_iso_date(a.get("loi_process_letter_date")),
            "loi_bid_received_date": to_iso_date(a.get("loi_bid_received_date")),
            "management_presentation_date": to_iso_date(a.get("management_presentation_date")),
            "revised_bid_received_date": to_iso_date(a.get("revised_bid_received_date")),
            "declined_date": to_iso_date(a.get("declined_date")),
            "declined_reason": norm_text(a.get("declined_reason")),
            "declined_comments": norm_text(a.get("declined_comments")),
            "last_activity_type": norm_text(a.get("last_activity_type")),
            "last_activity_date": to_iso_date(a.get("last_activity_date")),
            "last_activity_comments": norm_text(a.get("last_activity_comments")),
            "latest_deal_stage": norm_text(a.get("latest_deal_stage")),
            "pipeline_state": pipeline_state(a),
            "ioi_bid_low": norm_text(a.get("ioi_bid_low")),
            "ioi_bid_high": norm_text(a.get("ioi_bid_high")),
            "loi_bid_low": norm_text(a.get("loi_bid_low")),
            "loi_bid_high": norm_text(a.get("loi_bid_high")),
            "revised_bid_low": norm_text(a.get("revised_bid_low")),
            "revised_bid_high": norm_text(a.get("revised_bid_high")),
            "removed_flag": norm_text(a.get("removed_flag")),
            "on_hold_flag": norm_text(a.get("on_hold_flag")),
            "client_buyer_id": norm_text(a.get("client_buyer_id")),
        }
        if ckey in by_company_act:
            activities_existing[by_company_act[ckey]] = row
        else:
            by_company_act[ckey] = len(activities_existing)
            activities_existing.append(row)

    return {
        "batch_id": batch_id,
        "source_file": path.name,
        "rows_seen": len(parsed.contacts),
        "rows_inserted": inserted,
        "rows_updated": updated,
        "rows_unchanged": unchanged,
        "rows_invalid": len(invalid_rows),
        "company_records_seen": len(parsed.companies),
        "activity_records_seen": len(parsed.activities),
        "duplicate_decisions": dup_decisions,
        "schema_mapping_rows": parsed.schema_map_rows,
        "invalid_rows": invalid_rows,
    }


def reclassify_master_rows(
    master_rows: list[dict[str, Any]],
    companies_rows: list[dict[str, Any]],
) -> int:
    """Re-run derived classifiers against existing master rows.

    Used after the canonical mapping rules (INVESTOR_TYPE_CANON,
    HOTEL_FOCUS_HINT, SENIORITY_PATTERNS) are updated. Reads each row
    plus the company's raw notes/description (loaded from Companies
    sheet, which preserves the original Datasite raw values) and
    re-computes investor_type + hotel_focus in place.

    Returns the number of rows whose investor_type changed.
    """
    company_idx = {company_key(c.get("company")): c for c in companies_rows}
    changed = 0
    for row in master_rows:
        ckey = company_key(row.get("company"))
        company = company_idx.get(ckey, {})
        raw = company.get("investor_type_raw")
        subtype = company.get("investor_subtype")
        description = company.get("description")
        notes = " ".join(str(company.get(k, "") or "") for k in (
            "company_notes", "internal_notes", "external_notes", "keyword",
        ))
        new_investor = canonical_investor_type(raw, subtype, description, notes)
        new_hotel = detect_hotel_focus(
            row.get("industry"), description, notes, company.get("association"),
        )
        if row.get("investor_type") != new_investor:
            row["investor_type"] = new_investor
            changed += 1
        row["hotel_focus"] = new_hotel
        row["last_updated_at"] = now_iso()
    return changed


def main() -> int:
    reclassify = "--reclassify" in sys.argv

    incoming_files = sorted(
        [p for p in INCOMING.iterdir() if p.is_file() and p.suffix.lower() in (".xlsm", ".xlsx")],
        key=lambda p: p.stat().st_mtime,
    )
    print(f"→ CONTACTOS DATASITE ingest · {len(incoming_files)} file(s) in incoming/ · reclassify={reclassify}")

    master_rows, contacts_rows, companies_rows, activities_rows = load_existing_master()
    print(f"  existing master: {len(master_rows)} contacts · {len(companies_rows)} companies · {len(activities_rows)} activities")

    state = {
        "master": master_rows,
        "contacts": contacts_rows,
        "companies": companies_rows,
        "activities": activities_rows,
    }

    last_batch_id = ""

    if reclassify and state["master"]:
        n = reclassify_master_rows(state["master"], state["companies"])
        print(f"  reclassified · {n} master rows had investor_type change")
        write_jsonl(REPORTS / "ingestion-log.jsonl", {
            "batch_id": "reclassify-" + datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ"),
            "source_file": "(reclassify · no source)",
            "rows_seen": 0,
            "rows_inserted": 0,
            "rows_updated": n,
            "rows_unchanged": len(state["master"]) - n,
            "rows_invalid": 0,
            "company_records_seen": 0,
            "activity_records_seen": 0,
            "completed_at": now_iso(),
        })

    if not incoming_files:
        print("  no new files · regenerating Summary against current master")
        summary = build_summary(state["master"], state["companies"], state["activities"], last_batch_id)
        write_master_workbook(state["master"], state["contacts"], state["companies"], state["activities"], summary)
        print(f"✓ master rewritten · {MASTER_FILE}")
        return 0

    for src in incoming_files:
        batch_id = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
        last_batch_id = batch_id
        print(f"\n→ processing {src.name} · batch_id={batch_id}")
        result = process_one_file(src, batch_id, state)
        print(f"  seen={result['rows_seen']} insert={result['rows_inserted']} "
              f"update={result['rows_updated']} unchanged={result['rows_unchanged']} "
              f"invalid={result['rows_invalid']}")

        # Per-batch reports
        write_jsonl(REPORTS / "ingestion-log.jsonl", {
            "batch_id": batch_id,
            "source_file": result["source_file"],
            "rows_seen": result["rows_seen"],
            "rows_inserted": result["rows_inserted"],
            "rows_updated": result["rows_updated"],
            "rows_unchanged": result["rows_unchanged"],
            "rows_invalid": result["rows_invalid"],
            "company_records_seen": result["company_records_seen"],
            "activity_records_seen": result["activity_records_seen"],
            "completed_at": now_iso(),
        })
        write_csv(REPORTS / f"duplicate-resolution_{batch_id}.csv",
                  result["duplicate_decisions"],
                  ["decision", "strategy", "score", "master_id", "email",
                   "full_name", "company", "fields_changed", "source_file"])
        write_csv(REPORTS / f"schema-mapping_{batch_id}.csv",
                  result["schema_mapping_rows"],
                  ["source_file", "sheet", "source_column", "canonical_field", "mapped"])
        write_csv(REPORTS / f"invalid-missing_{batch_id}.csv",
                  result["invalid_rows"],
                  ["source_file", "datasite_contact_number", "reason",
                   "company", "first_name", "last_name"])

        # Move source → old/
        dest = OLD / src.name
        if dest.exists():
            stem = dest.stem
            dest = OLD / f"{stem}.{batch_id}{src.suffix}"
        shutil.move(str(src), str(dest))
        print(f"  moved {src.name} → old/{dest.name}")

    summary = build_summary(state["master"], state["companies"], state["activities"], last_batch_id)
    write_master_workbook(state["master"], state["contacts"], state["companies"], state["activities"], summary)
    print(f"\n✓ master written · {MASTER_FILE}")
    print(f"  totals · contacts={len(state['master'])} · companies={len(state['companies'])} · activities={len(state['activities'])}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
