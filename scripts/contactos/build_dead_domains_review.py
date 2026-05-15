#!/usr/bin/env python3
"""
Phase 2.B.3 · Dead-domains review generator.

For every domain with ≥ 50% bounce rate AND ≥ 2 contacts seen, surface ALL
the institutional context Miguel needs to investigate the rebrand · merger ·
domain change · or company shutdown:

  - domain · bounced/total · ratio
  - sample contacts (full_name + company + last_email_date)
  - distinct companies represented at this domain
  - latest known activity (max last_email_date across all domain contacts)
  - latest bounce date (max last_bounce_date)
  - has_active_relationships? (any contact with band ∈ {active, strategic, warm})
  - investigation_priority: H/M/L based on (latest activity recency · band)
  - rebrand_hint: best-guess domain candidates from Master that share
    a normalised company token (e.g. metrovacesa.com → metrovacesa.es)

This is human-driven · NEVER auto-applied. Operator opens the CSV, looks up
LinkedIn, and decides per row whether to update the corporate-domain mapping.

Output (overwritten on each run):
  CONTACTOS DATASITE/reports/dead-domains-review_<batch>.csv
"""
from __future__ import annotations

import csv
import json
import re
import sys
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except (AttributeError, ValueError):
    pass

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl is required · pip install openpyxl", file=sys.stderr)
    sys.exit(2)

REPO = Path(__file__).resolve().parent.parent.parent
ROOT = REPO / "CONTACTOS DATASITE"
MASTER_PATH = ROOT / "master" / "metcub-contacts-master.xlsx"
SIGNAL_DIRS = [ROOT / "incoming" / "gmail-signals", ROOT / "old" / "gmail-signals"]
REPORTS = ROOT / "reports"
REPORTS.mkdir(parents=True, exist_ok=True)

PERSONAL_DOMAINS = {
    "gmail.com", "googlemail.com", "hotmail.com", "outlook.com", "outlook.es",
    "live.com", "yahoo.com", "yahoo.es", "icloud.com", "me.com", "mac.com",
    "protonmail.com", "proton.me", "yandex.com", "mail.ru", "aol.com",
    "msn.com", "ymail.com", "telefonica.net", "terra.es",
}

DEAD_RATIO = 0.50
MIN_CONTACTS = 2
ACTIVE_BANDS = {"active", "strategic", "warm"}
HIGH_PRIORITY_BANDS = {"active", "strategic"}
RECENT_DAYS = 730  # 24 mo


def domain_of(email: str | None) -> str:
    if not email or "@" not in email:
        return ""
    return email.split("@", 1)[1].strip().lower()


def normalise_company_token(name: str | None) -> str:
    if not name:
        return ""
    s = re.sub(r"[^a-z0-9]+", "", name.lower())
    return s


def parse_iso(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value if value.tzinfo else value.replace(tzinfo=timezone.utc)
    if not isinstance(value, str) or not value.strip():
        return None
    s = value.strip().replace("Z", "+00:00")
    try:
        d = datetime.fromisoformat(s)
        return d if d.tzinfo else d.replace(tzinfo=timezone.utc)
    except ValueError:
        pass
    try:
        d = datetime.strptime(value.strip(), "%Y-%m-%d")
        return d.replace(tzinfo=timezone.utc)
    except ValueError:
        return None


def fmt_date(value: Any) -> str:
    d = parse_iso(value)
    return d.date().isoformat() if d else ""


def collect_master(path: Path) -> tuple[
    dict[str, list[dict]],          # domain → [contact dict, ...]
    dict[str, set[str]],            # company_token → {domains}
]:
    wb = load_workbook(path, read_only=True)
    by_domain: dict[str, list[dict]] = defaultdict(list)
    company_token_to_domains: dict[str, set[str]] = defaultdict(set)

    for sheet in ("Master", "INVALID_ARCHIVE"):
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        idx = {h: i for i, h in enumerate(header) if h}
        if "email" not in idx:
            continue
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not r:
                continue
            em = r[idx["email"]]
            if not isinstance(em, str) or "@" not in em:
                continue
            dom = domain_of(em)
            if not dom or dom in PERSONAL_DOMAINS:
                continue
            entry = {
                "email": em.strip().lower(),
                "full_name": r[idx["full_name"]] if "full_name" in idx else "",
                "company": r[idx["company"]] if "company" in idx else "",
                "contact_category": r[idx["contact_category"]] if "contact_category" in idx else "",
                "relationship_band": r[idx["relationship_band"]] if "relationship_band" in idx else "",
                "last_email_date": r[idx["last_email_date"]] if "last_email_date" in idx else None,
                "last_bounce_date": r[idx["last_bounce_date"]] if "last_bounce_date" in idx else None,
                "email_validity": r[idx["email_validity"]] if "email_validity" in idx else "",
                "source_sheet": sheet,
            }
            by_domain[dom].append(entry)
            tok = normalise_company_token(entry["company"])
            if tok:
                company_token_to_domains[tok].add(dom)

    wb.close()
    return by_domain, company_token_to_domains


def collect_gmail_bounce_stats() -> dict[str, dict[str, int]]:
    stats: dict[str, dict[str, int]] = defaultdict(lambda: {"total": 0, "bounced": 0})
    seen: dict[str, set[str]] = defaultdict(set)
    bounced_emails: dict[str, set[str]] = defaultdict(set)
    for d in SIGNAL_DIRS:
        if not d.exists():
            continue
        for f in sorted(d.glob("*.jsonl")):
            with f.open("r", encoding="utf-8") as fh:
                for line in fh:
                    line = line.strip()
                    if not line:
                        continue
                    try:
                        rec = json.loads(line)
                    except json.JSONDecodeError:
                        continue
                    em = (rec.get("email") or "").strip().lower()
                    if not em or "@" not in em:
                        continue
                    dom = em.split("@", 1)[1]
                    if dom in PERSONAL_DOMAINS:
                        continue
                    seen[dom].add(em)
                    bc = int(rec.get("bounce_count") or 0)
                    ic = int(rec.get("inbound_count") or 0)
                    if bc >= 2 or (bc >= 1 and ic == 0):
                        bounced_emails[dom].add(em)
    for dom, emails in seen.items():
        stats[dom]["total"] = len(emails)
        stats[dom]["bounced"] = len(bounced_emails.get(dom, set()))
    return stats


def main() -> int:
    if not MASTER_PATH.exists():
        print(f"ERROR: Master not found · {MASTER_PATH}", file=sys.stderr)
        return 1

    print(f"→ Dead-domains review · loading sources")
    by_domain, token_to_doms = collect_master(MASTER_PATH)
    gmail_stats = collect_gmail_bounce_stats()

    # Pick dead domains using gmail-signal stats (authoritative on bounces)
    dead: list[dict] = []
    now = datetime.now(timezone.utc)
    for dom, stats in gmail_stats.items():
        if stats["total"] < MIN_CONTACTS:
            continue
        ratio = stats["bounced"] / stats["total"] if stats["total"] else 0.0
        if ratio < DEAD_RATIO:
            continue

        contacts = by_domain.get(dom, [])
        companies = sorted({(c["company"] or "").strip() for c in contacts if c["company"]})
        sample = sorted(
            contacts,
            key=lambda c: (
                0 if (c["relationship_band"] or "").lower() in HIGH_PRIORITY_BANDS else 1,
                -((parse_iso(c["last_email_date"]) or datetime(1970, 1, 1, tzinfo=timezone.utc)).timestamp()),
            ),
        )[:5]

        latest_activity = None
        latest_bounce = None
        bands_seen: set[str] = set()
        for c in contacts:
            d = parse_iso(c["last_email_date"])
            if d and (latest_activity is None or d > latest_activity):
                latest_activity = d
            d2 = parse_iso(c["last_bounce_date"])
            if d2 and (latest_bounce is None or d2 > latest_bounce):
                latest_bounce = d2
            if c["relationship_band"]:
                bands_seen.add(c["relationship_band"].lower())

        is_recent = bool(latest_activity and (now - latest_activity).days <= RECENT_DAYS)
        has_active = bool(bands_seen & ACTIVE_BANDS)
        if has_active and is_recent:
            priority = "H"
        elif has_active or is_recent:
            priority = "M"
        else:
            priority = "L"

        # Rebrand hint: same normalised company token mapped to other domains
        rebrand_hints: set[str] = set()
        for c in contacts:
            tok = normalise_company_token(c["company"])
            if not tok:
                continue
            for cand_dom in token_to_doms.get(tok, set()):
                if cand_dom != dom and cand_dom not in PERSONAL_DOMAINS:
                    cand_stats = gmail_stats.get(cand_dom)
                    if cand_stats and cand_stats["total"] >= 1:
                        cand_ratio = cand_stats["bounced"] / cand_stats["total"] if cand_stats["total"] else 0
                        if cand_ratio < DEAD_RATIO:
                            rebrand_hints.add(cand_dom)
                    else:
                        rebrand_hints.add(cand_dom)

        dead.append({
            "domain": dom,
            "bounced": stats["bounced"],
            "total_contacts_seen": stats["total"],
            "bounce_ratio_pct": round(ratio * 100, 1),
            "investigation_priority": priority,
            "has_active_relationships": "yes" if has_active else "no",
            "bands_seen": ",".join(sorted(bands_seen)) if bands_seen else "",
            "latest_activity_date": fmt_date(latest_activity),
            "latest_bounce_date": fmt_date(latest_bounce),
            "distinct_companies_count": len(companies),
            "distinct_companies": " | ".join(companies[:5]),
            "sample_contacts": " | ".join(
                f"{c['full_name'] or '?'} <{c['email']}> [{c['relationship_band'] or '-'}]"
                for c in sample
            ),
            "rebrand_hint_domains": " | ".join(sorted(rebrand_hints)) if rebrand_hints else "",
            "investigation_action": "",   # operator fills: rebrand · shutdown · merger · skip
            "new_corporate_domain": "",   # operator fills if rebrand confirmed
        })

    # Sort by priority H>M>L · then by ratio desc · then by total contacts desc
    pri_rank = {"H": 0, "M": 1, "L": 2}
    dead.sort(key=lambda x: (pri_rank.get(x["investigation_priority"], 9), -x["bounce_ratio_pct"], -x["total_contacts_seen"]))

    batch_id = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    out_csv = REPORTS / f"dead-domains-review_{batch_id}.csv"
    if dead:
        with out_csv.open("w", encoding="utf-8", newline="") as f:
            w = csv.DictWriter(f, fieldnames=list(dead[0].keys()))
            w.writeheader()
            for row in dead:
                w.writerow(row)
    else:
        out_csv.write_text("# no dead domains identified at threshold\n", encoding="utf-8")

    h_count = sum(1 for d in dead if d["investigation_priority"] == "H")
    m_count = sum(1 for d in dead if d["investigation_priority"] == "M")
    l_count = sum(1 for d in dead if d["investigation_priority"] == "L")
    print(f"  · dead domains: {len(dead)} (H={h_count} M={m_count} L={l_count})")
    print(f"\n✓ {out_csv.relative_to(REPO)}")
    if dead:
        print(f"\nTop 10 high-priority:")
        for d in dead[:10]:
            print(f"  [{d['investigation_priority']}] {d['domain']:30s} {d['bounced']}/{d['total_contacts_seen']} "
                  f"({d['bounce_ratio_pct']:.0f}%) · companies={d['distinct_companies_count']} · "
                  f"hint={d['rebrand_hint_domains'] or '—'}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
