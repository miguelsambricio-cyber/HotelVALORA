#!/usr/bin/env python3
"""
Phase 2.B.3 · Replacement-suggestion + invalid-archive engine.

Goal
----
For every Master contact flagged `email_validity = invalid`, look for a likely
replacement *within the same domain* using these signals (priority order):
  1. Same surname (exact normalised match)
  2. Same surname-initial pattern (e.g. j.smith vs john.smith → smith)
  3. Recent Gmail activity (last_email_date within last 24 months)
  4. Strong relationship_band (active / strategic > warm > cold > dormant)
  5. Same contact_category (Principal · Broker · Lender · Developer · Proveedor)

Invalids fall in two lanes:
  - WITH ≥ 1 candidate above MIN_SCORE → emit row in replacement-suggestions.csv
    (NEVER auto-applied · operator decides)
  - WITHOUT any candidate → move to INVALID_ARCHIVE sheet inside the same xlsx
    with full provenance (archived_at · archived_batch_id · archived_reason)

The Master-active row count drops by N (where N = invalids without candidate).
All 63 canonical columns are preserved on the archive side.

Modes
-----
  python build_replacement_suggestions.py
    → generates fresh suggestions CSV + archives no-candidate invalids

  python build_replacement_suggestions.py --apply <suggestions_csv_path>
    → reads CSV with apply_decision column filled, applies approved replacements
    → preserves original_email + audit trail (source, replaced_by, replaced_at)

Outputs
-------
  CONTACTOS DATASITE/reports/replacement-suggestions_<batch>.csv
  CONTACTOS DATASITE/reports/invalid-archive-manifest_<batch>.csv
  CONTACTOS DATASITE/reports/invalid-archive-log.jsonl  (append-only audit)
  CONTACTOS DATASITE/master/metcub-contacts-master.xlsx (INVALID_ARCHIVE sheet
                                                         created/extended, or
                                                         replacements applied)

Re-runnable: archive sheet appends · suggestions CSV is per-batch · --apply is
idempotent (same CSV applied twice → no-op on already-applied rows).
"""
from __future__ import annotations

import csv
import json
import re
import sys
import unicodedata
from collections import defaultdict
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any

try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except (AttributeError, ValueError):
    pass

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill
except ImportError:
    print("ERROR: openpyxl is required · pip install openpyxl", file=sys.stderr)
    sys.exit(2)

REPO = Path(__file__).resolve().parent.parent.parent
ROOT = REPO / "CONTACTOS DATASITE"
MASTER_PATH = ROOT / "master" / "metcub-contacts-master.xlsx"
REPORTS = ROOT / "reports"
REPORTS.mkdir(parents=True, exist_ok=True)

ARCHIVE_LOG = REPORTS / "invalid-archive-log.jsonl"

# Personal-domain emails are NEVER valid replacement candidates · they belong
# to a person, not an institution, so the dead address can't be replaced by a
# colleague at the same domain.
PERSONAL_DOMAINS = {
    "gmail.com", "googlemail.com", "hotmail.com", "outlook.com", "outlook.es",
    "live.com", "yahoo.com", "yahoo.es", "icloud.com", "me.com", "mac.com",
    "protonmail.com", "proton.me", "yandex.com", "mail.ru", "aol.com",
    "msn.com", "ymail.com", "telefonica.net", "terra.es",
}

MIN_SCORE = 30        # below this · candidate is not surfaced
RECENT_MONTHS = 24    # for "recent activity" boost
ACTIVE_BANDS = {"active", "strategic"}
WARM_BANDS = {"warm"}


def norm_text(value: Any) -> str:
    if not isinstance(value, str):
        return ""
    s = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode().lower()
    return re.sub(r"\s+", " ", s).strip()


def domain_of(email: str | None) -> str:
    if not email or "@" not in email:
        return ""
    return email.split("@", 1)[1].strip().lower()


def local_of(email: str | None) -> str:
    if not email or "@" not in email:
        return ""
    return email.split("@", 1)[0].strip().lower()


def surname_from_full_name(full_name: str | None) -> str:
    if not full_name:
        return ""
    parts = norm_text(full_name).split()
    if not parts:
        return ""
    # Spanish names often have two surnames · last token is the most reliable.
    return parts[-1]


def surname_from_local(local: str) -> str:
    """Heuristic: in firstname.lastname / f.lastname / firstinitial-lastname."""
    if not local:
        return ""
    cleaned = re.split(r"[._\-+]", local)
    cleaned = [c for c in cleaned if c]
    if not cleaned:
        return ""
    return cleaned[-1]


def parse_date_safe(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value if value.tzinfo else value.replace(tzinfo=timezone.utc)
    if not isinstance(value, str) or not value.strip():
        return None
    s = value.strip()
    for fmt in ("%Y-%m-%d", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%dT%H:%M:%S%z", "%Y-%m-%d %H:%M:%S"):
        try:
            d = datetime.strptime(s.replace("Z", "+0000") if "Z" in s and "%z" in fmt else s, fmt)
            return d if d.tzinfo else d.replace(tzinfo=timezone.utc)
        except ValueError:
            continue
    try:
        d = datetime.fromisoformat(s.replace("Z", "+00:00"))
        return d if d.tzinfo else d.replace(tzinfo=timezone.utc)
    except ValueError:
        return None


def is_recent(value: Any, months: int = RECENT_MONTHS) -> bool:
    d = parse_date_safe(value)
    if d is None:
        return False
    cutoff = datetime.now(timezone.utc) - timedelta(days=months * 30)
    return d >= cutoff


def score_candidate(invalid: dict, candidate: dict) -> tuple[int, list[str]]:
    """Return (score, reasoning_tokens). Same-domain is a precondition."""
    score = 0
    reasons: list[str] = []

    inv_surname = invalid["surname"]
    cand_surname = candidate["surname"]
    inv_local_surname = surname_from_local(invalid["local"])
    cand_local_surname = surname_from_local(candidate["local"])

    # 1. Surname match (full-name or derived from email local part)
    if inv_surname and cand_surname and inv_surname == cand_surname:
        score += 50
        reasons.append("same_surname_full_name")
    elif inv_local_surname and cand_local_surname and inv_local_surname == cand_local_surname:
        score += 35
        reasons.append("same_surname_email_local")
    elif inv_surname and cand_local_surname and inv_surname == cand_local_surname:
        score += 35
        reasons.append("same_surname_cross_match")
    elif inv_local_surname and cand_surname and inv_local_surname == cand_surname:
        score += 35
        reasons.append("same_surname_cross_match")

    # 2. Recent Gmail activity boost
    if is_recent(candidate.get("last_email_date")):
        score += 20
        reasons.append(f"recent_activity_<{RECENT_MONTHS}mo")

    # 3. Strong relationship band
    band = (candidate.get("relationship_band") or "").lower()
    if band in ACTIVE_BANDS:
        score += 30
        reasons.append(f"band={band}")
    elif band in WARM_BANDS:
        score += 15
        reasons.append(f"band={band}")

    # 4. Same contact_category (taxonomy continuity)
    if invalid.get("contact_category") and candidate.get("contact_category") \
            and invalid["contact_category"] == candidate["contact_category"]:
        score += 10
        reasons.append("same_category")

    # 5. Inbound thread evidence (someone is actually answering)
    try:
        threads = int(candidate.get("active_threads") or 0)
    except (TypeError, ValueError):
        threads = 0
    if threads >= 2:
        score += 10
        reasons.append(f"threads={threads}")

    return score, reasons


def apply_replacements(suggestions_csv_path: str) -> int:
    """Apply approved replacements from CSV with apply_decision column filled."""
    csv_path = Path(suggestions_csv_path)
    if not csv_path.exists():
        print(f"ERROR: CSV not found · {csv_path}", file=sys.stderr)
        return 1
    if not MASTER_PATH.exists():
        print(f"ERROR: Master not found · {MASTER_PATH}", file=sys.stderr)
        return 1

    print(f"→ --apply mode · loading suggestions CSV")
    suggestions: list[dict] = []
    with csv_path.open("r", encoding="utf-8") as f:
        r = csv.DictReader(f)
        for row in r:
            if row:
                suggestions.append(row)

    approved = [s for s in suggestions if (s.get("apply_decision") or "").strip().lower() == "apply"]
    print(f"  · total rows in CSV: {len(suggestions)}")
    print(f"  · marked 'apply': {len(approved)}")

    if not approved:
        print("  · nothing to apply · exiting")
        return 0

    print(f"\n→ Loading Master")
    wb = load_workbook(MASTER_PATH)
    ws = wb["Master"]

    # Read original header and data
    header = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(header) if h}
    orig_idx = dict(idx)  # Keep original index for reading rows

    # Build new header in-memory (no worksheet modification yet)
    new_header = list(header)
    if "original_email" not in idx:
        new_header.insert(0, "original_email")
    if "replacement_source" not in new_header:
        new_header.append("replacement_source")
    if "replaced_by_master_id" not in new_header:
        new_header.append("replaced_by_master_id")
    if "replaced_at" not in new_header:
        new_header.append("replaced_at")

    # New index mapping
    idx = {h: i for i, h in enumerate(new_header) if h}

    # Build map: master_id → (sheet_row, row_data)
    by_master_id: dict[str, tuple[int, list[Any]]] = {}
    rows_for_sheet: list[list[Any]] = []
    for sheet_row, r in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if r:
            row_data = list(r)
            # If original_email wasn't in header, insert at position 0 FIRST
            if "original_email" not in orig_idx:
                row_data.insert(0, None)

            # Then extend to new header length
            while len(row_data) < len(new_header):
                row_data.append(None)

            # Map master_id using ORIGINAL index (before we inserted original_email)
            if "master_id" in orig_idx:
                mid_idx_orig = orig_idx["master_id"]
                # After we inserted at position 0, master_id shifts
                mid_idx_adjusted = mid_idx_orig + 1 if "original_email" not in orig_idx else mid_idx_orig
                mid = row_data[mid_idx_adjusted] if mid_idx_adjusted < len(row_data) else None
                if mid:
                    by_master_id[str(mid)] = (sheet_row, row_data)
            rows_for_sheet.append(row_data)

    # Apply approved replacements
    replaced_count = 0
    already_applied = 0
    errors = 0
    replaced_at_iso = datetime.now(timezone.utc).replace(microsecond=0).isoformat()

    for sugg in approved:
        invalid_id = sugg.get("invalid_master_id", "").strip()
        new_email = sugg.get("candidate_email", "").strip().lower()
        replaced_by_id = sugg.get("candidate_master_id", "").strip()

        if not invalid_id or not new_email or not replaced_by_id:
            print(f"  · SKIP invalid suggestion row (missing fields): {invalid_id}", file=sys.stderr)
            errors += 1
            continue

        if invalid_id not in by_master_id:
            print(f"  · SKIP {invalid_id}: not found in Master", file=sys.stderr)
            errors += 1
            continue

        sheet_row, row_data = by_master_id[invalid_id]
        row_data = list(row_data)

        # Check if already applied (original_email is set)
        orig_email_idx = idx.get("original_email")
        if orig_email_idx is not None and orig_email_idx < len(row_data) and row_data[orig_email_idx]:
            print(f"  · SKIP {invalid_id}: already applied")
            already_applied += 1
            continue

        # Preserve original email before overwriting
        email_idx_orig = orig_idx.get("email", 0)
        email_idx_adjusted = email_idx_orig + 1 if "original_email" not in orig_idx else email_idx_orig
        old_email = row_data[email_idx_adjusted] if email_idx_adjusted < len(row_data) else ""
        email_idx = idx.get("email")  # New index for writing

        # Ensure all indices exist in row_data
        while len(row_data) <= max(idx.get("replacement_source", 0), idx.get("replaced_by_master_id", 0), idx.get("replaced_at", 0)):
            row_data.append(None)

        row_data[email_idx] = new_email
        row_data[idx["original_email"]] = old_email
        row_data[idx["replacement_source"]] = "replacement_heuristic"
        row_data[idx["replaced_by_master_id"]] = replaced_by_id
        row_data[idx["replaced_at"]] = replaced_at_iso

        # Update in-memory row
        by_master_id[invalid_id] = (sheet_row, row_data)
        print(f"  ✓ {invalid_id}: {old_email} → {new_email}")
        replaced_count += 1

    print(f"\n→ Writing updated rows to Master")
    # Update header with new audit columns
    for col_idx, header_val in enumerate(new_header, start=1):
        ws.cell(row=1, column=col_idx).value = header_val

    # Delete all data rows at once (keep header)
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row)

    for row_data in rows_for_sheet:
        # Find the updated version if it was modified
        mid = row_data[idx.get("master_id", 0)] if "master_id" in idx else None
        if mid and str(mid) in by_master_id:
            _, updated_row = by_master_id[str(mid)]
            ws.append(updated_row)
        else:
            ws.append(row_data)

    print(f"\n→ Saving Master")
    wb.save(MASTER_PATH)
    wb.close()

    print(f"\n✓ Replacements applied: {replaced_count}")
    if already_applied:
        print(f"  · already applied (skipped): {already_applied}")
    if errors:
        print(f"  · errors/skipped: {errors}")

    print(f"\nNext steps:")
    print(f"  1. Verify Master has audit columns: original_email, replacement_source, replaced_by_master_id, replaced_at")
    print(f"  2. Regenerate: relationship-health-report.md (engagement bands, strategic contacts)")
    print(f"  3. Regenerate: campaign-ready contact lists")
    print(f"  4. Re-run extract_gmail_signals.py (blocklist will skip archived emails)")

    # Append audit log
    with ARCHIVE_LOG.open("a", encoding="utf-8") as f:
        for sugg in approved:
            if sugg.get("invalid_master_id") in by_master_id:
                f.write(json.dumps({
                    "event": "replacement_applied",
                    "at": replaced_at_iso,
                    **sugg,
                }, ensure_ascii=False, default=str) + "\n")

    return 0


def main() -> int:
    # Command-line dispatch
    if len(sys.argv) > 1 and sys.argv[1] == "--apply":
        if len(sys.argv) < 3:
            print("ERROR: --apply requires CSV path argument", file=sys.stderr)
            return 1
        return apply_replacements(sys.argv[2])

    # Default: generate fresh suggestions
    if not MASTER_PATH.exists():
        print(f"ERROR: Master not found at {MASTER_PATH}", file=sys.stderr)
        return 1

    print(f"→ Loading Master · {MASTER_PATH}")
    wb = load_workbook(MASTER_PATH)
    ws = wb["Master"]

    header = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(header) if h}
    required = ["email_validity", "email", "full_name", "company", "master_id",
                "contact_category", "relationship_band", "last_email_date",
                "bounce_count", "last_bounce_date", "active_threads", "bucket"]
    missing = [c for c in required if c not in idx]
    if missing:
        print(f"ERROR: missing required columns · {missing}", file=sys.stderr)
        return 2

    # Pull every row into memory · 4547 rows × 63 cols is trivial.
    rows: list[list[Any]] = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        rows.append(list(r))

    invalids: list[dict] = []
    by_domain: dict[str, list[dict]] = defaultdict(list)

    for row_idx, r in enumerate(rows, start=2):
        email = (r[idx["email"]] or "").strip().lower() if isinstance(r[idx["email"]], str) else ""
        validity = (r[idx["email_validity"]] or "").strip().lower() if isinstance(r[idx["email_validity"]], str) else ""
        dom = domain_of(email)
        record = {
            "sheet_row": row_idx,
            "row_data": r,
            "master_id": r[idx["master_id"]],
            "email": email,
            "domain": dom,
            "local": local_of(email),
            "full_name": r[idx["full_name"]] or "",
            "surname": surname_from_full_name(r[idx["full_name"]]),
            "company": r[idx["company"]] or "",
            "contact_category": r[idx["contact_category"]] or "",
            "relationship_band": r[idx["relationship_band"]] or "",
            "last_email_date": r[idx["last_email_date"]],
            "active_threads": r[idx["active_threads"]],
            "bounce_count": r[idx["bounce_count"]],
            "last_bounce_date": r[idx["last_bounce_date"]],
            "bucket": r[idx["bucket"]] or "",
            "validity": validity,
        }
        if validity == "invalid":
            invalids.append(record)
        elif dom:
            by_domain[dom].append(record)

    print(f"  · invalid rows: {len(invalids)}")
    print(f"  · candidate domains (live): {len(by_domain)}")

    # Score each invalid against every same-domain live contact
    suggestions: list[dict] = []          # one row per (invalid, candidate) above MIN_SCORE
    invalids_with_candidate: set[str] = set()
    invalids_no_candidate: list[dict] = []

    for inv in invalids:
        if not inv["domain"] or inv["domain"] in PERSONAL_DOMAINS:
            invalids_no_candidate.append(inv)
            continue
        pool = by_domain.get(inv["domain"], [])
        scored: list[tuple[int, list[str], dict]] = []
        for cand in pool:
            if cand["email"] == inv["email"]:
                continue
            if (cand.get("validity") or "").lower() == "invalid":
                continue
            score, reasons = score_candidate(inv, cand)
            if score >= MIN_SCORE:
                scored.append((score, reasons, cand))
        if not scored:
            invalids_no_candidate.append(inv)
            continue
        scored.sort(key=lambda x: -x[0])
        invalids_with_candidate.add(str(inv["master_id"]))
        # Keep top 3 candidates per invalid
        for rank, (score, reasons, cand) in enumerate(scored[:3], start=1):
            suggestions.append({
                "invalid_master_id": inv["master_id"],
                "invalid_email": inv["email"],
                "invalid_full_name": inv["full_name"],
                "invalid_company": inv["company"],
                "invalid_category": inv["contact_category"],
                "invalid_bounce_count": inv["bounce_count"],
                "invalid_last_bounce_date": inv["last_bounce_date"],
                "rank": rank,
                "score": score,
                "candidate_master_id": cand["master_id"],
                "candidate_email": cand["email"],
                "candidate_full_name": cand["full_name"],
                "candidate_band": cand["relationship_band"],
                "candidate_category": cand["contact_category"],
                "candidate_last_email_date": cand["last_email_date"],
                "candidate_active_threads": cand["active_threads"],
                "reasoning": " · ".join(reasons),
                "apply_decision": "",   # operator fills "apply" / "skip"
            })

    print(f"  · invalids WITH candidate: {len(invalids_with_candidate)}")
    print(f"  · invalids WITHOUT candidate: {len(invalids_no_candidate)}")
    print(f"  · total suggestion rows: {len(suggestions)}")

    batch_id = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")

    # ── 1. Emit suggestions CSV ─────────────────────────────────────────────
    sug_csv = REPORTS / f"replacement-suggestions_{batch_id}.csv"
    with sug_csv.open("w", encoding="utf-8", newline="") as f:
        if suggestions:
            w = csv.DictWriter(f, fieldnames=list(suggestions[0].keys()))
            w.writeheader()
            for s in suggestions:
                w.writerow(s)
        else:
            f.write("# no suggestions generated\n")
    print(f"\n✓ replacement-suggestions: {sug_csv.name} · {len(suggestions)} row(s)")

    # ── 2. Move sin-candidato → INVALID_ARCHIVE sheet ───────────────────────
    if "INVALID_ARCHIVE" not in wb.sheetnames:
        archive_ws = wb.create_sheet("INVALID_ARCHIVE")
        archive_header = list(header) + [
            "archived_at", "archived_batch_id", "archived_reason",
            "archived_source_sheet_row",
        ]
        archive_ws.append(archive_header)
        for cell in archive_ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="FFE4E1")
        print(f"  · created INVALID_ARCHIVE sheet ({len(archive_header)} cols)")
    else:
        archive_ws = wb["INVALID_ARCHIVE"]
        print(f"  · existing INVALID_ARCHIVE sheet · {archive_ws.max_row - 1} prior row(s)")

    # Sort candidate-less invalids by sheet_row DESC so deletion indices stay valid
    invalids_no_candidate.sort(key=lambda r: -r["sheet_row"])
    archived_count = 0
    archive_manifest: list[dict] = []

    archived_at_iso = datetime.now(timezone.utc).replace(microsecond=0).isoformat()
    for inv in invalids_no_candidate:
        archive_row = list(inv["row_data"]) + [
            archived_at_iso,
            batch_id,
            "invalid_email · no_replacement_candidate_in_domain",
            inv["sheet_row"],
        ]
        archive_ws.append(archive_row)
        ws.delete_rows(inv["sheet_row"], 1)
        archived_count += 1
        archive_manifest.append({
            "master_id": inv["master_id"],
            "email": inv["email"],
            "full_name": inv["full_name"],
            "company": inv["company"],
            "contact_category": inv["contact_category"],
            "bounce_count": inv["bounce_count"],
            "last_bounce_date": inv["last_bounce_date"],
            "source_sheet_row": inv["sheet_row"],
            "archived_at": archived_at_iso,
            "archived_batch_id": batch_id,
            "reason": "no_replacement_candidate_in_domain",
        })

    print(f"  · moved {archived_count} invalid row(s) → INVALID_ARCHIVE")

    # ── 3. Manifest CSV (per-batch) ─────────────────────────────────────────
    man_csv = REPORTS / f"invalid-archive-manifest_{batch_id}.csv"
    with man_csv.open("w", encoding="utf-8", newline="") as f:
        if archive_manifest:
            w = csv.DictWriter(f, fieldnames=list(archive_manifest[0].keys()))
            w.writeheader()
            for m in archive_manifest:
                w.writerow(m)
        else:
            f.write("# nothing archived in this batch\n")
    print(f"✓ invalid-archive-manifest: {man_csv.name} · {len(archive_manifest)} row(s)")

    # ── 4. Append-only JSONL audit log ──────────────────────────────────────
    with ARCHIVE_LOG.open("a", encoding="utf-8") as f:
        for m in archive_manifest:
            f.write(json.dumps(m, ensure_ascii=False, default=str) + "\n")
        for s in suggestions:
            f.write(json.dumps({
                "event": "suggestion_emitted",
                "batch_id": batch_id,
                "at": archived_at_iso,
                **s,
            }, ensure_ascii=False, default=str) + "\n")
    print(f"✓ audit log: {ARCHIVE_LOG.name} (append)")

    # ── 5. Save Master ──────────────────────────────────────────────────────
    print(f"\n→ Saving Master · this rewrites the xlsx, may take a few seconds")
    wb.save(MASTER_PATH)
    wb.close()
    print(f"✓ Master saved · active rows: {ws.max_row - 1} · archived: {archived_count}")
    print(f"\nNext steps:")
    print(f"  1. Open {sug_csv.name} · fill apply_decision = apply | skip per row")
    print(f"  2. (Future pass) re-run with --apply to perform approved swaps")
    print(f"  3. Re-run extract_gmail_signals · the bounce blocklist will skip archived emails")
    return 0


if __name__ == "__main__":
    sys.exit(main())
