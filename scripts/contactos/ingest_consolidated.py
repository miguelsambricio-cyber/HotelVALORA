#!/usr/bin/env python3
"""
CONTACTOS DATASITE · Consolidated ingestion lane (β · Phase 2.B).

Fourth lane of the institutional relationship graph. Digests the XLSX outputs
of Capa A (Desktop / Contactos Linkedin):

  - outlook_linkedin.xlsx (FASE 1 · structured Outlook + LinkedIn + VCF + Access)
  - msgs_sueltos.xlsx     (FASE 2 · 1.153 .msg files · Brokers/Personal/etc.)
  - psts.xlsx             (FASE 3 · 26 PSTs · 60 GB · pending generation)

All three share a unified "Capa A" schema (19-24 cols, superset of FASE 1).
This handler maps that schema to the canonical MASTER_SCHEMA, applies the
same identity-resolution engine used by ingest_google.py, and merges with
strict gap-fill discipline + paranoid mode for Datasite-seeded rows.

Operator workflow
-----------------
1. Drop a FASE 1/2/3 .xlsx into CONTACTOS DATASITE/incoming/consolidated/
2. Run:
       python scripts/contactos/pipeline.py        (dispatched automatically)
   or:
       python scripts/contactos/ingest_consolidated.py [--dry-run] [--limit N]

Pipeline behaviour
------------------
1. Parse each .xlsx → list of raw row dicts (tolerant of missing cols)
2. build_candidate_from_raw → normalised candidate per raw row
3. aggregate_candidates → collapse multi-row identities (FASE 2/3 emit
   one row per person×message×role) · hard caps on notes payload
4. Master dedup engine (reused from ingest_google.py) · 4 strategies:
     email_exact · phone_exact · linkedin_exact · name_company_exact · fuzzy
5. Merge into Master:
     - strict gap-fill (only write empty fields)
     - DATASITE_AUTHORITATIVE_FIELDS protected (imported from ingest_google.py)
     - paranoid mode: if Master row is Datasite-seeded, only touch
       notes_consolidated + last_seen_batch_id + last_updated_at
     - notes anexes prefixed with [YYYY-MM-DD lane]
6. UNMATCHED candidates → logged to consolidated-unmatched-contacts_<batch>.csv
   · NEVER auto-inserted (institutional discipline)
7. Five reports written under reports/ (see writers below)
8. Archive each processed .xlsx → old/consolidated/

Flags
-----
  --dry-run    Run all stages and write reports, but DO NOT mutate the Master
               nor archive the input. Allows operator to inspect what WOULD
               happen before committing.
  --limit N    After parsing, truncate to the first N raw candidates. Debug
               aid for subset runs (~100) before a full 33K+ run.

Dependencies: openpyxl + stdlib. Reuses helpers from ingest.py + ingest_google.py.
"""

from __future__ import annotations

import argparse
import csv
import difflib
import json
import re
import shutil
import sys
import unicodedata
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except (AttributeError, ValueError):
    pass

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl", file=sys.stderr)
    sys.exit(2)

sys.path.insert(0, str(Path(__file__).resolve().parent))
try:
    from ingest import (  # type: ignore[import-not-found]
        MASTER_SCHEMA,
        load_existing_master,
        write_master_workbook,
        build_summary,
        consolidate_notes,
        canonical_investor_type,
        detect_hotel_focus,
        norm_text,
        norm_email,
        norm_phone,
        norm_linkedin,
        company_key,
        now_iso,
        to_iso_date,
    )
    from ingest_google import (  # type: ignore[import-not-found]
        build_master_indices,
        resolve_identity,
        DATASITE_AUTHORITATIVE_FIELDS,
        PERSONAL_EMAIL_DOMAINS,
        email_domain_kind,
    )
except ImportError as err:
    print(f"ERROR: cannot import from siblings: {err}", file=sys.stderr)
    sys.exit(2)


# ── paths ───────────────────────────────────────────────────────────────────

REPO = Path(__file__).resolve().parent.parent.parent
ROOT = REPO / "CONTACTOS DATASITE"
INCOMING = ROOT / "incoming" / "consolidated"
OLD = ROOT / "old" / "consolidated"
REPORTS = ROOT / "reports"
MASTER_FILE = ROOT / "master" / "metcub-contacts-master.xlsx"

for d in (INCOMING, OLD, REPORTS):
    d.mkdir(parents=True, exist_ok=True)


# ── source taxonomy · Categoria field → source_capa_a slug ─────────────────

CATEGORIA_TO_LANE: dict[str, str] = {
    "Outlook Contactos": "outlook_libreta",
    "Outlook Histórico": "outlook_libreta",
    "Outlook Histórico v2": "outlook_libreta",
    "LinkedIn": "linkedin_csv",
    "LinkedIn (Outlook fmt)": "linkedin_outlook_fmt",
    "LinkedIn Imported": "linkedin_imported",
    "VCF Movil": "vcf",
    "Access Contactos": "access",
    "Access Histórico": "access",
    # FASE 2 (.msg) folder buckets
    "Brokers": "outlook_msg",
    "Correo Personal": "outlook_msg",
    "Desarrolladores": "outlook_msg",
    "Financiadores": "outlook_msg",
}


def categoria_to_lane(categoria: str) -> str:
    """Map a FASE 1/2/3 Categoria value to a source_capa_a slug."""
    if not categoria:
        return "unknown"
    cat = categoria.strip()
    # FASE 3 uses "PST: <folder>" prefix
    if cat.lower().startswith("pst:") or cat.lower().startswith("pst "):
        return "outlook_pst"
    return CATEGORIA_TO_LANE.get(cat, "unknown")


# ── parse FASE 1/2/3 xlsx into a flat list of raw row dicts ────────────────


def parse_consolidated_xlsx(path: Path) -> list[dict[str, str]]:
    """Read a FASE 1/2/3 export · return list of dicts (values as strings).

    Tolerates missing columns · FASE 2 schema is a subset of FASE 1/3 schema.
    """
    wb = load_workbook(path, data_only=True, read_only=True)
    if not wb.sheetnames:
        return []
    ws = wb[wb.sheetnames[0]]
    if ws.max_row < 2:
        return []
    header = [norm_text(c.value) for c in ws[1]]
    rows: list[dict[str, str]] = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        if all(cell in (None, "") for cell in raw):
            continue
        rec: dict[str, str] = {}
        for i, h in enumerate(header):
            if i >= len(raw):
                break
            val = raw[i]
            if val is None:
                rec[h] = ""
            elif isinstance(val, datetime):
                rec[h] = val.date().isoformat()
            else:
                rec[h] = str(val).strip()
        rows.append(rec)
    return rows


# ── per-raw-row candidate (one fila bruta) ─────────────────────────────────


def build_candidate_from_raw(raw: dict[str, str], source_file: str) -> dict[str, Any] | None:
    """Map one FASE 1/2/3 raw row to a normalised candidate dict.

    Returns None when no usable identifier is present:
      (no email) AND (no linkedin) AND (no full_name+company)
    """
    nombre = norm_text(raw.get("Nombre"))
    apellidos = norm_text(raw.get("Apellidos"))
    full_name = " ".join(p for p in (nombre, apellidos) if p)

    email = norm_email(raw.get("Email"))
    email_2 = norm_email(raw.get("Email_2"))
    email_3 = norm_email(raw.get("Email_3"))

    # Phone precedence: Telefono → Movil → Otro_Tel
    phone_main = (norm_phone(raw.get("Telefono"))
                  or norm_phone(raw.get("Movil"))
                  or norm_phone(raw.get("Otro_Tel")))
    phone_alts: list[tuple[str, str]] = []
    for kind, src_col in (("telefono", "Telefono"), ("movil", "Movil"), ("otro", "Otro_Tel")):
        p = norm_phone(raw.get(src_col))
        if p and p != phone_main:
            phone_alts.append((kind, p))

    company = norm_text(raw.get("Empresa"))
    title = norm_text(raw.get("Cargo"))
    departamento = norm_text(raw.get("Departamento"))
    direccion = norm_text(raw.get("Direccion"))
    city = norm_text(raw.get("Ciudad"))
    country = norm_text(raw.get("Pais"))
    web = norm_text(raw.get("Web"))
    notas = norm_text(raw.get("Notas"))
    fecha_conexion = norm_text(raw.get("Fecha_Conexion"))

    linkedin = ""
    web_extra = ""
    if web:
        if "linkedin" in web.lower():
            linkedin = norm_linkedin(web)
        else:
            web_extra = web

    rol = norm_text(raw.get("Rol"))
    asunto = norm_text(raw.get("Asunto"))
    fecha = to_iso_date(raw.get("Fecha"))
    carpeta_pst = norm_text(raw.get("Carpeta_PST"))
    pst_file = norm_text(raw.get("PST"))

    categoria = norm_text(raw.get("Categoria"))
    fuente_archivo = norm_text(raw.get("Fuente_Archivo"))
    lane = categoria_to_lane(categoria)

    if not (email or linkedin or (full_name and company)):
        return None

    return {
        "full_name": full_name,
        "primary_email": email,
        "secondary_emails_list": [e for e in (email_2, email_3) if e and e != email],
        "primary_phone": phone_main,
        "secondary_phones_list": phone_alts,
        "linkedin": linkedin,
        "company": company,
        "title": title,
        "departamento": departamento,
        "direccion": direccion,
        "city": city,
        "country": country,
        "notas": notas,
        "web_extra": web_extra,
        "fecha_conexion": fecha_conexion,
        "rol": rol,
        "asunto": asunto,
        "fecha": fecha,
        "carpeta_pst": carpeta_pst,
        "pst_file": pst_file,
        "categoria": categoria,
        "source_capa_a": lane,
        "fuente_archivo": fuente_archivo,
        "source_file_xlsx": source_file,
    }


# ── pre-merge internal aggregation ────────────────────────────────────────
# FASE 2/3 emits one row per person×message×role. The same identity can show
# up dozens of times. Aggregate to ONE candidate per identity BEFORE going
# to the Master dedup engine. Hard caps on notes payload prevent explosion.

MAX_SUBJECTS = 5
MAX_PST_FOLDERS = 3
MAX_ALT_EMAILS = 5
MAX_ALT_PHONES = 5
MAX_ROLES_SHOWN = 4
MAX_PST_FILES = 3


def candidate_identity_key(c: dict[str, Any]) -> tuple[str, str, str]:
    """Identity key for INTERNAL deduplication across raw rows.

    Mirrors the Master dedup engine priority:
      1. email      → ("email", email, "")
      2. linkedin   → ("linkedin", li, "")
      3. name+co    → ("name_company", full_name.lower(), company_key)
      4. fallback   → ("fallback", best-effort, "")
    """
    if c["primary_email"]:
        return ("email", c["primary_email"], "")
    if c["linkedin"]:
        return ("linkedin", c["linkedin"], "")
    if c["full_name"] and c["company"]:
        return ("name_company", c["full_name"].lower(), company_key(c["company"]))
    return ("fallback", (c["full_name"].lower() or c["primary_phone"] or ""), "")


def aggregate_candidates(raw_candidates: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Collapse multiple raw rows with the same identity into ONE candidate."""
    groups: dict[tuple[str, str, str], list[dict[str, Any]]] = defaultdict(list)
    for c in raw_candidates:
        groups[candidate_identity_key(c)].append(c)

    aggregated: list[dict[str, Any]] = []
    for key, members in groups.items():
        if not members:
            continue

        source_xlsx_set: set[str] = set()
        fuente_archivo_set: set[str] = set()
        categoria_set: set[str] = set()
        lane_counter: Counter[str] = Counter()
        roles_counter: Counter[str] = Counter()
        subjects_seen: list[tuple[str, str]] = []
        pst_folders_set: set[str] = set()
        pst_files_set: set[str] = set()
        alt_emails: set[str] = set()
        alt_phones: list[tuple[str, str]] = []
        max_fecha = ""

        best_full_name = ""
        best_company = ""
        best_title = ""
        best_email = ""
        best_phone = ""
        best_linkedin = ""
        best_city = ""
        best_country = ""
        best_notas_pieces: list[str] = []
        best_web_extra: set[str] = set()
        best_direccion_set: set[str] = set()
        best_departamento_set: set[str] = set()
        best_fecha_conexion = ""

        for m in members:
            source_xlsx_set.add(m["source_file_xlsx"])
            if m["fuente_archivo"]:
                fuente_archivo_set.add(m["fuente_archivo"])
            if m["categoria"]:
                categoria_set.add(m["categoria"])
            lane_counter[m["source_capa_a"]] += 1

            if m["rol"]:
                roles_counter[m["rol"]] += 1
            if m["asunto"]:
                subjects_seen.append((m["fecha"], m["asunto"]))
            if m["carpeta_pst"]:
                pst_folders_set.add(m["carpeta_pst"])
            if m["pst_file"]:
                pst_files_set.add(m["pst_file"])

            for e in [m["primary_email"]] + m["secondary_emails_list"]:
                if e:
                    alt_emails.add(e)

            existing_phone_vals = {p[1] for p in alt_phones}
            if m["primary_phone"] and m["primary_phone"] not in existing_phone_vals:
                alt_phones.append(("main", m["primary_phone"]))
                existing_phone_vals.add(m["primary_phone"])
            for kind, val in m["secondary_phones_list"]:
                if val and val not in existing_phone_vals:
                    alt_phones.append((kind, val))
                    existing_phone_vals.add(val)

            if m["fecha"] and m["fecha"] > max_fecha:
                max_fecha = m["fecha"]

            if m["full_name"] and len(m["full_name"]) > len(best_full_name):
                best_full_name = m["full_name"]
            if m["company"] and len(m["company"]) > len(best_company):
                best_company = m["company"]
            if m["title"] and len(m["title"]) > len(best_title):
                best_title = m["title"]
            if not best_email and m["primary_email"]:
                best_email = m["primary_email"]
            if not best_phone and m["primary_phone"]:
                best_phone = m["primary_phone"]
            if not best_linkedin and m["linkedin"]:
                best_linkedin = m["linkedin"]
            if not best_city and m["city"]:
                best_city = m["city"]
            if not best_country and m["country"]:
                best_country = m["country"]
            if m["notas"]:
                best_notas_pieces.append(m["notas"])
            if m["web_extra"]:
                best_web_extra.add(m["web_extra"])
            if m["direccion"]:
                best_direccion_set.add(m["direccion"])
            if m["departamento"]:
                best_departamento_set.add(m["departamento"])
            if not best_fecha_conexion and m["fecha_conexion"]:
                best_fecha_conexion = m["fecha_conexion"]

        primary_email = best_email
        alt_emails_clean = [e for e in alt_emails if e and e != primary_email][:MAX_ALT_EMAILS]

        primary_phone = best_phone
        alt_phones_clean = [p[1] for p in alt_phones if p[1] and p[1] != primary_phone][:MAX_ALT_PHONES]

        subjects_sorted = sorted(subjects_seen, key=lambda t: t[0] or "", reverse=True)
        seen_subj: set[str] = set()
        top_subjects: list[str] = []
        for _f, s in subjects_sorted:
            k = s.lower().strip()
            if k and k not in seen_subj:
                seen_subj.add(k)
                top_subjects.append(s)
            if len(top_subjects) >= MAX_SUBJECTS:
                break

        top_pst_folders = list(pst_folders_set)[:MAX_PST_FOLDERS]
        top_pst_files = list(pst_files_set)[:MAX_PST_FILES]
        roles_display = "; ".join(f"{r}({roles_counter[r]})"
                                   for r, _ in roles_counter.most_common(MAX_ROLES_SHOWN))

        dominant_lane = lane_counter.most_common(1)[0][0] if lane_counter else "unknown"
        all_lanes = "+".join(sorted(lane_counter.keys()))
        source_capa_a = dominant_lane if len(lane_counter) == 1 else all_lanes

        aggregated.append({
            "identity_key": key,
            "full_name": best_full_name,
            "primary_email": primary_email,
            "secondary_emails": "; ".join(alt_emails_clean),
            "secondary_emails_list": alt_emails_clean,
            "primary_phone": primary_phone,
            "secondary_phones": "; ".join(alt_phones_clean),
            "secondary_phones_list": alt_phones_clean,
            "linkedin": best_linkedin,
            "company": best_company,
            "title": best_title,
            "city": best_city,
            "country": best_country,
            "notas_combined": " | ".join(best_notas_pieces),
            "departamento_set": sorted(best_departamento_set),
            "direccion_set": sorted(best_direccion_set),
            "web_extra_set": sorted(best_web_extra),
            "fecha_conexion": best_fecha_conexion,
            "roles_display": roles_display,
            "top_subjects": top_subjects,
            "top_pst_folders": top_pst_folders,
            "pst_files_list": top_pst_files,
            "max_fecha": max_fecha,
            "source_capa_a": source_capa_a,
            "fuente_archivos": "; ".join(sorted(fuente_archivo_set)),
            "categorias": "; ".join(sorted(categoria_set)),
            "source_files_xlsx": sorted(source_xlsx_set),
            "raw_count": len(members),
        })
    return aggregated


# ── notes payload builder ──────────────────────────────────────────────────


def build_notes_for_master(candidate: dict[str, Any], ts: str) -> str:
    """Build the unconditional notes string · prefixed with [date lane].

    Always-anexed fields (enrichment with no Master column equivalent):
    notas · secondary_emails · secondary_phones · departamento · direccion ·
    web (non-LinkedIn) · fecha_conexion · PST-roles/subjects/folders/files.
    """
    parts: list[str] = []
    if candidate["notas_combined"]:
        parts.append(candidate["notas_combined"])
    pre = f"[{ts} {candidate['source_capa_a']}]"
    if candidate["secondary_emails_list"]:
        parts.append(f"{pre} alt-emails: {'; '.join(candidate['secondary_emails_list'])}")
    if candidate["secondary_phones_list"]:
        parts.append(f"{pre} alt-phones: {'; '.join(candidate['secondary_phones_list'])}")
    if candidate["departamento_set"]:
        parts.append(f"{pre} dept: {'; '.join(candidate['departamento_set'])}")
    if candidate["direccion_set"]:
        parts.append(f"{pre} addr: {'; '.join(candidate['direccion_set'])}")
    if candidate["web_extra_set"]:
        parts.append(f"{pre} web: {'; '.join(candidate['web_extra_set'])}")
    if candidate["fecha_conexion"]:
        parts.append(f"{pre} linkedin-connected: {candidate['fecha_conexion']}")
    if candidate["roles_display"]:
        parts.append(f"{pre} PST-roles: {candidate['roles_display']}")
    if candidate["top_subjects"]:
        parts.append(f"{pre} PST-subjects: [{'; '.join(candidate['top_subjects'])}]")
    if candidate["top_pst_folders"]:
        parts.append(f"{pre} PST-folders: [{'; '.join(candidate['top_pst_folders'])}]")
    if candidate["pst_files_list"]:
        parts.append(f"{pre} PST-files: [{'; '.join(candidate['pst_files_list'])}]")
    return consolidate_notes(*parts)


def build_paranoid_gap_notes(candidate: dict[str, Any], existing: dict[str, Any], ts: str) -> str:
    """In paranoid mode, anex fields that have a Master column equivalent
    BUT only when that Master column is empty · so no data loss occurs even
    when we refuse to gap-fill the Master column itself.

    Mirrors the addr/dept pattern from build_notes_for_master: same prefix,
    same separator, idempotent via consolidate_notes downstream.
    """
    parts: list[str] = []
    pre = f"[{ts} {candidate['source_capa_a']}]"
    if candidate["title"] and not str(existing.get("title", "") or "").strip():
        parts.append(f"{pre} title: {candidate['title']}")
    if candidate["primary_phone"] and not str(existing.get("phone", "") or "").strip():
        parts.append(f"{pre} phone: {candidate['primary_phone']}")
    if candidate["city"] and not str(existing.get("city", "") or "").strip():
        parts.append(f"{pre} city: {candidate['city']}")
    if candidate["country"] and not str(existing.get("country", "") or "").strip():
        parts.append(f"{pre} country: {candidate['country']}")
    if candidate["linkedin"] and not str(existing.get("linkedin", "") or "").strip():
        parts.append(f"{pre} linkedin: {candidate['linkedin']}")
    return consolidate_notes(*parts) if parts else ""


# ── merge into Master · strict gap-fill + paranoid mode ────────────────────

CONSOLIDATED_AUTHORITATIVE_FIELDS: set[str] = set()


def is_datasite_master_row(row: dict[str, Any]) -> bool:
    """True if this Master row was originally seeded by the Datasite ingester."""
    if row.get("datasite_contact_number"):
        return True
    src = str(row.get("source_file", "") or "").lower()
    return "datasite-" in src or "metcub" in src


def merge_source_capa_a(existing_val: Any, candidate_val: str) -> str:
    """Idempotent merge of source_capa_a · union deduplicated · alphabetic-sorted.

    NOTE · source_capa_a is PROVENANCE METADATA, not institutional identity.
    It captures only the Capa A lanes that touched this row across all β runs.
    Datasite origin remains identifiable via source_file (which starts with
    'datasite-' for Datasite-seeded rows). This column is safe to merge in
    paranoid mode because it does NOT alter Datasite-canonical data.
    """
    existing_lanes = {x.strip() for x in str(existing_val or "").split("+") if x.strip()}
    new_lanes = {x.strip() for x in str(candidate_val or "").split("+") if x.strip()}
    union = existing_lanes | new_lanes
    return "+".join(sorted(union))


def merge_consolidated_into_master_row(existing: dict[str, Any],
                                       candidate: dict[str, Any],
                                       batch_id: str) -> list[str]:
    """Strict gap-fill + paranoid mode if row is Datasite-seeded.

    Paranoid mode: only touch notes_consolidated + last_seen_batch_id + last_updated_at.
    Normal mode: gap-fill non-authoritative fields when Master slot is empty.
    """
    ts = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    paranoid = is_datasite_master_row(existing)
    changed: list[str] = []
    new_notes = build_notes_for_master(candidate, ts)

    # source_capa_a is PROVENANCE METADATA · merged in BOTH modes (paranoid
    # and normal) because it tracks "which Capa A lanes touched this row"
    # · NOT Datasite-curated data · safe to write under paranoid discipline.
    merged_lane = merge_source_capa_a(existing.get("source_capa_a", ""), candidate["source_capa_a"])
    if merged_lane != str(existing.get("source_capa_a", "") or ""):
        existing["source_capa_a"] = merged_lane
        changed.append("source_capa_a")

    if paranoid:
        # Paranoid mode · cero Master canonical column writes (only notes +
        # provenance metadata). All descriptive data is preserved in
        # notes_consolidated: unconditional anexes from build_notes_for_master
        # + gap-anexes for fields with a Master column equivalent (only when
        # that column is empty).
        gap_notes = build_paranoid_gap_notes(candidate, existing, ts)
        combined = consolidate_notes(new_notes, gap_notes) if (new_notes or gap_notes) else ""
        # Idempotency · skip notes write if the block is already present
        # in notes_consolidated (re-runs on the same day produce identical
        # prefixes and consolidate_notes dedupes only on whole-part match).
        existing_notes = str(existing.get("notes_consolidated", "") or "")
        if combined and combined not in existing_notes:
            merged_notes = consolidate_notes(existing_notes, combined)
            if merged_notes != existing_notes:
                existing["notes_consolidated"] = merged_notes
                changed.append("notes_consolidated")
        existing["last_seen_batch_id"] = batch_id
        existing["last_updated_at"] = now_iso()
        return changed

    # Normal mode: prepare gap-fill candidates
    cand_geo = ", ".join(p for p in (candidate["city"], candidate["country"]) if p)
    cand_inv = canonical_investor_type(candidate["company"], "",
                                        candidate["notas_combined"], candidate["title"])
    if cand_inv in ("Unknown", ""):
        cand_inv = ""
    cand_hf = detect_hotel_focus("", candidate["notas_combined"], "", "")
    if cand_hf == "Unknown":
        cand_hf = ""

    field_map: dict[str, Any] = {
        "full_name": candidate["full_name"],
        "email": candidate["primary_email"],
        "phone": candidate["primary_phone"],
        "linkedin": candidate["linkedin"],
        "title": candidate["title"],
        "company": candidate["company"],
        "city": candidate["city"],
        "country": candidate["country"],
        "geography": cand_geo,
        "hotel_focus": cand_hf,
        "investor_type": cand_inv,
    }
    # source_capa_a is handled separately by merge_source_capa_a above ·
    # NOT a strict-gap-fill candidate (provenance · always merged)
    # last_email_date update IF fresher than existing
    cand_fecha = candidate.get("max_fecha", "") or ""
    exist_led = str(existing.get("last_email_date", "") or "")
    if cand_fecha and (not exist_led or cand_fecha > exist_led):
        field_map["last_email_date"] = cand_fecha

    # Notes write · idempotent (skip if block already present)
    existing_notes = str(existing.get("notes_consolidated", "") or "")
    if new_notes and new_notes not in existing_notes:
        merged_notes = consolidate_notes(existing_notes, new_notes)
        if merged_notes != existing_notes:
            existing["notes_consolidated"] = merged_notes
            changed.append("notes_consolidated")

    for field, new_val in field_map.items():
        if field in DATASITE_AUTHORITATIVE_FIELDS:
            continue
        if field in CONSOLIDATED_AUTHORITATIVE_FIELDS:
            continue
        if not new_val:
            continue
        old_val = existing.get(field, "")
        if not old_val or old_val == 0:
            existing[field] = new_val
            changed.append(field)

    existing["last_seen_batch_id"] = batch_id
    existing["last_updated_at"] = now_iso()

    # Append source_file trail
    existing_src = str(existing.get("source_file", "") or "")
    for fname in candidate["source_files_xlsx"]:
        token = f"consolidated:{fname}"
        if token and token not in existing_src:
            existing_src = f"{existing_src}; {token}".strip("; ")
    existing["source_file"] = existing_src

    return changed


# ── report writers ──────────────────────────────────────────────────────────


def write_csv_report(path: Path, rows: list[dict[str, Any]], fieldnames: list[str]) -> None:
    with path.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        w.writeheader()
        for r in rows:
            w.writerow({k: ("" if r.get(k) is None else r.get(k)) for k in fieldnames})


def write_jsonl_line(path: Path, record: dict[str, Any]) -> None:
    with path.open("a", encoding="utf-8") as f:
        f.write(json.dumps(record, ensure_ascii=False, default=str) + "\n")


# ── suspect false-negatives · 3 relaxed lenses on unmatched ────────────────
# Backlog #29 · operator review caso-por-caso. Engine intact · these are
# CSV-only hints, not auto-merges.

DIMINUTIVES_MAP: dict[str, str] = {
    "bob": "robert", "bobby": "robert",
    "bill": "william", "billy": "william", "will": "william",
    "jim": "james", "jimmy": "james",
    "mike": "michael", "mick": "michael",
    "tony": "anthony",
    "steve": "steven",
    "joe": "joseph", "joey": "joseph",
    "tom": "thomas", "tommy": "thomas",
    "rick": "richard", "rich": "richard", "dick": "richard",
    "dave": "david",
    "chris": "christopher",
    "alex": "alexander",
    "matt": "matthew",
    "ben": "benjamin",
    "nick": "nicholas",
    "dan": "daniel", "danny": "daniel",
    "ed": "edward", "eddie": "edward",
    "ken": "kenneth",
    "larry": "lawrence",
    "pat": "patrick",
    "ron": "ronald",
    "sam": "samuel",
    # Spanish common diminutives
    "pepe": "jose", "paco": "francisco", "pancho": "francisco",
    "lola": "dolores", "loli": "dolores",
    "nacho": "ignacio",
    "manolo": "manuel", "manu": "manuel",
    "rafa": "rafael",
    "javi": "javier",
    "fran": "francisco",
    "lupe": "guadalupe",
    "concha": "concepcion",
    "merche": "mercedes",
}

PERSONAL_DOMAINS_EXTENDED = PERSONAL_EMAIL_DOMAINS

LENS_HARD_CAP = 5  # if a master_id is matched by >N unmatched, drop those hits


def _strip_accents(s: str) -> str:
    if not s:
        return s
    return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")


def _norm_name_basic(name: str) -> str:
    if not name:
        return ""
    s = _strip_accents(name.lower())
    s = re.sub(r"[^\w\s]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _norm_name_strict(name: str) -> str:
    base = _norm_name_basic(name)
    if not base:
        return ""
    tokens = [DIMINUTIVES_MAP.get(t, t) for t in base.split()]
    return " ".join(sorted(tokens))


def _email_domain(email: str) -> str:
    if not email or "@" not in email:
        return ""
    return email.split("@", 1)[-1].lower().strip()


def _company_tokens(c: str) -> set[str]:
    if not c:
        return set()
    base = _norm_name_basic(c)
    stop = {"the", "and", "of", "for", "y", "de", "la", "el", "los", "las",
            "ltd", "inc", "llc", "sa", "sl", "sarl", "gmbh", "kg", "ag",
            "nv", "bv", "plc", "limited", "llp", "corp", "corporation",
            "group", "grupo", "company", "hotels", "hotel"}
    return {t for t in base.split() if len(t) >= 3 and t not in stop}


def detect_suspect_false_negatives(unmatched_candidates: list[dict[str, Any]],
                                    master_rows: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], int, int]:
    """Apply 3 relaxed lenses on unmatched · return (hits_post_cap, hits_dropped_by_cap, master_ids_capped).

    LENS A · name fuzzy ≥0.92 IGNORING company
    LENS B · name fuzzy ≥0.90 + same email domain
    LENS C · normalized name (no accents + diminutives) + (geo match OR ≥2 company tokens)
    Hard cap: drop hits where the same Master row is matched by > LENS_HARD_CAP candidates.
    """
    # Master indices
    by_firstletter: dict[str, list[int]] = defaultdict(list)
    m_nb: list[str] = []
    m_ns: list[str] = []
    m_dom: list[str] = []
    m_co_tokens: list[set[str]] = []
    m_city: list[str] = []
    m_country: list[str] = []
    for i, m in enumerate(master_rows):
        nb = _norm_name_basic(str(m.get("full_name", "") or ""))
        ns = _norm_name_strict(str(m.get("full_name", "") or ""))
        m_nb.append(nb)
        m_ns.append(ns)
        if nb:
            by_firstletter[nb[0]].append(i)
        m_dom.append(_email_domain(str(m.get("email", "") or "")))
        m_co_tokens.append(_company_tokens(str(m.get("company", "") or "")))
        m_city.append(_norm_name_basic(str(m.get("city", "") or "")))
        m_country.append(_norm_name_basic(str(m.get("country", "") or "")))

    # Collect raw hits: per (candidate_idx, master_idx), which lenses fired
    raw_hits: dict[tuple[int, int], dict[str, Any]] = {}

    for ui, u in enumerate(unmatched_candidates):
        u_name = u.get("full_name", "") or ""
        u_email = u.get("primary_email", "") or ""
        u_company = u.get("company", "") or ""
        u_city = _norm_name_basic(u.get("city", "") or "")
        u_country = _norm_name_basic(u.get("country", "") or "")
        u_nb = _norm_name_basic(u_name)
        u_ns = _norm_name_strict(u_name)
        u_dom = _email_domain(u_email)
        u_co_tokens = _company_tokens(u_company)

        if not u_nb:
            continue
        candidate_idxs = by_firstletter.get(u_nb[0], [])

        for mi in candidate_idxs:
            mn_b = m_nb[mi]
            if not mn_b:
                continue
            lenses_hit: set[str] = set()
            # LENS A
            score_a = difflib.SequenceMatcher(None, u_nb, mn_b).ratio()
            if score_a >= 0.92:
                lenses_hit.add("A")
            # LENS B (same domain · skip pure-personal generic domains to reduce noise)
            if u_dom and m_dom[mi] == u_dom and u_dom not in PERSONAL_DOMAINS_EXTENDED:
                score_b = difflib.SequenceMatcher(None, u_nb, mn_b).ratio()
                if score_b >= 0.90:
                    lenses_hit.add("B")
            # LENS C (strict normalized + geo OR company overlap)
            if u_ns and m_ns[mi] == u_ns:
                geo_ok = (u_city and m_city[mi] == u_city) or (u_country and m_country[mi] == u_country)
                co_overlap = len(u_co_tokens & m_co_tokens[mi]) >= 2
                if geo_ok or co_overlap:
                    lenses_hit.add("C")
            if not lenses_hit:
                continue
            key = (ui, mi)
            if key in raw_hits:
                raw_hits[key]["lenses"].update(lenses_hit)
            else:
                raw_hits[key] = {"ui": ui, "mi": mi, "lenses": lenses_hit, "score_a": score_a}

    # Hard cap: drop hits where the same Master row is matched by > LENS_HARD_CAP candidates
    by_master: dict[int, list[tuple[int, int]]] = defaultdict(list)
    for (ui, mi), _hit in raw_hits.items():
        by_master[mi].append((ui, mi))

    capped_master_ids: set[int] = {mi for mi, hits in by_master.items() if len(hits) > LENS_HARD_CAP}
    dropped_by_cap = sum(len(by_master[mi]) for mi in capped_master_ids)

    hits_post_cap: list[dict[str, Any]] = []
    for (ui, mi), hit in raw_hits.items():
        if mi in capped_master_ids:
            continue
        u = unmatched_candidates[ui]
        m = master_rows[mi]
        lenses = "".join(sorted(hit["lenses"]))  # e.g. "A", "AB", "ABC"
        if "C" in lenses:
            confidence = "high"
        elif "B" in lenses:
            confidence = "medium"
        else:
            confidence = "low"
        hits_post_cap.append({
            "candidate_full_name": u.get("full_name", "") or "",
            "candidate_email": u.get("primary_email", "") or "",
            "candidate_company": u.get("company", "") or "",
            "candidate_source_capa_a": u.get("source_capa_a", "") or "",
            "matched_master_id": m.get("master_id", "") or "",
            "matched_master_name": m.get("full_name", "") or "",
            "matched_master_email": m.get("email", "") or "",
            "matched_master_company": m.get("company", "") or "",
            "lens_triggered": lenses,
            "confidence": confidence,
            "score_a": f"{hit['score_a']:.3f}" if "A" in lenses else "",
        })
    return hits_post_cap, dropped_by_cap, len(capped_master_ids)


# ── main ────────────────────────────────────────────────────────────────────


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true",
                    help="Run all stages and write reports, but do NOT mutate the Master nor archive inputs.")
    ap.add_argument("--limit", type=int, default=0,
                    help="After parsing, truncate to the first N raw candidates (debug aid).")
    args = ap.parse_args()

    files = sorted(
        [p for p in INCOMING.iterdir() if p.is_file() and p.suffix.lower() == ".xlsx"],
        key=lambda p: p.stat().st_mtime,
    )
    print(f"→ Consolidated lane · {len(files)} xlsx in incoming/consolidated/ "
          f"· dry_run={args.dry_run} · limit={args.limit or 'none'}")
    if not files:
        print("  nothing to do · drop a FASE 1/2/3 xlsx into incoming/consolidated/")
        return 0

    batch_id = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")

    raw_candidates: list[dict[str, Any]] = []
    files_seen: list[str] = []
    invalid_no_identifier = 0
    total_raw_rows = 0
    for src in files:
        print(f"\n→ parsing {src.name}")
        try:
            rows = parse_consolidated_xlsx(src)
        except Exception as e:
            print(f"  ✗ parse failed: {e}")
            continue
        print(f"  parsed {len(rows)} raw rows")
        total_raw_rows += len(rows)
        files_seen.append(src.name)
        for raw in rows:
            cand = build_candidate_from_raw(raw, src.name)
            if cand is None:
                invalid_no_identifier += 1
                continue
            raw_candidates.append(cand)
    print(f"\n  valid raw candidates: {len(raw_candidates)} "
          f"· invalid (no identifier): {invalid_no_identifier}")

    aggregated = aggregate_candidates(raw_candidates)
    print(f"  aggregated to {len(aggregated)} unique identities "
          f"(from {len(raw_candidates)} raw)")

    if args.limit and args.limit > 0:
        aggregated = aggregated[:args.limit]
        print(f"  --limit {args.limit} active · truncated to {len(aggregated)} unique identities")

    within_dups_rows: list[dict[str, Any]] = []
    for agg in aggregated:
        if agg["raw_count"] > 1:
            within_dups_rows.append({
                "identity_kind": agg["identity_key"][0],
                "identity_value": agg["identity_key"][1],
                "full_name": agg["full_name"],
                "primary_email": agg["primary_email"],
                "company": agg["company"],
                "raw_rows_collapsed": agg["raw_count"],
                "lanes": agg["source_capa_a"],
                "source_xlsx": "; ".join(agg["source_files_xlsx"]),
            })

    master_rows, contacts_rows, companies_rows, activities_rows = load_existing_master()
    print(f"  Master loaded: {len(master_rows)} contacts")
    indices = build_master_indices(master_rows)

    resolution_rows: list[dict[str, Any]] = []
    matched_for_apply: list[tuple[int, dict[str, Any], str, float]] = []
    unmatched_candidates: list[dict[str, Any]] = []
    by_strategy: Counter[str] = Counter()
    for agg in aggregated:
        idx, strategy, score = resolve_identity(agg, indices, master_rows)
        by_strategy[strategy] += 1
        master_row = master_rows[idx] if idx is not None else None
        resolution_rows.append({
            "identity_kind": agg["identity_key"][0],
            "full_name": agg["full_name"],
            "primary_email": agg["primary_email"],
            "company": agg["company"],
            "linkedin": agg["linkedin"],
            "source_capa_a": agg["source_capa_a"],
            "raw_count": agg["raw_count"],
            "match_status": "matched" if idx is not None else "unmatched",
            "match_strategy": strategy,
            "match_score": f"{score:.3f}" if idx is not None else "",
            "master_id": (master_row or {}).get("master_id", ""),
            "master_full_name": (master_row or {}).get("full_name", ""),
            "master_email": (master_row or {}).get("email", ""),
            "master_company": (master_row or {}).get("company", ""),
            "master_is_datasite_seeded": (str(is_datasite_master_row(master_row))
                                          if master_row is not None else ""),
        })
        if idx is not None:
            matched_for_apply.append((idx, agg, strategy, score))
        else:
            unmatched_candidates.append(agg)

    matched_count = sum(1 for r in resolution_rows if r["match_status"] == "matched")
    print(f"  resolution · matched={matched_count} · unmatched={len(unmatched_candidates)}")
    print(f"  strategy distribution: {dict(by_strategy)}")

    apply_log: list[dict[str, str]] = []
    merged_count = 0
    skipped_count = 0
    paranoid_count = 0
    for idx, agg, strategy, score in matched_for_apply:
        target_row = master_rows[idx]
        paranoid = is_datasite_master_row(target_row)
        if paranoid:
            paranoid_count += 1
        if args.dry_run:
            simulated = dict(target_row)
            changed = merge_consolidated_into_master_row(simulated, agg, batch_id)
        else:
            changed = merge_consolidated_into_master_row(target_row, agg, batch_id)
        if changed:
            merged_count += 1
            apply_log.append({
                "action": "MERGE" + (" (paranoid)" if paranoid else ""),
                "master_id": target_row.get("master_id", ""),
                "master_full_name": target_row.get("full_name", ""),
                "master_email": target_row.get("email", ""),
                "master_company": target_row.get("company", ""),
                "strategy": strategy,
                "score": f"{score:.3f}",
                "fields_changed": "|".join(changed),
                "source_capa_a": agg["source_capa_a"],
                "raw_count_aggregated": str(agg["raw_count"]),
            })
        else:
            skipped_count += 1

    print(f"  apply · merged={merged_count} · skipped={skipped_count} "
          f"· paranoid_rows={paranoid_count}")

    if not args.dry_run:
        summary = build_summary(master_rows, companies_rows, activities_rows, batch_id)
        write_master_workbook(master_rows, contacts_rows, companies_rows, activities_rows, summary)
        print(f"  Master rewritten · {MASTER_FILE}")

    # Backlog #29 · suspect false-negatives detection BEFORE writing the log
    # so we can include its counts in the ingestion-log
    print("\n→ scanning unmatched for suspect false-negatives (3 relaxed lenses)")
    suspect_hits, dropped_by_cap, capped_master_ids = detect_suspect_false_negatives(
        unmatched_candidates, master_rows,
    )
    by_lens_post = Counter(h["lens_triggered"] for h in suspect_hits)
    by_conf_post = Counter(h["confidence"] for h in suspect_hits)
    print(f"  suspect false-negatives · hits={len(suspect_hits)} "
          f"· dropped_by_cap={dropped_by_cap} · capped_master_ids={capped_master_ids}")
    print(f"  by lens combo: {dict(by_lens_post)}")
    print(f"  by confidence: {dict(by_conf_post)}")

    print("\n→ writing reports")
    write_jsonl_line(REPORTS / "consolidated-ingestion-log.jsonl", {
        "batch_id": batch_id,
        "dry_run": args.dry_run,
        "limit": args.limit,
        "source_files": files_seen,
        "raw_rows_parsed": total_raw_rows,
        "invalid_no_identifier": invalid_no_identifier,
        "valid_raw_candidates": len(raw_candidates),
        "aggregated_unique_identities": len(aggregated),
        "matched": matched_count,
        "unmatched": len(unmatched_candidates),
        "merged": merged_count,
        "skipped": skipped_count,
        "paranoid_rows": paranoid_count,
        "strategy_distribution": dict(by_strategy),
        "suspect_false_negatives_kept": len(suspect_hits),
        "suspect_false_negatives_dropped_by_cap": dropped_by_cap,
        "suspect_false_negatives_capped_master_ids": capped_master_ids,
        "suspect_false_negatives_by_lens": dict(by_lens_post),
        "suspect_false_negatives_by_confidence": dict(by_conf_post),
        "completed_at": now_iso(),
    })
    write_csv_report(REPORTS / f"consolidated-within-input-duplicates_{batch_id}.csv",
                     within_dups_rows,
                     ["identity_kind", "identity_value", "full_name", "primary_email",
                      "company", "raw_rows_collapsed", "lanes", "source_xlsx"])
    write_csv_report(REPORTS / f"consolidated-identity-resolution_{batch_id}.csv",
                     resolution_rows,
                     ["identity_kind", "full_name", "primary_email", "company", "linkedin",
                      "source_capa_a", "raw_count",
                      "match_status", "match_strategy", "match_score",
                      "master_id", "master_full_name", "master_email", "master_company",
                      "master_is_datasite_seeded"])
    write_csv_report(REPORTS / f"consolidated-applied-to-master_{batch_id}.csv",
                     apply_log,
                     ["action", "master_id", "master_full_name", "master_email",
                      "master_company", "strategy", "score", "fields_changed",
                      "source_capa_a", "raw_count_aggregated"])
    unmatched_rows = []
    for agg in unmatched_candidates:
        unmatched_rows.append({
            "full_name": agg["full_name"],
            "primary_email": agg["primary_email"],
            "primary_phone": agg["primary_phone"],
            "linkedin": agg["linkedin"],
            "company": agg["company"],
            "title": agg["title"],
            "city": agg["city"],
            "country": agg["country"],
            "source_capa_a": agg["source_capa_a"],
            "raw_count_aggregated": agg["raw_count"],
            "source_files_xlsx": "; ".join(agg["source_files_xlsx"]),
            "categorias": agg["categorias"],
            "email_domain_kind": email_domain_kind(agg["primary_email"]),
            "reason": "no Master row matches · candidate for manual onboarding",
        })
    write_csv_report(REPORTS / f"consolidated-unmatched-contacts_{batch_id}.csv",
                     unmatched_rows,
                     ["full_name", "primary_email", "primary_phone", "linkedin", "company",
                      "title", "city", "country", "source_capa_a", "raw_count_aggregated",
                      "source_files_xlsx", "categorias", "email_domain_kind", "reason"])
    write_csv_report(REPORTS / f"consolidated-suspect-false-negatives_{batch_id}.csv",
                     suspect_hits,
                     ["candidate_full_name", "candidate_email", "candidate_company",
                      "candidate_source_capa_a",
                      "matched_master_id", "matched_master_name", "matched_master_email",
                      "matched_master_company",
                      "lens_triggered", "confidence", "score_a"])
    print("  6 reports written under reports/")

    if not args.dry_run:
        for src in files:
            # Structured archive name · same pattern as pipeline.py wrapper:
            # consolidated-<slugified-stem>-<batch_id>.<ext>
            # Ensures consistent naming whether the handler is invoked
            # directly OR via pipeline.py orchestrator.
            stem_slug = re.sub(r"[^A-Za-z0-9]+", "_", Path(src.name).stem)
            stem_slug = re.sub(r"_+", "_", stem_slug).strip("_")[:60] or "file"
            archive_name = f"consolidated-{stem_slug}-{batch_id}{src.suffix.lower()}"
            dest = OLD / archive_name
            if dest.exists():
                # extremely defensive · batch_id should make this unreachable
                dest = OLD / f"{dest.stem}.dup{dest.suffix}"
            shutil.move(str(src), str(dest))
            print(f"  moved {src.name} → old/consolidated/{dest.name}")

    print(f"\n✓ consolidated lane done · batch_id={batch_id} · dry_run={args.dry_run}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
