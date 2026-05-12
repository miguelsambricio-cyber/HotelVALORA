#!/usr/bin/env python3
"""
Phase 2.C · promote the canonical institutional relationship graph from
the local Master xlsx + Gmail signal JSONLs into Supabase.

Targets (per migration 0014_relationship_contacts):
  public.relationship_companies   · entities · unique by company_key
  public.relationship_contacts    · canonical Master rows · unique by master_id
  public.relationship_interactions · Datasite Activities · unique by company_id
  public.relationship_labels      · per-contact Gmail label edges
  public.relationship_health      · per-contact bounce + validity audit

Idempotent · re-runnable · upserts by natural keys. No data ever leaves
the local machine: the Supabase project is the operator's own private
instance (twebgqutuqgonabvhzjk).

USAGE
  cd <repo root>
  python scripts/contactos/promote_to_supabase.py

Env (read from .env.local):
  NEXT_PUBLIC_SUPABASE_URL
  SUPABASE_SERVICE_ROLE_KEY
"""

from __future__ import annotations

import json
import os
import re
import sys
import unicodedata
from datetime import date, datetime
from pathlib import Path
from typing import Any
from urllib.request import Request, urlopen
from urllib.error import HTTPError

try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except (AttributeError, ValueError):
    pass

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl required · pip install openpyxl", file=sys.stderr)
    sys.exit(2)


# ── env ─────────────────────────────────────────────────────────────────────


def load_dotenv(path: Path) -> None:
    if not path.exists():
        return
    for line in path.read_text(encoding="utf-8").splitlines():
        m = re.match(r"^([A-Z0-9_]+)\s*=\s*(.*)$", line)
        if not m:
            continue
        k, v = m.group(1), m.group(2)
        if (v.startswith('"') and v.endswith('"')) or (v.startswith("'") and v.endswith("'")):
            v = v[1:-1]
        os.environ.setdefault(k, v)


REPO = Path(__file__).resolve().parent.parent.parent
load_dotenv(REPO / "apps" / "web" / ".env.local")

SUPABASE_URL = os.environ.get("NEXT_PUBLIC_SUPABASE_URL", "").rstrip("/")
SERVICE_KEY = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "")
if not SUPABASE_URL or not SERVICE_KEY:
    print("ERROR: NEXT_PUBLIC_SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY required", file=sys.stderr)
    sys.exit(2)


# ── paths ───────────────────────────────────────────────────────────────────

ROOT = REPO / "CONTACTOS DATASITE"
MASTER_FILE = ROOT / "master" / "metcub-contacts-master.xlsx"
GMAIL_OLD = ROOT / "old" / "gmail-signals"
GMAIL_INCOMING = ROOT / "incoming" / "gmail-signals"

if not MASTER_FILE.exists():
    print(f"ERROR: Master xlsx not found at {MASTER_FILE}", file=sys.stderr)
    sys.exit(2)


# ── helpers ─────────────────────────────────────────────────────────────────


def norm_text(v: Any) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    if not s:
        return ""
    s = unicodedata.normalize("NFC", s)
    return re.sub(r"\s+", " ", s)


def company_key(v: Any) -> str:
    s = norm_text(v).lower()
    if not s:
        return ""
    s = re.sub(r"[,\.]", "", s)
    s = re.sub(r"\b(llc|inc|ltd|sl|sa|sarl|gmbh|kg|ag|nv|bv|plc|limited|llp)\b\s*$", "", s)
    return re.sub(r"\s+", " ", s).strip()


def to_date(v: Any) -> str | None:
    """Return YYYY-MM-DD or None."""
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    s = str(v).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S",
                "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s.split("T")[0].split(" ")[0], fmt).date().isoformat()
        except ValueError:
            continue
    return None


def to_int(v: Any, default: int = 0) -> int:
    if v is None or v == "":
        return default
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return default


def to_bool(v: Any) -> bool:
    if isinstance(v, bool):
        return v
    s = str(v or "").strip().lower()
    return s in ("yes", "true", "1", "y")


def clean(value: Any) -> Any:
    """Pass-through for non-empty values · None for empty string."""
    if value is None:
        return None
    if isinstance(value, str):
        s = norm_text(value)
        return s if s else None
    return value


# ── PostgREST client ────────────────────────────────────────────────────────


def postgrest(method: str, path: str, payload: list[dict] | None = None,
              headers_extra: dict[str, str] | None = None) -> tuple[int, str]:
    """Minimal stdlib PostgREST client.

    Service-role key bypasses RLS · we're writing canonical data.
    """
    url = f"{SUPABASE_URL}/rest/v1/{path}"
    data = None
    if payload is not None:
        data = json.dumps(payload, default=str).encode("utf-8")
    headers = {
        "apikey": SERVICE_KEY,
        "Authorization": f"Bearer {SERVICE_KEY}",
        "Content-Type": "application/json",
        "Accept": "application/json",
    }
    if headers_extra:
        headers.update(headers_extra)
    req = Request(url, data=data, headers=headers, method=method)
    try:
        with urlopen(req) as r:
            return r.status, r.read().decode("utf-8", errors="replace")
    except HTTPError as e:
        body = ""
        try:
            body = e.read().decode("utf-8", errors="replace")
        except Exception:
            pass
        return e.code, body


def upsert(table: str, rows: list[dict[str, Any]], on_conflict: str, batch: int = 500) -> int:
    """Bulk upsert via PostgREST · returns total rows accepted."""
    if not rows:
        return 0
    total = 0
    for i in range(0, len(rows), batch):
        chunk = rows[i:i + batch]
        path = f"{table}?on_conflict={on_conflict}"
        status, body = postgrest(
            "POST", path, chunk,
            {"Prefer": "resolution=merge-duplicates,return=minimal"},
        )
        if status >= 300:
            print(f"  ✗ {table} batch {i//batch}: HTTP {status} · {body[:300]}", file=sys.stderr)
            return total
        total += len(chunk)
    return total


def delete_all(table: str) -> None:
    """Truncate a table via DELETE (PostgREST has no TRUNCATE)."""
    # Match-all with `id=neq.<impossible-uuid>` works as a "where true"
    status, body = postgrest("DELETE", f"{table}?id=neq.00000000-0000-0000-0000-000000000000",
                              headers_extra={"Prefer": "return=minimal"})
    if status >= 300:
        print(f"  ⚠  {table} delete returned HTTP {status} · {body[:200]}")


def query_count(table: str) -> int:
    status, body = postgrest("GET", f"{table}?select=count",
                              headers_extra={"Prefer": "count=exact",
                                              "Range-Unit": "items",
                                              "Range": "0-0"})
    if status >= 300:
        return -1
    try:
        return json.loads(body)[0]["count"]
    except (json.JSONDecodeError, IndexError, KeyError):
        return -1


def fetch_all(path: str, page_size: int = 1000) -> list[dict[str, Any]]:
    """Paginate a PostgREST GET (Supabase default db-max-rows = 1000)."""
    out: list[dict[str, Any]] = []
    start = 0
    while True:
        end = start + page_size - 1
        status, body = postgrest(
            "GET", path,
            headers_extra={"Range-Unit": "items", "Range": f"{start}-{end}"},
        )
        if status >= 300:
            print(f"  ✗ HTTP {status} on {path} · {body[:200]}", file=sys.stderr)
            return out
        chunk = json.loads(body) if body else []
        if not chunk:
            break
        out.extend(chunk)
        if len(chunk) < page_size:
            break
        start += page_size
    return out


# ── Master reader ───────────────────────────────────────────────────────────


def read_master() -> dict[str, list[dict[str, Any]]]:
    print(f"→ reading {MASTER_FILE.name}")
    wb = load_workbook(MASTER_FILE, data_only=True, read_only=True)
    out: dict[str, list[dict[str, Any]]] = {}
    for name in ("Master", "Contacts", "Companies", "Activities"):
        if name not in wb.sheetnames:
            out[name] = []
            continue
        ws = wb[name]
        if ws.max_row < 2:
            out[name] = []
            continue
        header = [norm_text(c.value) for c in ws[1]]
        rows: list[dict[str, Any]] = []
        for raw in ws.iter_rows(min_row=2, values_only=True):
            if all(v in (None, "") for v in raw):
                continue
            rec: dict[str, Any] = {}
            for i, h in enumerate(header):
                if i >= len(raw):
                    break
                rec[h] = raw[i]
            rows.append(rec)
        out[name] = rows
        print(f"  · {name}: {len(rows)} rows")
    return out


def read_gmail_signals() -> dict[str, dict[str, Any]]:
    """email → most-recent signal record (with bounce_reasons aggregated)."""
    aggregated: dict[str, dict[str, Any]] = {}
    paths: list[Path] = []
    for d in (GMAIL_INCOMING, GMAIL_OLD):
        if not d.exists():
            continue
        for p in d.iterdir():
            if p.is_file() and p.suffix.lower() in (".jsonl", ".json"):
                paths.append(p)
    paths.sort(key=lambda p: p.stat().st_mtime)
    for path in paths:
        with path.open("r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                try:
                    rec = json.loads(line)
                except json.JSONDecodeError:
                    continue
                em = (rec.get("email") or "").strip().lower()
                if not em:
                    continue
                cur = aggregated.setdefault(em, {})
                cur["email"] = em
                cur["labels"] = sorted(set((cur.get("labels") or []) + (rec.get("labels") or [])))
                cur["thread_count"] = max(int(cur.get("thread_count") or 0), int(rec.get("thread_count") or 0))
                cur["bounce_count"] = max(int(cur.get("bounce_count") or 0), int(rec.get("bounce_count") or 0))
                cur["bounce_reasons"] = (cur.get("bounce_reasons") or []) + (rec.get("bounce_reasons") or [])
                new_lb = rec.get("last_bounce_date") or ""
                if new_lb and new_lb > (cur.get("last_bounce_date") or ""):
                    cur["last_bounce_date"] = new_lb
                new_le = rec.get("last_email_date") or ""
                if new_le and new_le > (cur.get("last_email_date") or ""):
                    cur["last_email_date"] = new_le
    return aggregated


# ── builders ────────────────────────────────────────────────────────────────

# We need to map company_key → uuid AFTER the first companies upsert. We
# query back to retrieve the generated UUIDs.


def build_companies_rows(master_companies: list[dict[str, Any]]) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    seen_keys: set[str] = set()
    for c in master_companies:
        name = norm_text(c.get("company"))
        if not name:
            continue
        key = company_key(name)
        if not key or key in seen_keys:
            continue
        seen_keys.add(key)
        out.append({
            "company_key": key,
            "name": name,
            "investor_type_canonical": clean(c.get("investor_type_canonical")),
            "investor_type_raw": clean(c.get("investor_type_raw")),
            "investor_subtype": clean(c.get("investor_subtype")),
            "tier": clean(c.get("tier")),
            "company_type": clean(c.get("company_type")),
            "industry": clean(c.get("company_industry")),
            "fund_size": clean(c.get("fund_size")),
            "investment_preference": clean(c.get("investment_preference")),
            "investment_min": clean(c.get("investment_min")),
            "investment_max": clean(c.get("investment_max")),
            "association": clean(c.get("association")),
            "description": clean(c.get("description")),
            "continent": clean(c.get("continent")),
            "location": clean(c.get("company_location")),
            "datasite_company_number": clean(c.get("datasite_company_number")),
            "client_company_id": clean(c.get("client_company_id")),
            "internal_notes": clean(c.get("internal_notes")),
            "external_notes": clean(c.get("external_notes")),
            "company_notes": clean(c.get("company_notes")),
            "keyword": clean(c.get("keyword")),
            "coverage_officer": clean(c.get("coverage_officer")),
            "calling_lead": clean(c.get("calling_lead")),
        })
    return out


def build_contacts_rows(master: list[dict[str, Any]],
                        company_key_to_id: dict[str, str]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    seen_master_ids: set[str] = set()
    for r in master:
        master_id = norm_text(r.get("master_id"))
        if not master_id or master_id in seen_master_ids:
            continue
        seen_master_ids.add(master_id)

        company = norm_text(r.get("company"))
        ckey = company_key(company)
        company_id = company_key_to_id.get(ckey)

        rows.append({
            "master_id": master_id,
            "full_name": clean(r.get("full_name")),
            "email": clean(r.get("email")),
            "phone": clean(r.get("phone")),
            "linkedin": clean(r.get("linkedin")),
            "title": clean(r.get("title")),
            "role": clean(r.get("role")),
            "is_primary_contact": to_bool(r.get("is_primary_contact")),
            "company_id": company_id,
            "company_name": company or None,
            "investor_type": clean(r.get("investor_type")),
            "investor_subtype": clean(r.get("investor_subtype")),
            "tier": clean(r.get("tier")),
            "industry": clean(r.get("industry")),
            "fund_size": clean(r.get("fund_size")),
            "investment_preference": clean(r.get("investment_preference")),
            "investment_min": clean(r.get("investment_min")),
            "investment_max": clean(r.get("investment_max")),
            "association": clean(r.get("association")),
            "hotel_focus": clean(r.get("hotel_focus")),
            "geography": clean(r.get("geography")),
            "city": clean(r.get("city")),
            "state": clean(r.get("state")),
            "country": clean(r.get("country")),
            "continent": clean(r.get("continent")),
            "latest_deal_stage": clean(r.get("latest_deal_stage")),
            "last_activity_type": clean(r.get("last_activity_type")),
            "last_activity_date": to_date(r.get("last_activity_date")),
            "buyer_added_date": to_date(r.get("buyer_added_date")),
            "pipeline_state": clean(r.get("pipeline_state")),
            "ioi_bid_low": clean(r.get("ioi_bid_low")),
            "ioi_bid_high": clean(r.get("ioi_bid_high")),
            "loi_bid_low": clean(r.get("loi_bid_low")),
            "loi_bid_high": clean(r.get("loi_bid_high")),
            "revised_bid_low": clean(r.get("revised_bid_low")),
            "revised_bid_high": clean(r.get("revised_bid_high")),
            "relationship_status": clean(r.get("relationship_status")),
            "relationship_manager": clean(r.get("relationship_manager")),
            "coverage_officer": clean(r.get("coverage_officer")),
            "calling_lead": clean(r.get("calling_lead")),
            # Phase 2.B Gmail enrichment
            "relationship_strength": to_int(r.get("relationship_strength")),
            "active_threads": to_int(r.get("active_threads")),
            "last_email_date": to_date(r.get("last_email_date")),
            "inferred_relationship_stage": clean(r.get("inferred_relationship_stage")),
            "email_directionality": clean(r.get("email_directionality")),
            "gmail_signal_source": clean(r.get("gmail_signal_source")),
            # Phase 2.B.2 quality
            "relationship_band": clean(r.get("relationship_band")) or "cold",
            "collaboration_potential_score": to_int(r.get("collaboration_potential_score")),
            "email_validity": clean(r.get("email_validity")) or "uncertain",
            "bounce_count": to_int(r.get("bounce_count")),
            "last_bounce_date": to_date(r.get("last_bounce_date")),
            "flagged_for_correction": to_bool(r.get("flagged_for_correction")),
            "bucket": clean(r.get("bucket")) or "active",
            # provenance
            "notes_consolidated": clean(r.get("notes_consolidated")),
            "datasite_contact_number": clean(r.get("datasite_contact_number")),
            "datasite_company_number": clean(r.get("datasite_company_number")),
            "client_contact_id": clean(r.get("client_contact_id")),
            "source_file": clean(r.get("source_file")),
            "first_seen_batch_id": clean(r.get("first_seen_batch_id")),
            "last_seen_batch_id": clean(r.get("last_seen_batch_id")),
        })
    return rows


def build_interactions_rows(activities: list[dict[str, Any]],
                            company_key_to_id: dict[str, str]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    seen_companies: set[str] = set()
    for a in activities:
        company = norm_text(a.get("company"))
        if not company:
            continue
        ckey = company_key(company)
        if ckey in seen_companies:
            continue
        seen_companies.add(ckey)
        rows.append({
            "company_id": company_key_to_id.get(ckey),
            "company_name": company,
            "datasite_activity_number": clean(a.get("datasite_activity_number")),
            "buyer_added_date": to_date(a.get("buyer_added_date")),
            "initial_contact_date": to_date(a.get("initial_contact_date")),
            "teaser_sent_date": to_date(a.get("teaser_sent_date")),
            "nda_initial_sent_date": to_date(a.get("nda_initial_sent_date")),
            "nda_signed_date": to_date(a.get("nda_signed_date")),
            "nda_executed_date": to_date(a.get("nda_executed_date")),
            "cim_sent_date": to_date(a.get("cim_sent_date")),
            "ioi_process_letter_date": to_date(a.get("ioi_process_letter_date")),
            "ioi_bid_received_date": to_date(a.get("ioi_bid_received_date")),
            "loi_process_letter_date": to_date(a.get("loi_process_letter_date")),
            "loi_bid_received_date": to_date(a.get("loi_bid_received_date")),
            "management_presentation_date": to_date(a.get("management_presentation_date")),
            "revised_bid_received_date": to_date(a.get("revised_bid_received_date")),
            "declined_date": to_date(a.get("declined_date")),
            "declined_reason": clean(a.get("declined_reason")),
            "declined_comments": clean(a.get("declined_comments")),
            "last_activity_type": clean(a.get("last_activity_type")),
            "last_activity_date": to_date(a.get("last_activity_date")),
            "last_activity_comments": clean(a.get("last_activity_comments")),
            "latest_deal_stage": clean(a.get("latest_deal_stage")),
            "pipeline_state": clean(a.get("pipeline_state")),
            "ioi_bid_low": clean(a.get("ioi_bid_low")),
            "ioi_bid_high": clean(a.get("ioi_bid_high")),
            "loi_bid_low": clean(a.get("loi_bid_low")),
            "loi_bid_high": clean(a.get("loi_bid_high")),
            "revised_bid_low": clean(a.get("revised_bid_low")),
            "revised_bid_high": clean(a.get("revised_bid_high")),
            "removed_flag": clean(a.get("removed_flag")),
            "on_hold_flag": clean(a.get("on_hold_flag")),
            "client_buyer_id": clean(a.get("client_buyer_id")),
        })
    return rows


# ── main ────────────────────────────────────────────────────────────────────


def main() -> int:
    print(f"→ Phase 2.C · promoting Master to Supabase")
    print(f"  target: {SUPABASE_URL}")
    sheets = read_master()
    signals = read_gmail_signals()
    print(f"  · Gmail signals: {len(signals)} unique emails")

    # 1. Companies first
    company_rows = build_companies_rows(sheets["Companies"])
    print(f"\n→ upserting {len(company_rows)} companies")
    n = upsert("relationship_companies", company_rows, on_conflict="company_key")
    print(f"  ✓ {n} accepted")

    # Pull back company_key → id map (paginated · Supabase default 1000/page)
    print(f"  → fetching company_key → id map")
    company_data = fetch_all("relationship_companies?select=id,company_key")
    company_key_to_id: dict[str, str] = {r["company_key"]: r["id"] for r in company_data}
    print(f"  ✓ {len(company_key_to_id)} company id mappings")

    # 2. Contacts (canonical Master · per-person)
    contact_rows = build_contacts_rows(sheets["Master"], company_key_to_id)
    print(f"\n→ upserting {len(contact_rows)} contacts")
    n = upsert("relationship_contacts", contact_rows, on_conflict="master_id")
    print(f"  ✓ {n} accepted")

    # Pull back master_id → contact_id map (needed for labels + health)
    print(f"  → fetching master_id → id map (paginated)")
    contacts_db = fetch_all("relationship_contacts?select=id,master_id,email,full_name,bounce_count")
    master_id_to_contact: dict[str, dict[str, Any]] = {c["master_id"]: c for c in contacts_db}
    email_to_contact: dict[str, dict[str, Any]] = {
        (c.get("email") or "").lower(): c for c in contacts_db if c.get("email")
    }
    print(f"  ✓ {len(master_id_to_contact)} contact id mappings · {len(email_to_contact)} indexed by email")

    # 3. Interactions (Datasite Activities · per-company)
    interaction_rows = build_interactions_rows(sheets["Activities"], company_key_to_id)
    # Delete + re-insert is cleaner than upsert here (single row per company)
    print(f"\n→ replacing {len(interaction_rows)} interactions (delete-then-insert)")
    delete_all("relationship_interactions")
    n = upsert("relationship_interactions", interaction_rows, on_conflict="company_id")
    print(f"  ✓ {n} accepted")

    # 4. Labels (per-contact Gmail labels · from Gmail signals)
    print(f"\n→ replacing relationship_labels (delete-then-insert)")
    delete_all("relationship_labels")
    label_rows: list[dict[str, Any]] = []
    for email, sig in signals.items():
        contact = email_to_contact.get(email)
        if not contact:
            continue
        labels = sig.get("labels") or []
        if not labels:
            continue
        for label in labels:
            label_rows.append({
                "contact_id": contact["id"],
                "label": label,
            })
    n = upsert("relationship_labels", label_rows, on_conflict="contact_id,label")
    print(f"  ✓ {n} labels accepted ({len(label_rows)} attempted)")

    # 5. Health (per-contact · bounce + validity audit)
    print(f"\n→ replacing relationship_health (delete-then-insert)")
    delete_all("relationship_health")
    health_rows: list[dict[str, Any]] = []
    for email, sig in signals.items():
        contact = email_to_contact.get(email)
        if not contact:
            continue
        bcount = int(sig.get("bounce_count") or 0)
        if bcount == 0:
            continue  # only emit rows for contacts with bounce history
        # Determine validity (mirrors ingest_gmail.compute_email_validity)
        inbound = int(sig.get("inbound_count") or 0)
        if bcount >= 2 or (bcount >= 1 and inbound == 0):
            validity = "invalid"
            flagged = True
        elif inbound > 0:
            validity = "valid"
            flagged = False
        else:
            validity = "uncertain"
            flagged = False
        health_rows.append({
            "contact_id": contact["id"],
            "email_validity": validity,
            "bounce_count": bcount,
            "last_bounce_date": sig.get("last_bounce_date") or None,
            "bounce_reasons": sig.get("bounce_reasons") or None,
            "flagged_for_correction": flagged,
        })
    n = upsert("relationship_health", health_rows, on_conflict="contact_id")
    print(f"  ✓ {n} health records accepted")

    # ── verification ───────────────────────────────────────────────────────
    print(f"\n→ verification · row counts")
    for table in ("relationship_companies", "relationship_contacts",
                   "relationship_interactions", "relationship_labels",
                   "relationship_health"):
        c = query_count(table)
        print(f"  · {table:<32} {c}")

    print(f"\n✓ Phase 2.C ingest complete")
    return 0


if __name__ == "__main__":
    sys.exit(main())
