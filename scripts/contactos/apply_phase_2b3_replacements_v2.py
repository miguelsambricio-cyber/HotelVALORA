#!/usr/bin/env python3
"""
apply_phase_2b3_replacements_v2.py · Phase B-Repair · 2026-05-15

Re-applies the two approved Phase 2.B.3 replacements correctly, AFTER the
header/data offset corruption has been repaired by fix_master_alignment.py.

History:
  - Phase 2.B.3 --apply (2026-05-14 23:54) claimed to apply these
    replacements but the apply silently failed: row email column was
    untouched, original_email audit field stayed None, no audit JSONL
    written. The script ALSO corrupted the Master header/data alignment.
  - fix_master_alignment.py (2026-05-15) restored the canonical 64-col
    schema with zero data loss.
  - This script now applies the replacements properly against the
    repaired Master.

Approved replacements (from reports/replacement-suggestions_20260514T232135Z.csv):
  1. f193186dd9eb0c22  crocher@bancsabadell.com    → prietose@bancsabadell.com
  2. 596a76514db8d527  p.j.rodera@reyalurbis.com   → gestiondeactivos2@reyalurbis.com

Per-row mutations:
  email          ← new_email
  original_email ← old_email   (was None · now populated for audit)

Side log:
  reports/phase_2b3_apply_log.jsonl — one JSONL line per attempt with
  timestamp, master_id, old_email, new_email, replaced_by_master_id,
  reason = "re-application after Phase B-Repair offset corruption fix".

Idempotency:
  Refuses to overwrite if `original_email` is already populated on the
  target row (means the script ran before and the audit trail exists).

Scope:
  File-only. Supabase frozen by Phase B-Repair sentinel. Operator
  promotes manually after lifting the freeze.
"""
from __future__ import annotations

import json
import sys
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

sys.stdout.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parent.parent.parent
MASTER_PATH = ROOT / "CONTACTOS DATASITE" / "master" / "metcub-contacts-master.xlsx"
LOG_PATH = ROOT / "CONTACTOS DATASITE" / "reports" / "phase_2b3_apply_log.jsonl"
LOG_PATH.parent.mkdir(parents=True, exist_ok=True)

APPROVED_REPLACEMENTS = [
    {
        "master_id": "f193186dd9eb0c22",
        "old_email": "crocher@bancsabadell.com",
        "new_email": "prietose@bancsabadell.com",
        "replaced_by_master_id": "9bf709fb1abda6e5",
        "candidate_full_name": "Sergio Prieto",
        "score": 50,
        "reasoning": "band=strategic · same_category=Lender · threads=5",
    },
    {
        "master_id": "596a76514db8d527",
        "old_email": "p.j.rodera@reyalurbis.com",
        "new_email": "gestiondeactivos2@reyalurbis.com",
        "replaced_by_master_id": "a27778b5f04425ea",
        "candidate_full_name": "Reyal Urbis (gestion de activos)",
        "score": 60,
        "reasoning": "recent_activity_<24mo · band=strategic · threads=2",
    },
]


def main() -> int:
    if not MASTER_PATH.exists():
        print(f"ERROR: Master not found at {MASTER_PATH}")
        return 1

    print(f"→ Loading Master from {MASTER_PATH}")
    wb = load_workbook(MASTER_PATH)
    ws = wb["Master"]
    header = [c.value for c in ws[1]]
    print(f"  · header cols: {len(header)} · expected 64")
    if len(header) != 64:
        print(f"!! UNEXPECTED · refusing to run against non-canonical schema")
        return 2

    try:
        col_master_id = header.index("master_id") + 1
        col_email = header.index("email") + 1
        col_original_email = header.index("original_email") + 1
    except ValueError as e:
        print(f"!! UNEXPECTED · missing required header: {e}")
        return 2
    print(f"  · master_id at col {col_master_id} · email at col {col_email} · original_email at col {col_original_email}")

    # Locate target rows
    target_ids = {r["master_id"]: r for r in APPROVED_REPLACEMENTS}
    rows_by_mid: dict[str, int] = {}
    for row_idx in range(2, ws.max_row + 1):
        mid = ws.cell(row=row_idx, column=col_master_id).value
        if mid in target_ids:
            rows_by_mid[str(mid)] = row_idx

    print(f"\n→ Located {len(rows_by_mid)}/{len(target_ids)} target rows")
    missing = [mid for mid in target_ids if mid not in rows_by_mid]
    if missing:
        print(f"!! MISSING in Master: {missing}")
        print("   Aborting · refusing to partial-apply")
        return 3

    # Apply per row
    applied = []
    skipped_already = []
    ts_iso = datetime.now(timezone.utc).replace(microsecond=0).isoformat()
    print(f"\n→ Applying {len(APPROVED_REPLACEMENTS)} replacements (timestamp = {ts_iso})")
    for rep in APPROVED_REPLACEMENTS:
        mid = rep["master_id"]
        row_idx = rows_by_mid[mid]
        cur_email = ws.cell(row=row_idx, column=col_email).value
        cur_orig = ws.cell(row=row_idx, column=col_original_email).value

        if cur_orig:
            print(f"  · SKIP {mid} row {row_idx}: original_email already populated ({cur_orig!r}) · idempotent skip")
            skipped_already.append({**rep, "row": row_idx, "current_email": cur_email, "current_original_email": cur_orig})
            continue

        if (cur_email or "").strip().lower() != rep["old_email"].strip().lower():
            print(f"  !! REFUSE {mid} row {row_idx}: current email {cur_email!r} != expected old_email {rep['old_email']!r}")
            print(f"     This row may have been modified outside this pipeline · manual review required")
            return 4

        # Write new state
        ws.cell(row=row_idx, column=col_email).value = rep["new_email"]
        ws.cell(row=row_idx, column=col_original_email).value = rep["old_email"]
        print(f"  ✓ {mid} row {row_idx}: {rep['old_email']} → {rep['new_email']} (original_email preserved)")
        applied.append({
            "row": row_idx,
            "event": "replacement_applied",
            "at": ts_iso,
            "reason": "re-application after Phase B-Repair offset corruption fix · Phase 2.B.3-correction",
            **rep,
        })

    if applied:
        print(f"\n→ Saving Master")
        wb.save(MASTER_PATH)
        wb.close()

        print(f"→ Appending audit log to {LOG_PATH}")
        with LOG_PATH.open("a", encoding="utf-8") as f:
            for entry in applied:
                f.write(json.dumps(entry, ensure_ascii=False, default=str) + "\n")
    else:
        wb.close()

    print(f"\n→ Summary")
    print(f"  · applied        : {len(applied)}")
    print(f"  · skipped (idem) : {len(skipped_already)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
