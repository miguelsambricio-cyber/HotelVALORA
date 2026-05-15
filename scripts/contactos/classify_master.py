#!/usr/bin/env python3
"""
Classify all Master contacts by contact_category (v1) or
contact_category_v2 (Phase B · 2026-05-15 · new canonical taxonomy).

Schemes:
  v1 (legacy compat · default): Principal · Broker · Lender · Developer ·
       Proveedor · IA aplicaciones · Uncategorized · writes to existing
       contact_category column (header lookup, no hardcoded position).
  v2 (Phase B canonical): Principal · Broker · Lender · Operator ·
       Developer · Hotel Supply · IA Supply · Uncategorized · writes to
       contact_category_v2 column (creates if missing).

Phase B additions (only when --scheme=v2):
  - contact_category_v2          (the new canonical bucket)
  - original_category_raw        (NULL · only populated if a future ingest
                                  step provides a real source-of-record
                                  category snapshot · NEVER inferred)
  - original_category_source     (NULL · or one of: 'source_field_<name>'
                                  · 'inferred_backfill' if explicitly
                                  marked · NEVER fabricated)

v1 column `contact_category` is NEVER touched by --scheme=v2 runs ·
preserved for backward compatibility with existing readers.

CLI:
  python classify_master.py                 # default · v1
  python classify_master.py --scheme=v1     # explicit · v1
  python classify_master.py --scheme=v2     # Phase B canonical taxonomy

Re-runnable: idempotent · updates in place if columns already exist.
"""
from __future__ import annotations

import argparse
import re
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

from _phase_b_repair_freeze import abort_if_frozen

abort_if_frozen("classify_master.py")

sys.stdout.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parent.parent.parent
MASTER_PATH = ROOT / "CONTACTOS DATASITE" / "master" / "metcub-contacts-master.xlsx"


# ─── v1 detection (legacy · UNCHANGED) ────────────────────────────────

TECH_DOMAINS_V1 = {
    "rapidapi.com", "mapbox.com", "supabase.com", "resend.com", "vercel.com",
    "posthog.com", "sentry.io", "mailchimp.com", "namecheap.com", "datasite.com",
    "salesforce.com", "zapier.com", "anthropic.com", "openai.com", "github.com",
    "wordpress.com", "hubspot.com", "intercom.com", "stripe.com", "figma.com",
    "claude.ai",
}

LENDER_DOMAINS_V1 = {
    "bbva.com", "bbva.es", "bancsabadell.com", "caixabank.com", "caixabank.es",
    "bankinter.com", "santander.com", "gruposantander.es", "ing.com", "novobanco.es",
    "ibercaja.es", "sareb.es", "libertymutual.com", "sabadell.com", "bancamarch.es",
    "bancomarch.es",
}

BROKER_DOMAINS_V1 = {
    "cbre.com", "cbrehotels.com", "jll.com", "eu.jll.com", "savills.com",
    "savills-aguirrenewman.es", "colliers.com", "email.colliers.com", "christie.com",
    "labordemarcet.com", "engelvoelkers.com", "berkadia.com", "e.berkadia.com",
    "forcadell.com", "mapesainversiones.com", "dparealestate.com", "hlb.com.pt",
    "iberolat.com", "tcapital.es", "pierrecelestin-group.com", "qualyx.es",
    "arnold.investments", "cushwake.com", "matamalayasociados.es", "foodbox.es",
    "arcofood.com", "rcm1.com", "datasite.com",
}

DEVELOPER_PATTERN_V1 = re.compile(
    r"\b(promotor|constructor|construccion|construcción|aedas|metrovacesa|neinor|"
    r"grupo lar|colonial|monthisa|inmoglaciar|fernandez molina|strabag|drees|"
    r"gettys|ubm|reig|exclusive|ardid|ard-id)\b", re.IGNORECASE)

PRINCIPAL_PATTERN_V1 = re.compile(
    r"\b(hotel|hoteles|resort|melia|nh|hyatt|marriott|hilton|accor|wyndham|"
    r"iberostar|hotusa|citizenm|catalonia|barcelo|riu|grupo posadas|kempinski|"
    r"mandarin|virgin|h10|hospes|alda|ihg|fattal|covivio|invesco)\b", re.IGNORECASE)

PROVEEDOR_PATTERN_V1 = re.compile(
    r"\b(foodbox|restalia|mentidero|grupo azotea|brunchit|grupo escondite|"
    r"streetsense|gestoria|abogados|legal|notario|arquitecto|architects|"
    r"architecture|interiorismo|dipsa|workmanager|mazars|deloitte)\b", re.IGNORECASE)

LENDER_PATTERN_V1 = re.compile(
    r"\b(bank|banc|banco|caixa|sabadell|santander|liberty|sareb|march)\b", re.IGNORECASE)

BROKER_PATTERN_V1 = re.compile(
    r"\b(cbre|jll|colliers|savills|christie|engel|laborde|berkadia|forcadell|"
    r"cushman|wakefield)\b", re.IGNORECASE)


def extract_domain(email: str) -> str:
    if "@" not in email: return ""
    return email.split("@", 1)[1].lower() if email else ""


def classify_row_v1(row: dict[str, str | None]) -> str:
    """Original v1 classifier · preserved verbatim for backward compat."""
    email = (row.get("email") or "").strip()
    company = (row.get("company") or "").strip()
    role = (row.get("role") or "").strip()
    investor_type = (row.get("investor_type") or "").strip()
    title = (row.get("title") or "").strip()
    gmail_labels = (row.get("gmail_labels") or "").strip()
    domain = extract_domain(email)

    if domain in TECH_DOMAINS_V1 or "IA APLICACIONES" in gmail_labels.upper():
        return "IA aplicaciones"
    if "FINANCIADOR" in gmail_labels.upper():
        return "Lender"
    if domain in LENDER_DOMAINS_V1 or LENDER_PATTERN_V1.search(company):
        return "Lender"
    if "INTERMEDIARIO" in gmail_labels.upper():
        return "Broker"
    if domain in BROKER_DOMAINS_V1 or BROKER_PATTERN_V1.search(company):
        return "Broker"
    if re.search(r"\b(broker|agent|consultant|advisor|asesor|intermediario|consultor)\b", title, re.IGNORECASE):
        return "Broker"
    gmail_upper = gmail_labels.upper()
    if "PROMOTOR" in gmail_upper or "CONSTRUCTOR" in gmail_upper:
        return "Developer"
    if DEVELOPER_PATTERN_V1.search(company):
        return "Developer"
    gmail_check = any(tag in gmail_upper for tag in ["INVERSOR", "INVERSORES", "CADENA", "PROPIETARIO", "BRANDED", "RONDA", "LOI", "MoU"])
    if gmail_check:
        return "Principal"
    if investor_type and investor_type.lower() not in ("broker", "lender", "developer"):
        return "Principal"
    if PRINCIPAL_PATTERN_V1.search(company):
        return "Principal"
    if re.search(r"\b(principal|investor|owner|propietario)\b", role, re.IGNORECASE):
        return "Principal"
    if "F&B" in gmail_labels.upper() or "PROVEEDOR" in gmail_labels.upper():
        return "Proveedor"
    if PROVEEDOR_PATTERN_V1.search(company) and domain not in TECH_DOMAINS_V1:
        return "Proveedor"
    return "Uncategorized"


# ─── v2 detection (Phase B · NEW canonical taxonomy) ──────────────────

# IA Supply: TECH_DOMAINS expanded with hospitality-tech vendors +
# data-intelligence providers. Detection by domain OR company-name
# keyword to catch vendors whose domain isn't an obvious "tech" TLD.
IA_SUPPLY_DOMAINS = TECH_DOMAINS_V1 | {
    # Hospitality PMS / RMS / channel managers / CRM
    "cloudbeds.com", "mews.com", "mews.co", "opera.com", "oraclehospitality.com",
    "siteminder.com", "rategain.com", "ideas.com", "duetto.com", "atomize.io",
    "revinate.com", "profitroom.com", "yardi.com", "shiji-group.com", "shijigroup.com",
    "sas.com", "cvent.com", "infor.com", "ezee-frontdesk.com", "hotelogix.com",
    "innroad.com", "stayntouch.com", "guestcentric.com", "cendyn.com",
    # Data / market intelligence vendors
    "costar.com", "str.com", "smithtravelresearch.com", "trepp.com", "msci.com",
    "statista.com", "kalibri-labs.com", "kalibrilabs.com", "lighthouse.io",
    "hotelnewsnow.com", "skift.com", "phocuswright.com", "phocuswire.com",
    # AI / automation / general SaaS infrastructure
    "notion.so", "atlassian.com", "linear.app", "monday.com", "asana.com",
    "slack.com", "zoom.us", "microsoft.com", "amazonaws.com", "aws.amazon.com",
    "googlecloud.com", "azure.com", "databricks.com", "snowflake.com",
    # Phase B-tuning B-conservative additions (2026-05-15)
    "calendly.com", "send.calendly.com",
}

IA_SUPPLY_PATTERN = re.compile(
    r"\b("
    r"saas|pms|property management system|rms|revenue management system|"
    r"channel manager|booking engine|hospitality tech|hotel tech|"
    r"hospitality software|hotel software|"
    r"crm hospitality|"
    r"ai (vendor|platform|integration|workflow|workflows|automation)|"
    r"\bai\b|machine learning|\bml\b|"
    r"data intelligence|business intelligence|market intelligence|"
    r"data platform|analytics platform|data provider|"
    r"workflow automation|integration platform|api platform|"
    r"revenue tech|reservation system|distribution system|"
    r"hospitality data|underwriting (tech|platform|software)|"
    r"smart hotel|hotel digital|guest tech|"
    r"hospitality crm|guest crm|"
    # Phase B-tuning B-conservative additions (2026-05-15) · high-confidence only
    r"smart systems|smart hospitality|"
    r"data analytics|real estate analytics|"
    r"hospitality analytics|hotel analytics"
    r")\b", re.IGNORECASE)

IA_SUPPLY_COMPANY_HINTS = re.compile(
    r"\b("
    r"cloudbeds|mews|opera cloud|oracle hospitality|siteminder|rategain|"
    r"ideas revenue|duetto|atomize|revinate|profitroom|yardi|shiji|"
    r"costar|smith travel|str global|trepp|msci real assets|kalibri|"
    r"databricks|snowflake|infor|cvent|stayntouch|cendyn|guestcentric|"
    r"hotelogix|innroad|ezee"
    r")\b", re.IGNORECASE)

# Operator: Hotel Chain · Operator · Brand · White Label · Aparthotel ·
# Hostel · F&B · Resort · Branded Residence
OPERATOR_INVESTOR_TYPES = {
    "hotel chain", "operator", "brand", "white label", "white label operator",
    "f&b operator", "f&b", "aparthotel operator", "hostel operator",
    "resort operator", "branded residence operator",
}

OPERATOR_PATTERN = re.compile(
    r"\b("
    r"hotel chain|hotel group|hospitality group|"
    r"\boperator\b|brand|aparthotel|apart hotel|hostel|"
    r"resort operator|branded residence|residence operator|"
    r"f&b operator|food & beverage|restaurant group|"
    r"white label operator|white-label operator"
    r")\b", re.IGNORECASE)

# Specific operator brands (when company name matches, classify as Operator)
OPERATOR_BRANDS = re.compile(
    r"\b(marriott|hilton|hyatt|accor|wyndham|ihg|four seasons|kempinski|"
    r"mandarin oriental|peninsula|melia|melia hotels|nh hotel|nh hoteles|"
    r"hotusa|citizenm|catalonia hotels|barcelo|riu hotels|h10 hotels|"
    r"hospes|alda hoteles|fattal|invesco hotels|iberostar hotels|"
    r"radisson|best western|choice hotels|hyatt place|kimpton|"
    r"yotel|pestana|virgin hotels|edition hotels|aman resorts|"
    r"banyan tree|six senses|rosewood|st\. regis|ritz-carlton|"
    r"covivio|atlantica|grupo posadas|hyperion hotels)\b", re.IGNORECASE)

# Lender: bank + debt fund + alternative lender + financial advisor +
# structured finance
LENDER_DOMAINS_V2 = LENDER_DOMAINS_V1 | {
    "deutsche-bank.com", "db.com", "jpmorgan.com", "jpmorganchase.com",
    "morganstanley.com", "morgan-stanley.com", "goldmansachs.com", "gs.com",
    "wellsfargo.com", "barclays.com", "creditsuisse.com", "cs.com",
    "ubs.com", "hsbc.com", "rbc.com", "rbcwealthmanagement.com",
    "bankofamerica.com", "bofa.com", "natixis.com", "bnpparibas.com",
    "societegenerale.com", "credit-agricole.com",
}

LENDER_PATTERN_V2 = re.compile(
    r"\b(bank|banc|banco|caixa|sabadell|santander|liberty|sareb|march|"
    r"debt fund|alternative lender|structured finance|"
    r"mortgage|credit fund|private debt|"
    r"financial advisory|financial advisor)\b", re.IGNORECASE)

# Broker: + Investment Sales + Capital Markets + hospitality consultant
BROKER_DOMAINS_V2 = BROKER_DOMAINS_V1 | {
    "cushmanwakefield.com", "knightfrank.com", "newmark.com", "avison-young.com",
    "marcusmillichap.com", "hodgesward.com", "hwecapital.com",
}

BROKER_PATTERN_V2 = re.compile(
    r"\b(cbre|jll|colliers|savills|christie|engel|laborde|berkadia|forcadell|"
    r"cushman|wakefield|knight frank|newmark|avison young|marcus & millichap|"
    r"investment sales|capital markets|hospitality consultant|"
    r"hospitality advisory|hotel advisory|hotel brokerage)\b", re.IGNORECASE)

# Developer: + project manager + architecture/engineering
DEVELOPER_PATTERN_V2 = re.compile(
    r"\b(promotor|constructor|construccion|construcción|aedas|metrovacesa|neinor|"
    r"grupo lar|colonial|monthisa|inmoglaciar|fernandez molina|strabag|drees|"
    r"gettys|ubm|reig|exclusive|ardid|ard-id|"
    r"developer|promoter|project manager|"
    r"architecture|architects|engineering|engineering firm)\b", re.IGNORECASE)

# Hotel Supply: FF&E · furniture · technology (non-software) · hospitality
# equipment · procurement · interior design · materials · hospitality services
HOTEL_SUPPLY_PATTERN = re.compile(
    r"\b("
    r"ff&e|ffe|furniture|hospitality equipment|hotel equipment|"
    r"interior design|interiorism|interiorismo|hotel design|hospitality design|"
    r"procurement|hotel procurement|hospitality procurement|"
    r"materials supply|hotel materials|"
    r"hospitality services|hotel services|"
    r"linens|hotel linens|amenities|hotel amenities|"
    r"foodbox|restalia|mentidero|grupo azotea|brunchit|grupo escondite|"
    r"streetsense|gestoria|abogados|legal services|notario|arquitecto interiorista|"
    r"dipsa|workmanager|mazars|deloitte hospitality|"
    r"f&b services|catering|food service"
    r")\b", re.IGNORECASE)

# Principal v2: pure investors / owners / family offices / funds (no ops)
PRINCIPAL_INVESTOR_TYPES_V2 = {
    "investor", "fund", "family office", "reit/socimi", "reit", "socimi",
    "owner", "asset manager", "hotel owner group", "sovereign wealth",
    "institutional investor", "insurance", "pension fund",
}

PRINCIPAL_PATTERN_V2 = re.compile(
    r"\b(investments|investment group|capital partners|"
    r"family office|fund|funds|"
    r"reit|socimi|asset management|asset managers|equity partners|"
    r"private equity|hotel investments|hospitality investments|"
    r"hotel owner|hospitality owner|sovereign wealth)\b", re.IGNORECASE)


def classify_row_v2(row: dict[str, str | None]) -> str:
    """v2 classifier · 8 buckets · operator split + IA Supply expanded.

    Priority order chosen to minimise false positives:
      1. IA Supply  (high specificity · check first)
      2. Lender
      3. Broker
      4. Operator   (split from v1 Principal · explicit hotel-ops bucket)
      5. Developer
      6. Principal  (narrowed · pure investors only · no operators)
      7. Hotel Supply
      8. Uncategorized
    """
    email = (row.get("email") or "").strip()
    company = (row.get("company") or "").strip()
    role = (row.get("role") or "").strip()
    investor_type = (row.get("investor_type") or "").strip()
    title = (row.get("title") or "").strip()
    gmail_labels = (row.get("gmail_labels") or "").strip()
    domain = extract_domain(email)
    gmail_upper = gmail_labels.upper()
    inv_lower = investor_type.lower()
    company_lower = company.lower()

    # 1. IA Supply
    if domain in IA_SUPPLY_DOMAINS or "IA APLICACIONES" in gmail_upper:
        return "IA Supply"
    if IA_SUPPLY_COMPANY_HINTS.search(company):
        return "IA Supply"
    if IA_SUPPLY_PATTERN.search(company) or IA_SUPPLY_PATTERN.search(title):
        return "IA Supply"

    # 2. Lender
    if "FINANCIADOR" in gmail_upper:
        return "Lender"
    if inv_lower == "lender":
        return "Lender"
    if domain in LENDER_DOMAINS_V2 or LENDER_PATTERN_V2.search(company):
        return "Lender"

    # 3. Broker
    if "INTERMEDIARIO" in gmail_upper:
        return "Broker"
    if inv_lower == "broker":
        return "Broker"
    if domain in BROKER_DOMAINS_V2 or BROKER_PATTERN_V2.search(company):
        return "Broker"
    if re.search(r"\b(broker|investment sales|capital markets|"
                 r"hospitality consultant|hospitality advisor)\b",
                 title, re.IGNORECASE):
        return "Broker"

    # 4. Operator (split from v1 Principal)
    if "CADENA" in gmail_upper or "BRANDED" in gmail_upper:
        return "Operator"
    if inv_lower in OPERATOR_INVESTOR_TYPES:
        return "Operator"
    # Phase B-tuning D (2026-05-15): when investor_type is explicitly
    # 'Investor', do NOT poach via company-name pattern. Such rows are
    # investors in hospitality assets, not operators of them — Principal
    # owns this classification.
    if inv_lower != "investor":
        if OPERATOR_BRANDS.search(company):
            return "Operator"
        if OPERATOR_PATTERN.search(company):
            return "Operator"

    # 5. Developer
    if "PROMOTOR" in gmail_upper or "CONSTRUCTOR" in gmail_upper:
        return "Developer"
    if inv_lower == "developer":
        return "Developer"
    if DEVELOPER_PATTERN_V2.search(company):
        return "Developer"

    # 6. Principal (narrowed v2)
    if any(tag in gmail_upper for tag in ["INVERSOR", "INVERSORES", "PROPIETARIO", "RONDA", "LOI", "MoU"]):
        return "Principal"
    if inv_lower in PRINCIPAL_INVESTOR_TYPES_V2:
        return "Principal"
    if PRINCIPAL_PATTERN_V2.search(company):
        return "Principal"
    if re.search(r"\b(principal|investor|owner|propietario|asset manager)\b",
                 role, re.IGNORECASE):
        return "Principal"

    # 7. Hotel Supply
    if "F&B" in gmail_upper or "PROVEEDOR" in gmail_upper:
        return "Hotel Supply"
    if HOTEL_SUPPLY_PATTERN.search(company):
        return "Hotel Supply"
    # Phase B-tuning A (2026-05-15): default for investor_type
    # 'Service Provider' and 'Media' (operator decision: media is treated
    # as a service-provider category, not its own bucket). Reaches here
    # only AFTER IA Supply / Lender / Broker / Operator / Developer /
    # Principal had a chance to claim the row, so this default doesn't
    # over-classify obvious tech/finance/hospitality companies.
    if inv_lower in {"service provider", "media"}:
        return "Hotel Supply"

    return "Uncategorized"


# ─── pipeline ────────────────────────────────────────────────────────


def parse_args() -> argparse.Namespace:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--scheme", choices=["v1", "v2"], default="v1",
                    help="v1 = legacy (default · writes contact_category) · "
                         "v2 = Phase B canonical (writes contact_category_v2 + audit cols)")
    return ap.parse_args()


def main() -> int:
    args = parse_args()

    if not MASTER_PATH.exists():
        print(f"ERROR: Master file not found at {MASTER_PATH}")
        return 1

    print(f"→ Reading Master from {MASTER_PATH}")
    print(f"→ Scheme: {args.scheme}")

    wb_read = load_workbook(MASTER_PATH, read_only=True)
    ws_read = wb_read["Master"]
    header_row = [c.value for c in ws_read[1]]
    print(f"  · header cols: {len(header_row)}")
    print(f"  · data rows  : {ws_read.max_row - 1}")

    data_rows = []
    for raw in ws_read.iter_rows(min_row=2, values_only=True):
        d = {h: raw[i] for i, h in enumerate(header_row) if h and i < len(raw)}
        data_rows.append(d)
    wb_read.close()
    print(f"  · loaded {len(data_rows)} rows")

    # Decide which column(s) to write
    wb = load_workbook(MASTER_PATH)
    ws = wb["Master"]

    if args.scheme == "v1":
        target_col = _ensure_column(ws, header_row, "contact_category")
        print(f"\n→ v1 · writing to column {target_col} (contact_category)")
        classifications: defaultdict[str, int] = defaultdict(int)
        for row_idx, row_dict in enumerate(data_rows, start=2):
            cat = classify_row_v1(row_dict)
            ws.cell(row=row_idx, column=target_col).value = cat
            classifications[cat] += 1
        v1_buckets = ["Principal", "Broker", "Lender", "Developer",
                      "Proveedor", "IA aplicaciones", "Uncategorized"]
        _print_distribution("v1 distribution", classifications, v1_buckets)

    elif args.scheme == "v2":
        col_v2 = _ensure_column(ws, header_row, "contact_category_v2")
        col_raw = _ensure_column(ws, header_row, "original_category_raw")
        col_src = _ensure_column(ws, header_row, "original_category_source")
        print(f"\n→ v2 · writing to columns:")
        print(f"    contact_category_v2     : col {col_v2}")
        print(f"    original_category_raw   : col {col_raw} (NULL for all rows · provenance integrity)")
        print(f"    original_category_source: col {col_src} (NULL for all rows · no inference)")
        print(f"\n  v1 column 'contact_category' is NOT modified.")

        classifications = defaultdict(int)
        # Track: for each row, whether original_category_raw already had a value
        preserved_raw = 0
        preserved_src = 0
        for row_idx, row_dict in enumerate(data_rows, start=2):
            cat = classify_row_v2(row_dict)
            ws.cell(row=row_idx, column=col_v2).value = cat
            classifications[cat] += 1

            # Preserve any pre-existing original_category_raw value
            existing_raw = row_dict.get("original_category_raw")
            if existing_raw not in (None, ""):
                preserved_raw += 1
                # don't overwrite
            else:
                ws.cell(row=row_idx, column=col_raw).value = None

            existing_src = row_dict.get("original_category_source")
            if existing_src not in (None, ""):
                preserved_src += 1
            else:
                ws.cell(row=row_idx, column=col_src).value = None

        v2_buckets = ["Principal", "Broker", "Lender", "Operator",
                      "Developer", "Hotel Supply", "IA Supply", "Uncategorized"]
        _print_distribution("v2 distribution", classifications, v2_buckets)
        print(f"\n  · original_category_raw    : {preserved_raw} preserved · {len(data_rows) - preserved_raw} kept NULL")
        print(f"  · original_category_source : {preserved_src} preserved · {len(data_rows) - preserved_src} kept NULL")

    else:
        print(f"!! Unknown scheme: {args.scheme}")
        return 2

    print(f"\n→ Saving Master")
    wb.save(MASTER_PATH)
    wb.close()
    print(f"\n✓ classify_master.py {args.scheme} complete")
    return 0


def _ensure_column(ws, header_row, col_name: str) -> int:
    """Return 1-indexed column number for col_name. Append at end if missing."""
    if col_name in header_row:
        return header_row.index(col_name) + 1
    new_col = len(header_row) + 1
    ws.cell(row=1, column=new_col).value = col_name
    header_row.append(col_name)
    print(f"  · added new header col {new_col}: {col_name!r}")
    return new_col


def _print_distribution(title: str, counts: dict, ordered_buckets: list[str]) -> None:
    total = sum(counts.values())
    print("\n" + "=" * 70)
    print(title)
    print("=" * 70)
    for b in ordered_buckets:
        n = counts.get(b, 0)
        pct = (n / total * 100) if total else 0.0
        print(f"  {b:20s}: {n:5d} ({pct:5.1f}%)")
    print(f"  {'TOTAL':20s}: {total:5d} (100.0%)")
    print("=" * 70)


if __name__ == "__main__":
    sys.exit(main())
