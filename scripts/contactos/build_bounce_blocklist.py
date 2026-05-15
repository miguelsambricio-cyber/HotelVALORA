#!/usr/bin/env python3
"""
Phase 2.B.3 · Build Gmail bounce + dead-domain blocklists.

Inputs (read-only)
------------------
  CONTACTOS DATASITE/master/metcub-contacts-master.xlsx
      - Master sheet: email_validity = "invalid" rows
      - INVALID_ARCHIVE sheet (if present): every archived contact
  CONTACTOS DATASITE/incoming/gmail-signals/*.jsonl
  CONTACTOS DATASITE/old/gmail-signals/*.jsonl
      - any signal record with bounce_count >= 1 AND inbound_count = 0
        (i.e. it bounced and nobody at that address ever replied · DEAD)

Outputs (idempotent · overwritten on each run)
----------------------------------------------
  CONTACTOS DATASITE/master/blocklists/gmail-bounce-blocklist.txt
      one normalised email per line · "# comment" lines allowed
  CONTACTOS DATASITE/master/blocklists/dead-domains-blocklist.txt
      one domain per line · ≥50% bounce rate AND ≥2 contacts seen

Both files are consumed by:
  - scripts/contactos/extract_gmail_signals.py
  - scripts/contactos/harvest_untagged.py
…via the shared loader in scripts/contactos/_blocklist.py.

Re-runnable: blocklists are regenerated from scratch each run.
"""
from __future__ import annotations

import json
import sys
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except (AttributeError, ValueError):
    pass

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl is required · pip install openpyxl", file=sys.stderr)
    sys.exit(2)

REPO = Path(__file__).resolve().parent.parent.parent
ROOT = REPO / "CONTACTOS DATASITE"
MASTER_PATH = ROOT / "master" / "metcub-contacts-master.xlsx"
SIGNAL_DIRS = [ROOT / "incoming" / "gmail-signals", ROOT / "old" / "gmail-signals"]
BLOCKLIST_DIR = ROOT / "master" / "blocklists"
BLOCKLIST_DIR.mkdir(parents=True, exist_ok=True)

EMAIL_BLOCKLIST = BLOCKLIST_DIR / "gmail-bounce-blocklist.txt"
DOMAIN_BLOCKLIST = BLOCKLIST_DIR / "dead-domains-blocklist.txt"

# Personal domains can't be blocklisted at the domain level · only specific addresses
PERSONAL_DOMAINS = {
    "gmail.com", "googlemail.com", "hotmail.com", "outlook.com", "outlook.es",
    "live.com", "yahoo.com", "yahoo.es", "icloud.com", "me.com", "mac.com",
    "protonmail.com", "proton.me", "yandex.com", "mail.ru", "aol.com",
    "msn.com", "ymail.com", "telefonica.net", "terra.es",
}

DEAD_DOMAIN_BOUNCE_RATIO = 0.50
DEAD_DOMAIN_MIN_CONTACTS = 2


def collect_master_invalids(path: Path) -> tuple[set[str], set[str]]:
    """Return (invalid_emails_active, invalid_emails_archived)."""
    if not path.exists():
        return set(), set()
    wb = load_workbook(path, read_only=True)
    active: set[str] = set()
    archived: set[str] = set()

    if "Master" in wb.sheetnames:
        ws = wb["Master"]
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        idx = {h: i for i, h in enumerate(header) if h}
        if "email_validity" in idx and "email" in idx:
            for r in ws.iter_rows(min_row=2, values_only=True):
                if not r:
                    continue
                v = r[idx["email_validity"]]
                if isinstance(v, str) and v.strip().lower() == "invalid":
                    em = (r[idx["email"]] or "")
                    if isinstance(em, str) and "@" in em:
                        active.add(em.strip().lower())

    if "INVALID_ARCHIVE" in wb.sheetnames:
        ws = wb["INVALID_ARCHIVE"]
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        idx = {h: i for i, h in enumerate(header) if h}
        if "email" in idx:
            for r in ws.iter_rows(min_row=2, values_only=True):
                if not r:
                    continue
                em = r[idx["email"]]
                if isinstance(em, str) and "@" in em:
                    archived.add(em.strip().lower())

    wb.close()
    return active, archived


def collect_gmail_bounces() -> tuple[set[str], dict[str, dict[str, int]]]:
    """
    Walk every gmail-signals JSONL and return:
      - dead_emails: address with bounce_count >= 2
                     OR (bounce_count >= 1 AND inbound_count == 0)
      - per_domain stats: {domain: {"total": n, "bounced": n}}
    """
    dead: set[str] = set()
    per_domain: dict[str, dict[str, int]] = defaultdict(lambda: {"total": 0, "bounced": 0})
    seen_emails_per_domain: dict[str, set[str]] = defaultdict(set)
    bounced_emails_per_domain: dict[str, set[str]] = defaultdict(set)

    for d in SIGNAL_DIRS:
        if not d.exists():
            continue
        for f in sorted(d.glob("*.jsonl")):
            try:
                with f.open("r", encoding="utf-8") as fh:
                    for line in fh:
                        line = line.strip()
                        if not line:
                            continue
                        try:
                            rec = json.loads(line)
                        except json.JSONDecodeError:
                            continue
                        em = (rec.get("email") or "").strip().lower()
                        if not em or "@" not in em:
                            continue
                        bc = int(rec.get("bounce_count") or 0)
                        ic = int(rec.get("inbound_count") or 0)
                        dom = em.split("@", 1)[1]
                        seen_emails_per_domain[dom].add(em)
                        if bc >= 2 or (bc >= 1 and ic == 0):
                            dead.add(em)
                            bounced_emails_per_domain[dom].add(em)
            except OSError as err:
                print(f"  ⚠ could not read {f.name} · {err}", file=sys.stderr)

    for dom, emails in seen_emails_per_domain.items():
        per_domain[dom]["total"] = len(emails)
        per_domain[dom]["bounced"] = len(bounced_emails_per_domain.get(dom, set()))

    return dead, per_domain


def main() -> int:
    print(f"→ Bounce blocklist generator · {datetime.now(timezone.utc).isoformat(timespec='seconds')}")

    master_active, master_archived = collect_master_invalids(MASTER_PATH)
    print(f"  · Master invalid (active sheet): {len(master_active)}")
    print(f"  · Master invalid (archived sheet): {len(master_archived)}")

    gmail_dead, domain_stats = collect_gmail_bounces()
    print(f"  · Gmail-signal dead addresses: {len(gmail_dead)}")
    print(f"  · Domains with bounce data: {len(domain_stats)}")

    all_blocked = master_active | master_archived | gmail_dead
    print(f"  · Union (final email blocklist): {len(all_blocked)}")

    # ── Email blocklist file ────────────────────────────────────────────────
    sorted_emails = sorted(all_blocked)
    with EMAIL_BLOCKLIST.open("w", encoding="utf-8") as f:
        f.write(f"# gmail-bounce-blocklist · regenerated {datetime.now(timezone.utc).isoformat(timespec='seconds')}\n")
        f.write(f"# sources: Master[email_validity=invalid] · INVALID_ARCHIVE · gmail-signals (bc>=2 or bc>=1+inbound=0)\n")
        f.write(f"# total: {len(sorted_emails)}\n")
        f.write(f"# DO NOT EDIT BY HAND · regenerate via scripts/contactos/build_bounce_blocklist.py\n\n")
        for em in sorted_emails:
            f.write(em + "\n")
    print(f"\n✓ {EMAIL_BLOCKLIST.relative_to(REPO)} · {len(sorted_emails)} address(es)")

    # ── Dead-domain blocklist file ──────────────────────────────────────────
    dead_domains: list[tuple[str, int, int, float]] = []
    for dom, stats in domain_stats.items():
        if dom in PERSONAL_DOMAINS:
            continue
        total = stats["total"]
        bounced = stats["bounced"]
        if total < DEAD_DOMAIN_MIN_CONTACTS:
            continue
        ratio = bounced / total if total else 0.0
        if ratio >= DEAD_DOMAIN_BOUNCE_RATIO:
            dead_domains.append((dom, bounced, total, ratio))
    dead_domains.sort(key=lambda x: (-x[3], -x[1], x[0]))

    with DOMAIN_BLOCKLIST.open("w", encoding="utf-8") as f:
        f.write(f"# dead-domains-blocklist · regenerated {datetime.now(timezone.utc).isoformat(timespec='seconds')}\n")
        f.write(f"# threshold: bounce_ratio >= {int(DEAD_DOMAIN_BOUNCE_RATIO * 100)}% AND total_contacts >= {DEAD_DOMAIN_MIN_CONTACTS}\n")
        f.write(f"# format: domain  # bounced/total  ratio%\n")
        f.write(f"# total: {len(dead_domains)}\n")
        f.write(f"# DO NOT EDIT BY HAND · regenerate via scripts/contactos/build_bounce_blocklist.py\n\n")
        for dom, bounced, total, ratio in dead_domains:
            f.write(f"{dom}  # {bounced}/{total}  {ratio:.0%}\n")
    print(f"✓ {DOMAIN_BLOCKLIST.relative_to(REPO)} · {len(dead_domains)} domain(s)")

    print(f"\nNext: re-run extract_gmail_signals.py and harvest_untagged.py · they will honour these lists.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
